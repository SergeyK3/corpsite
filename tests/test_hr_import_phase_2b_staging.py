"""Service tests for ADR-038 Phase 2B (parse → staging)."""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from sqlalchemy import text

from app.db.engine import engine
from app.db.models.hr_import import (
    BATCH_STATUS_IN_REVIEW,
    BATCH_STATUS_UPLOADED,
    CLASSIFICATION_CATEGORY_ROW,
    CLASSIFICATION_DUPLICATE_IIN,
    CLASSIFICATION_INVALID_IIN,
    CLASSIFICATION_NORMAL,
    MATCH_STATUS_AUTO,
    MATCH_STATUS_NO_MATCH,
    MATCH_STATUS_NOT_PROCESSED,
    MATCH_STATUS_REVIEW,
    SOURCE_TYPE_HR_CONTROL_LIST,
)
from app.services.hr_import_service import (
    BATCH_STATUS_CREATED,
    BATCH_STATUS_REVIEW_READY,
    classify_row,
    create_batch,
    import_control_list,
    summarize_batch,
)
from scripts.import_hr_control_list import ParsedRow, build_audit, parse_workbook
from tests.conftest import table_exists
from tests.hr_import_fixtures import cleanup_import_batch, write_control_list_workbook
from tests.test_import_hr_control_list import _build_sample_workbook

CORE_SUMMARY_KEYS = (
    "total_rows",
    "valid_iin",
    "invalid_iin",
    "duplicate_iin_groups",
    "duplicate_iin_rows",
    "missing_full_name",
    "missing_department",
    "with_training",
    "with_certification",
)

_EMPLOYEE_ROSTER_ROW_FILTER = """
    normalized_payload->'metadata'->>'is_employee_roster' = 'true'
    OR normalized_payload->'metadata'->>'classification' = :normal_classification
"""


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_2a_available() -> bool:
    with engine.begin() as conn:
        return all(
            table_exists(conn, table)
            for table in (
                "hr_import_batches",
                "hr_import_rows",
            )
        )


def _phase_2b_available() -> bool:
    with engine.begin() as conn:
        row = conn.execute(
            text(
                """
                SELECT 1
                FROM pg_constraint
                WHERE conname = 'chk_hr_import_rows_match_status'
                  AND pg_get_constraintdef(oid) LIKE '%NOT_PROCESSED%'
                LIMIT 1
                """
            )
        ).first()
        return row is not None


def _normalized_records_available() -> bool:
    with engine.begin() as conn:
        return table_exists(conn, "hr_import_normalized_records")


def _require_phase_2b() -> None:
    if not _phase_2a_available():
        pytest.skip("Phase 2A tables missing — run: alembic upgrade head")
    if not _phase_2b_available():
        pytest.skip("Phase 2B migration missing — run: alembic upgrade head (d3e4f5a6b7c8)")


def _delete_batch(conn, batch_id: int) -> None:
    cleanup_import_batch(conn, batch_id)


def _cleanup_import_batch(batch_id: int | None) -> None:
    if batch_id is None:
        return
    with engine.begin() as conn:
        exists = conn.execute(
            text(
                """
                SELECT 1 FROM public.hr_import_batches
                WHERE batch_id = :batch_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).first()
        if exists:
            _delete_batch(conn, batch_id)


def _assert_core_summary(summary: dict, expected: dict) -> None:
    for key in CORE_SUMMARY_KEYS:
        assert summary[key] == expected[key], key


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_create_import_batch(seed):
    _require_phase_2b()

    suffix = uuid4().hex[:8]
    batch_id: int | None = None
    try:
        with engine.begin() as conn:
            batch_id = create_batch(
                conn,
                source_type=SOURCE_TYPE_HR_CONTROL_LIST,
                file_name=f"pytest_{suffix}.xlsx",
                imported_by=int(seed["initiator_user_id"]),
            )
            batch = conn.execute(
                text(
                    """
                    SELECT source_type, file_name, imported_by, status,
                           total_rows, valid_rows, error_rows
                    FROM public.hr_import_batches
                    WHERE batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().one()

        assert batch_id > 0
        assert batch["source_type"] == SOURCE_TYPE_HR_CONTROL_LIST
        assert batch["file_name"] == f"pytest_{suffix}.xlsx"
        assert int(batch["imported_by"]) == int(seed["initiator_user_id"])
        assert batch["status"] == BATCH_STATUS_CREATED == BATCH_STATUS_UPLOADED
        assert batch["total_rows"] == 0
        assert batch["valid_rows"] == 0
        assert batch["error_rows"] == 0
    finally:
        _cleanup_import_batch(batch_id)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_import_control_list_stages_rows_and_resolves_employee_binding(seed, tmp_path: Path):
    """import_control_list persists rows then runs employee binding during normalization."""
    _require_phase_2b()

    source = write_control_list_workbook(tmp_path, yymm="2606")
    _build_sample_workbook(source)

    batch_id: int | None = None
    try:
        with engine.begin() as conn:
            batch_id, summary, warnings = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            batch = conn.execute(
                text(
                    """
                    SELECT status, total_rows, valid_rows, error_rows
                    FROM public.hr_import_batches
                    WHERE batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().one()
            row_count = conn.execute(
                text("SELECT COUNT(*) FROM public.hr_import_rows WHERE batch_id = :batch_id"),
                {"batch_id": batch_id},
            ).scalar_one()
            employee_row = conn.execute(
                text(
                    f"""
                    SELECT source_sheet, source_row_number, raw_payload, normalized_payload,
                           match_status, review_status, error_codes
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                      AND ({_EMPLOYEE_ROSTER_ROW_FILTER})
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id, "normal_classification": CLASSIFICATION_NORMAL},
            ).mappings().one()

        assert batch_id > 0
        assert batch["status"] == BATCH_STATUS_REVIEW_READY == BATCH_STATUS_IN_REVIEW
        assert row_count == summary["total_rows"]
        assert batch["total_rows"] == summary["total_rows"]
        assert employee_row["review_status"] == "PENDING"
        assert "metadata" in employee_row["normalized_payload"]
        assert "classification" in employee_row["normalized_payload"]["metadata"]
        assert isinstance(employee_row["raw_payload"], dict)
        assert any(w.startswith("skip_unknown_sheet:") for w in warnings)

        metadata = employee_row["normalized_payload"]["metadata"]
        if _normalized_records_available():
            assert employee_row["match_status"] in {
                MATCH_STATUS_NO_MATCH,
                MATCH_STATUS_AUTO,
                MATCH_STATUS_REVIEW,
            }
            assert metadata.get("employee_binding_status") in {"unbound", "bound", "conflict"}
        else:
            assert employee_row["match_status"] == MATCH_STATUS_NOT_PROCESSED
    finally:
        _cleanup_import_batch(batch_id)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_batch_summary(seed, tmp_path: Path):
    _require_phase_2b()

    source = write_control_list_workbook(tmp_path, yymm="2606")
    _build_sample_workbook(source)

    rows, _ = parse_workbook(source)
    expected = build_audit(rows)
    expected_summary = {
        "total_rows": expected["total_rows"],
        "valid_iin": expected["valid_iin"],
        "invalid_iin": expected["invalid_iin"],
        "duplicate_iin_groups": expected["duplicate_iin"],
        "duplicate_iin_rows": expected["duplicate_iin_rows"],
        "missing_full_name": expected["missing_full_name"],
        "missing_department": expected["missing_department"],
        "with_training": expected["with_training"],
        "with_certification": expected["with_certification"],
    }

    batch_id: int | None = None
    try:
        with engine.begin() as conn:
            batch_id, summary, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            persisted = summarize_batch(conn, batch_id)

        _assert_core_summary(summary, expected_summary)
        _assert_core_summary(persisted, expected_summary)

        if "monthly_diff" in summary:
            diff = summary["monthly_diff"]
            assert diff["batch_id"] == batch_id
            assert isinstance(diff.get("summary"), dict)
            assert "NEW" in diff["summary"]
    finally:
        _cleanup_import_batch(batch_id)


def test_duplicate_iin_classification():
    row_a = ParsedRow(
        data={"full_name": "A", "department": "D", "source_sheet": "s", "source_row_number": "1", "iin": "900101300123"},
        sheet_type="doctors",
        iin_valid=True,
        iin_digits="900101300123",
        errors=[],
    )
    row_b = ParsedRow(
        data={"full_name": "B", "department": "D", "source_sheet": "s", "source_row_number": "2", "iin": "900101300123"},
        sheet_type="doctors",
        iin_valid=True,
        iin_digits="900101300123",
        errors=[],
    )
    duplicate_iins = {"900101300123"}

    assert classify_row(row_a, duplicate_iins) == CLASSIFICATION_DUPLICATE_IIN
    assert classify_row(row_b, duplicate_iins) == CLASSIFICATION_DUPLICATE_IIN


def test_invalid_iin_classification():
    row = ParsedRow(
        data={"full_name": "Test", "department": "D", "source_sheet": "s", "source_row_number": "3", "iin": "12345"},
        sheet_type="doctors",
        iin_valid=False,
        iin_digits="12345",
        errors=["invalid_iin_length:5"],
    )

    assert classify_row(row, set()) == CLASSIFICATION_INVALID_IIN


def test_category_row_classification_takes_priority_over_invalid_iin():
    row = ParsedRow(
        data={
            "full_name": "Невалидный ИИН",
            "department": "ОТДЕЛ",
            "source_sheet": "s",
            "source_row_number": "1",
            "iin": "12345",
        },
        sheet_type="doctors",
        row_type="CATEGORY_ROW",
        iin_valid=False,
        iin_digits="12345",
        errors=["invalid_iin_length:5"],
    )

    assert classify_row(row, set()) == CLASSIFICATION_CATEGORY_ROW


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_duplicate_iin_classification_persisted(seed, tmp_path: Path):
    _require_phase_2b()

    source = tmp_path / "контрольный2607.xlsx"
    _build_sample_workbook(source)

    batch_id: int | None = None
    try:
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            dup_rows = conn.execute(
                text(
                    """
                    SELECT normalized_payload->'metadata'->>'classification' AS classification
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                      AND normalized_payload->>'iin' = '900101300123'
                    """
                ),
                {"batch_id": batch_id},
            ).scalars().all()

        assert len(dup_rows) >= 2
        assert all(item == CLASSIFICATION_DUPLICATE_IIN for item in dup_rows)
    finally:
        _cleanup_import_batch(batch_id)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_invalid_iin_employee_row_classification_and_errors_persisted(seed, tmp_path: Path):
    _require_phase_2b()

    from datetime import datetime
    from openpyxl import Workbook

    path = tmp_path / "контрольный2608.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    for _ in range(6):
        ws.append([""] * 20)
    ws.append(
        [
            "№",
            "",
            "Фамилия, имя, отчество",
            "Год рождения",
            "ИИН",
            "пол",
        ]
        + [""] * 14
    )
    ws.append(
        [
            1,
            "Отдел терапии",
            "Иванов Иван Иванович",
            datetime(1990, 1, 1),
            "12345",
            "муж",
        ]
        + [""] * 14
    )
    wb.save(path)

    batch_id: int | None = None
    try:
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=path,
                imported_by=int(seed["initiator_user_id"]),
            )
            row = conn.execute(
                text(
                    """
                    SELECT normalized_payload->'metadata'->>'classification' AS classification,
                           error_codes
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                      AND normalized_payload->>'full_name' = 'Иванов Иван Иванович'
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().one()

        assert row["classification"] == CLASSIFICATION_INVALID_IIN
        assert row["error_codes"] is not None
        assert any(str(code).startswith("invalid_iin") for code in row["error_codes"])
    finally:
        _cleanup_import_batch(batch_id)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_category_row_with_invalid_iin_persisted_as_category_row(seed, tmp_path: Path):
    _require_phase_2b()

    from datetime import datetime
    from openpyxl import Workbook

    path = tmp_path / "контрольный2609.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    for _ in range(6):
        ws.append([""] * 20)
    ws.append(
        [
            "№",
            "",
            "Фамилия, имя, отчество",
            "Год рождения",
            "ИИН",
            "пол",
        ]
        + [""] * 14
    )
    ws.append(
        [
            1,
            "ОТДЕЛ",
            "Невалидный ИИН",
            datetime(1990, 1, 1),
            "12345",
            "муж",
        ]
        + [""] * 14
    )
    wb.save(path)

    batch_id: int | None = None
    try:
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=path,
                imported_by=int(seed["initiator_user_id"]),
            )
            classification = conn.execute(
                text(
                    """
                    SELECT normalized_payload->'metadata'->>'classification' AS classification
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).scalar_one()

        assert classification == CLASSIFICATION_CATEGORY_ROW
    finally:
        _cleanup_import_batch(batch_id)
