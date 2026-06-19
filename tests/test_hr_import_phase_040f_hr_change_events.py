"""Tests for ADR-040 Phase F — materialized HR change events between snapshots."""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_canonical_snapshot_service import (
    build_canonical_snapshot_from_batch,
    canonical_snapshot_available,
)
from app.services.hr_import_service import import_control_list
from app.services.hr_snapshot_comparison_service import (
    EVENT_TYPE_DEPARTMENT_CHANGED,
    EVENT_TYPE_NEW,
    EVENT_TYPE_POSITION_CHANGED,
    EVENT_TYPE_REMOVED,
    compare_snapshots,
    hr_change_events_available,
    list_hr_change_events,
    materialize_snapshot_change_events,
)
from tests.test_hr_import_phase_040a_canonical_snapshot import (
    _cleanup_promotion_batch,
    _cyrillic_full_name,
    _ensure_department_mapping,
    _ensure_roster_employee_metadata,
    _prepare_roster_row,
    _require_phase_040a,
    _test_iin,
)
from tests.test_import_hr_control_list import _build_doctors_sheet


def _require_phase_040f() -> None:
    _require_phase_040a()
    with engine.begin() as conn:
        if not hr_change_events_available(conn):
            pytest.skip("ADR-040 Phase F migration not applied — run alembic upgrade head")


def _unique_iin(seed: str | None = None) -> str:
    token = seed or uuid4().hex
    return _test_iin(token)


def _build_workbook(
    path: Path,
    *,
    full_name: str,
    iin: str,
    position: str = "",
    department: str = "",
) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=iin)
    if position:
        ws.cell(row=8, column=7, value=position)
    if department:
        ws.cell(row=8, column=6, value=department)
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)
    wb.save(path)
    wb.close()


def _import_prepared_batch(
    seed,
    tmp_path: Path,
    *,
    full_name: str,
    iin: str,
    department: str,
    org_unit_id: int,
    position: str = "",
) -> int:
    source = tmp_path / f"phase040f_{uuid4().hex[:8]}.xlsx"
    _build_workbook(source, full_name=full_name, iin=iin, position=position, department=department)
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row_id = conn.execute(
            text(
                """
                SELECT row_id
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).scalar_one()
        _prepare_roster_row(
            conn,
            batch_id=batch_id,
            row_id=int(row_id),
            full_name=full_name,
            iin=iin,
            department=department,
            org_unit_id=org_unit_id,
        )
        if position:
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET normalized_payload = jsonb_set(
                        normalized_payload,
                        '{position_raw}',
                        to_jsonb(CAST(:position AS text)),
                        true
                    )
                    WHERE row_id = :row_id
                    """
                ),
                {"row_id": int(row_id), "position": position},
            )
            _ensure_roster_employee_metadata(conn, batch_id=batch_id, row_id=int(row_id))
    return int(batch_id)


def _set_recoding_org_unit_name(conn, *, department: str, org_unit_name: str) -> None:
    conn.execute(
        text(
            """
            UPDATE public.department_recoding
            SET org_unit_name = :org_unit_name
            WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:department))
            """
        ),
        {"department": department, "org_unit_name": org_unit_name},
    )


def _create_snapshot_from_batch(seed, batch_id: int) -> dict:
    with engine.begin() as conn:
        return build_canonical_snapshot_from_batch(
            conn,
            batch_id,
            promoted_by=int(seed["initiator_user_id"]),
        )


def _get_snapshots_for_batches(batch_id_1: int, batch_id_2: int) -> tuple[int, int]:
    with engine.begin() as conn:
        prior = conn.execute(
            text(
                """
                SELECT snapshot_id
                FROM public.hr_canonical_snapshots
                WHERE source_batch_id = :batch_id
                """
            ),
            {"batch_id": batch_id_1},
        ).scalar_one()
        new = conn.execute(
            text(
                """
                SELECT snapshot_id
                FROM public.hr_canonical_snapshots
                WHERE source_batch_id = :batch_id
                """
            ),
            {"batch_id": batch_id_2},
        ).scalar_one()
    return int(prior), int(new)


def _events_for_pair(prior_snapshot_id: int, new_snapshot_id: int) -> list[dict]:
    with engine.begin() as conn:
        rows = conn.execute(
            text(
                """
                SELECT event_type, match_key, field_name, old_value, new_value,
                       full_name, department, org_unit_id
                FROM public.hr_change_events
                WHERE prior_snapshot_id = :prior_snapshot_id
                  AND new_snapshot_id = :new_snapshot_id
                ORDER BY change_event_id
                """
            ),
            {
                "prior_snapshot_id": prior_snapshot_id,
                "new_snapshot_id": new_snapshot_id,
            },
        ).mappings().all()
    return [dict(row) for row in rows]


def test_new_employee_event(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    org_unit_id = int(seed["unit_id"])
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Base{suffix}"),
        iin=_unique_iin(f"1{suffix}"),
        department=department,
        org_unit_id=org_unit_id,
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"New{suffix}"),
        iin=_unique_iin(f"2{suffix}"),
        department=department,
        org_unit_id=org_unit_id,
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        _create_snapshot_from_batch(seed, batch_id_2)
        prior_id, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
        events = _events_for_pair(prior_id, new_id)
        event_types = {event["event_type"] for event in events}
        assert EVENT_TYPE_NEW in event_types
        assert EVENT_TYPE_REMOVED in event_types
        new_events = [event for event in events if event["event_type"] == EVENT_TYPE_NEW]
        assert len(new_events) >= 1
        assert new_events[0]["full_name"] is not None
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_removed_employee_event(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    org_unit_id = int(seed["unit_id"])
    full_name = _cyrillic_full_name(f"Rem{suffix}")
    iin = _unique_iin(f"1{suffix}")
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Other{suffix}"),
        iin=_unique_iin(f"2{suffix}"),
        department=department,
        org_unit_id=org_unit_id,
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        _create_snapshot_from_batch(seed, batch_id_2)
        prior_id, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
        events = _events_for_pair(prior_id, new_id)
        removed = [event for event in events if event["event_type"] == EVENT_TYPE_REMOVED]
        assert len(removed) >= 1
        assert removed[0]["full_name"] == full_name
        assert removed[0]["old_value"] is None
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_position_changed_event(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    org_unit_id = int(seed["unit_id"])
    full_name = _cyrillic_full_name(f"Pos{suffix}")
    iin = _unique_iin(suffix)
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
        position="Медицинский техник",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
        position="Инженер по медицинскому оборудованию",
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        _create_snapshot_from_batch(seed, batch_id_2)
        prior_id, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
        events = _events_for_pair(prior_id, new_id)
        position_events = [event for event in events if event["event_type"] == EVENT_TYPE_POSITION_CHANGED]
        assert len(position_events) == 1
        assert position_events[0]["field_name"] == "position_raw"
        assert position_events[0]["old_value"] == "Медицинский техник"
        assert position_events[0]["new_value"] == "Инженер по медицинскому оборудованию"
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_department_changed_event(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    dept_a = f"Pytest Dept A {suffix}"
    dept_b = f"Pytest Dept B {suffix}"
    org_unit_a = int(seed["unit_id"])
    org_unit_b = org_unit_a
    full_name = _cyrillic_full_name(f"Dep{suffix}")
    iin = _unique_iin(suffix)
    with engine.begin() as conn:
        _ensure_department_mapping(
            conn,
            department=dept_a,
            org_unit_id=org_unit_a,
        )
        conn.execute(
            text(
                """
                UPDATE public.department_recoding
                SET org_unit_name = :org_unit_name
                WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:name))
                """
            ),
            {"name": dept_a, "org_unit_name": f"Unit Alpha {suffix}"},
        )
        _ensure_department_mapping(
            conn,
            department=dept_b,
            org_unit_id=org_unit_b,
        )
        conn.execute(
            text(
                """
                UPDATE public.department_recoding
                SET org_unit_name = :org_unit_name
                WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:name))
                """
            ),
            {"name": dept_b, "org_unit_name": f"Unit Beta {suffix}"},
        )
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=dept_a,
        org_unit_id=org_unit_a,
        position="Врач",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=dept_b,
        org_unit_id=org_unit_b,
        position="Врач",
    )
    try:
        with engine.begin() as conn:
            _set_recoding_org_unit_name(conn, department=dept_a, org_unit_name=f"Unit Alpha {suffix}")
            _set_recoding_org_unit_name(conn, department=dept_b, org_unit_name=f"Unit Beta {suffix}")
        _create_snapshot_from_batch(seed, batch_id_1)
        _create_snapshot_from_batch(seed, batch_id_2)
        prior_id, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
        events = _events_for_pair(prior_id, new_id)
        dept_events = [event for event in events if event["event_type"] == EVENT_TYPE_DEPARTMENT_CHANGED]
        assert len(dept_events) == 1
        assert dept_events[0]["old_value"] == f"Unit Alpha {suffix}"
        assert dept_events[0]["new_value"] == f"Unit Beta {suffix}"
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_multiple_changes_generate_multiple_events(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    dept_a = f"Pytest Multi A {suffix}"
    dept_b = f"Pytest Multi B {suffix}"
    org_unit_a = int(seed["unit_id"])
    org_unit_b = org_unit_a
    full_name = _cyrillic_full_name(f"Multi{suffix}")
    iin = _unique_iin(suffix)
    with engine.begin() as conn:
        _ensure_department_mapping(conn, department=dept_a, org_unit_id=org_unit_a)
        conn.execute(
            text(
                """
                UPDATE public.department_recoding
                SET org_unit_name = :org_unit_name
                WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:name))
                """
            ),
            {"name": dept_a, "org_unit_name": f"Multi Alpha {suffix}"},
        )
        _ensure_department_mapping(conn, department=dept_b, org_unit_id=org_unit_b)
        conn.execute(
            text(
                """
                UPDATE public.department_recoding
                SET org_unit_name = :org_unit_name
                WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:name))
                """
            ),
            {"name": dept_b, "org_unit_name": f"Multi Beta {suffix}"},
        )
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=dept_a,
        org_unit_id=org_unit_a,
        position="Медсестра",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=dept_b,
        org_unit_id=org_unit_b,
        position="Старшая медсестра",
    )
    try:
        with engine.begin() as conn:
            _set_recoding_org_unit_name(conn, department=dept_a, org_unit_name=f"Multi Alpha {suffix}")
            _set_recoding_org_unit_name(conn, department=dept_b, org_unit_name=f"Multi Beta {suffix}")
        _create_snapshot_from_batch(seed, batch_id_1)
        _create_snapshot_from_batch(seed, batch_id_2)
        prior_id, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
        events = _events_for_pair(prior_id, new_id)
        event_types = {event["event_type"] for event in events}
        assert EVENT_TYPE_POSITION_CHANGED in event_types
        assert EVENT_TYPE_DEPARTMENT_CHANGED in event_types
        assert len(events) >= 2
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_list_hr_change_events_filters(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    department = f"Pytest Filter {suffix}"
    org_unit_id = int(seed["unit_id"])
    full_name = _cyrillic_full_name(f"Filter{suffix}")
    iin = _unique_iin(suffix)
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
        position="Терапевт",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
        position="Заведующий отделением",
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        _create_snapshot_from_batch(seed, batch_id_2)
        with engine.begin() as conn:
            stored_department = conn.execute(
                text(
                    """
                    SELECT department
                    FROM public.hr_change_events
                    WHERE event_type = :event_type
                    ORDER BY change_event_id DESC
                    LIMIT 1
                    """
                ),
                {"event_type": EVENT_TYPE_POSITION_CHANGED},
            ).scalar_one()
            by_type = list_hr_change_events(
                conn,
                event_type=EVENT_TYPE_POSITION_CHANGED,
                department=stored_department,
            )
            assert by_type["total"] >= 1
            assert all(item["event_type"] == EVENT_TYPE_POSITION_CHANGED for item in by_type["items"])
            by_dept = list_hr_change_events(conn, department=stored_department)
            assert by_dept["total"] >= 1
            assert all(item["department"] == stored_department for item in by_dept["items"])
            by_org = list_hr_change_events(conn, org_unit_id=org_unit_id)
            assert by_org["total"] >= 1
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_compare_snapshots_without_persist(seed, tmp_path: Path) -> None:
    _require_phase_040f()
    suffix = uuid4().hex[:8]
    department = f"Pytest Compare {suffix}"
    org_unit_id = int(seed["unit_id"])
    full_name = _cyrillic_full_name(f"Cmp{suffix}")
    iin = _unique_iin(suffix)
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
        position="A",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        org_unit_id=org_unit_id,
        position="B",
    )
    try:
        with engine.begin() as conn:
            first = build_canonical_snapshot_from_batch(
                conn,
                batch_id_1,
                promoted_by=int(seed["initiator_user_id"]),
            )
            second = build_canonical_snapshot_from_batch(
                conn,
                batch_id_2,
                promoted_by=int(seed["initiator_user_id"]),
            )
            prior_id = int(first["snapshot_id"])
            new_id = int(second["snapshot_id"])
            events = compare_snapshots(conn, prior_snapshot_id=prior_id, new_snapshot_id=new_id)
            assert any(event["event_type"] == EVENT_TYPE_POSITION_CHANGED for event in events)
            result = materialize_snapshot_change_events(
                conn,
                prior_snapshot_id=prior_id,
                new_snapshot_id=new_id,
            )
            assert result["event_count"] == len(events)
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)
