from __future__ import annotations

from contextlib import contextmanager
from datetime import date
from hashlib import sha256
from io import BytesIO
from uuid import uuid4

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook
from sqlalchemy import create_engine, text

import app.control_list_export.router as export_router_module
import app.services.access_resolver_service as access_resolver_module
from app.auth import get_current_user
from app.control_list_export.audit import (
    ControlListAuditScope,
    write_control_list_export_audit,
)
from app.control_list_export.router import router
from app.db.engine import engine as project_test_engine
from tests.test_control_list_projection_service import (
    _insert_assignment,
    _insert_employee,
    _projection_database,
    _seed_projection_reference_data,
)


def _postgres_available() -> bool:
    try:
        with project_test_engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


@contextmanager
def _audit_database():
    database_name = f"corpsite_wpcl003_audit_{uuid4().hex[:10]}_test"
    admin_url = (
        str(project_test_engine.url.render_as_string(hide_password=False)).rsplit("/", 1)[0]
        + "/postgres"
    )
    database_url = admin_url.rsplit("/", 1)[0] + f"/{database_name}"
    admin_engine = create_engine(admin_url, isolation_level="AUTOCOMMIT")
    audit_engine = create_engine(database_url, hide_parameters=True)
    with admin_engine.connect() as conn:
        conn.execute(text(f'CREATE DATABASE "{database_name}"'))
    try:
        with audit_engine.begin() as conn:
            conn.execute(
                text(
                    """
                    CREATE TABLE public.security_audit_log (
                        audit_id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                        event_type TEXT NOT NULL,
                        happened_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                        actor_user_id BIGINT NULL,
                        target_user_id BIGINT NULL,
                        target_person_id BIGINT NULL,
                        target_employee_id BIGINT NULL,
                        ip_address INET NULL,
                        user_agent TEXT NULL,
                        success BOOLEAN NOT NULL,
                        failure_reason TEXT NULL,
                        metadata JSONB NOT NULL,
                        request_id TEXT NULL,
                        CONSTRAINT chk_sal_event_type
                            CHECK (event_type IN ('CONTROL_LIST_EXPORT'))
                    )
                    """
                )
            )
        yield audit_engine
    finally:
        audit_engine.dispose()
        with admin_engine.connect() as conn:
            conn.execute(
                text(
                    """
                    SELECT pg_terminate_backend(pid)
                    FROM pg_stat_activity
                    WHERE datname=:database_name AND pid <> pg_backend_pid()
                    """
                ),
                {"database_name": database_name},
            )
            conn.execute(text(f'DROP DATABASE IF EXISTS "{database_name}"'))
        admin_engine.dispose()


@pytest.mark.skipif(not _postgres_available(), reason="PostgreSQL not available")
def test_error_audit_commits_after_independent_read_only_rollback() -> None:
    with _audit_database() as audit_engine:
        with pytest.raises(RuntimeError):
            with audit_engine.connect() as projection_conn:
                with projection_conn.begin():
                    projection_conn.exec_driver_sql(
                        "SET TRANSACTION ISOLATION LEVEL REPEATABLE READ READ ONLY"
                    )
                    projection_conn.execute(text("SELECT 1"))
                    raise RuntimeError("projection failed")

        audit_id = write_control_list_export_audit(
            audit_engine,
            actor_user_id=7,
            request_id="req-pg-error",
            result="ERROR",
            scope=ControlListAuditScope(
                organization_wide=False,
                org_unit_ids=(10,),
                resolution="RESOLVED",
            ),
            schema_version="CONTROL_LIST_EXPORT_V1",
            error_code="CONTROL_LIST_PROJECTION_ERROR",
        )
        with audit_engine.connect() as conn:
            row = conn.execute(
                text(
                    """
                    SELECT audit_id, event_type, success, failure_reason, metadata, request_id
                    FROM public.security_audit_log
                    """
                )
            ).mappings().one()
        assert row["audit_id"] == audit_id
        assert row["event_type"] == "CONTROL_LIST_EXPORT"
        assert row["success"] is False
        assert row["failure_reason"] == "CONTROL_LIST_PROJECTION_ERROR"
        assert row["request_id"] == "req-pg-error"
        assert row["metadata"]["result"] == "ERROR"
        assert row["metadata"]["scope"]["org_unit_ids"] == [10]
        serialized = str(row["metadata"])
        assert "iin" not in serialized.lower()
        assert "phone" not in serialized.lower()
        assert "employee_id" not in serialized.lower()


@pytest.mark.skipif(not _postgres_available(), reason="PostgreSQL not available")
def test_all_audit_results_persist_without_projection_pii() -> None:
    with _audit_database() as audit_engine:
        for index, result in enumerate(
            ("SUCCESS", "FORBIDDEN", "CONFLICT", "ERROR"), start=1
        ):
            write_control_list_export_audit(
                audit_engine,
                actor_user_id=7,
                request_id=f"req-result-{index}",
                result=result,
                scope=ControlListAuditScope(
                    organization_wide=True,
                    org_unit_ids=None,
                    resolution="RESOLVED",
                ),
                schema_version="CONTROL_LIST_EXPORT_V1",
                as_of_date=date(2026, 9, 3) if result != "FORBIDDEN" else None,
                row_count=593 if result == "SUCCESS" else None,
                error_code=None if result == "SUCCESS" else f"SAFE_{result}",
                sha256="a" * 64 if result == "SUCCESS" else None,
            )

        with audit_engine.connect() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT success, failure_reason, metadata
                    FROM public.security_audit_log
                    ORDER BY audit_id
                    """
                )
            ).mappings().all()
        assert [row["metadata"]["result"] for row in rows] == [
            "SUCCESS",
            "FORBIDDEN",
            "CONFLICT",
            "ERROR",
        ]
        assert rows[0]["success"] is True
        assert rows[0]["metadata"]["row_count"] == 593
        assert rows[0]["metadata"]["sha256"] == "a" * 64
        assert all(row["failure_reason"] is not None for row in rows[1:])
        serialized = str([row["metadata"] for row in rows]).lower()
        assert "iin" not in serialized
        assert "phone" not in serialized
        assert "employee_id" not in serialized


def _install_endpoint_integration_schema(integration_engine) -> None:
    with integration_engine.begin() as conn:
        conn.execute(
            text(
                """
                CREATE TABLE public.roles (
                    role_id BIGINT PRIMARY KEY,
                    code TEXT NOT NULL UNIQUE,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE
                );
                CREATE TABLE public.users (
                    user_id BIGINT PRIMARY KEY,
                    login TEXT NOT NULL UNIQUE,
                    role_id BIGINT NULL,
                    employee_id BIGINT NULL,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE
                );
                CREATE TABLE public.access_roles (
                    access_role_id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    code TEXT NOT NULL UNIQUE,
                    name TEXT NOT NULL,
                    description TEXT NULL,
                    access_level TEXT NOT NULL,
                    level_rank INTEGER NOT NULL,
                    is_system BOOLEAN NOT NULL DEFAULT FALSE,
                    is_active BOOLEAN NOT NULL DEFAULT TRUE,
                    created_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                    updated_at TIMESTAMPTZ NOT NULL DEFAULT now()
                );
                CREATE TABLE public.access_grants (
                    grant_id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    access_role_id BIGINT NOT NULL,
                    target_type TEXT NOT NULL,
                    target_id BIGINT NOT NULL,
                    resource_key TEXT NOT NULL DEFAULT '*',
                    scope_type TEXT NOT NULL DEFAULT 'GLOBAL',
                    scope_id BIGINT NULL,
                    include_subtree BOOLEAN NOT NULL DEFAULT FALSE,
                    starts_at TIMESTAMPTZ NOT NULL DEFAULT statement_timestamp(),
                    ends_at TIMESTAMPTZ NULL,
                    active_flag BOOLEAN NOT NULL DEFAULT TRUE,
                    granted_by_user_id BIGINT NOT NULL,
                    reason TEXT NULL
                );
                CREATE TABLE public.security_audit_log (
                    audit_id BIGINT GENERATED ALWAYS AS IDENTITY PRIMARY KEY,
                    event_type TEXT NOT NULL,
                    happened_at TIMESTAMPTZ NOT NULL DEFAULT now(),
                    actor_user_id BIGINT NULL,
                    target_user_id BIGINT NULL,
                    target_person_id BIGINT NULL,
                    target_employee_id BIGINT NULL,
                    ip_address INET NULL,
                    user_agent TEXT NULL,
                    success BOOLEAN NOT NULL,
                    failure_reason TEXT NULL,
                    metadata JSONB NOT NULL,
                    request_id TEXT NULL,
                    CONSTRAINT chk_sal_event_type
                        CHECK (event_type IN ('CONTROL_LIST_EXPORT'))
                );

                INSERT INTO public.roles (role_id, code) VALUES (2, 'ADMIN');
                INSERT INTO public.users (user_id, login, role_id)
                VALUES (700, 'wpcl003-admin', 2);
                INSERT INTO public.access_roles
                    (code, name, access_level, level_rank, is_system)
                VALUES
                    ('CONTROL_LIST_EXPORT', 'Control-list export', 'READ', 10, TRUE);
                """
            )
        )
        _seed_projection_reference_data(conn)
        _insert_employee(conn, 42)
        _insert_assignment(conn, 420, 42, 10)


@pytest.mark.skipif(not _postgres_available(), reason="PostgreSQL not available")
def test_endpoint_uses_actual_permission_projection_workbook_and_audit(
    monkeypatch,
) -> None:
    """Exercise the production chain without replacing RBAC, projection or audit."""

    with _projection_database() as integration_engine:
        _install_endpoint_integration_schema(integration_engine)
        monkeypatch.setenv("ORGANIZATION_TIMEZONE", "Asia/Almaty")
        monkeypatch.setattr(export_router_module, "engine", integration_engine)
        monkeypatch.setattr(access_resolver_module, "engine", integration_engine)

        application = FastAPI()
        application.include_router(router, prefix="/directory")
        application.dependency_overrides[get_current_user] = lambda: {
            "user_id": 700,
            "role_id": 2,
        }

        with TestClient(application) as http:
            forbidden = http.post(
                "/directory/personnel/control-list/export",
                headers={"X-Request-ID": "actual-forbidden"},
            )
            assert forbidden.status_code == 403
            assert forbidden.json()["detail"]["code"] == "CONTROL_LIST_EXPORT_FORBIDDEN"
            assert "content-disposition" not in forbidden.headers

            with integration_engine.begin() as conn:
                conn.execute(
                    text(
                        """
                        INSERT INTO public.access_grants
                            (access_role_id, target_type, target_id,
                             granted_by_user_id, reason)
                        SELECT access_role_id, 'USER', 700, 700, 'WP-CL-003 test'
                        FROM public.access_roles
                        WHERE code = 'CONTROL_LIST_EXPORT'
                        """
                    )
                )

            success = http.post(
                "/directory/personnel/control-list/export",
                headers={"X-Request-ID": "actual-success"},
            )
            assert success.status_code == 200
            assert success.headers["x-content-sha256"] == sha256(success.content).hexdigest()
            assert load_workbook(BytesIO(success.content)).sheetnames == [
                "Контрольный список",
                "Метаданные",
            ]

            with integration_engine.begin() as conn:
                _insert_assignment(conn, 421, 42, 20)

            conflict = http.post(
                "/directory/personnel/control-list/export",
                headers={"X-Request-ID": "actual-conflict"},
            )
            assert conflict.status_code == 409
            assert conflict.json()["detail"]["conflicts"] == [
                {"employee_id": 42, "violation": "MULTIPLE_PRIMARY_ASSIGNMENTS"}
            ]
            assert "content-disposition" not in conflict.headers

            monkeypatch.setenv("ORGANIZATION_TIMEZONE", "invalid//iana-zone")
            failure = http.post(
                "/directory/personnel/control-list/export",
                headers={"X-Request-ID": "actual-error"},
            )
            assert failure.status_code == 500
            assert failure.json()["detail"]["code"] == "CONTROL_LIST_CONFIGURATION_ERROR"
            assert "invalid//iana-zone" not in failure.text
            assert "content-disposition" not in failure.headers

        with integration_engine.connect() as conn:
            rows = conn.execute(
                text(
                    """
                    SELECT success, failure_reason, metadata, request_id
                    FROM public.security_audit_log
                    ORDER BY audit_id
                    """
                )
            ).mappings().all()

        assert [row["metadata"]["result"] for row in rows] == [
            "FORBIDDEN",
            "SUCCESS",
            "CONFLICT",
            "ERROR",
        ]
        assert [row["request_id"] for row in rows] == [
            "actual-forbidden",
            "actual-success",
            "actual-conflict",
            "actual-error",
        ]
        assert rows[1]["success"] is True
        assert rows[1]["metadata"]["sha256"] == sha256(success.content).hexdigest()
        assert rows[1]["metadata"]["row_count"] == 1
        assert all(row["failure_reason"] for row in (rows[0], rows[2], rows[3]))
        serialized = str(rows).lower()
        assert "900102300042" not in serialized
        assert "person 42" not in serialized
        assert "invalid//iana-zone" not in serialized
