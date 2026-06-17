"""ADR-038 Phase A.1 — import integrity hardening tests."""
from __future__ import annotations

import logging
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import load_workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.department_recoding_service import seed_department_recoding
from app.services.employee_import_profile_override_service import (
    employee_overrides_available,
    load_employee_override,
    resolve_directory_employee_id,
)
from app.services.hr_import_education_profile_service import update_education_profile
from app.services.hr_import_employee_card_service import get_employee_import_card, save_employee_import_card
from app.services.hr_import_profile_override_service import apply_profile_override
from app.services.hr_import_service import import_control_list
from tests.conftest import get_columns, insert_returning_id
from tests.test_import_hr_control_list import _build_sample_workbook


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _require_phase_a1() -> None:
    with engine.connect() as conn:
        if not employee_overrides_available(conn):
            pytest.skip("employee_import_profile_overrides not available — run alembic upgrade head")
        row = conn.execute(
            text(
                """
                SELECT 1 FROM information_schema.columns
                WHERE table_schema = 'public'
                  AND table_name = 'employee_import_profile_overrides'
                  AND column_name = 'base_batch_id'
                """
            )
        ).first()
        if not row:
            pytest.skip("Phase A.1 columns not available — run alembic upgrade head")


def _create_employee(conn, *, full_name: str, org_unit_id: int) -> int:
    cols = get_columns(conn, "employees")
    values = {"full_name": full_name}
    if "org_unit_id" in cols:
        values["org_unit_id"] = org_unit_id
    if "is_active" in cols:
        values["is_active"] = True
    return insert_returning_id(conn, table="employees", id_col="employee_id", values=values)


def _build_workbook_with_cert(path: Path, cert_text: str) -> None:
    _build_sample_workbook(path)
    wb = load_workbook(path)
    wb["врачи"].cell(row=8, column=14, value=cert_text)
    wb.save(path)


def _build_workbook_without_ivanov(path: Path) -> None:
    _build_sample_workbook(path)
    wb = load_workbook(path)
    wb["врачи"].delete_rows(8)
    wb.save(path)


@pytest.fixture
def seed_user_unit(seed):
    with engine.connect() as conn:
        user_id = int(seed["initiator_user_id"])
        unit_id = int(seed["unit_id"])
    return {"user_id": user_id, "unit_id": unit_id}


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_save_records_provenance_and_updated_by(seed_user_unit, tmp_path: Path):
    _require_phase_a1()
    source = tmp_path / f"a1_prov_{uuid4().hex[:8]}.xlsx"
    _build_sample_workbook(source)
    employee_id = None
    batch_id = None
    try:
        with engine.begin() as conn:
            seed_department_recoding(conn)
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=seed_user_unit["user_id"],
            )
            employee_id = _create_employee(
                conn,
                full_name="Иванов Иван Иванович",
                org_unit_id=seed_user_unit["unit_id"],
            )
            detail = get_employee_import_card(conn, employee_id)
            save_employee_import_card(
                conn,
                employee_id,
                profile=apply_profile_override(detail["profile"], {"notes": "provenance test"}),
                updated_by=seed_user_unit["user_id"],
            )
            card = get_employee_import_card(conn, employee_id)
            stored = load_employee_override(conn, employee_id)

        assert stored is not None
        assert stored["base_batch_id"] == batch_id
        assert stored["base_row_id"] == detail["row_id"]
        assert stored["base_imported_at"] is not None
        assert stored["created_by"] == seed_user_unit["user_id"]
        assert stored["updated_by"] == seed_user_unit["user_id"]
        assert card["base_batch_id"] == batch_id
        assert card["base_imported_at"] is not None
    finally:
        with engine.begin() as conn:
            if employee_id:
                conn.execute(
                    text("DELETE FROM public.employee_import_profile_overrides WHERE employee_id = :id"),
                    {"id": employee_id},
                )
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": employee_id})
            if batch_id:
                conn.execute(text("DELETE FROM public.hr_import_batches WHERE batch_id = :id"), {"id": batch_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_missing_from_latest_import_when_absent_from_new_batch(seed_user_unit, tmp_path: Path):
    _require_phase_a1()
    june = tmp_path / f"june_{uuid4().hex[:8]}.xlsx"
    july = tmp_path / f"july_{uuid4().hex[:8]}.xlsx"
    _build_sample_workbook(june)
    _build_workbook_without_ivanov(july)
    employee_id = None
    batch_ids: list[int] = []
    try:
        with engine.begin() as conn:
            seed_department_recoding(conn)
            b1, _, _ = import_control_list(conn, file_path=june, imported_by=seed_user_unit["user_id"])
            batch_ids.append(int(b1))
            employee_id = _create_employee(
                conn,
                full_name="Иванов Иван Иванович",
                org_unit_id=seed_user_unit["unit_id"],
            )
            b2, _, _ = import_control_list(conn, file_path=july, imported_by=seed_user_unit["user_id"])
            batch_ids.append(int(b2))
            card = get_employee_import_card(conn, employee_id)

        assert card["latest_batch_id"] == b2
        assert card["card_batch_id"] == b1
        assert card["missing_from_latest_import"] is True
    finally:
        with engine.begin() as conn:
            if employee_id:
                conn.execute(
                    text("DELETE FROM public.employee_import_profile_overrides WHERE employee_id = :id"),
                    {"id": employee_id},
                )
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": employee_id})
            for bid in batch_ids:
                conn.execute(text("DELETE FROM public.hr_import_batches WHERE batch_id = :id"), {"id": bid})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_not_missing_when_present_in_latest_batch(seed_user_unit, tmp_path: Path):
    _require_phase_a1()
    june = tmp_path / f"j1_{uuid4().hex[:8]}.xlsx"
    july = tmp_path / f"j2_{uuid4().hex[:8]}.xlsx"
    _build_sample_workbook(june)
    _build_sample_workbook(july)
    employee_id = None
    batch_ids: list[int] = []
    try:
        with engine.begin() as conn:
            seed_department_recoding(conn)
            b1, _, _ = import_control_list(conn, file_path=june, imported_by=seed_user_unit["user_id"])
            batch_ids.append(int(b1))
            employee_id = _create_employee(
                conn,
                full_name="Иванов Иван Иванович",
                org_unit_id=seed_user_unit["unit_id"],
            )
            b2, _, _ = import_control_list(conn, file_path=july, imported_by=seed_user_unit["user_id"])
            batch_ids.append(int(b2))
            card = get_employee_import_card(conn, employee_id)

        assert card["latest_batch_id"] == b2
        assert card["card_batch_id"] == b2
        assert card["missing_from_latest_import"] is False
    finally:
        with engine.begin() as conn:
            if employee_id:
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": employee_id})
            for bid in batch_ids:
                conn.execute(text("DELETE FROM public.hr_import_batches WHERE batch_id = :id"), {"id": bid})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_ambiguous_name_does_not_create_employee_override(seed_user_unit, tmp_path: Path, caplog):
    _require_phase_a1()
    source = tmp_path / f"amb_{uuid4().hex[:8]}.xlsx"
    _build_sample_workbook(source)
    batch_id = None
    employee_ids: list[int] = []
    try:
        with engine.begin() as conn:
            seed_department_recoding(conn)
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=seed_user_unit["user_id"],
            )
            employee_ids.append(
                _create_employee(
                    conn,
                    full_name="Дубликат ФИО Тест",
                    org_unit_id=seed_user_unit["unit_id"],
                )
            )
            employee_ids.append(
                _create_employee(
                    conn,
                    full_name="Дубликат ФИО Тест",
                    org_unit_id=seed_user_unit["unit_id"],
                )
            )
            profile_id = conn.execute(
                text(
                    """
                    SELECT row_id FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).scalar_one()

            with caplog.at_level(logging.WARNING):
                conn.execute(
                    text(
                        """
                        UPDATE public.hr_import_rows
                        SET normalized_payload = jsonb_set(
                            normalized_payload,
                            '{full_name}',
                            '"Дубликат ФИО Тест"'::jsonb,
                            true
                        )
                        WHERE batch_id = :batch_id AND row_id = :row_id
                        """
                    ),
                    {"batch_id": batch_id, "row_id": profile_id},
                )
                update_education_profile(
                    conn,
                    int(batch_id),
                    int(profile_id),
                    profile={"notes": "batch only"},
                    updated_by=seed_user_unit["user_id"],
                )
                resolved = resolve_directory_employee_id(
                    conn,
                    payload={"full_name": "Дубликат ФИО Тест"},
                )

            row_override = conn.execute(
                text(
                    """
                    SELECT profile_override->>'notes'
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id AND row_id = :row_id
                    """
                ),
                {"batch_id": batch_id, "row_id": profile_id},
            ).scalar_one()
            employee_override_count = conn.execute(
                text(
                    """
                    SELECT COUNT(*) FROM public.employee_import_profile_overrides
                    WHERE employee_id = ANY(:ids)
                    """
                ),
                {"ids": employee_ids},
            ).scalar_one()

        assert resolved is None
        assert row_override == "batch only"
        assert int(employee_override_count) == 0
        assert any("ambiguous full_name match" in rec.message for rec in caplog.records)
    finally:
        with engine.begin() as conn:
            for eid in employee_ids:
                conn.execute(
                    text("DELETE FROM public.employee_import_profile_overrides WHERE employee_id = :id"),
                    {"id": eid},
                )
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": eid})
            if batch_id:
                conn.execute(text("DELETE FROM public.hr_import_batches WHERE batch_id = :id"), {"id": batch_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_certificate_override_hides_new_import_certificate(seed_user_unit, tmp_path: Path):
    """Known limitation: section-level replace — override certificates hide new import data."""
    _require_phase_a1()
    june = tmp_path / f"cert_j_{uuid4().hex[:8]}.xlsx"
    july = tmp_path / f"cert_jl_{uuid4().hex[:8]}.xlsx"
    _build_workbook_with_cert(june, "сертификат специалиста по терапии 2020")
    _build_workbook_with_cert(july, "сертификат специалиста по кардиологии 2024")
    employee_id = None
    batch_ids: list[int] = []
    try:
        with engine.begin() as conn:
            seed_department_recoding(conn)
            b1, _, _ = import_control_list(conn, file_path=june, imported_by=seed_user_unit["user_id"])
            batch_ids.append(int(b1))
            employee_id = _create_employee(
                conn,
                full_name="Иванов Иван Иванович",
                org_unit_id=seed_user_unit["unit_id"],
            )
            detail = get_employee_import_card(conn, employee_id)
            june_topics = [c.get("topic") for c in detail["profile"].get("certificate_records", [])]
            assert june_topics
            assert any("терапии" in (t or "").lower() for t in june_topics)

            save_employee_import_card(
                conn,
                employee_id,
                profile=apply_profile_override(
                    detail["profile"],
                    {
                        "certificates": [
                            {
                                "kind": "Сертификат",
                                "topic": "сертификат B",
                                "date": "2021-01-01",
                                "specialty": "",
                            }
                        ]
                    },
                ),
                updated_by=seed_user_unit["user_id"],
            )
            b2, _, _ = import_control_list(conn, file_path=july, imported_by=seed_user_unit["user_id"])
            batch_ids.append(int(b2))
            card = get_employee_import_card(conn, employee_id)

        display_topics = [c.get("topic") for c in card["profile"].get("certificate_records", [])]
        assert display_topics == ["сертификат B"]
        assert card["batch_id"] == b2
        assert card["missing_from_latest_import"] is False
        assert "кардиологии" not in " ".join(display_topics).lower()
    finally:
        with engine.begin() as conn:
            if employee_id:
                conn.execute(
                    text("DELETE FROM public.employee_import_profile_overrides WHERE employee_id = :id"),
                    {"id": employee_id},
                )
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": employee_id})
            for bid in batch_ids:
                conn.execute(text("DELETE FROM public.hr_import_batches WHERE batch_id = :id"), {"id": bid})
