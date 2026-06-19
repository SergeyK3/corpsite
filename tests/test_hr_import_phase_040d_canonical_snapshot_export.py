"""Tests for ADR-040 Phase D — canonical snapshot Excel export."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_canonical_snapshot_service import (
    build_canonical_snapshot_from_batch,
    canonical_snapshot_available,
)
from app.services.hr_canonical_snapshot_export_service import (
    BASE_COLUMNS,
    METADATA_COLUMNS,
    CanonicalSnapshotExportError,
    export_canonical_snapshot_xlsx,
    normalize_export_source_type,
)
from app.services.hr_import_service import import_control_list
from tests.test_hr_import_phase_040a_canonical_snapshot import (
    _cleanup_promotion_batch,
    _cyrillic_full_name,
    _prepare_roster_row,
    _require_phase_040a,
    _test_iin,
)
from tests.test_import_hr_control_list import _build_doctors_sheet


def _require_phase_040d() -> None:
    _require_phase_040a()


def _unique_iin(seed: str | None = None) -> str:
    token = seed or uuid4().hex
    return _test_iin(token)


def _build_workbook(path: Path, *, full_name: str, iin: str, position: str = "") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=iin)
    if position:
        ws.cell(row=8, column=7, value=position)
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)
    wb.save(path)
    wb.close()


def _import_prepared_batch(
    seed,
    tmp_path: Path,
    *,
    full_name: str,
    iin: str,
    department: str,
    position: str = "",
) -> int:
    source = tmp_path / f"phase040d_{uuid4().hex[:8]}.xlsx"
    _build_workbook(source, full_name=full_name, iin=iin, position=position)
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row_id = conn.execute(
            text(
                """
                SELECT row_id
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).scalar_one()
        _prepare_roster_row(
            conn,
            batch_id=batch_id,
            row_id=int(row_id),
            full_name=full_name,
            iin=iin,
            department=department,
            org_unit_id=int(seed["unit_id"]),
        )
        if position:
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET normalized_payload = jsonb_set(
                        normalized_payload,
                        '{position_raw}',
                        to_jsonb(CAST(:position AS text)),
                        true
                    )
                    WHERE row_id = :row_id
                    """
                ),
                {"row_id": int(row_id), "position": position},
            )
    return int(batch_id)


def _read_sheet(content: bytes):
    return load_workbook(BytesIO(content), data_only=True).active


def _header_row(content: bytes) -> list[str]:
    ws = _read_sheet(content)
    return [str(ws.cell(row=1, column=i).value or "") for i in range(1, ws.max_column + 1)]


def _data_rows(content: bytes) -> list[list]:
    ws = _read_sheet(content)
    rows: list[list] = []
    for row_idx in range(2, ws.max_row + 1):
        rows.append([ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)])
    return rows


def test_normalize_export_source_type_aliases_roster() -> None:
    assert normalize_export_source_type("roster") == "HR_CONTROL_LIST"
    assert normalize_export_source_type(None) == "HR_CONTROL_LIST"


def test_export_active_snapshot(seed, tmp_path: Path) -> None:
    _require_phase_040d()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"Exp{suffix}")
    iin = _unique_iin(suffix)
    department = f"Pytest Export {suffix}"
    position = "Инженер по медицинскому оборудованию"
    batch_id = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position=position,
    )
    try:
        with engine.begin() as conn:
            build_canonical_snapshot_from_batch(
                conn,
                batch_id,
                promoted_by=int(seed["initiator_user_id"]),
            )
            content, filename = export_canonical_snapshot_xlsx(conn, source_type="roster")
            assert filename.endswith(".xlsx")
            headers = _header_row(content)
            assert headers[: len(BASE_COLUMNS)] == [label for _, label in BASE_COLUMNS]
            rows = _data_rows(content)
            assert len(rows) >= 1
            full_name_idx = headers.index("ФИО")
            iin_idx = headers.index("ИИН")
            position_idx = headers.index("Должность")
            assert rows[0][full_name_idx] == full_name
            assert rows[0][iin_idx] == iin
            assert rows[0][position_idx] == position
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_export_selected_snapshot_id(seed, tmp_path: Path) -> None:
    _require_phase_040d()
    suffix = uuid4().hex[:8]
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Snap1{suffix}"),
        iin=_unique_iin(f"1{suffix}"),
        department=f"Dept A {suffix}",
        position="Врач",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Snap2{suffix}"),
        iin=_unique_iin(f"2{suffix}"),
        department=f"Dept B {suffix}",
        position="Медсестра",
    )
    try:
        with engine.begin() as conn:
            first = build_canonical_snapshot_from_batch(
                conn,
                batch_id_1,
                promoted_by=int(seed["initiator_user_id"]),
            )
            build_canonical_snapshot_from_batch(
                conn,
                batch_id_2,
                promoted_by=int(seed["initiator_user_id"]),
            )
            content, _ = export_canonical_snapshot_xlsx(
                conn,
                snapshot_id=int(first["snapshot_id"]),
            )
            headers = _header_row(content)
            rows = _data_rows(content)
            position_idx = headers.index("Должность")
            assert any(row[position_idx] == "Врач" for row in rows)
            assert all(row[position_idx] != "Медсестра" for row in rows)
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_export_with_metadata_columns(seed, tmp_path: Path) -> None:
    _require_phase_040d()
    suffix = uuid4().hex[:8]
    batch_id = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Meta{suffix}"),
        iin=_unique_iin(suffix),
        department=f"Meta Dept {suffix}",
    )
    try:
        with engine.begin() as conn:
            snapshot = build_canonical_snapshot_from_batch(
                conn,
                batch_id,
                promoted_by=int(seed["initiator_user_id"]),
            )
            content, _ = export_canonical_snapshot_xlsx(
                conn,
                include_metadata=True,
            )
            headers = _header_row(content)
            for _, label in METADATA_COLUMNS:
                assert label in headers
            rows = _data_rows(content)
            snapshot_id_idx = headers.index("snapshot_id")
            assert rows[0][snapshot_id_idx] == int(snapshot["snapshot_id"])
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_empty_snapshot_exports_headers_only(seed, tmp_path: Path) -> None:
    _require_phase_040d()
    suffix = uuid4().hex[:8]
    source = tmp_path / f"empty_{suffix}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    if ws.max_row >= 8:
        ws.delete_rows(8, ws.max_row - 7)
    wb.save(source)
    wb.close()

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
    try:
        with engine.begin() as conn:
            build_canonical_snapshot_from_batch(
                conn,
                batch_id,
                promoted_by=int(seed["initiator_user_id"]),
            )
            content, _ = export_canonical_snapshot_xlsx(conn)
            headers = _header_row(content)
            assert headers[0] == "ИИН"
            assert _data_rows(content) == []
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_export_missing_snapshot_id_raises(seed, tmp_path: Path) -> None:
    _require_phase_040d()
    with engine.begin() as conn:
        with pytest.raises(CanonicalSnapshotExportError) as exc:
            export_canonical_snapshot_xlsx(conn, snapshot_id=9_999_999)
        assert exc.value.status_code == 404
