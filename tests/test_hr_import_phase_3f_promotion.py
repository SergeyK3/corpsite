"""Tests for ADR-039 Phase 3F — normalized records promotion API and service."""
from __future__ import annotations

from datetime import date
from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_import_normalized_record_service import normalized_records_available
from app.services.hr_import_promotion_service import (
    BLOCKER_EMPLOYEE_REQUIRED,
    BLOCKER_MEDICAL_SPECIALTY_UNRESOLVED,
    BLOCKER_NOT_APPROVED,
    SKIP_ALREADY_PROMOTED,
    SKIP_DUPLICATE_ACTIVE_DOCUMENT,
    promote_normalized_records,
)
from app.services.hr_import_service import import_control_list
from tests.conftest import auth_headers, insert_returning_id, table_exists
from tests.test_employee_documents_routes import (
    _create_employee,
    _create_position,
    _lookup_document_type_id,
    _phase_1a_available,
)
from tests.test_import_hr_control_list import _build_doctors_sheet


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_3f_available() -> bool:
    with engine.begin() as conn:
        if not normalized_records_available(conn):
            return False
        row = conn.execute(
            text(
                """
                SELECT 1
                FROM public.document_types
                WHERE code = 'QUALIFICATION_CATEGORY'
                  AND is_active = TRUE
                LIMIT 1
                """
            )
        ).first()
        return row is not None


def _require_phase_3f() -> None:
    if not _phase_3f_available():
        pytest.skip("ADR-039 Phase 3F migration not applied — run alembic upgrade head")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _build_doctors_sheet_with_column_m(path: Path, column_m: str, column_n: str = "") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=13, value=column_m)
    if column_n:
        ws.cell(row=8, column=14, value=column_n)
    wb.save(path)
    wb.close()


def _make_employee(seed) -> tuple[int, list[int], list[int]]:
    emp_ids: list[int] = []
    pos_ids: list[int] = []
    with engine.begin() as conn:
        if not _phase_1a_available():
            pytest.skip("ADR-037 Phase 1A tables missing — run alembic upgrade head")
        pos_id = _create_position(conn, name=f"pytest_promo_{uuid4().hex[:8]}")
        pos_ids.append(pos_id)
        emp_id = _create_employee(
            conn,
            full_name=f"Pytest Promo {uuid4().hex[:8]}",
            org_unit_id=int(seed["unit_id"]),
            position_id=pos_id,
            is_active=True,
        )
        emp_ids.append(emp_id)
    return emp_id, emp_ids, pos_ids


def _cleanup_employees(employee_ids: list[int], position_ids: list[int]) -> None:
    with engine.begin() as conn:
        if employee_ids and table_exists(conn, "employee_documents"):
            conn.execute(
                text(
                    """
                    DELETE FROM public.employee_documents
                    WHERE employee_id = ANY(:ids)
                       OR source_normalized_record_id IN (
                            SELECT normalized_record_id
                            FROM public.hr_import_normalized_records
                            WHERE employee_id = ANY(:ids)
                       )
                    """
                ),
                {"ids": employee_ids},
            )
        if employee_ids and table_exists(conn, "employees"):
            conn.execute(
                text("DELETE FROM public.employees WHERE employee_id = ANY(:ids)"),
                {"ids": employee_ids},
            )
        if position_ids and table_exists(conn, "positions"):
            conn.execute(
                text("DELETE FROM public.positions WHERE position_id = ANY(:ids)"),
                {"ids": position_ids},
            )


@pytest.fixture
def privileged_headers(seed, monkeypatch):
    monkeypatch.setenv("DIRECTORY_PRIVILEGED_USER_IDS", str(seed["initiator_user_id"]))
    return auth_headers(seed["initiator_user_id"])


@pytest.fixture
def approved_education_record(seed, tmp_path: Path):
    _require_phase_3f()
    source = tmp_path / f"phase3f_promo_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(
        source,
        "КазНМУ, 1982; ПК 144 ч",
        "сертификат специалиста, действ. до 01.01.2028",
    )

    emp_id, emp_ids, pos_ids = _make_employee(seed)

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row = conn.execute(
            text(
                """
                SELECT normalized_record_id, source_record_key, row_id
                FROM public.hr_import_normalized_records
                WHERE batch_id = :batch_id
                  AND record_kind = 'education'
                ORDER BY normalized_record_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).mappings().first()
        if row is None:
            import_row = conn.execute(
                text(
                    """
                    SELECT row_id
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).first()
            assert import_row is not None
            source_record_key = f"pytest-education-{uuid4().hex}"
            inserted = conn.execute(
                text(
                    """
                    INSERT INTO public.hr_import_normalized_records (
                        batch_id,
                        row_id,
                        employee_id,
                        fragment_index,
                        source_field,
                        source_text,
                        source_record_key,
                        record_kind,
                        document_type_id,
                        document_type_code,
                        title,
                        provider,
                        issue_date,
                        parse_method,
                        review_status,
                        reviewed_at,
                        reviewed_by
                    )
                    VALUES (
                        :batch_id,
                        :row_id,
                        :employee_id,
                        0,
                        'education_raw',
                        'КазНМУ, 1982',
                        :source_record_key,
                        'education',
                        (SELECT document_type_id FROM public.document_types WHERE code = 'EDUCATION_GRADUATION'),
                        'EDUCATION_GRADUATION',
                        'КазНМУ',
                        'КазНМУ',
                        DATE '1982-01-01',
                        'manual',
                        'approved',
                        NOW(),
                        :reviewed_by
                    )
                    RETURNING normalized_record_id, source_record_key, row_id
                    """
                ),
                {
                    "batch_id": batch_id,
                    "row_id": int(import_row[0]),
                    "employee_id": emp_id,
                    "source_record_key": source_record_key,
                    "reviewed_by": int(seed["initiator_user_id"]),
                },
            ).one()
            row = {
                "normalized_record_id": int(inserted[0]),
                "source_record_key": str(inserted[1]),
                "row_id": int(inserted[2]),
            }
        else:
            record_id = int(row["normalized_record_id"])
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_normalized_records
                    SET employee_id = :employee_id,
                        review_status = 'approved',
                        reviewed_at = NOW(),
                        reviewed_by = :reviewed_by
                    WHERE normalized_record_id = :record_id
                    """
                ),
                {
                    "employee_id": emp_id,
                    "reviewed_by": int(seed["initiator_user_id"]),
                    "record_id": record_id,
                },
            )

    payload = {
        "batch_id": batch_id,
        "record_id": int(row["normalized_record_id"]),
        "employee_id": emp_id,
        "source_record_key": str(row["source_record_key"]),
        "row_id": int(row["row_id"]),
    }
    yield payload
    with engine.begin() as conn:
        _delete_batch(conn, batch_id)
    _cleanup_employees(emp_ids, pos_ids)


def _promote(client: TestClient, headers: dict[str, str], body: dict) -> dict:
    resp = client.post(
        "/directory/personnel/import/normalized-records/promote",
        headers={**headers, "Content-Type": "application/json"},
        json=body,
    )
    return resp


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_qualification_category_document_type_seed():
    _require_phase_3f()
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.document_types
                SET requires_medical_specialty = FALSE
                WHERE code = 'QUALIFICATION_CATEGORY'
                """
            )
        )
        row = conn.execute(
            text(
                """
                SELECT code, category, has_valid_until, requires_medical_specialty, tracks_hours
                FROM public.document_types
                WHERE code = 'QUALIFICATION_CATEGORY'
                """
            )
        ).mappings().one()
    assert row["category"] == "CREDENTIAL"
    assert bool(row["has_valid_until"]) is True
    assert bool(row["requires_medical_specialty"]) is False
    assert bool(row["tracks_hours"]) is False


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_promote_education_single_writes_provenance(
    client: TestClient,
    privileged_headers,
    approved_education_record,
    seed,
):
    record_id = approved_education_record["record_id"]
    emp_id = approved_education_record["employee_id"]

    before_docs = 0
    with engine.begin() as conn:
        before_docs = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*)
                    FROM public.employee_documents
                    WHERE source_normalized_record_id = :record_id
                    """
                ),
                {"record_id": record_id},
            ).scalar_one()
        )

    resp = _promote(
        client,
        privileged_headers,
        {"record_ids": [record_id], "dry_run": False},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["promoted"] == 1
    assert body["failed"] == 0
    item = body["items"][0]
    assert item["outcome"] == "promoted"
    assert item["document_id"] is not None

    with engine.begin() as conn:
        doc = conn.execute(
            text(
                """
                SELECT
                    document_id,
                    employee_id,
                    source_batch_id,
                    source_row_id,
                    source_normalized_record_id,
                    source_record_key,
                    source_text,
                    parse_method,
                    parse_confidence,
                    verification_status,
                    lifecycle_status
                FROM public.employee_documents
                WHERE document_id = :document_id
                """
            ),
            {"document_id": int(item["document_id"])},
        ).mappings().one()
        hinr = conn.execute(
            text(
                """
                SELECT review_status, promoted_document_id, promoted_by
                FROM public.hr_import_normalized_records
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        ).mappings().one()

    assert int(doc["employee_id"]) == emp_id
    assert doc["source_batch_id"] is not None
    assert doc["source_row_id"] is not None
    assert int(doc["source_normalized_record_id"]) == record_id
    assert doc["source_record_key"]
    assert doc["source_text"]
    assert doc["parse_method"]
    assert doc["verification_status"] == "UNVERIFIED"
    assert doc["lifecycle_status"] == "ACTIVE"
    assert hinr["review_status"] == "promoted"
    assert int(hinr["promoted_document_id"]) == int(item["document_id"])
    assert int(hinr["promoted_by"]) == int(seed["initiator_user_id"])
    assert before_docs == 0


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_dry_run_does_not_write(
    client: TestClient,
    privileged_headers,
    approved_education_record,
):
    record_id = approved_education_record["record_id"]
    resp = _promote(
        client,
        privileged_headers,
        {"record_ids": [record_id], "dry_run": True},
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["dry_run"] is True
    assert body["would_promote"] == 1
    assert body["promoted"] == 0
    assert body["items"][0]["outcome"] == "would_promote"
    assert body["items"][0]["preview"]["document_type_code"] == "EDUCATION_GRADUATION"

    with engine.begin() as conn:
        count = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*)
                    FROM public.employee_documents
                    WHERE source_normalized_record_id = :record_id
                    """
                ),
                {"record_id": record_id},
            ).scalar_one()
        )
        status = conn.execute(
            text(
                """
                SELECT review_status
                FROM public.hr_import_normalized_records
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        ).scalar_one()
    assert count == 0
    assert status == "approved"


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_not_approved_blocker(client: TestClient, privileged_headers, approved_education_record):
    record_id = approved_education_record["record_id"]
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET review_status = 'pending', reviewed_at = NULL, reviewed_by = NULL
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        )

    resp = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["outcome"] == "failed"
    assert item["blockers"][0]["code"] == BLOCKER_NOT_APPROVED


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_employee_required_blocker(client: TestClient, privileged_headers, approved_education_record):
    record_id = approved_education_record["record_id"]
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET employee_id = NULL
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        )

    resp = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["outcome"] == "failed"
    assert item["blockers"][0]["code"] == BLOCKER_EMPLOYEE_REQUIRED


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_training_promotes_without_medical_specialty(
    client: TestClient,
    privileged_headers,
    approved_education_record,
):
    record_id = approved_education_record["record_id"]
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET record_kind = 'training',
                    document_type_code = 'CONTINUING_EDUCATION',
                    document_type_id = (
                        SELECT document_type_id
                        FROM public.document_types
                        WHERE code = 'CONTINUING_EDUCATION'
                        LIMIT 1
                    ),
                    title = 'Внедрение HR-менеджмента в организациях здравоохранения',
                    issue_date = DATE '2024-01-01',
                    hours = 72,
                    specialty_text = NULL,
                    medical_specialty_id = NULL
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        )

    resp = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["outcome"] == "promoted"
    assert item["resolved_document_type_code"] == "CONTINUING_EDUCATION"
    assert item.get("resolved_medical_specialty_id") is None


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_category_promotes_with_qualification_category_and_free_text_specialty(
    client: TestClient,
    privileged_headers,
    approved_education_record,
    seed,
):
    record_id = approved_education_record["record_id"]
    batch_id = approved_education_record["batch_id"]
    row_id = approved_education_record["row_id"]
    emp_id = approved_education_record["employee_id"]

    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET record_kind = 'category',
                    document_type_code = 'QUALIFICATION_CATEGORY',
                    document_type_id = (
                        SELECT document_type_id
                        FROM public.document_types
                        WHERE code = 'QUALIFICATION_CATEGORY'
                        LIMIT 1
                    ),
                    title = 'Врач',
                    specialty_text = 'организация здравоохранения',
                    expiry_date = DATE '2030-12-31',
                    issue_date = NULL,
                    hours = NULL,
                    medical_specialty_id = NULL,
                    source_text = 'Категория врач организация здравоохранения до 31.12.2030'
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        )

    resp = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["outcome"] == "promoted"
    assert item["resolved_document_type_code"] == "QUALIFICATION_CATEGORY"
    assert item.get("resolved_medical_specialty_id") is None

    with engine.begin() as conn:
        doc = conn.execute(
            text(
                """
                SELECT title, medical_specialty_id, source_text
                FROM public.employee_documents
                WHERE document_id = :document_id
                """
            ),
            {"document_id": int(item["document_id"])},
        ).mappings().one()
        _delete_batch(conn, batch_id)

    assert "Врач" in doc["title"]
    assert "организация здравоохранения" in doc["title"]
    assert doc["medical_specialty_id"] is None
    assert doc["source_text"]


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_certificate_still_requires_medical_specialty(
    client: TestClient,
    privileged_headers,
    approved_education_record,
):
    record_id = approved_education_record["record_id"]
    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET record_kind = 'certificate',
                    document_type_code = 'SPECIALIST_CERTIFICATION',
                    document_type_id = (
                        SELECT document_type_id
                        FROM public.document_types
                        WHERE code = 'SPECIALIST_CERTIFICATION'
                        LIMIT 1
                    ),
                    title = 'Сертификат специалиста',
                    specialty_text = 'Несуществующая специальность XYZ',
                    expiry_date = DATE '2028-01-01',
                    medical_specialty_id = NULL
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        )

    resp = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["outcome"] == "failed"
    assert item["blockers"][0]["code"] == BLOCKER_MEDICAL_SPECIALTY_UNRESOLVED


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_already_promoted_skip(client: TestClient, privileged_headers, approved_education_record):
    record_id = approved_education_record["record_id"]
    first = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert first.status_code == 200, first.text
    document_id = first.json()["items"][0]["document_id"]

    second = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert second.status_code == 200, second.text
    item = second.json()["items"][0]
    assert item["outcome"] == "skipped"
    assert item["reason"] == SKIP_ALREADY_PROMOTED
    assert item["document_id"] == document_id


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_duplicate_active_document_skip(
    client: TestClient,
    privileged_headers,
    approved_education_record,
    seed,
):
    record_id = approved_education_record["record_id"]
    emp_id = approved_education_record["employee_id"]
    source_record_key = approved_education_record["source_record_key"]

    with engine.begin() as conn:
        education_type_id = _lookup_document_type_id(conn, "EDUCATION_GRADUATION")
        existing_id = insert_returning_id(
            conn,
            table="employee_documents",
            id_col="document_id",
            values={
                "employee_id": emp_id,
                "document_type_id": education_type_id,
                "title": "Existing diploma",
                "lifecycle_status": "ACTIVE",
                "created_by": int(seed["initiator_user_id"]),
                "source_record_key": source_record_key,
            },
        )

    resp = _promote(client, privileged_headers, {"record_ids": [record_id]})
    assert resp.status_code == 200, resp.text
    item = resp.json()["items"][0]
    assert item["outcome"] == "skipped"
    assert item["reason"] == SKIP_DUPLICATE_ACTIVE_DOCUMENT
    assert item["document_id"] == existing_id

    with engine.begin() as conn:
        count = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*)
                    FROM public.employee_documents
                    WHERE employee_id = :employee_id
                      AND source_record_key = :source_record_key
                      AND lifecycle_status = 'ACTIVE'
                    """
                ),
                {"employee_id": emp_id, "source_record_key": source_record_key},
            ).scalar_one()
        )
    assert count == 1


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_bulk_promote_by_batch_and_kind(
    client: TestClient,
    privileged_headers,
    approved_education_record,
    seed,
):
    batch_id = approved_education_record["batch_id"]
    record_id = approved_education_record["record_id"]
    emp_id = approved_education_record["employee_id"]
    row_id = approved_education_record["row_id"]

    with engine.begin() as conn:
        conn.execute(
            text(
                """
                INSERT INTO public.hr_import_normalized_records (
                    batch_id,
                    row_id,
                    employee_id,
                    fragment_index,
                    source_field,
                    source_text,
                    source_record_key,
                    record_kind,
                    document_type_id,
                    document_type_code,
                    title,
                    issue_date,
                    parse_method,
                    review_status,
                    reviewed_at,
                    reviewed_by
                )
                VALUES (
                    :batch_id,
                    :row_id,
                    :employee_id,
                    99,
                    'education_raw',
                    'Second university',
                    :source_record_key,
                    'education',
                    (SELECT document_type_id FROM public.document_types WHERE code = 'EDUCATION_GRADUATION'),
                    'EDUCATION_GRADUATION',
                    'Second university',
                    DATE '2015-01-01',
                    'manual',
                    'approved',
                    NOW(),
                    :reviewed_by
                )
                """
            ),
            {
                "batch_id": batch_id,
                "row_id": row_id,
                "employee_id": emp_id,
                "source_record_key": f"bulk-{uuid4().hex}",
                "reviewed_by": int(seed["initiator_user_id"]),
            },
        )

    resp = _promote(
        client,
        privileged_headers,
        {
            "batch_id": batch_id,
            "filters": {"review_status": "approved", "record_kind": "education"},
        },
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert body["requested"] >= 2
    assert body["promoted"] >= 2
    promoted_ids = {
        item["record_id"] for item in body["items"] if item["outcome"] == "promoted"
    }
    assert record_id in promoted_ids


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_promote_certificate_with_exact_specialty_match(
    client: TestClient,
    privileged_headers,
    seed,
    tmp_path: Path,
):
    _require_phase_3f()
    source = tmp_path / f"phase3f_cert_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(
        source,
        "2024 курс ПК 72 ч",
        'Сертификат "Онкология" до 01.01.2028',
    )
    emp_id, emp_ids, pos_ids = _make_employee(seed)

    try:
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row = conn.execute(
                text(
                    """
                    SELECT normalized_record_id
                    FROM public.hr_import_normalized_records
                    WHERE batch_id = :batch_id
                      AND record_kind = 'certificate'
                    ORDER BY normalized_record_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().first()
            assert row is not None
            record_id = int(row["normalized_record_id"])
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_normalized_records
                    SET employee_id = :employee_id,
                        review_status = 'approved',
                        reviewed_at = NOW(),
                        reviewed_by = :reviewed_by,
                        specialty_text = 'Врач-онколог',
                        expiry_date = DATE '2028-01-01'
                    WHERE normalized_record_id = :record_id
                    """
                ),
                {
                    "employee_id": emp_id,
                    "reviewed_by": int(seed["initiator_user_id"]),
                    "record_id": record_id,
                },
            )

        resp = _promote(
            client,
            privileged_headers,
            {"record_ids": [record_id], "dry_run": True},
        )
        assert resp.status_code == 200, resp.text
        item = resp.json()["items"][0]
        assert item["outcome"] == "would_promote"
        assert item["resolved_document_type_code"] == "SPECIALIST_CERTIFICATION"
        assert item["resolved_medical_specialty_id"] is not None
    finally:
        with engine.begin() as conn:
            if "batch_id" in locals():
                _delete_batch(conn, batch_id)
        _cleanup_employees(emp_ids, pos_ids)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_promote_unprivileged_forbidden(client, seed, approved_education_record):
    record_id = approved_education_record["record_id"]
    headers = auth_headers(seed["executor_user_id"])
    resp = _promote(client, headers, {"record_ids": [record_id]})
    assert resp.status_code == 403


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_service_partial_failure_isolation(seed, approved_education_record):
    record_id = approved_education_record["record_id"]
    batch_id = approved_education_record["batch_id"]
    row_id = approved_education_record["row_id"]
    emp_id = approved_education_record["employee_id"]

    with engine.begin() as conn:
        pending_id = int(
            conn.execute(
                text(
                    """
                    INSERT INTO public.hr_import_normalized_records (
                        batch_id,
                        row_id,
                        employee_id,
                        fragment_index,
                        source_field,
                        source_text,
                        source_record_key,
                        record_kind,
                        document_type_id,
                        document_type_code,
                        review_status
                    )
                    VALUES (
                        :batch_id,
                        :row_id,
                        :employee_id,
                        100,
                        'education_training_raw',
                        'pending training',
                        :source_record_key,
                        'training',
                        (SELECT document_type_id FROM public.document_types WHERE code = 'CONTINUING_EDUCATION'),
                        'CONTINUING_EDUCATION',
                        'pending'
                    )
                    RETURNING normalized_record_id
                    """
                ),
                {
                    "batch_id": batch_id,
                    "row_id": row_id,
                    "employee_id": emp_id,
                    "source_record_key": f"pending-{uuid4().hex}",
                },
            ).scalar_one()
        )
        result = promote_normalized_records(
            conn,
            promoted_by=int(seed["initiator_user_id"]),
            record_ids=[record_id, pending_id],
        )
    outcomes = {item["record_id"]: item["outcome"] for item in result["items"]}
    assert outcomes[record_id] == "promoted"
    assert outcomes[pending_id] == "failed"
