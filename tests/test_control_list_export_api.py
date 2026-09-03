from __future__ import annotations

from datetime import date
from hashlib import sha256
from io import BytesIO

import pytest
from fastapi import FastAPI
from fastapi.testclient import TestClient
from openpyxl import load_workbook

import app.control_list_export.router as export_router_module
from app.auth import get_current_user
from app.control_list_export.router import router
from app.control_list_export.workbook import (
    ControlListExportLimitError,
    ControlListWorkbookError,
)
from app.control_list_projection import (
    ControlListAssignmentConflict,
    ControlListAuthorizationError,
)
from app.control_list_projection.schemas import (
    ControlListAssignmentConflictDetail,
    ControlListAssignmentConflictItem,
)
from app.control_list_projection.service import ControlListConfigurationError
from tests.test_control_list_xlsx_export import _projection


@pytest.fixture
def client(monkeypatch):
    application = FastAPI()
    application.include_router(router, prefix="/directory")
    application.dependency_overrides[get_current_user] = lambda: {
        "user_id": 7,
        "role_id": 5,
    }
    audits: list[dict] = []
    monkeypatch.setattr(
        export_router_module,
        "write_control_list_export_audit",
        lambda _engine, **kwargs: audits.append(kwargs) or 1,
    )
    return TestClient(application), audits


def test_success_download_headers_workbook_and_audit(client, monkeypatch) -> None:
    http, audits = client
    monkeypatch.setattr(
        export_router_module, "build_control_list_projection", lambda *_a, **_kw: _projection()
    )
    response = http.post(
        "/directory/personnel/control-list/export",
        headers={"X-Request-ID": "req-7"},
    )
    assert response.status_code == 200
    assert response.headers["content-type"].startswith(
        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
    )
    disposition = response.headers["content-disposition"]
    assert 'filename="control-list-2026-09-03.xlsx"' in disposition
    assert "filename*=UTF-8''%D0%9A" in disposition
    assert response.headers["cache-control"] == "private, no-store, max-age=0"
    assert response.headers["pragma"] == "no-cache"
    assert response.headers["expires"] == "0"
    assert response.headers["x-content-type-options"] == "nosniff"
    assert response.headers["x-request-id"] == "req-7"
    assert response.headers["x-content-sha256"] == sha256(response.content).hexdigest()
    assert load_workbook(BytesIO(response.content)).sheetnames == [
        "Контрольный список",
        "Метаданные",
    ]
    assert len(audits) == 1
    assert audits[0]["result"] == "SUCCESS"
    assert audits[0]["row_count"] == 1
    assert audits[0]["request_id"] == "req-7"
    assert audits[0]["sha256"] == sha256(response.content).hexdigest()
    assert audits[0]["scope"].organization_wide is False
    assert audits[0]["scope"].org_unit_ids == (10, 20)


@pytest.mark.parametrize(
    ("exception", "status", "result", "code"),
    [
        (
            ControlListAuthorizationError(
                "CONTROL_LIST_EXPORT_FORBIDDEN",
                "Control-list export permission is required.",
            ),
            403,
            "FORBIDDEN",
            "CONTROL_LIST_EXPORT_FORBIDDEN",
        ),
        (
            ControlListConfigurationError("secret timezone value"),
            500,
            "ERROR",
            "CONTROL_LIST_CONFIGURATION_ERROR",
        ),
        (
            RuntimeError("SQL failed for 001234567890"),
            500,
            "ERROR",
            "CONTROL_LIST_PROJECTION_ERROR",
        ),
    ],
)
def test_safe_projection_errors_are_audited(
    client, monkeypatch, exception, status, result, code
) -> None:
    http, audits = client

    def fail(*_args, **_kwargs):
        raise exception

    monkeypatch.setattr(export_router_module, "build_control_list_projection", fail)
    monkeypatch.setattr(
        export_router_module,
        "compute_scope",
        lambda *_a, **_kw: {
            "privileged": False,
            "has_personnel_visibility": True,
            "scope_unit_ids": [10],
        },
    )
    response = http.post("/directory/personnel/control-list/export")
    assert response.status_code == status
    assert response.json()["detail"]["code"] == code
    assert response.headers["cache-control"] == "private, no-store, max-age=0"
    assert "secret timezone value" not in response.text
    assert "001234567890" not in response.text
    assert audits[0]["result"] == result
    assert audits[0]["error_code"] == code
    assert audits[0]["row_count"] is None


def test_conflict_returns_no_file_and_audits_only_technical_ids(client, monkeypatch) -> None:
    http, audits = client
    conflict = ControlListAssignmentConflict(
        ControlListAssignmentConflictDetail(
            code="CONTROL_LIST_ASSIGNMENT_CONFLICT",
            message="Active primary assignment invariant is violated.",
            schema_version="CONTROL_LIST_EXPORT_V1",
            as_of_date=date(2026, 9, 3),
            conflicts=[
                ControlListAssignmentConflictItem(
                    employee_id=42, violation="MULTIPLE_PRIMARY_ASSIGNMENTS"
                )
            ],
        )
    )

    def fail(*_args, **_kwargs):
        raise conflict

    monkeypatch.setattr(export_router_module, "build_control_list_projection", fail)
    monkeypatch.setattr(
        export_router_module,
        "compute_scope",
        lambda *_a, **_kw: {
            "privileged": False,
            "has_personnel_visibility": True,
            "scope_unit_ids": [10],
        },
    )
    response = http.post("/directory/personnel/control-list/export")
    assert response.status_code == 409
    assert response.headers["cache-control"] == "private, no-store, max-age=0"
    assert response.headers["content-type"].startswith("application/json")
    assert "application/vnd.openxmlformats" not in response.headers["content-type"]
    assert response.json()["detail"]["conflicts"] == [
        {"employee_id": 42, "violation": "MULTIPLE_PRIMARY_ASSIGNMENTS"}
    ]
    assert audits[0]["result"] == "CONFLICT"
    serialized_audit = str(audits[0])
    assert "001234567890" not in serialized_audit
    assert "+7 701" not in serialized_audit
    assert "Иванов" not in serialized_audit


@pytest.mark.parametrize(
    ("exception", "status", "code"),
    [
        (ControlListWorkbookError("PII =001234567890"), 500, "CONTROL_LIST_XLSX_BUILD_ERROR"),
        (ControlListExportLimitError("too large"), 413, "CONTROL_LIST_EXPORT_TOO_LARGE"),
    ],
)
def test_workbook_failures_return_no_partial_file_and_safe_error(
    client, monkeypatch, exception, status, code
) -> None:
    http, audits = client
    monkeypatch.setattr(
        export_router_module, "build_control_list_projection", lambda *_a, **_kw: _projection()
    )

    def fail(*_args, **_kwargs):
        raise exception

    monkeypatch.setattr(export_router_module, "build_control_list_workbook", fail)
    response = http.post("/directory/personnel/control-list/export")
    assert response.status_code == status
    assert response.json()["detail"]["code"] == code
    assert "001234567890" not in response.text
    assert "content-disposition" not in response.headers
    assert audits[0]["result"] == "ERROR"
    assert audits[0]["error_code"] == code


def test_invalid_request_id_is_not_reflected(client, monkeypatch) -> None:
    http, _audits = client
    monkeypatch.setattr(
        export_router_module, "build_control_list_projection", lambda *_a, **_kw: _projection()
    )
    response = http.post(
        "/directory/personnel/control-list/export",
        headers={"X-Request-ID": "bad\r\nvalue"},
    )
    assert response.status_code == 200
    assert response.headers["x-request-id"] != "bad value"
    assert len(response.headers["x-request-id"]) == 32


def test_audit_failure_is_fail_closed_and_does_not_return_file(client, monkeypatch) -> None:
    http, _audits = client
    monkeypatch.setattr(
        export_router_module, "build_control_list_projection", lambda *_a, **_kw: _projection()
    )

    def audit_failure(*_args, **_kwargs):
        raise RuntimeError("database details must not escape")

    monkeypatch.setattr(
        export_router_module, "write_control_list_export_audit", audit_failure
    )
    response = http.post("/directory/personnel/control-list/export")
    assert response.status_code == 500
    assert response.json()["detail"]["code"] == "CONTROL_LIST_AUDIT_ERROR"
    assert "database details" not in response.text
    assert "content-disposition" not in response.headers


def test_no_public_projection_json_route_exists() -> None:
    from app.main import app

    route_methods = {(route.path, tuple(sorted(route.methods or []))) for route in app.routes}
    assert not any(
        path == "/directory/personnel/control-list/projection"
        for path, _methods in route_methods
    )
    assert (
        "/directory/personnel/control-list/export",
        ("POST",),
    ) in route_methods


def test_export_does_not_authenticate_from_cookie() -> None:
    """Bearer auth is not an ambient browser credential, so a cookie cannot CSRF the POST."""

    application = FastAPI()
    application.include_router(router, prefix="/directory")
    with TestClient(application) as http:
        http.cookies.set("access_token", "not-a-bearer-token")
        response = http.post("/directory/personnel/control-list/export")
    assert response.status_code == 401
    assert response.json()["detail"] == "Missing Authorization: Bearer token"
