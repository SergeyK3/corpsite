from __future__ import annotations

from datetime import datetime, timezone
from io import BytesIO

import pytest
from openpyxl import load_workbook

from app.main import app
from app.services.personnel_reports_service import (
    PersonnelReportAccessError,
    PersonnelReportFilterError,
    build_personnel_roster,
    build_personnel_roster_xlsx,
    list_report_org_options,
    normalize_position_name,
    personnel_position_rank,
    roster_filename,
)


class _MappingsResult:
    def __init__(self, rows):
        self._rows = rows

    def mappings(self):
        return self

    def all(self):
        return self._rows


class _Connection:
    def __init__(self, option_rows, employee_rows=None):
        self.option_rows = option_rows
        self.employee_rows = employee_rows or []
        self.statements: list[str] = []
        self.params: list[dict] = []

    def execute(self, statement, params=None):
        sql = str(statement)
        self.statements.append(sql)
        self.params.append(params or {})
        rows = self.employee_rows if "WITH current_roster" in sql else self.option_rows
        return _MappingsResult(rows)

    def __enter__(self):
        return self

    def __exit__(self, *_args):
        return False


class _Engine:
    def __init__(self, option_rows, employee_rows=None):
        self.connection = _Connection(option_rows, employee_rows)

    def connect(self):
        return self.connection


OPTIONS = [
    {"unit_id": 10, "unit_name": "Хирургия", "group_id": 1, "group_name": "Клинические"},
    {"unit_id": 11, "unit_name": "Терапия", "group_id": 1, "group_name": "Клинические"},
    {"unit_id": 20, "unit_name": "Лаборатория", "group_id": 2, "group_name": "Параклинические"},
]

EMPLOYEES = [
    {
        "employee_id": 3,
        "group_id": 1,
        "group_name": "Клинические",
        "unit_id": 11,
        "unit_name": "Терапия",
        "full_name": "Абдулова А.А.",
        "position_name": None,
        "position_category": None,
        "rate": 1,
    },
    {
        "employee_id": 1,
        "group_id": 1,
        "group_name": "Клинические",
        "unit_id": 10,
        "unit_name": "Хирургия",
        "full_name": "Беков Б.Б.",
        "position_name": "Врач",
        "position_category": "medical",
        "rate": None,
    },
    {
        "employee_id": 2,
        "group_id": 1,
        "group_name": "Клинические",
        "unit_id": 10,
        "unit_name": "Хирургия",
        "full_name": "Волкова В.В.",
        "position_name": "Врач",
        "position_category": "medical",
        "rate": 0.5,
    },
    {
        "employee_id": 4,
        "group_id": 2,
        "group_name": "Параклинические",
        "unit_id": 20,
        "unit_name": "Лаборатория",
        "full_name": "Громов Г.Г.",
        "position_name": "Лаборант",
        "position_category": "medical",
        "rate": 1,
    },
]


def _build(engine, **filters):
    return build_personnel_roster(
        engine,
        scope_unit_ids=[10, 11, 20],
        generated_at=datetime(2026, 8, 29, 9, 30, tzinfo=timezone.utc),
        **filters,
    )


def test_removed_declaration_export_api_route_is_not_registered():
    paths = {route.path for route in app.routes}
    assert "/directory/personnel/import/batches/{batch_id}/declarations/export" not in paths


def test_options_only_include_visible_departments_in_business_group_order():
    engine = _Engine([OPTIONS[2], OPTIONS[0], OPTIONS[1]])

    result = list_report_org_options(engine, scope_unit_ids=[10, 11, 20])

    assert result["groups"] == [
        {"group_id": 1, "group_name": "Клинические"},
        {"group_id": 2, "group_name": "Параклинические"},
    ]
    assert engine.connection.params == [{"scope_unit_ids": [10, 11, 20]}]
    assert "ou.unit_id = ANY(:scope_unit_ids)" in engine.connection.statements[0]
    assert "ORDER BY ou.group_id" in engine.connection.statements[0]


def test_without_filters_returns_all_visible_groups_summary_and_unique_total():
    report = _build(_Engine(OPTIONS, EMPLOYEES))

    assert [group["id"] for group in report["groups"]] == [1, 2]
    assert [row["department"]["id"] for row in report["summary"]] == [11, 10, 20]
    assert [row["employee_count"] for row in report["summary"]] == [1, 2, 1]
    assert [row["rate_total"] for row in report["summary"]] == [1.0, 0.5, 1.0]
    assert report["total"] == 4
    assert report["total_rate"] == 2.5
    assert report["missing_rate_count"] == 1
    assert report["total"] == len({item["employee_id"] for item in report["items"]})
    assert [
        [item["number"] for item in department["items"]]
        for group in report["groups"]
        for department in group["departments"]
    ] == [[1], [1, 2], [1]]
    assert report["filters"] == {"group": None, "department": None}


def test_group_filter_excludes_other_groups_and_sends_only_group_departments_to_query():
    engine = _Engine(OPTIONS, EMPLOYEES[:3])

    report = _build(engine, group_id=1)

    assert [group["id"] for group in report["groups"]] == [1]
    assert engine.connection.params[1] == {"org_unit_ids": [10, 11]}
    assert report["filters"]["group"] == {"id": 1, "name": "Клинические"}
    assert (report["total"], report["total_rate"]) == (3, 1.5)


def test_department_filter_excludes_other_departments_and_restarts_numbering():
    engine = _Engine(OPTIONS, EMPLOYEES[1:3])

    report = _build(engine, org_unit_id=10)

    assert len(report["summary"]) == 1
    assert report["summary"][0]["employee_count"] == 2
    assert report["summary"][0]["rate_total"] == 0.5
    assert (report["total"], report["total_rate"]) == (2, 0.5)
    assert [item["number"] for item in report["groups"][0]["departments"][0]["items"]] == [1, 2]
    assert engine.connection.params[1] == {"org_unit_ids": [10]}


def test_department_from_another_selected_group_is_rejected():
    with pytest.raises(PersonnelReportFilterError, match="не относится"):
        _build(_Engine(OPTIONS), group_id=1, org_unit_id=20)


@pytest.mark.parametrize("filters", [{"group_id": 99}, {"org_unit_id": 99}])
def test_inaccessible_group_or_department_is_rejected_before_employee_query(filters):
    engine = _Engine(OPTIONS)

    with pytest.raises(PersonnelReportAccessError, match="недоступ"):
        _build(engine, **filters)

    assert len(engine.connection.statements) == 1


def test_roster_query_excludes_inactive_candidates_and_historical_assignments():
    engine = _Engine(OPTIONS, EMPLOYEES)
    _build(engine)

    sql = engine.connection.statements[1]
    assert "e.operational_status = 'active'" in sql
    assert "COALESCE(e.is_active, TRUE) IS TRUE" in sql
    assert "COALESCE(p.person_status, 'active') = 'active'" in sql
    assert "pa.active_flag IS TRUE" in sql
    assert "pa.is_primary IS TRUE" in sql
    assert "pa.lifecycle_status = 'active'" in sql
    assert "pos.category" in sql
    assert "PARTITION BY e.employee_id" in sql
    assert "ORDER BY group_id, LOWER(unit_name)" in sql
    assert "LOWER(full_name)" in sql
    assert "hr_import" not in sql


def test_missing_position_and_rate_are_not_hidden():
    report = _build(_Engine(OPTIONS, EMPLOYEES))

    assert report["items"][0]["position"] == "Не указано"
    assert report["items"][1]["rate"] == "Не указано"
    assert report["items"][1]["rate_value"] is None


def _ranking_rows(group_id, group_name):
    positions = [
        (7, "Яковлев Я.Я.", "Неизвестная должность", "other"),
        (6, "Зорин З.З.", "Санитар", "medical"),
        (5, "Егорова Е.Е.", "Сестра-хозяйка", "technical"),
        (4, "Громов Г.Г.", "Медицинский брат", "medical"),
        (3, "Волкова В.В.", "СТАРШАЯ  МЕДИЦИНСКАЯ-СЕСТРА", "medical"),
        (2, "Беков Б.Б.", "Врач-терапевт", "medical"),
        (1, "Абдулов А.А.", "Врач — заведующий отделением", "leaders"),
    ]
    return [
        {
            "employee_id": employee_id,
            "group_id": group_id,
            "group_name": group_name,
            "unit_id": 10,
            "unit_name": "Отделение",
            "full_name": full_name,
            "position_name": position_name,
            "position_category": category,
            "rate": 0.25,
        }
        for employee_id, full_name, position_name, category in positions
    ]


@pytest.mark.parametrize(
    ("group_id", "group_name"),
    [(1, "Клинические"), (2, "Параклинические")],
)
def test_medical_groups_use_all_seven_position_ranks(group_id, group_name):
    options = [
        {"unit_id": 10, "unit_name": "Отделение", "group_id": group_id, "group_name": group_name}
    ]
    report = build_personnel_roster(
        _Engine(options, _ranking_rows(group_id, group_name)),
        scope_unit_ids=[10],
    )

    assert [item["position"] for item in report["items"]] == [
        "Врач — заведующий отделением",
        "Врач-терапевт",
        "СТАРШАЯ  МЕДИЦИНСКАЯ-СЕСТРА",
        "Медицинский брат",
        "Сестра-хозяйка",
        "Санитар",
        "Неизвестная должность",
    ]
    assert report["total_rate"] == 1.75


def test_specific_nurse_categories_and_unknown_fallback_are_stable():
    assert normalize_position_name("  СТАРШАЯ-МЕДИЦИНСКАЯ  СЕСТРА ") == "старшая медицинская сестра"
    assert personnel_position_rank(
        group_id=1,
        position_name="Старшая медицинская сестра",
        position_category="medical",
    ) == 2
    assert personnel_position_rank(
        group_id=1,
        position_name="Сестра-хозяйка",
        position_category="technical",
    ) == 4
    assert personnel_position_rank(
        group_id=1,
        position_name="Неизвестная должность",
        position_category=None,
    ) == 6


def test_administrative_report_uses_shared_eight_rank_order():
    options = [
        {"unit_id": 30, "unit_name": "Администрация", "group_id": 3, "group_name": "АХО"}
    ]
    rows = [
        {
            "employee_id": employee_id,
            "group_id": 3,
            "group_name": "АХО",
            "unit_id": 30,
            "unit_name": "Администрация",
            "full_name": full_name,
            "position_name": position_name,
            "position_category": category,
            "rate": 1,
        }
        for employee_id, full_name, position_name, category in [
            (1, "Андреев А.А.", "Неизвестная должность", "other"),
            (2, "Борисов Б.Б.", "Уборщик помещений", "service"),
            (3, "Волков В.В.", "Техник", "technical"),
            (4, "Громова Г.Г.", "Секретарь-референт", "admin"),
            (5, "Денисов Д.Д.", "Менеджер", "admin"),
            (6, "Егоров Е.Е.", "Главный специалист", "admin"),
            (7, "Жуков Ж.Ж.", "Заместитель руководителя", "leaders"),
            (8, "Яковлев Я.Я.", "Руководитель отдела", "leaders"),
        ]
    ]

    report = build_personnel_roster(_Engine(options, rows), scope_unit_ids=[30])

    assert [item["full_name"] for item in report["items"]] == [
        "Яковлев Я.Я.",
        "Жуков Ж.Ж.",
        "Егоров Е.Е.",
        "Денисов Д.Д.",
        "Громова Г.Г.",
        "Волков В.В.",
        "Борисов Б.Б.",
        "Андреев А.А.",
    ]


def test_excel_contains_same_summary_totals_and_grouped_employee_rows_as_preview():
    report = _build(_Engine(OPTIONS, EMPLOYEES))
    workbook = load_workbook(BytesIO(build_personnel_roster_xlsx(report)))
    sheet = workbook["Личный состав"]
    values = [tuple(cell.value for cell in row) for row in sheet.iter_rows(min_col=1, max_col=5)]

    assert sheet["A1"].value == "Личный состав"
    assert (
        "№",
        "Группа отделений",
        "Отделение",
        "Количество человек",
        "Количество ставок",
    ) in values
    assert ("ВСЕГО", None, None, report["total"], report["total_rate"]) in values
    assert values.count(("№", "ФИО", "Должность", "Ставка", None)) == len(report["summary"])
    for item in report["items"]:
        expected_rate = item["rate_value"] if item["rate_value"] is not None else "Не указано"
        assert (item["number"], item["full_name"], item["position"], expected_rate, None) in values
    report_names = [item["full_name"] for item in report["items"]]
    excel_names = [
        sheet.cell(row, 2).value
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 2).value in set(report_names)
    ]
    assert excel_names == report_names
    numeric_rate_item = next(item for item in report["items"] if item["rate_value"] == 0.5)
    numeric_rate_row = next(
        row
        for row in range(1, sheet.max_row + 1)
        if sheet.cell(row, 2).value == numeric_rate_item["full_name"]
    )
    assert isinstance(sheet.cell(numeric_rate_row, 4).value, (int, float))
    assert sheet.cell(numeric_rate_row, 4).number_format == "0.##"
    assert any(
        cell.value == "Ставка не указана у 1 сотрудников"
        for row in sheet.iter_rows()
        for cell in row
    )
    assert sheet.page_setup.orientation == "portrait"
    assert sheet.page_setup.fitToWidth == 1
    assert len(sheet.row_breaks.brk) == len(report["summary"]) - 1
    assert "Все_доступные" in roster_filename(report)
