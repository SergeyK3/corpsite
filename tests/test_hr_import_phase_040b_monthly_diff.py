"""Tests for ADR-040 Phase B — monthly diff engine."""
from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_canonical_snapshot_service import (
    build_canonical_snapshot_from_batch,
    canonical_snapshot_available,
)
from app.services.hr_import_monthly_diff_service import (
    DIFF_STATUS_CHANGED,
    DIFF_STATUS_CONFLICT,
    DIFF_STATUS_NEW,
    DIFF_STATUS_REMOVED,
    DIFF_STATUS_UNCHANGED,
    compute_batch_monthly_diff,
    compute_field_diffs,
    get_batch_diff_summary,
    monthly_diff_available,
)
from app.services.hr_import_roster_promotion_service import promote_roster_batch
from app.services.hr_import_service import import_control_list
from tests.conftest import table_exists
from tests.test_hr_import_phase_040a_canonical_snapshot import (
    _cleanup_promotion_batch,
    _cyrillic_full_name,
    _ensure_department_mapping,
    _ensure_roster_employee_metadata,
    _prepare_roster_row,
    _require_phase_040a,
    _test_iin,
)
from tests.test_import_hr_control_list import _build_doctors_sheet


def _require_phase_040b() -> None:
    _require_phase_040a()
    with engine.begin() as conn:
        if not monthly_diff_available(conn):
            pytest.skip("ADR-040 Phase B migration not applied — run alembic upgrade head")


def _unique_iin(seed: str | None = None) -> str:
    token = seed or uuid4().hex
    return _test_iin(token)


def _build_workbook(path: Path, *, full_name: str, iin: str, position: str = "") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=iin)
    if position:
        ws.cell(row=8, column=7, value=position)
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)
    wb.save(path)
    wb.close()


def _import_prepared_batch(
    seed,
    tmp_path: Path,
    *,
    full_name: str,
    iin: str,
    department: str,
    position: str = "",
) -> int:
    source = tmp_path / f"phase040b_{uuid4().hex[:8]}.xlsx"
    _build_workbook(source, full_name=full_name, iin=iin, position=position)
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row_id = conn.execute(
            text(
                """
                SELECT row_id
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).scalar_one()
        _prepare_roster_row(
            conn,
            batch_id=batch_id,
            row_id=int(row_id),
            full_name=full_name,
            iin=iin,
            department=department,
            org_unit_id=int(seed["unit_id"]),
        )
        if position:
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET normalized_payload = jsonb_set(
                        normalized_payload,
                        '{position_raw}',
                        to_jsonb(CAST(:position AS text)),
                        true
                    )
                    WHERE row_id = :row_id
                    """
                ),
                {"row_id": int(row_id), "position": position},
            )
            _ensure_roster_employee_metadata(conn, batch_id=batch_id, row_id=int(row_id))
    return int(batch_id)


def _create_snapshot_from_batch(seed, batch_id: int) -> None:
    with engine.begin() as conn:
        promote_roster_batch(
            conn,
            batch_id,
            created_by=int(seed["initiator_user_id"]),
            dry_run=False,
        )


def test_no_active_snapshot_marks_incoming_as_new(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"NoSnap{suffix}")
    iin = _unique_iin(suffix)
    department = f"Pytest Dept {suffix}"
    batch_id = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
    )
    try:
        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_id)
            assert result["snapshot_id"] is None
            assert result["summary"][DIFF_STATUS_NEW] >= 1
            row = conn.execute(
                text(
                    """
                    SELECT diff_status
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                      AND diff_status IS NOT NULL
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).one()
            assert row[0] == DIFF_STATUS_NEW
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_repeat_import_is_unchanged(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"Same{suffix}")
    iin = _unique_iin(suffix)
    department = f"Pytest Dept {suffix}"
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position="Врач терапевт",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position="Врач терапевт",
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_id_2)
            assert result["summary"][DIFF_STATUS_UNCHANGED] >= 1
            row = conn.execute(
                text(
                    """
                    SELECT diff_status, field_diffs
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id_2},
            ).mappings().one()
            assert row["diff_status"] == DIFF_STATUS_UNCHANGED
            assert row["field_diffs"] is None
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_changed_field_sets_changed_status_and_field_diffs(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"Chg{suffix}")
    iin = _unique_iin(suffix)
    department = f"Pytest Dept {suffix}"
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position="Медицинский техник",
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position="Инженер по медицинскому оборудованию",
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_id_2)
            assert result["summary"][DIFF_STATUS_CHANGED] >= 1
            row = conn.execute(
                text(
                    """
                    SELECT diff_status, field_diffs
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id_2},
            ).mappings().one()
            assert row["diff_status"] == DIFF_STATUS_CHANGED
            field_diffs = row["field_diffs"]
            if isinstance(field_diffs, str):
                field_diffs = json.loads(field_diffs)
            assert "position_raw" in field_diffs
            assert field_diffs["position_raw"]["canonical"] == "Медицинский техник"
            assert field_diffs["position_raw"]["incoming"] == "Инженер по медицинскому оборудованию"
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_new_incoming_record_is_new(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Base{suffix}"),
        iin=_unique_iin(f"1{suffix}"),
        department=department,
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"New{suffix}"),
        iin=_unique_iin(f"2{suffix}"),
        department=department,
    )
    try:
        _create_snapshot_from_batch(seed, batch_id_1)
        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_id_2)
            assert result["summary"][DIFF_STATUS_NEW] >= 1
            assert result["summary"][DIFF_STATUS_REMOVED] >= 1
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_missing_canonical_entry_is_removed(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    full_name = _cyrillic_full_name(f"Rem{suffix}")
    iin = _unique_iin(f"1{suffix}")
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Other{suffix}"),
        iin=_unique_iin(f"2{suffix}"),
        department=department,
    )
    _create_snapshot_from_batch(seed, batch_id_1)
    try:
        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_id_2)
            removed = conn.execute(
                text(
                    """
                    SELECT match_key, record_kind, diff_status, payload
                    FROM public.hr_import_diff_removals
                    WHERE batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id_2},
            ).mappings().all()
            assert result["summary"][DIFF_STATUS_REMOVED] >= 1
            assert len(removed) >= 1
            assert removed[0]["diff_status"] == DIFF_STATUS_REMOVED
            assert removed[0]["record_kind"] == "roster"
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_field_diffs_shape() -> None:
    diffs = compute_field_diffs(
        canonical_payload={"position_raw": "A", "full_name": "Same"},
        incoming_payload={"position_raw": "B", "full_name": "Same"},
        compare_fields=frozenset({"position_raw", "full_name"}),
    )
    assert diffs == {
        "position_raw": {"canonical": "A", "incoming": "B"},
    }


def _patch_snapshot_roster_correction(
    conn,
    *,
    batch_id: int,
    corrected_iin: str,
) -> None:
    updated = conn.execute(
        text(
            """
            UPDATE public.hr_canonical_snapshot_entries AS e
            SET payload = jsonb_set(
                    jsonb_set(e.payload, '{iin}', to_jsonb(CAST(:corrected_iin AS text)), true),
                    '{_canonical_correction_fields}',
                    '["iin"]'::jsonb,
                    true
                )
            FROM public.hr_canonical_snapshots AS s
            WHERE s.snapshot_id = e.snapshot_id
              AND s.source_batch_id = :batch_id
              AND e.record_kind = 'roster'
            RETURNING e.entry_id, e.payload
            """
        ),
        {"batch_id": batch_id, "corrected_iin": corrected_iin},
    ).mappings().all()
    if not updated:
        raise AssertionError(f"no roster snapshot entry found for batch_id={batch_id}")
    entry = updated[0]
    from app.services.hr_canonical_snapshot_service import (
        RECORD_KIND_ROSTER,
        compute_canonical_hash,
    )

    payload = dict(entry["payload"])
    entity_scope = str(
        conn.execute(
            text(
                """
                SELECT match_key
                FROM public.hr_canonical_snapshot_entries
                WHERE entry_id = :entry_id
                """
            ),
            {"entry_id": int(entry["entry_id"])},
        ).scalar_one()
    )
    canonical_hash = compute_canonical_hash(
        record_kind=RECORD_KIND_ROSTER,
        entity_scope=entity_scope,
        payload=payload,
    )
    conn.execute(
        text(
            """
            UPDATE public.hr_canonical_snapshot_entries
            SET canonical_hash = :canonical_hash
            WHERE entry_id = :entry_id
            """
        ),
        {"entry_id": int(entry["entry_id"]), "canonical_hash": canonical_hash},
    )


def test_classify_marks_conflict_on_corrected_fields() -> None:
    from app.services.hr_import_monthly_diff_service import _classify_incoming_status

    status = _classify_incoming_status(
        canonical_entry={
            "canonical_hash": "aaa",
            "payload": {
                "iin": "111111111111",
                "_canonical_correction_fields": ["iin"],
            },
        },
        incoming_hash="bbb",
        canonical_hash="aaa",
        field_diffs={"iin": {"canonical": "111111111111", "incoming": "999999999999"}},
        incoming_base={"iin": "999999999999"},
        incoming_effective={"iin": "999999999999"},
        duplicate_incoming=False,
    )
    assert status == DIFF_STATUS_CONFLICT


def test_conflict_when_excel_disagrees_with_canonical_correction(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"Cfl{suffix}")
    wrong_iin = _unique_iin(f"1{suffix}")
    corrected_iin = _unique_iin(f"9{suffix}")
    department = f"Pytest Dept {suffix}"
    batch_id_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=wrong_iin,
        department=department,
    )
    batch_id_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=wrong_iin,
        department=department,
    )
    try:
        with engine.begin() as conn:
            build_canonical_snapshot_from_batch(
                conn,
                batch_id_1,
                promoted_by=int(seed["initiator_user_id"]),
            )
            _patch_snapshot_roster_correction(
                conn,
                batch_id=batch_id_1,
                corrected_iin=corrected_iin,
            )
            result = compute_batch_monthly_diff(conn, batch_id_2)
            row = conn.execute(
                text(
                    """
                    SELECT diff_status, field_diffs
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id_2},
            ).mappings().one()
            assert row["diff_status"] == DIFF_STATUS_CONFLICT, result
            assert result["summary"][DIFF_STATUS_CONFLICT] >= 1
            field_diffs = row["field_diffs"]
            if isinstance(field_diffs, str):
                field_diffs = json.loads(field_diffs)
            assert "iin" in field_diffs
            assert field_diffs["iin"]["canonical"] == corrected_iin
            assert field_diffs["iin"]["incoming"] == wrong_iin
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_import_runs_diff_on_upload_without_active_snapshot(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    batch_id = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Auto{suffix}"),
        iin=_unique_iin(suffix),
        department=f"Pytest Dept {suffix}",
    )
    try:
        with engine.begin() as conn:
            row = conn.execute(
                text(
                    """
                    SELECT diff_status, diff_computed_at
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().one()
            assert row["diff_status"] == DIFF_STATUS_NEW
            assert row["diff_computed_at"] is not None
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_roster_promotion_dry_run_still_works(seed, tmp_path: Path) -> None:
    _require_phase_040b()
    suffix = uuid4().hex[:8]
    batch_id = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=_cyrillic_full_name(f"Dry{suffix}"),
        iin=_test_iin(suffix),
        department=f"Pytest Dept {suffix}",
    )
    try:
        with engine.begin() as conn:
            preview = promote_roster_batch(
                conn,
                batch_id,
                created_by=int(seed["initiator_user_id"]),
                dry_run=True,
            )
            assert preview["dry_run"] is True
            assert "canonical_snapshot" not in preview
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)
