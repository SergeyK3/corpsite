"""Tests for HR Import Phase 2D/2E — education + training from column M."""
from __future__ import annotations

from decimal import Decimal
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_import_document_parser import (
    parse_education_training_raw,
    parse_training_hours_fragment,
)
from app.services.hr_import_document_parser import parse_education_graduation_fragment
from app.services.hr_import_document_candidate_service import (
    list_document_candidates,
    parse_and_persist_document_candidates,
)
from app.services.hr_import_service import import_control_list
from tests.conftest import table_exists
from tests.test_import_hr_control_list import _build_doctors_sheet, get_layout_profile, parse_sheet_with_profile


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_2d_available() -> bool:
    with engine.begin() as conn:
        if not table_exists(conn, "hr_import_document_candidates"):
            return False
        cols = {
            c["name"]
            for c in conn.execute(
                text(
                    """
                    SELECT column_name AS name
                    FROM information_schema.columns
                    WHERE table_schema = 'public'
                      AND table_name = 'hr_import_document_candidates'
                    """
                )
            ).mappings()
        }
        return "source_field" in cols and "batch_id" in cols


def _require_phase_2d() -> None:
    if not _phase_2d_available():
        pytest.skip("Phase 2D migration not applied — run alembic upgrade head")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _employee_doc_count(conn) -> int:
    if not table_exists(conn, "employee_documents"):
        return 0
    return int(conn.execute(text("SELECT COUNT(*) FROM public.employee_documents")).scalar_one())


def _build_doctors_sheet_with_column_m(path: Path, column_m: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=13, value=column_m)
    wb.save(path)
    wb.close()


def test_doctors_column_m_maps_to_education_training_raw(tmp_path: Path):
    path = tmp_path / "doctors_m.xlsx"
    _build_doctors_sheet_with_column_m(path, "2024 курс ПК 72 ч")

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    rows = parse_sheet_with_profile(wb["врачи"], sheet_type="doctors", profile=get_layout_profile("doctors"))
    wb.close()

    first = rows[0]
    assert first.education_training_raw == "2024 курс ПК 72 ч"
    assert first.training_raw == "2024 курс ПК 72 ч"


def test_parser_education_graduation_from_university_year():
    parsed = parse_education_graduation_fragment("КазНМУ, 1982, специальность Лечебное дело", fragment_index=0)
    assert parsed.proposed_document_type == "EDUCATION_GRADUATION"
    assert parsed.document_kind == "education"
    assert parsed.category == "EDUCATION"
    assert parsed.organization == "КазНМУ"
    assert parsed.parsed_issued_at is not None
    assert parsed.parsed_issued_at.year == 1982


def test_parser_training_hours_from_144_hours():
    parsed = parse_training_hours_fragment("ПК 144 часа", fragment_index=0)
    assert parsed.proposed_document_type == "TRAINING_HOURS"
    assert parsed.document_kind == "training"
    assert parsed.category == "TRAINING"
    assert parsed.parsed_hours == Decimal("144")


def test_parser_mixed_cell_creates_two_candidates():
    parsed = parse_education_training_raw(
        "КазНМУ, 1982, Лечебное дело; повышение квалификации 144 ч"
    )
    assert len(parsed) == 2
    kinds = {item.proposed_document_type for item in parsed}
    assert kinds == {"EDUCATION_GRADUATION", "TRAINING_HOURS"}
    assert all(item.source_field == "education_training_raw" for item in parsed)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_column_m_in_staging_payload(seed, tmp_path: Path):
    _require_phase_2d()
    source = tmp_path / f"phase2d_staging_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(source, "КазНМУ, 1982; ПК 144 ч")

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row = conn.execute(
            text(
                """
                SELECT normalized_payload
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                  AND normalized_payload->>'full_name' = 'Иванов Иван Иванович'
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).mappings().first()
        _delete_batch(conn, batch_id)

    payload = dict(row["normalized_payload"])
    assert payload.get("education_training_raw") == "КазНМУ, 1982; ПК 144 ч"
    assert payload.get("training_raw") == payload.get("education_training_raw")


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_candidates_from_column_m_no_employee_documents(seed, tmp_path: Path):
    _require_phase_2d()
    source = tmp_path / f"phase2d_candidates_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(source, "КазНМУ, 1982; ПК 144 ч")

    with engine.begin() as conn:
        before_docs = _employee_doc_count(conn)
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        after_docs = _employee_doc_count(conn)

        education = list_document_candidates(
            conn, batch_id, document_kind="education", limit=50
        )
        training = list_document_candidates(
            conn, batch_id, document_kind="training", limit=50
        )
        _delete_batch(conn, batch_id)

    assert after_docs == before_docs
    assert education["total"] >= 1
    assert training["total"] >= 1
    assert any(
        item["document_type"] == "EDUCATION_GRADUATION" for item in education["items"]
    )
    assert any(item["document_type"] == "TRAINING_HOURS" for item in training["items"])
    assert all(item["source_field"] == "education_training_raw" for item in education["items"])
    assert all(item["source_field"] == "education_training_raw" for item in training["items"])


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_rebuild_candidates_from_column_m_idempotent(seed, tmp_path: Path):
    _require_phase_2d()
    source = tmp_path / f"phase2d_rebuild_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(source, "ПК 144 часа")

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        first = parse_and_persist_document_candidates(conn, batch_id)
        second = parse_and_persist_document_candidates(conn, batch_id)
        _delete_batch(conn, batch_id)

    assert first["total_candidates"] == second["total_candidates"]
    assert first["training_candidates"] >= 1
