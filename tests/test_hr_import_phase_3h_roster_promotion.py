"""Tests for ADR-039 Phase 3H — roster promotion to directory employees."""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.main import app
from app.services.department_recoding_service import seed_department_recoding
from app.services.hr_import_employee_binding_service import repair_batch_employee_bindings
from app.services.hr_import_normalized_record_service import populate_normalized_records
from app.services.hr_import_promotion_service import BLOCKER_EMPLOYEE_REQUIRED, promote_normalized_records
from app.services.hr_import_roster_promotion_service import (
    OUTCOME_BLOCKED,
    OUTCOME_WOULD_CREATE,
    OUTCOME_WOULD_UPDATE,
    evaluate_roster_promotion,
    promote_roster_batch,
)
from app.services.hr_import_service import import_control_list
from tests.conftest import auth_headers, insert_returning_id, table_exists
from tests.test_employee_documents_routes import _create_employee, _create_position, _phase_1a_available
from tests.test_import_hr_control_list import _build_doctors_sheet


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _require_phase_3h() -> None:
    with engine.begin() as conn:
        if not table_exists(conn, "department_recoding"):
            pytest.skip("department_recoding table missing")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _build_workbook(path: Path, *, full_name: str, iin: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=iin)
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)
    wb.save(path)
    wb.close()


def _test_iin(seed: str) -> str:
    digits = "".join(ch for ch in seed if ch.isdigit())
    return f"{digits:0>12}"[-12:]


def _ensure_department_mapping(conn, *, department: str, org_unit_id: int) -> None:
    seed_department_recoding(conn)
    existing = conn.execute(
        text(
            """
            SELECT id FROM public.department_recoding
            WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:name))
            LIMIT 1
            """
        ),
        {"name": department},
    ).first()
    if existing:
        conn.execute(
            text(
                """
                UPDATE public.department_recoding
                SET org_unit_id = :org_unit_id, org_unit_name = :org_unit_name, is_active = TRUE
                WHERE id = :id
                """
            ),
            {
                "id": int(existing[0]),
                "org_unit_id": org_unit_id,
                "org_unit_name": f"Pytest Unit {org_unit_id}",
            },
        )
    else:
        conn.execute(
            text(
                """
                INSERT INTO public.department_recoding (
                    import_department_name, org_unit_id, org_unit_name, department_group, is_active
                )
                VALUES (:department, :org_unit_id, :org_unit_name, 'CLINICAL', TRUE)
                """
            ),
            {
                "department": department,
                "org_unit_id": org_unit_id,
                "org_unit_name": f"Pytest Unit {org_unit_id}",
            },
        )


def _cyrillic_full_name(prefix: str) -> str:
    """Unique Cyrillic FIO so import classifies the row as EMPLOYEE."""
    token = int(uuid4().hex[:6], 16)
    patronymics = [
        "Петрович",
        "Иванович",
        "Сидорович",
        "Алексеевич",
        "Дмитриевич",
        "Николаевич",
        "Сергеевич",
        "Викторович",
    ]
    return f"{prefix} Тест {patronymics[token % len(patronymics)]}"


def _ensure_roster_employee_metadata(conn, *, batch_id: int, row_id: int) -> None:
    conn.execute(
        text(
            """
            UPDATE public.hr_import_rows
            SET normalized_payload = jsonb_set(
                normalized_payload,
                '{metadata}',
                COALESCE(normalized_payload->'metadata', '{}'::jsonb)
                    || jsonb_build_object(
                        'row_type', 'EMPLOYEE',
                        'is_employee_roster', TRUE,
                        'classification', 'NORMAL',
                        'sheet_type', 'doctors'
                    ),
                true
            )
            WHERE batch_id = :batch_id AND row_id = :row_id
            """
        ),
        {"batch_id": batch_id, "row_id": row_id},
    )


def _prepare_roster_row(
    conn,
    *,
    batch_id: int,
    row_id: int,
    full_name: str,
    iin: str,
    department: str,
    org_unit_id: int,
) -> None:
    _ensure_department_mapping(conn, department=department, org_unit_id=org_unit_id)
    conn.execute(
        text(
            """
            UPDATE public.hr_import_rows
            SET normalized_payload = jsonb_set(
                jsonb_set(
                    jsonb_set(normalized_payload, '{full_name}', to_jsonb(CAST(:full_name AS text)), true),
                    '{iin}',
                    to_jsonb(CAST(:iin AS text)),
                    true
                ),
                '{department}',
                to_jsonb(CAST(:department AS text)),
                true
            )
            WHERE row_id = :row_id AND batch_id = :batch_id
            """
        ),
        {
            "row_id": row_id,
            "batch_id": batch_id,
            "full_name": full_name,
            "iin": iin,
            "department": department,
        },
    )
    _ensure_roster_employee_metadata(conn, batch_id=batch_id, row_id=row_id)
    populate_normalized_records(conn, batch_id)


def _first_roster_row_id(conn, batch_id: int) -> int:
    row = conn.execute(
        text(
            """
            SELECT row_id
            FROM public.hr_import_rows
            WHERE batch_id = :batch_id
            ORDER BY row_id
            LIMIT 1
            """
        ),
        {"batch_id": batch_id},
    ).scalar_one()
    return int(row)


@pytest.fixture
def privileged_headers(seed, monkeypatch):
    monkeypatch.setenv("DIRECTORY_PRIVILEGED_USER_IDS", str(seed["initiator_user_id"]))
    return auth_headers(seed["initiator_user_id"])


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_roster_promotion_creates_employee_and_identity(seed, tmp_path: Path):
    _require_phase_3h()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name("Создание")
    iin = _test_iin(f"1{suffix}")
    department = f"Pytest Dept {suffix}"
    batch_id = None
    emp_id = None

    try:
        source = tmp_path / f"roster_create_{suffix}.xlsx"
        _build_workbook(source, full_name=full_name, iin=iin)
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id = _first_roster_row_id(conn, batch_id)
            _prepare_roster_row(
                conn,
                batch_id=batch_id,
                row_id=row_id,
                full_name=full_name,
                iin=iin,
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )
            preview = evaluate_roster_promotion(conn, batch_id)
            assert preview["items"][0]["outcome"] == OUTCOME_WOULD_CREATE
            result = promote_roster_batch(
                conn,
                batch_id,
                created_by=int(seed["initiator_user_id"]),
                dry_run=False,
            )
            emp_id = int(
                conn.execute(
                    text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                    {"row_id": row_id},
                ).scalar_one()
            )
            identity = conn.execute(
                text(
                    """
                    SELECT identity_value
                    FROM public.employee_identities
                    WHERE employee_id = :employee_id AND identity_type = 'IIN'
                    """
                ),
                {"employee_id": emp_id},
            ).scalar_one()

        assert emp_id > 0
        assert identity == iin
        assert result["dry_run"] is False
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_roster_promotion_updates_existing_employee_by_iin(seed, tmp_path: Path):
    _require_phase_3h()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name("Обновление")
    iin = _test_iin(f"2{suffix}")
    department = f"Pytest Dept Upd {suffix}"
    batch_id = None
    emp_id = None

    try:
        with engine.begin() as conn:
            pos_id = _create_position(conn, name=f"pytest_roster_upd_{suffix}")
            emp_id = _create_employee(
                conn,
                full_name="Old Name",
                org_unit_id=int(seed["unit_id"]),
                position_id=pos_id,
                is_active=True,
            )
            insert_returning_id(
                conn,
                table="employee_identities",
                id_col="identity_id",
                values={
                    "employee_id": emp_id,
                    "identity_type": "IIN",
                    "identity_value": iin,
                    "is_primary": True,
                    "created_by": int(seed["initiator_user_id"]),
                },
            )

        source = tmp_path / f"roster_update_{suffix}.xlsx"
        _build_workbook(source, full_name=full_name, iin=iin)
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id = _first_roster_row_id(conn, batch_id)
            _prepare_roster_row(
                conn,
                batch_id=batch_id,
                row_id=row_id,
                full_name=full_name,
                iin=iin,
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET employee_id = NULL, match_status = 'NOT_PROCESSED'
                    WHERE batch_id = :batch_id AND row_id = :row_id
                    """
                ),
                {"batch_id": batch_id, "row_id": row_id},
            )
            preview = evaluate_roster_promotion(conn, batch_id, row_ids=[row_id])
            assert preview["items"][0]["outcome"] == OUTCOME_WOULD_UPDATE
            promote_roster_batch(
                conn,
                batch_id,
                created_by=int(seed["initiator_user_id"]),
                dry_run=False,
                row_ids=[row_id],
            )
            updated_name = conn.execute(
                text("SELECT full_name FROM public.employees WHERE employee_id = :id"),
                {"id": emp_id},
            ).scalar_one()

        assert updated_name == full_name
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_roster_promotion_blocks_invalid_iin(seed, tmp_path: Path):
    _require_phase_3h()
    suffix = uuid4().hex[:8]
    source = tmp_path / f"roster_block_{suffix}.xlsx"
    blocked_name = _cyrillic_full_name("Блокировка")
    _build_workbook(source, full_name=blocked_name, iin="123")
    batch_id = None
    try:
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id = _first_roster_row_id(conn, batch_id)
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET normalized_payload = jsonb_set(
                        jsonb_set(normalized_payload, '{iin}', to_jsonb('123'::text), true),
                        '{metadata}',
                        COALESCE(normalized_payload->'metadata', '{}'::jsonb)
                            || jsonb_build_object(
                                'row_type', 'EMPLOYEE',
                                'is_employee_roster', TRUE,
                                'classification', 'NORMAL',
                                'sheet_type', 'doctors'
                            ),
                        true
                    )
                    WHERE batch_id = :batch_id AND row_id = :row_id
                    """
                ),
                {"batch_id": batch_id, "row_id": row_id},
            )
            preview = evaluate_roster_promotion(conn, batch_id, row_ids=[row_id])
        assert preview["items"][0]["outcome"] == OUTCOME_BLOCKED
        assert "ИИН" in (preview["items"][0].get("reason") or "")
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_roster_promotion_propagates_employee_id_to_normalized_records(seed, tmp_path: Path):
    _require_phase_3h()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name("Пропагация")
    iin = _test_iin(f"3{suffix}")
    department = f"Pytest Dept Prop {suffix}"
    batch_id = None
    emp_id = None

    try:
        source = tmp_path / f"roster_prop_{suffix}.xlsx"
        _build_workbook(source, full_name=full_name, iin=iin)
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id = _first_roster_row_id(conn, batch_id)
            _prepare_roster_row(
                conn,
                batch_id=batch_id,
                row_id=row_id,
                full_name=full_name,
                iin=iin,
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )
            promote_roster_batch(
                conn,
                batch_id,
                created_by=int(seed["initiator_user_id"]),
                dry_run=False,
            )
            emp_id = conn.execute(
                text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                {"row_id": row_id},
            ).scalar_one()
            records = conn.execute(
                text(
                    """
                    SELECT employee_id
                    FROM public.hr_import_normalized_records
                    WHERE row_id = :row_id
                    """
                ),
                {"row_id": row_id},
            ).fetchall()
        assert records
        assert all(row[0] is not None for row in records)
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_document_dry_run_passes_after_roster_promotion(seed, tmp_path: Path, privileged_headers):
    _require_phase_3h()
    if not _phase_1a_available():
        pytest.skip("employee_documents tables missing")

    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name("Документ")
    iin = _test_iin(f"4{suffix}")
    department = f"Pytest Dept Doc {suffix}"
    batch_id = None
    record_id = None
    emp_id = None

    try:
        source = tmp_path / f"roster_doc_{suffix}.xlsx"
        _build_workbook(source, full_name=full_name, iin=iin)
        with engine.begin() as conn:
            batch_id, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id = _first_roster_row_id(conn, batch_id)
            _prepare_roster_row(
                conn,
                batch_id=batch_id,
                row_id=row_id,
                full_name=full_name,
                iin=iin,
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )
            promote_roster_batch(
                conn,
                batch_id,
                created_by=int(seed["initiator_user_id"]),
                dry_run=False,
            )
            emp_id = conn.execute(
                text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                {"row_id": row_id},
            ).scalar_one()
            repair_batch_employee_bindings(conn, batch_id)
            record_id = conn.execute(
                text(
                    """
                    SELECT normalized_record_id
                    FROM public.hr_import_normalized_records
                    WHERE row_id = :row_id
                    ORDER BY normalized_record_id
                    LIMIT 1
                    """
                ),
                {"row_id": row_id},
            ).scalar_one_or_none()
            if record_id is None:
                pytest.skip("no normalized records for batch row")
            record_id = int(record_id)
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_normalized_records
                    SET review_status = 'approved', reviewed_at = NOW(), reviewed_by = :uid
                    WHERE normalized_record_id = :record_id
                    """
                ),
                {"record_id": record_id, "uid": int(seed["initiator_user_id"])},
            )
            blocked = promote_normalized_records(
                conn,
                promoted_by=int(seed["initiator_user_id"]),
                dry_run=True,
                record_ids=[record_id],
            )

        assert not any(
            blocker.get("code") == BLOCKER_EMPLOYEE_REQUIRED
            for blocker in blocked["items"][0].get("blockers") or []
        )
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})
