"""Tests for ADR-039 Phase 3F.3 — normalized record review override layer."""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_import_normalized_record_service import (
    normalized_records_available,
    review_override_available,
)
from app.services.hr_import_service import import_control_list
from tests.conftest import auth_headers, table_exists
from tests.test_employee_documents_routes import (
    _create_employee,
    _create_position,
    _phase_1a_available,
)
from tests.test_import_hr_control_list import _build_doctors_sheet


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_3f3_available() -> bool:
    with engine.begin() as conn:
        if not normalized_records_available(conn):
            return False
        return review_override_available(conn)


def _require_phase_3f3() -> None:
    if not _phase_3f3_available():
        pytest.skip("ADR-039 Phase 3F.3 migration not applied — run alembic upgrade head")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _build_doctors_sheet_with_column_m(path: Path, column_m: str, column_n: str = "") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=13, value=column_m)
    if column_n:
        ws.cell(row=8, column=14, value=column_n)
    wb.save(path)
    wb.close()


@pytest.fixture
def privileged_headers(seed, monkeypatch):
    monkeypatch.setenv("DIRECTORY_PRIVILEGED_USER_IDS", str(seed["initiator_user_id"]))
    return auth_headers(seed["initiator_user_id"])


@pytest.fixture
def pending_education_record(seed, tmp_path: Path):
    _require_phase_3f3()
    source = tmp_path / f"phase3f3_override_{uuid4().hex[:8]}.xlsx"
    source_text = "КазНМУ, 1982; ПК 144 ч"
    _build_doctors_sheet_with_column_m(
        source,
        source_text,
        "сертификат специалиста, действ. до 01.01.2028",
    )

    emp_ids: list[int] = []
    pos_ids: list[int] = []
    with engine.begin() as conn:
        if not _phase_1a_available():
            pytest.skip("ADR-037 Phase 1A tables missing — run alembic upgrade head")
        pos_id = _create_position(conn, name=f"pytest_override_{uuid4().hex[:8]}")
        pos_ids.append(pos_id)
        emp_id = _create_employee(
            conn,
            full_name=f"Pytest Override {uuid4().hex[:8]}",
            org_unit_id=int(seed["unit_id"]),
            position_id=pos_id,
            is_active=True,
        )
        emp_ids.append(emp_id)
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row = conn.execute(
            text(
                """
                SELECT normalized_record_id, source_text, title, provider, row_id
                FROM public.hr_import_normalized_records
                WHERE batch_id = :batch_id
                  AND record_kind = 'education'
                ORDER BY normalized_record_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).mappings().first()
        if row is None:
            import_row = conn.execute(
                text(
                    """
                    SELECT row_id
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).first()
            assert import_row is not None
            source_record_key = f"pytest-education-override-{uuid4().hex}"
            inserted = conn.execute(
                text(
                    """
                    INSERT INTO public.hr_import_normalized_records (
                        batch_id,
                        row_id,
                        employee_id,
                        fragment_index,
                        source_field,
                        source_text,
                        source_record_key,
                        record_kind,
                        document_type_id,
                        document_type_code,
                        title,
                        provider,
                        issue_date,
                        parse_method,
                        review_status
                    )
                    VALUES (
                        :batch_id,
                        :row_id,
                        :employee_id,
                        0,
                        'education_raw',
                        :source_text,
                        :source_record_key,
                        'education',
                        (SELECT document_type_id FROM public.document_types WHERE code = 'EDUCATION_GRADUATION'),
                        'EDUCATION_GRADUATION',
                        'КазНМУ',
                        'КазНМУ',
                        DATE '1982-01-01',
                        'regex_v1',
                        'pending'
                    )
                    RETURNING normalized_record_id, source_text, title, provider, row_id
                    """
                ),
                {
                    "batch_id": batch_id,
                    "row_id": int(import_row[0]),
                    "employee_id": emp_id,
                    "source_text": source_text,
                    "source_record_key": source_record_key,
                },
            ).mappings().one()
            row = dict(inserted)
        else:
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_normalized_records
                    SET employee_id = :employee_id,
                        review_status = 'pending'
                    WHERE normalized_record_id = :record_id
                    """
                ),
                {"employee_id": emp_id, "record_id": int(row["normalized_record_id"])},
            )

    payload = {
        "batch_id": batch_id,
        "record_id": int(row["normalized_record_id"]),
        "employee_id": emp_id,
        "source_text": str(row["source_text"]),
        "parsed_title": str(row["title"]),
        "parsed_provider": str(row["provider"]),
    }
    yield payload
    with engine.begin() as conn:
        if table_exists(conn, "hr_import_normalized_records"):
            conn.execute(
                text(
                    """
                    DELETE FROM public.hr_import_normalized_records
                    WHERE batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id},
            )
        if emp_ids and table_exists(conn, "employee_documents"):
            conn.execute(
                text("DELETE FROM public.employee_documents WHERE employee_id = ANY(:ids)"),
                {"ids": emp_ids},
            )
        _delete_batch(conn, batch_id)
        if emp_ids and table_exists(conn, "employees"):
            conn.execute(
                text("DELETE FROM public.employees WHERE employee_id = ANY(:ids)"),
                {"ids": emp_ids},
            )
        if pos_ids and table_exists(conn, "positions"):
            conn.execute(
                text("DELETE FROM public.positions WHERE position_id = ANY(:ids)"),
                {"ids": pos_ids},
            )


def _patch_override(
    client: TestClient,
    headers: dict[str, str],
    record_id: int,
    review_override: dict,
) -> dict:
    resp = client.patch(
        f"/directory/personnel/import/normalized-records/{record_id}",
        headers={**headers, "Content-Type": "application/json"},
        json={"review_override": review_override},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


def _patch_review(
    client: TestClient,
    headers: dict[str, str],
    record_id: int,
    review_status: str,
) -> dict:
    resp = client.patch(
        f"/directory/personnel/import/normalized-records/{record_id}",
        headers={**headers, "Content-Type": "application/json"},
        json={"review_status": review_status},
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


def _promote(
    client: TestClient,
    headers: dict[str, str],
    body: dict,
) -> dict:
    resp = client.post(
        "/directory/personnel/import/normalized-records/promote",
        headers={**headers, "Content-Type": "application/json"},
        json=body,
    )
    assert resp.status_code == 200, resp.text
    return resp.json()


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_save_override_for_education(
    client: TestClient,
    privileged_headers,
    pending_education_record,
    seed,
):
    record_id = pending_education_record["record_id"]
    body = _patch_override(
        client,
        privileged_headers,
        record_id,
        {
            "title": "Исправленное название вуза",
            "provider": "Исправленная организация",
            "issue_date": "1982-06-15",
            "document_number": "DOC-123",
        },
    )
    assert body["review_status"] == "pending"
    assert body["title"] == "Исправленное название вуза"
    assert body["provider"] == "Исправленная организация"
    assert body["issue_date"].startswith("1982-06-15")
    assert body["document_number"] == "DOC-123"
    assert body["parsed_values"]["title"] == pending_education_record["parsed_title"]
    assert body["review_override"]["title"] == "Исправленное название вуза"
    assert body["review_override_updated_by"] == int(seed["initiator_user_id"])
    assert body["review_override_updated_at"] is not None


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_list_api_shows_effective_override_values(
    client: TestClient,
    privileged_headers,
    pending_education_record,
):
    record_id = pending_education_record["record_id"]
    _patch_override(
        client,
        privileged_headers,
        record_id,
        {"title": "Drawer Effective Title", "document_number": "OV-1"},
    )
    resp = client.get(
        f"/directory/personnel/import/normalized-records?batch_id={pending_education_record['batch_id']}&limit=50",
        headers=privileged_headers,
    )
    assert resp.status_code == 200, resp.text
    items = resp.json()["items"]
    item = next(row for row in items if row["record_id"] == record_id)
    assert item["title"] == "Drawer Effective Title"
    assert item["document_number"] == "OV-1"
    assert item["parsed_values"]["title"] == pending_education_record["parsed_title"]


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_approval_keeps_override(
    client: TestClient,
    privileged_headers,
    pending_education_record,
):
    record_id = pending_education_record["record_id"]
    _patch_override(
        client,
        privileged_headers,
        record_id,
        {"title": "Approved Override Title"},
    )
    approved = _patch_review(client, privileged_headers, record_id, "approved")
    assert approved["review_status"] == "approved"
    assert approved["title"] == "Approved Override Title"
    assert approved["review_override"]["title"] == "Approved Override Title"


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_promotion_uses_override_instead_of_parsed_value(
    client: TestClient,
    privileged_headers,
    pending_education_record,
):
    record_id = pending_education_record["record_id"]
    _patch_override(
        client,
        privileged_headers,
        record_id,
        {
            "title": "Promoted Override Title",
            "provider": "Promoted Override Org",
        },
    )
    _patch_review(client, privileged_headers, record_id, "approved")

    body = _promote(
        client,
        privileged_headers,
        {"record_ids": [record_id], "dry_run": False},
    )
    assert body["promoted"] == 1
    document_id = body["items"][0]["document_id"]

    with engine.begin() as conn:
        doc = conn.execute(
            text(
                """
                SELECT title, issued_by
                FROM public.employee_documents
                WHERE document_id = :document_id
                """
            ),
            {"document_id": int(document_id)},
        ).mappings().one()
        parsed = conn.execute(
            text(
                """
                SELECT title, provider, review_override_json
                FROM public.hr_import_normalized_records
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        ).mappings().one()

    assert doc["title"] == "Promoted Override Title"
    assert doc["issued_by"] == "Promoted Override Org"
    assert parsed["title"] == pending_education_record["parsed_title"]
    assert parsed["provider"] == pending_education_record["parsed_provider"]
    assert parsed["review_override_json"]["title"] == "Promoted Override Title"


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_source_text_remains_unchanged_after_override_and_promotion(
    client: TestClient,
    privileged_headers,
    pending_education_record,
):
    record_id = pending_education_record["record_id"]
    original_source_text = pending_education_record["source_text"]
    _patch_override(
        client,
        privileged_headers,
        record_id,
        {"title": "Any Override"},
    )
    _patch_review(client, privileged_headers, record_id, "approved")
    _promote(client, privileged_headers, {"record_ids": [record_id], "dry_run": False})

    with engine.begin() as conn:
        row = conn.execute(
            text(
                """
                SELECT source_text
                FROM public.hr_import_normalized_records
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        ).one()
    assert str(row[0]) == original_source_text


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_dry_run_reflects_override_values(
    client: TestClient,
    privileged_headers,
    pending_education_record,
):
    record_id = pending_education_record["record_id"]
    _patch_override(
        client,
        privileged_headers,
        record_id,
        {
            "title": "Dry Run Override Title",
            "document_number": "DR-42",
        },
    )
    _patch_review(client, privileged_headers, record_id, "approved")

    body = _promote(
        client,
        privileged_headers,
        {"record_ids": [record_id], "dry_run": True},
    )
    assert body["would_promote"] == 1
    preview = body["items"][0]["preview"]
    assert preview["title"] == "Dry Run Override Title"
    assert preview["document_number"] == "DR-42"
