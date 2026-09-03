from __future__ import annotations

from datetime import date, datetime
from decimal import Decimal
from hashlib import sha256
from io import BytesIO
from zoneinfo import ZoneInfo

import pytest
from openpyxl import load_workbook

from app.control_list_export import workbook as workbook_module
from app.control_list_export.workbook import (
    ControlListExportLimitError,
    build_control_list_workbook,
)
from app.control_list_projection.schemas import (
    ControlListAcademicDegreeItem,
    ControlListAwardItem,
    ControlListEducationItem,
    ControlListPhoneItem,
    ControlListProjectionMetadata,
    ControlListProjectionResponse,
    ControlListProjectionRow,
    ControlListProjectionScope,
    ControlListTrainingItem,
)

EXPECTED_HEADERS = (
    "№",
    "Группа подразделений",
    "Подразделение",
    "Фамилия, имя, отчество",
    "Дата рождения",
    "ИИН",
    "Занимаемая должность",
    "Категория должности",
    "Ставка",
    "Дата начала в должности",
    "Образование",
    "Год окончания",
    "Специальность по диплому",
    "Повышение квалификации",
    "Учёная степень",
    "Награды",
    "Телефоны",
    "ID сотрудника",
)


def _projection(*, explicit_none: bool = False) -> ControlListProjectionResponse:
    row = ControlListProjectionRow(
        number=1,
        org_group="=External formula",
        org_unit="Отделение 1",
        full_name="+Иванов Иван Иванович",
        birth_date=date(1990, 1, 2),
        iin="001234567890",
        position="Врач",
        position_category="specialist",
        employment_rate=Decimal("0.75"),
        assignment_start_date=date(2024, 6, 7),
        education=[
            ControlListEducationItem(
                record_id=1,
                institution_name="Университет А",
                graduation_year=2012,
                specialty=None,
            ),
            ControlListEducationItem(
                record_id=2,
                institution_name=None,
                graduation_year=None,
                specialty="@Опасная специальность",
            ),
        ],
        training=[
            ControlListTrainingItem(
                record_id=3,
                title="Курс",
                organization_name="Центр",
                hours=Decimal("72"),
                completed_at=date(2025, 5, 6),
                certificate_number="=CERT",
            ),
            ControlListTrainingItem(record_id=4, title="Второй курс"),
        ],
        academic_degrees=[]
        if explicit_none
        else [
            ControlListAcademicDegreeItem(ordinal=0, label="Кандидат наук"),
            ControlListAcademicDegreeItem(ordinal=1, label="Доцент"),
        ],
        academic_degrees_none=explicit_none,
        awards=[]
        if explicit_none
        else [
            ControlListAwardItem(ordinal=0, name="Почётная грамота"),
            ControlListAwardItem(ordinal=1, name="Медаль"),
        ],
        awards_none=explicit_none,
        phones=[
            ControlListPhoneItem(contact_id=1, value="+7 701 000 00 01"),
            ControlListPhoneItem(contact_id=2, value="07010000002"),
        ],
        employee_id=42,
    )
    metadata = ControlListProjectionMetadata(
        schema_version="CONTROL_LIST_EXPORT_V1",
        as_of_date=date(2026, 9, 3),
        generated_at=datetime(2026, 9, 3, 10, 11, 12, tzinfo=ZoneInfo("Asia/Almaty")),
        timezone="Asia/Almaty",
        initiator_user_id=7,
        scope=ControlListProjectionScope(organization_wide=False, org_unit_ids=[10, 20]),
    )
    return ControlListProjectionResponse(metadata=metadata, total=1, items=[row])


def _open(projection: ControlListProjectionResponse | None = None):
    artifact = build_control_list_workbook(
        projection or _projection(), request_id="request-123"
    )
    return artifact, load_workbook(BytesIO(artifact.content), data_only=False)


def test_real_xlsx_has_exact_schema_types_and_formatting() -> None:
    artifact, workbook = _open()
    assert artifact.filename == "Контрольный_список_2026-09-03.xlsx"
    assert artifact.sha256 == sha256(artifact.content).hexdigest()
    assert workbook.sheetnames == ["Контрольный список", "Метаданные"]

    sheet = workbook["Контрольный список"]
    assert tuple(cell.value for cell in sheet[1]) == EXPECTED_HEADERS
    assert sheet.max_column == 18
    assert sheet.freeze_panes == "A2"
    assert sheet.auto_filter.ref == "A1:R2"
    assert sheet.column_dimensions["R"].hidden is True
    assert all(cell.alignment.wrap_text for cell in sheet[1])
    assert all(cell.font.bold for cell in sheet[1])

    assert sheet["A2"].value == 1 and sheet["A2"].data_type == "n"
    assert sheet["E2"].is_date and sheet["E2"].value.date() == date(1990, 1, 2)
    assert sheet["F2"].value == "001234567890" and sheet["F2"].data_type == "s"
    assert sheet["H2"].value == "—"
    assert sheet["I2"].value == 0.75 and sheet["I2"].data_type == "n"
    assert sheet["J2"].is_date and sheet["J2"].value.date() == date(2024, 6, 7)
    assert sheet["R2"].value == "42" and sheet["R2"].data_type == "s"

    assert sheet["K2"].value == "[1] Университет А\n[2] "
    assert sheet["L2"].value == "[1] 2012\n[2] "
    assert sheet["M2"].value == "[1] \n[2] @Опасная специальность"
    assert "\n" in sheet["N2"].value
    assert "\n" in sheet["O2"].value
    assert "\n" in sheet["P2"].value
    assert "\n" in sheet["Q2"].value
    assert sheet["Q2"].value.startswith("'+7 701 000 00 01")
    assert sheet["Q2"].data_type == "s"
    assert sheet["B2"].value == "'=External formula"
    assert sheet["D2"].value == "'+Иванов Иван Иванович"

    metadata = workbook["Метаданные"]
    assert metadata["B2"].is_date
    assert metadata["B3"].is_date
    assert metadata["B8"].value == 1
    assert metadata["B9"].value == "request-123"
    assert "персональные данные" in metadata["B10"].value
    assert "001234567890" not in " ".join(str(cell.value) for row in metadata for cell in row)


@pytest.mark.parametrize(
    ("category", "expected"),
    [
        ("leaders", "Руководители"),
        ("medical", "Медицинские"),
        ("admin", "Административные"),
        ("technical", "Технические"),
        ("other", "Прочие"),
        (None, None),
        ("", None),
        ("unsupported-category", "—"),
    ],
)
def test_position_category_uses_canonical_hr_labels(
    category: str | None,
    expected: str | None,
) -> None:
    projection = _projection()
    row = projection.items[0].model_copy(update={"position_category": category})
    workbook = _open(projection.model_copy(update={"items": [row]}))[1]
    assert workbook["Контрольный список"]["H2"].value == expected


def test_position_category_labels_cover_the_canonical_server_catalog() -> None:
    from app.directory.positions_routes import ALLOWED_CATEGORIES

    assert set(workbook_module._POSITION_CATEGORY_LABELS) == ALLOWED_CATEGORIES


def test_explicit_none_and_unfilled_collections_both_remain_empty_cells() -> None:
    explicit_projection = _projection(explicit_none=True)
    _, explicit_workbook = _open(explicit_projection)
    explicit_sheet = explicit_workbook["Контрольный список"]

    unfilled_row = explicit_projection.items[0].model_copy(
        update={"academic_degrees_none": False, "awards_none": False}
    )
    unfilled_projection = explicit_projection.model_copy(update={"items": [unfilled_row]})
    _, unfilled_workbook = _open(unfilled_projection)
    unfilled_sheet = unfilled_workbook["Контрольный список"]

    assert explicit_sheet["O2"].value is None
    assert explicit_sheet["P2"].value is None
    assert unfilled_sheet["O2"].value is None
    assert unfilled_sheet["P2"].value is None


def test_workbook_content_is_deterministic_at_the_cell_and_style_level() -> None:
    first_artifact, first = _open()
    second_artifact, second = _open()
    assert first_artifact.filename == second_artifact.filename
    for sheet_name in first.sheetnames:
        first_sheet = first[sheet_name]
        second_sheet = second[sheet_name]
        assert list(first_sheet.values) == list(second_sheet.values)
        assert first_sheet.freeze_panes == second_sheet.freeze_panes
        assert first_sheet.auto_filter.ref == second_sheet.auto_filter.ref


def test_export_row_limit_fails_without_truncation(monkeypatch) -> None:
    monkeypatch.setattr(workbook_module, "CONTROL_LIST_EXPORT_MAX_ROWS", 0)
    with pytest.raises(ControlListExportLimitError, match="row limit"):
        build_control_list_workbook(_projection(), request_id="request-123")


def test_export_file_size_limit_fails_without_returning_partial_bytes(monkeypatch) -> None:
    monkeypatch.setattr(workbook_module, "CONTROL_LIST_EXPORT_MAX_BYTES", 1)
    with pytest.raises(ControlListExportLimitError, match="file-size limit"):
        build_control_list_workbook(_projection(), request_id="request-123")


def test_formula_like_request_id_is_text_on_metadata_sheet() -> None:
    artifact = build_control_list_workbook(_projection(), request_id="-formula")
    workbook = load_workbook(BytesIO(artifact.content), data_only=False)
    assert workbook["Метаданные"]["B9"].value == "'-formula"
    assert workbook["Метаданные"]["B9"].data_type == "s"
