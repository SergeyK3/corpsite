"""Tests for ADR-039 Phase 3C — normalized record population from import profile/candidates."""
from __future__ import annotations

import json
from datetime import date
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_import_normalized_record_service import (
    _build_staging_rows_for_profile,
    compute_source_record_key,
    list_normalized_records,
    norm_source_text,
    normalized_records_available,
    normalized_records_summary,
    populate_normalized_records,
    RECORD_KIND_CATEGORY,
    RECORD_KIND_CERTIFICATE,
)
from app.services.hr_import_profile_service import build_import_profile
from app.services.hr_import_service import import_control_list
from tests.conftest import table_exists
from tests.test_import_hr_control_list import _build_doctors_sheet, get_layout_profile, parse_sheet_with_profile

MIXED_CATEGORY_CERT_TEXT = (
    'Высшая до 10.04.2023г. "Гигиена и эпидемиология" '
    'Сертификат "Гигиена-эпидемиология" 29.09.2028 г.'
)

MULTI_CERTIFICATE_TEXT = (
    'Сертификат "Общая врачебная практика" до 14.07.2027г. '
    '2. Сертификат "Онкология взрослаяг" до 01.08.2029'
)


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_3c_available() -> bool:
    with engine.begin() as conn:
        return normalized_records_available(conn)


def _require_phase_3c() -> None:
    if not _phase_3c_available():
        pytest.skip("ADR-039 Phase 3B migration not applied — run alembic upgrade head")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _employee_documents_with_source_count(conn) -> int:
    if not table_exists(conn, "employee_documents"):
        return 0
    return int(
        conn.execute(
            text(
                """
                SELECT COUNT(*)
                FROM public.employee_documents
                WHERE source_normalized_record_id IS NOT NULL
                   OR source_record_key IS NOT NULL
                """
            )
        ).scalar_one()
    )


def _build_doctors_sheet_with_column_m(path: Path, column_m: str, column_n: str = "") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=13, value=column_m)
    if column_n:
        ws.cell(row=8, column=14, value=column_n)
    wb.save(path)
    wb.close()


def test_compute_source_record_key_is_deterministic():
    kwargs = dict(
        row_id=10,
        employee_id=7,
        record_kind="training",
        title="ПК 72 ч",
        issue_date=date(2024, 6, 15),
        end_date=None,
        hours=72,
        document_number="",
        source_field="education_training_raw",
        fragment_index=0,
    )
    first = compute_source_record_key(**kwargs)
    second = compute_source_record_key(**kwargs)
    assert first == second
    assert len(first) == 64


def test_compute_source_record_key_scope_changes_with_employee():
    base = dict(
        row_id=10,
        record_kind="training",
        title="ПК",
        issue_date=None,
        end_date=None,
        hours=72,
        document_number="",
        source_field="education_training_raw",
        fragment_index=0,
    )
    row_scope = compute_source_record_key(**base, employee_id=None)
    emp_scope = compute_source_record_key(**base, employee_id=7)
    assert row_scope != emp_scope


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_populate_on_import_creates_training_and_education(seed, tmp_path: Path):
    _require_phase_3c()
    source = tmp_path / f"phase3c_populate_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(
        source,
        "КазНМУ, 1982; ПК 144 ч",
        "сертификат специалиста, действ. до 01.01.2028",
    )

    with engine.begin() as conn:
        before_docs = _employee_documents_with_source_count(conn)
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        summary = normalized_records_summary(conn, batch_id)
        training = list_normalized_records(conn, batch_id, record_kind="training")
        education = list_normalized_records(conn, batch_id, record_kind="education")
        _delete_batch(conn, batch_id)
        after_docs = _employee_documents_with_source_count(conn)

    assert summary["skipped"] is False
    assert summary["total_records"] >= 2
    assert summary["by_kind"].get("training", 0) >= 1
    assert (
        summary["by_kind"].get("education", 0) + summary["by_kind"].get("certificate", 0)
    ) >= 1
    assert training["total"] >= 1
    first_training = training["items"][0]
    assert first_training["hours"] == 144
    assert first_training["review_status"] == "pending"
    assert first_training["promoted_document_id"] is None
    assert first_training["source_record_key"]
    assert first_training["document_type_code"] == "CONTINUING_EDUCATION"
    assert after_docs == before_docs


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_populate_is_idempotent_for_batch(seed, tmp_path: Path):
    _require_phase_3c()
    source = tmp_path / f"phase3c_idempotent_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(source, "2024 курс ПК 72 ч")

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        first = populate_normalized_records(conn, batch_id)
        second = populate_normalized_records(conn, batch_id)
        total = int(
            conn.execute(
                text(
                    """
                    SELECT COUNT(*)
                    FROM public.hr_import_normalized_records
                    WHERE batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id},
            ).scalar_one()
        )
        _delete_batch(conn, batch_id)

    assert first["skipped"] is False
    assert second["skipped"] is False
    assert first["total_records"] == second["total_records"]
    assert total == first["total_records"]


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_populate_enriches_expiry_from_candidates(seed, tmp_path: Path):
    _require_phase_3c()
    source = tmp_path / f"phase3c_expiry_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(source, "ПК 72 ч", "сертификат, действ. до 15.06.2027")

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        certs = list_normalized_records(conn, batch_id, record_kind="certificate")
        _delete_batch(conn, batch_id)

    assert certs["total"] >= 1
    assert any(item.get("expiry_date") for item in certs["items"])


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_column_m_maps_before_populate(seed, tmp_path: Path):
    _require_phase_3c()
    path = tmp_path / f"phase3c_mapping_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(path, "2024 курс ПК 72 ч")

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    rows = parse_sheet_with_profile(wb["врачи"], sheet_type="doctors", profile=get_layout_profile("doctors"))
    wb.close()
    assert rows[0].education_training_raw == "2024 курс ПК 72 ч"

    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=path,
            imported_by=int(seed["initiator_user_id"]),
        )
        training = list_normalized_records(conn, batch_id, record_kind="training")
        _delete_batch(conn, batch_id)

    assert training["total"] >= 1
    assert training["items"][0]["hours"] == 72


def test_category_mixed_certification_text_staging_dates():
    profile = build_import_profile(
        {
            "full_name": "Test Employee",
            "certification_raw": MIXED_CATEGORY_CERT_TEXT,
        }
    )
    enrichment = {
        "fragment_index": 0,
        "parsed_issued_at": date(2028, 9, 29),
        "parsed_valid_until": date(2023, 4, 10),
        "parse_method": "regex_v1",
        "confidence_score": 0.85,
        "certificate_number": "123",
    }
    candidate_index = {
        (1, "certification_raw", norm_source_text(MIXED_CATEGORY_CERT_TEXT)): enrichment,
        (1, "certification_raw", 0): enrichment,
    }
    rows = _build_staging_rows_for_profile(
        batch_id=1,
        row_id=1,
        employee_id=None,
        profile=profile,
        candidate_index=candidate_index,
        document_type_ids={},
        open_employee_keys=set(),
    )
    category_rows = [row for row in rows if row["record_kind"] == RECORD_KIND_CATEGORY]
    assert category_rows
    cat = category_rows[0]
    assert cat["title"] == "highest"
    assert cat["expiry_date"] == date(2023, 4, 10)
    assert cat["issue_date"] is None


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_populate_category_mixed_certification_text_no_check_violation(seed):
    _require_phase_3c()
    suffix = uuid4().hex[:8]
    payload = {
        "full_name": "Category Mixed Cert",
        "certification_raw": MIXED_CATEGORY_CERT_TEXT,
        "metadata": {
            "sheet_type": "doctors",
            "row_type": "EMPLOYEE",
            "is_employee_roster": True,
            "classification": "NORMAL",
        },
    }
    with engine.begin() as conn:
        batch_id = conn.execute(
            text(
                """
                INSERT INTO public.hr_import_batches (
                    source_type, file_name, imported_by, status,
                    total_rows, valid_rows, error_rows
                )
                VALUES ('HR_CONTROL_LIST', :file_name, :uid, 'PARSED', 1, 1, 0)
                RETURNING batch_id
                """
            ),
            {
                "file_name": f"category_mixed_{suffix}.xlsx",
                "uid": int(seed["initiator_user_id"]),
            },
        ).scalar_one()
        conn.execute(
            text(
                """
                INSERT INTO public.hr_import_rows (
                    batch_id, source_sheet, source_row_number,
                    raw_payload, normalized_payload, match_status
                )
                VALUES (
                    :batch_id, 'doctors', 8,
                    CAST(:payload AS jsonb), CAST(:payload AS jsonb), 'NOT_PROCESSED'
                )
                """
            ),
            {"batch_id": batch_id, "payload": json.dumps(payload)},
        )
        populate_normalized_records(conn, batch_id)
        cat = conn.execute(
            text(
                """
                SELECT title, issue_date, expiry_date
                FROM public.hr_import_normalized_records
                WHERE batch_id = :batch_id
                  AND record_kind = 'category'
                ORDER BY normalized_record_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).mappings().first()
        _delete_batch(conn, batch_id)

    assert cat is not None
    assert cat["title"] == "highest"
    assert cat["issue_date"] is None
    assert cat["expiry_date"] == date(2023, 4, 10)


def test_multi_certificate_text_staging_splits_and_dates():
    profile = build_import_profile(
        {
            "full_name": "Test Employee",
            "certification_raw": MULTI_CERTIFICATE_TEXT,
        }
    )
    bad_enrichment = {
        "fragment_index": 0,
        "parsed_issued_at": date(2029, 1, 1),
        "parsed_valid_until": date(2027, 7, 14),
        "parse_method": "regex_v1",
        "confidence_score": 0.85,
    }
    candidate_index = {
        (1, "certification_raw", norm_source_text(MULTI_CERTIFICATE_TEXT)): bad_enrichment,
        (1, "certification_raw", 0): bad_enrichment,
    }
    rows = _build_staging_rows_for_profile(
        batch_id=1,
        row_id=1,
        employee_id=None,
        profile=profile,
        candidate_index=candidate_index,
        document_type_ids={},
        open_employee_keys=set(),
    )
    certs = [row for row in rows if row["record_kind"] == RECORD_KIND_CERTIFICATE]
    assert len(certs) == 2
    by_title = {row["title"]: row for row in certs}
    first = by_title["Общая врачебная практика"]
    second = by_title["Онкология взрослаяг"]
    assert first["expiry_date"] == date(2027, 7, 14)
    assert first["issue_date"] is None
    assert second["expiry_date"] == date(2029, 8, 1)
    assert second["issue_date"] is None


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_populate_multi_certificate_text_no_check_violation(seed):
    _require_phase_3c()
    suffix = uuid4().hex[:8]
    payload = {
        "full_name": "Multi Certificate",
        "certification_raw": MULTI_CERTIFICATE_TEXT,
        "metadata": {
            "sheet_type": "doctors",
            "row_type": "EMPLOYEE",
            "is_employee_roster": True,
            "classification": "NORMAL",
        },
    }
    with engine.begin() as conn:
        batch_id = conn.execute(
            text(
                """
                INSERT INTO public.hr_import_batches (
                    source_type, file_name, imported_by, status,
                    total_rows, valid_rows, error_rows
                )
                VALUES ('HR_CONTROL_LIST', :file_name, :uid, 'PARSED', 1, 1, 0)
                RETURNING batch_id
                """
            ),
            {
                "file_name": f"multi_cert_{suffix}.xlsx",
                "uid": int(seed["initiator_user_id"]),
            },
        ).scalar_one()
        conn.execute(
            text(
                """
                INSERT INTO public.hr_import_rows (
                    batch_id, source_sheet, source_row_number,
                    raw_payload, normalized_payload, match_status
                )
                VALUES (
                    :batch_id, 'doctors', 8,
                    CAST(:payload AS jsonb), CAST(:payload AS jsonb), 'NOT_PROCESSED'
                )
                """
            ),
            {"batch_id": batch_id, "payload": json.dumps(payload)},
        )
        populate_normalized_records(conn, batch_id)
        certs = conn.execute(
            text(
                """
                SELECT title, issue_date, expiry_date
                FROM public.hr_import_normalized_records
                WHERE batch_id = :batch_id
                  AND record_kind = 'certificate'
                ORDER BY expiry_date
                """
            ),
            {"batch_id": batch_id},
        ).mappings().all()
        _delete_batch(conn, batch_id)

    assert len(certs) == 2
    assert certs[0]["title"] == "Общая врачебная практика"
    assert certs[0]["issue_date"] is None
    assert certs[0]["expiry_date"] == date(2027, 7, 14)
    assert certs[1]["title"] == "Онкология взрослаяг"
    assert certs[1]["issue_date"] is None
    assert certs[1]["expiry_date"] == date(2029, 8, 1)
