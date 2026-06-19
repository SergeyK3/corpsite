"""Phase 2D — employee roster vs declaration sheet/row classification."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from app.services.hr_import_analytics_service import (
    _build_roster_department_index,
    _enrich_declaration_departments,
    _resolve_declaration_department,
    is_missing_iin_employee_row,
    is_real_employee_row,
)
from scripts.import_hr_control_list import (
    ROW_TYPE_CATEGORY_ROW,
    ROW_TYPE_DECLARATION_ROW,
    ROW_TYPE_EMPLOYEE,
    _build_parsed_row,
    infer_row_type,
    is_category_or_summary_label,
    looks_like_person_name,
    parse_workbook,
    resolve_declaration_group,
    resolve_sheet_type,
)


@pytest.mark.parametrize(
    "label",
    [
        "Высшая",
        "Первая",
        "Женщины-285",
        "всего 170",
        "Декрет",
        "Пенсионеры",
        "Доктор медицинских наук",
    ],
)
def test_category_labels_are_not_person_names(label: str):
    assert is_category_or_summary_label(label)
    assert not looks_like_person_name(label)


def test_person_name_detection():
    assert looks_like_person_name("Тулеутаев Мухтар Есетжанович")
    assert looks_like_person_name("Иванов Иван Иванович")
    assert not looks_like_person_name("Высшая")


def test_infer_row_type_category_on_doctors_sheet():
    row_type, is_roster = infer_row_type(full_name="Высшая", sheet_type="doctors", iin_digits="")
    assert row_type == ROW_TYPE_CATEGORY_ROW
    assert is_roster is False


def test_infer_row_type_employee_with_iin():
    row_type, is_roster = infer_row_type(
        full_name="Тулеутаев Мухтар Есетжанович",
        sheet_type="doctors",
        iin_digits="580131300091",
    )
    assert row_type == ROW_TYPE_EMPLOYEE
    assert is_roster is True


@pytest.mark.parametrize(
    "full_name,iin_digits",
    [
        ("Әбитаев Ерхан Сайлаубекулы", "800115300290"),
        ("Қайратов Нурлан", "900101300123"),
        ("Өмірбеков Серик", "900101300124"),
        ("Ілиясов Азамат", "900101300125"),
    ],
)
def test_infer_row_type_kazakh_fio_with_valid_iin(full_name: str, iin_digits: str):
    row_type, is_roster = infer_row_type(
        full_name=full_name,
        sheet_type="doctors",
        iin_digits=iin_digits,
    )
    assert row_type == ROW_TYPE_EMPLOYEE
    assert is_roster is True
    assert looks_like_person_name(full_name)


@pytest.mark.parametrize(
    "label",
    [
        "Категория врачей",
        "Администрация",
        "Отдел кадров",
    ],
)
def test_department_labels_stay_category_row_without_iin(label: str):
    row_type, is_roster = infer_row_type(full_name=label, sheet_type="doctors", iin_digits="")
    assert row_type == ROW_TYPE_CATEGORY_ROW
    assert is_roster is False
    assert not looks_like_person_name(label)


def test_declaration_sheet_type_from_title():
    assert resolve_sheet_type("врачи декларация") == "declaration"
    assert resolve_sheet_type("медсестра декларационные листы") == "declaration"
    assert resolve_sheet_type("врачи") == "doctors"


def test_declaration_group_from_sheet_title():
    assert resolve_declaration_group("врачи декларация") == "doctors"
    assert resolve_declaration_group("медсестра декларация") == "nurses"
    assert resolve_declaration_group("санитарки декларация") == "junior_staff"


def _build_mixed_roster_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)

    doctors = wb.create_sheet("врачи")
    for _ in range(6):
        doctors.append([""] * 14)
    doctors.append(
        [
            "№",
            "",
            "Фамилия, имя, отчество",
            "Год рождения",
            "ИИН",
            "пол",
            "",
            "",
            "",
            "Занимаемая должность",
        ]
        + [""] * 4
    )
    doctors.append(
        [
            1,
            "АДМИН",
            "Тулеутаев Мухтар Есетжанович",
            datetime(1958, 1, 31),
            "580131300091",
            "муж",
            "",
            "",
            "",
            "Директор",
        ]
        + [""] * 4
    )
    for label in ("Высшая", "Первая", "Женщины-285", "всего 170"):
        doctors.append([None, "", label, "", "", "", "", "", "", "", ""] + [""] * 4)

    decl = wb.create_sheet("врачи декларация")
    decl.append(["ФИО", "ИИН", "Категория"])
    decl.append(["Высшая", "", ""])
    decl.append(["Петрова Анна Сергеевна", "900101300123", "первая"])

    wb.save(path)


def test_declaration_sheet_rows_not_in_employee_roster(tmp_path: Path):
    path = tmp_path / "mixed.xlsx"
    _build_mixed_roster_workbook(path)
    rows, _ = parse_workbook(path)

    roster = [r for r in rows if r.is_employee_roster]
    declarations = [r for r in rows if r.sheet_type == "declaration"]

    assert len(roster) == 1
    assert roster[0].full_name == "Тулеутаев Мухтар Есетжанович"
    assert len(declarations) == 2
    assert all(r.declaration_group == "doctors" for r in declarations)
    assert declarations[0].row_type == ROW_TYPE_DECLARATION_ROW

    category_rows = [r for r in rows if r.row_type == ROW_TYPE_CATEGORY_ROW]
    assert len(category_rows) == 4


def test_category_rows_do_not_affect_missing_iin():
    employee = {
        "classification": "NORMAL",
        "sheet_type": "doctors",
        "row_type": ROW_TYPE_EMPLOYEE,
        "is_employee_roster": True,
        "iin": "580131300091",
        "has_training": False,
        "has_certification": False,
        "department": "X",
    }
    category = {
        "classification": "CATEGORY_ROW",
        "sheet_type": "doctors",
        "row_type": ROW_TYPE_CATEGORY_ROW,
        "is_employee_roster": False,
        "iin": "",
        "has_training": False,
        "has_certification": False,
        "department": "X",
        "full_name": "Высшая",
    }
    assert is_real_employee_row(employee)
    assert not is_real_employee_row(category)
    assert not is_missing_iin_employee_row(employee)
    assert not is_missing_iin_employee_row(category)


def test_declaration_department_resolved_from_roster_by_iin_and_name():
    roster_row = {
        "full_name": "Петрова Анна Сергеевна",
        "iin": "900101300123",
        "department": "ХИРУРГИЯ",
        "sheet_type": "doctors",
        "row_type": ROW_TYPE_EMPLOYEE,
        "is_employee_roster": True,
        "classification": "NORMAL",
    }
    declaration_row = {
        "full_name": "Петрова Анна Сергеевна",
        "iin": "900101300123",
        "department": "",
        "sheet_type": "declaration",
        "classification": "DECLARATION",
    }
    by_iin_name, by_iin, by_name = _build_roster_department_index([roster_row])
    resolved = _resolve_declaration_department(
        declaration_row,
        by_iin_name=by_iin_name,
        by_iin=by_iin,
        by_name=by_name,
    )
    assert resolved == "ХИРУРГИЯ"


def test_declaration_department_prefers_iin_name_over_ambiguous_iin():
    items = [
        {
            "full_name": "Иванов Иван Иванович",
            "iin": "900101300123",
            "department": "АДМИН",
            "sheet_type": "doctors",
            "row_type": ROW_TYPE_EMPLOYEE,
            "is_employee_roster": True,
            "classification": "NORMAL",
        },
        {
            "full_name": "Петрова Анна Сергеевна",
            "iin": "900101300123",
            "department": "ХИРУРГИЯ",
            "sheet_type": "doctors",
            "row_type": ROW_TYPE_EMPLOYEE,
            "is_employee_roster": True,
            "classification": "NORMAL",
        },
        {
            "full_name": "Петрова Анна Сергеевна",
            "iin": "900101300123",
            "department": "",
            "sheet_type": "declaration",
            "classification": "DECLARATION",
        },
    ]
    _enrich_declaration_departments(items)
    assert items[2]["department"] == "ХИРУРГИЯ"


def test_build_parsed_row_director_not_missing_iin():
    data = {
        "full_name": "Тулеутаев Мухтар Есетжанович",
        "department": "АДМИН",
        "source_sheet": "врачи",
        "source_row_number": "8",
    }
    for key in (
        "birth_date",
        "sex",
        "nationality",
        "position_raw",
        "education_raw",
        "diploma_specialty_raw",
        "qualification_raw",
        "experience_raw",
        "training_raw",
        "education_training_raw",
        "certification_raw",
        "degree_raw",
        "awards_raw",
        "note_raw",
        "phone_raw",
    ):
        data.setdefault(key, "")
    row = _build_parsed_row(data=data, sheet_type="doctors", iin_raw="580131300091")
    assert row.is_employee_roster
    assert row.iin_digits == "580131300091"
    assert "missing_iin" not in row.errors
