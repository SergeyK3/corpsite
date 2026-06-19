"""ADR-040 / ADR-041 — monthly diff business scenario regression tests.

Synthetic fixtures only (no real PII, no committed Excel files).
Scenarios:
  A — repeat same-month import after canonical snapshot → UNCHANGED, empty review
  B — next-month import → UNCHANGED / NEW / CHANGED / REMOVED
  C — manual review correction preserved → CONFLICT on stale external data
"""
from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_import_analytics_service import list_batch_rows
from app.services.hr_import_monthly_diff_service import (
    DIFF_STATUS_CHANGED,
    DIFF_STATUS_CONFLICT,
    DIFF_STATUS_NEW,
    DIFF_STATUS_REMOVED,
    DIFF_STATUS_UNCHANGED,
    compute_batch_monthly_diff,
    get_batch_diff_summary,
)
from app.services.hr_import_roster_promotion_service import promote_roster_batch
from tests.test_hr_import_phase_040a_canonical_snapshot import (
    _cleanup_promotion_batch,
    _cyrillic_full_name,
    _ensure_roster_employee_metadata,
    _prepare_roster_row,
)
from tests.test_hr_import_phase_040b_monthly_diff import (
    _import_prepared_batch,
    _patch_snapshot_roster_correction,
    _require_phase_040b,
    _unique_iin,
)
from tests.test_import_hr_control_list import _build_doctors_sheet
from app.services.hr_import_service import import_control_list


def _require_business_scenarios() -> None:
    _require_phase_040b()


def _employee_spec(
    *,
    label: str,
    suffix: str,
    position: str = "",
) -> dict[str, str]:
    return {
        "full_name": _cyrillic_full_name(f"{label}{suffix}"),
        "iin": _unique_iin(f"{label}_{suffix}_{uuid4().hex[:8]}"),
        "position": position,
    }


def _build_multi_row_workbook(path: Path, rows: list[dict[str, str]]) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    start_row = 8
    for offset, row in enumerate(rows):
        excel_row = start_row + offset
        ws.cell(row=excel_row, column=3, value=row["full_name"])
        ws.cell(row=excel_row, column=5, value=row["iin"])
        if row.get("position"):
            ws.cell(row=excel_row, column=10, value=row["position"])
    if ws.max_row > start_row + len(rows) - 1:
        ws.delete_rows(start_row + len(rows), ws.max_row - (start_row + len(rows) - 1))
    wb.save(path)
    wb.close()


def _import_multi_employee_batch(
    seed,
    tmp_path: Path,
    *,
    rows: list[dict[str, str]],
    department: str,
    file_tag: str,
) -> int:
    source = tmp_path / f"biz_{file_tag}_{uuid4().hex[:8]}.xlsx"
    _build_multi_row_workbook(source, rows)
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        db_rows = conn.execute(
            text(
                """
                SELECT row_id, normalized_payload->>'full_name' AS full_name
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                """
            ),
            {"batch_id": batch_id},
        ).mappings().all()
        name_to_spec = {r["full_name"]: r for r in rows}
        for db_row in db_rows:
            full_name = str(db_row["full_name"] or "").strip()
            spec = next(
                (row for row in rows if row["full_name"] == full_name),
                None,
            )
            if spec is None:
                continue
            _prepare_roster_row(
                conn,
                batch_id=batch_id,
                row_id=int(db_row["row_id"]),
                full_name=spec["full_name"],
                iin=spec["iin"],
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )
            if spec.get("position"):
                conn.execute(
                    text(
                        """
                        UPDATE public.hr_import_rows
                        SET normalized_payload = jsonb_set(
                            normalized_payload,
                            '{position_raw}',
                            to_jsonb(CAST(:position AS text)),
                            true
                        )
                        WHERE row_id = :row_id
                        """
                    ),
                    {"row_id": int(db_row["row_id"]), "position": spec["position"]},
                )
                _ensure_roster_employee_metadata(
                    conn,
                    batch_id=batch_id,
                    row_id=int(db_row["row_id"]),
                )
    return int(batch_id)


def _promote_and_snapshot(seed, batch_id: int) -> None:
    with engine.begin() as conn:
        promote_roster_batch(
            conn,
            batch_id,
            created_by=int(seed["initiator_user_id"]),
            dry_run=False,
        )


def _row_status_by_iin(conn, batch_id: int, iin: str) -> str:
    return str(
        conn.execute(
            text(
                """
                SELECT diff_status
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                  AND normalized_payload->>'iin' = :iin
                """
            ),
            {"batch_id": batch_id, "iin": iin},
        ).scalar_one()
    )


def test_scenario_a_repeat_june_import_unchanged_and_empty_review(seed, tmp_path: Path) -> None:
    """June original → promoted snapshot → repeat June file → all UNCHANGED, review by exception empty."""
    _require_business_scenarios()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    full_name = _cyrillic_full_name(f"JunA{suffix}")
    iin = _unique_iin(f"juna{suffix}")
    position = "Врач терапевт"

    batch_june_1 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position=position,
    )
    batch_june_2 = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position=position,
    )
    try:
        _promote_and_snapshot(seed, batch_june_1)
        with engine.begin() as conn:
            compute_batch_monthly_diff(conn, batch_june_2)
            diff_summary = get_batch_diff_summary(conn, batch_june_2)
            assert diff_summary["summary"].get(DIFF_STATUS_UNCHANGED, 0) >= 1

            roster_status = conn.execute(
                text(
                    """
                    SELECT diff_status
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_june_2},
            ).scalar_one()
            assert roster_status == DIFF_STATUS_UNCHANGED

            visible_rows = list_batch_rows(conn, batch_june_2, hide_unchanged=True, limit=500)
            assert visible_rows["total"] == 0
            assert visible_rows["items"] == []

            if diff_summary["review_visibility"]["visible_records"] == 0:
                assert diff_summary["review_visibility"]["no_changes_detected"] is True
                assert diff_summary["review_visibility"]["review_complete"] is True
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_june_1)
            _cleanup_promotion_batch(conn, batch_june_2)


def test_scenario_b_next_month_diff_classification(seed, tmp_path: Path) -> None:
    """June canonical → August file: UNCHANGED, NEW, CHANGED, REMOVED."""
    _require_business_scenarios()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"

    keep = _employee_spec(label="Keep", suffix=suffix, position="Врач терапевт")
    changed = _employee_spec(label="Chg", suffix=suffix, position="Медицинский техник")
    removed = _employee_spec(label="Rem", suffix=suffix, position="Врач-хирург")
    new_aug = _employee_spec(label="New", suffix=suffix, position="Старшая медсестра")

    june_rows = [keep, changed, removed]
    august_rows = [
        keep,
        {**changed, "position": "Инженер по медицинскому оборудованию"},
        new_aug,
    ]

    batch_june = _import_multi_employee_batch(
        seed,
        tmp_path,
        rows=june_rows,
        department=department,
        file_tag="june",
    )
    batch_august = _import_multi_employee_batch(
        seed,
        tmp_path,
        rows=august_rows,
        department=department,
        file_tag="august",
    )
    try:
        _promote_and_snapshot(seed, batch_june)
        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_august)
            summary = result["summary"]
            assert summary[DIFF_STATUS_UNCHANGED] >= 1
            assert summary[DIFF_STATUS_NEW] >= 1
            assert summary[DIFF_STATUS_CHANGED] >= 1
            assert summary[DIFF_STATUS_REMOVED] >= 1

            assert _row_status_by_iin(conn, batch_august, keep["iin"]) == DIFF_STATUS_UNCHANGED
            assert _row_status_by_iin(conn, batch_august, changed["iin"]) == DIFF_STATUS_CHANGED
            assert _row_status_by_iin(conn, batch_august, new_aug["iin"]) == DIFF_STATUS_NEW

            changed_row = conn.execute(
                text(
                    """
                    SELECT field_diffs
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                      AND normalized_payload->>'iin' = :iin
                    """
                ),
                {"batch_id": batch_august, "iin": changed["iin"]},
            ).mappings().one()
            field_diffs = changed_row["field_diffs"]
            if isinstance(field_diffs, str):
                field_diffs = json.loads(field_diffs)
            assert field_diffs["position_raw"]["canonical"] == "Медицинский техник"
            assert field_diffs["position_raw"]["incoming"] == "Инженер по медицинскому оборудованию"

            removed_entries = conn.execute(
                text(
                    """
                    SELECT payload->>'iin' AS iin, diff_status
                    FROM public.hr_import_diff_removals
                    WHERE batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_august},
            ).mappings().all()
            removed_iins = {str(r["iin"]) for r in removed_entries}
            assert removed["iin"] in removed_iins
            assert all(r["diff_status"] == DIFF_STATUS_REMOVED for r in removed_entries)
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_june)
            _cleanup_promotion_batch(conn, batch_august)


def test_scenario_c_manual_correction_survives_stale_reimport(seed, tmp_path: Path) -> None:
    """June review fixes position in canonical → promote → stale Excel → CONFLICT, canonical not overwritten."""
    _require_business_scenarios()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    full_name = _cyrillic_full_name(f"ManC{suffix}")
    iin = _unique_iin(f"manc_{suffix}_{uuid4().hex[:8]}")
    wrong_position = "Медицинский техник"
    corrected_position = "Врач терапевт"

    batch_june = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position=wrong_position,
    )
    batch_repeat = _import_prepared_batch(
        seed,
        tmp_path,
        full_name=full_name,
        iin=iin,
        department=department,
        position=wrong_position,
    )
    try:
        _promote_and_snapshot(seed, batch_june)
        with engine.begin() as conn:
            _patch_snapshot_roster_correction(
                conn,
                batch_id=batch_june,
                corrected_position=corrected_position,
            )

        with engine.begin() as conn:
            result = compute_batch_monthly_diff(conn, batch_repeat)
            status = _row_status_by_iin(conn, batch_repeat, iin)
            assert status == DIFF_STATUS_CONFLICT, result
            assert result["summary"][DIFF_STATUS_CONFLICT] >= 1

            row = conn.execute(
                text(
                    """
                    SELECT field_diffs
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                      AND normalized_payload->>'iin' = :iin
                    """
                ),
                {"batch_id": batch_repeat, "iin": iin},
            ).mappings().one()
            field_diffs = row["field_diffs"]
            if isinstance(field_diffs, str):
                field_diffs = json.loads(field_diffs)
            assert field_diffs["position_raw"]["canonical"] == corrected_position
            assert field_diffs["position_raw"]["incoming"] == wrong_position

            canonical_payload = conn.execute(
                text(
                    """
                    SELECT e.payload->>'position_raw' AS position_raw
                    FROM public.hr_canonical_snapshot_entries AS e
                    JOIN public.hr_canonical_snapshots AS s ON s.snapshot_id = e.snapshot_id
                    WHERE s.source_batch_id = :batch_id
                      AND e.record_kind = 'roster'
                      AND e.payload->>'iin' = :iin
                    """
                ),
                {"batch_id": batch_june, "iin": iin},
            ).scalar_one()
            assert canonical_payload == corrected_position
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_june)
            _cleanup_promotion_batch(conn, batch_repeat)
