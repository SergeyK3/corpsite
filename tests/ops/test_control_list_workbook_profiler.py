"""Unit tests for Control List read-only workbook profiler (WP-CL-001)."""
from __future__ import annotations

import json
import re
from datetime import datetime
from pathlib import Path

import pytest
from openpyxl import Workbook

from scripts.ops.control_list_import.header_aliases import match_semantic_field, normalize_header
from scripts.ops.control_list_import.inspect_workbook import main as cli_main
from scripts.ops.control_list_import.value_types import (
    analyze_composite_cell,
    classify_value_type,
    count_composite_records,
    detect_value_issues,
    extract_iin_digits,
    mask_full_name,
    mask_iin,
    mask_phone,
    mask_sample_value,
    sha256_file,
)
from scripts.ops.control_list_import.sheet_classification import classify_sheet_name, is_concurrent_sheet_name
from scripts.ops.control_list_import.workbook_profile import profile_workbook, sheet_exclusion_match


@pytest.fixture
def tmp_xlsx(tmp_path: Path):
    def _factory(name: str = "book.xlsx") -> Path:
        return tmp_path / name

    return _factory


def _save(wb: Workbook, path: Path) -> Path:
    wb.save(path)
    return path


def _profile(path: Path, **kwargs):
    defaults = {"exclusion_terms": ["декларация"]}
    defaults.update(kwargs)
    return profile_workbook(path, **defaults)


def _run_cli(path: Path, tmp_path: Path, **extra_args: str) -> int:
    out_json = tmp_path / "out.json"
    out_md = tmp_path / "out.md"
    args = [
        "--input",
        str(path),
        "--output-json",
        str(out_json),
        "--output-md",
        str(out_md),
    ]
    for key, value in extra_args.items():
        args.extend([f"--{key.replace('_', '-')}", value])
    return cli_main(args)


# --- 1. Header on first row ---


def test_header_on_first_row(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Врачи"
    ws.append(["Фамилия, имя, отчество", "ИИН", "Должность"])
    ws.append(["Тестов Тест Тестович", "900101300123", "Врач"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheet = report["workbook"]["sheets"][0]
    assert sheet["probable_header_row"] == 1
    assert sheet["statistics"]["probable_person_rows"] == 1


# --- 2. Header after title rows ---


def test_header_after_title_rows(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Медсёстры"
    ws.append(["Контрольный список"])
    ws.append(["Организация"])
    ws.append(["Фамилия, имя, отчество", "ИИН", "Подразделение", "Должность"])
    ws.append(["Тестова Анна Сергеевна", "850505400456", "Терапия", "Медсестра"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheet = report["workbook"]["sheets"][0]
    assert sheet["probable_header_row"] == 3


# --- 3. Different header aliases ---


@pytest.mark.parametrize(
    "header,expected_field",
    [
        ("Ф.И.О.", "person.full_name"),
        ("Год рождения", "person.birth_date"),
        ("Контактный телефон", "person.phone"),
        ("Повышение квалификации", "training.records"),
        ("Квалификационная категория", "qualification.category"),
    ],
)
def test_header_alias_variants(header: str, expected_field: str) -> None:
    field, _, confidence = match_semantic_field(header)
    assert field == expected_field
    assert confidence >= 0.55


# --- 4. Inherited department ---


def test_inherited_department_after_section(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.title = "Staff"
    ws.append(["ФИО", "ИИН", "Подразделение", "Должность"])
    ws.append(["Тестов Тест Тестович", "900101300123", "Хирургия", "Врач"])
    ws.append(["Тестова Анна Сергеевна", "850505400456", "", "Медсестра"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheet = report["workbook"]["sheets"][0]
    assert sheet["statistics"]["probable_inherited_section_rows"] == 1
    assert sheet["statistics"]["issue_counts"].get("probable_inherited_section_value") == 1


# --- 5-8. IIN variants ---


def test_iin_as_number(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "ИИН"])
    ws.append(["Тестов Тест Тестович", 900101300123])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][1]
    assert "iin_stored_as_number" in col["issue_counts"]


def test_iin_as_string(tmp_xlsx) -> None:
    digits, issues = extract_iin_digits("900101300123")
    assert digits == "900101300123"
    assert "iin_stored_as_number" not in issues


def test_iin_with_dot() -> None:
    digits, issues = extract_iin_digits("900101300123.0")
    assert digits == "900101300123"
    assert "iin_stored_as_number" in issues


def test_iin_with_spaces() -> None:
    digits, _ = extract_iin_digits("900 101 300 123")
    assert digits == "900101300123"


# --- 9-10. Dates ---


def test_date_as_datetime(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата рождения"])
    ws.append([datetime(1990, 1, 1)])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][0]
    assert col["value_type_distribution"].get("excel_date", 0) >= 1


def test_date_as_text(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Дата рождения"])
    ws.append(["01.01.1990"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][0]
    assert "date_stored_as_text" in col["issue_counts"]


# --- 11-12. Phone ---


def test_phone_as_number(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Телефон"])
    ws.append([77001234567])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][0]
    assert "phone_stored_as_number" in col["issue_counts"]


def test_phone_as_string(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Телефон"])
    ws.append(["+7 700 123 45 67"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][0]
    assert col["value_type_distribution"].get("phone_candidate", 0) >= 1


# --- 13. Composite training cell with dot in title ---


def test_composite_training_dot_in_title() -> None:
    text = (
        '1."Курс A. Продолжение". 60ч., 2021г.\n'
        '2."Курс B", 24ч., 2022г.'
    )
    assert count_composite_records(text) == 2
    assert classify_value_type(text) == "composite_text"


# --- 14. Multiple records without standard newline ---


def test_composite_without_standard_newline() -> None:
    text = '1."Курс A" 60ч., 2021г. 2."Курс B" 24ч., 2022г.'
    assert count_composite_records(text) >= 2


# --- 15-16. Inflated range ---


def test_inflated_max_column(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "ИИН"])
    ws.append(["Тестов Тест Тестович", "900101300123"])
    ws["Z1"].value = " "
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheet = report["workbook"]["sheets"][0]
    assert sheet["excel_max_column"] >= 26
    assert sheet["actual_last_column"] <= 2
    assert sheet["statistics"]["issue_counts"].get("inflated_excel_used_range", 0) >= 1


def test_inflated_max_row(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "ИИН"])
    ws.append(["Тестов Тест Тестович", "900101300123"])
    ws["A50"].value = " "
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheet = report["workbook"]["sheets"][0]
    assert sheet["excel_max_row"] >= 50
    assert sheet["actual_last_row"] == 2
    assert sheet["statistics"]["issue_counts"].get("inflated_excel_used_range", 0) >= 1


# --- 17. Formula ---


def test_formula_cell(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Sum"])
    ws["A2"] = "=1+1"
    path = _save(wb, tmp_xlsx("formula.xlsx"))
    # data_only workbook may show computed value; test typing helper directly
    assert classify_value_type("=1+1", is_formula=True) == "formula"


# --- 18. Excel error ---


def test_excel_error_cell(tmp_xlsx) -> None:
    assert classify_value_type("#N/A") == "error"


# --- 19. Mixed value types ---


def test_mixed_value_types_in_column(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Примечание", ""])
    ws.append(["текст", ""])
    ws.append([123, ""])
    ws.append([datetime(2020, 1, 1), ""])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][0]
    assert col["issue_counts"].get("mixed_value_types_in_column", 0) >= 1


# --- 20. Declaration sheet excluded ---


def test_declaration_sheet_excluded(tmp_xlsx) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Врачи"
    ws1.append(["ФИО", "ИИН"])
    ws1.append(["Тестов Тест Тестович", "900101300123"])
    ws2 = wb.create_sheet("Декларация врачей")
    ws2.append(["ФИО", "ИИН"])
    ws2.append(["Секретный Человек", "111111111111"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    assert report["workbook"]["sheet_count"] == 2
    assert report["workbook"]["excluded_sheet_count"] == 1
    assert report["workbook"]["analyzed_sheet_count"] == 1
    excluded = next(s for s in report["workbook"]["sheets"] if s["status"] == "excluded")
    assert excluded["exclusion_reason"] == "sheet_name_declaration"
    assert "probable_header_row" not in excluded
    assert "columns" not in excluded
    assert report["summary"]["probable_person_rows"] == 1
    assert report["summary"]["issues_by_code"].get("iin_stored_as_number", 0) == 0


# --- 21. Case-insensitive exclusion ---


@pytest.mark.parametrize(
    "sheet_name",
    ["ДЕКЛАРАЦИЯ", "декларация совместителей"],
)
def test_case_insensitive_exclusion(sheet_name: str) -> None:
    excluded, reason, term = sheet_exclusion_match(sheet_name, ["декларация"])
    assert excluded is True
    assert reason == "sheet_name_declaration"


# --- 22. Sheet without confident header ---


def test_sheet_without_confident_header(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["A", "B"])
    ws.append(["1", "2"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheet = report["workbook"]["sheets"][0]
    assert sheet["statistics"]["issue_counts"].get("header_not_confident", 0) >= 1


# --- 23-25. Masking ---


def test_mask_iin() -> None:
    masked = mask_iin("900101300123")
    assert masked == "******0123"
    assert "900101300123" not in masked


def test_mask_phone() -> None:
    masked = mask_phone("77001234567")
    assert "77001234567" not in masked
    assert masked.endswith("567")


def test_mask_full_name() -> None:
    masked = mask_full_name("Тестов Тест Тестович")
    assert "Тестов" not in masked
    assert masked.startswith("Т")


# --- 26. SHA-256 unchanged ---


def test_sha256_unchanged_via_cli(tmp_xlsx, tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "ИИН"])
    ws.append(["Тестов Тест Тестович", "900101300123"])
    path = _save(wb, tmp_xlsx())
    before = sha256_file(str(path))
    code = _run_cli(path, tmp_path)
    after = sha256_file(str(path))
    assert before == after
    assert code == 0
    out_json = tmp_path / "out.json"
    payload = json.loads(out_json.read_text(encoding="utf-8"))
    assert payload["source"]["unchanged"] is True


def test_report_json_has_no_full_pii(tmp_xlsx, tmp_path: Path) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Фамилия, имя, отчество", "ИИН", "Телефон"])
    ws.append(["Тестов Тест Тестович", "900101300123", "77001234567"])
    path = _save(wb, tmp_xlsx())
    _run_cli(path, tmp_path)
    raw = (tmp_path / "out.json").read_text(encoding="utf-8")
    assert "900101300123" not in raw
    assert "77001234567" not in raw
    assert "Тестов Тест Тестович" not in raw
    assert re.search(r"\*{3,}", raw)


def test_mask_iin_in_non_iin_column_sample() -> None:
    masked = mask_sample_value("990628450399", value_type="text", semantic_field="person.sex")
    assert "990628450399" not in masked
    assert masked.endswith("0399")


def test_normalize_header_whitespace() -> None:
    assert normalize_header("  ИИН  ") == "иин"


# --- Composite detection regressions ---


def test_position_two_lines_not_composite() -> None:
    text = "Заведующий\nхирургическим отделением"
    analysis = analyze_composite_cell(text)
    assert not analysis.is_composite
    assert classify_value_type(text) != "composite_text"


def test_qualification_category_with_year_not_composite() -> None:
    text = "высшая категория, 2020 г."
    assert not analyze_composite_cell(text).is_composite


def test_single_course_not_composite() -> None:
    text = '1."Курс A". 60ч., 2021г.'
    assert not analyze_composite_cell(text).is_composite


def test_two_numbered_courses_composite() -> None:
    text = '1."Курс A". 60ч., 2021г.\n2."Курс B", 24ч., 2022г.'
    analysis = analyze_composite_cell(text)
    assert analysis.is_composite
    assert analysis.record_count == 2
    assert "multiple_numbered_records_in_cell" in analysis.issues
    assert "composite_text_without_standard_line_breaks" not in analysis.issues


def test_two_numbered_inline_both_issues() -> None:
    text = '1."Курс A" 60ч., 2021г. 2."Курс B" 24ч., 2022г.'
    analysis = analyze_composite_cell(text)
    assert analysis.is_composite
    assert "multiple_numbered_records_in_cell" in analysis.issues
    assert "composite_text_without_standard_line_breaks" in analysis.issues


def test_category_line_with_leading_one_not_composite() -> None:
    text = "1 квалификационная категория"
    assert not analyze_composite_cell(text).is_composite


# --- IIN issue context ---


def test_iin_number_in_iin_column_single_issue(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ИИН"])
    ws.append([900101300123])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    col = report["workbook"]["sheets"][0]["columns"][0]
    assert col["issue_counts"].get("iin_stored_as_number", 0) == 1
    assert report["summary"]["issues_by_code"]["iin_stored_as_number"] == 1


def test_iin_in_sex_column_not_counted_as_iin_issue(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["пол", "Примечание"])
    ws.append(["муж", "900101300123"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    assert report["summary"]["issues_by_code"].get("iin_stored_as_number", 0) == 0
    assert report["summary"]["issues_by_code"].get("iin_not_12_digits", 0) == 0


def test_phone_in_phone_column_not_iin_issue(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["Телефон"])
    ws.append([77001234567])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    assert report["summary"]["issues_by_code"].get("iin_stored_as_number", 0) == 0
    assert report["summary"]["issues_by_code"].get("phone_stored_as_number", 0) == 1


def test_invalid_iin_in_iin_column(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ИИН"])
    ws.append(["12345"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    assert report["summary"]["issues_by_code"].get("iin_not_12_digits", 0) == 1


def test_unique_issue_per_cell(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ИИН"])
    ws.append([900101300123])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    diag = report["diagnostics"]
    assert diag["duplicate_issue_emissions"] == {}
    assert report["summary"]["issues_by_code"]["iin_stored_as_number"] == 1


def test_diagnostics_section_present(tmp_xlsx) -> None:
    wb = Workbook()
    ws = wb.active
    ws.append(["ФИО", "ИИН"])
    ws.append(["Тестов Тест Тестович", "900101300123"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    assert "diagnostics" in report
    assert "issues_by_sheet" in report["diagnostics"]
    assert "department_diagnostics" in report["diagnostics"]


# --- Sheet classification ---


@pytest.mark.parametrize(
    "sheet_name,category,mode",
    [
        ("врачи", "doctor", "primary"),
        ("врачи совместители", "doctor", "concurrent"),
        ("медсестра", "nursing_staff", "primary"),
        ("СМР совместители", "nursing_staff", "concurrent"),
        ("санитарки", "junior_medical_staff", "primary"),
        ("санитарки совмест", "junior_medical_staff", "concurrent"),
        ("прочее", "other_staff", "primary"),
        ("прочее совм.", "other_staff", "concurrent"),
    ],
)
def test_sheet_classification_expected_pairs(sheet_name: str, category: str, mode: str) -> None:
    cls = classify_sheet_name(sheet_name)
    assert cls.proposed_personnel_category == category
    assert cls.proposed_employment_mode == mode
    assert cls.proposed_sheet_purpose == "personnel_control_list"


def test_sheet_classification_case_insensitive() -> None:
    cls = classify_sheet_name("ВРАЧИ СОВМЕСТИТЕЛИ")
    assert cls.proposed_personnel_category == "doctor"
    assert cls.proposed_employment_mode == "concurrent"


def test_sheet_classification_whitespace_normalization() -> None:
    cls = classify_sheet_name("  врачи   совместители  ")
    assert cls.proposed_employment_mode == "concurrent"


def test_concurrent_not_matched_inside_unrelated_word() -> None:
    is_conc, _ = is_concurrent_sheet_name("пересмотр кадров")
    assert is_conc is False
    cls = classify_sheet_name("пересмотр кадров")
    assert cls.proposed_employment_mode == "unknown"


def test_unknown_sheet_classification() -> None:
    cls = classify_sheet_name("Sheet1")
    assert cls.proposed_personnel_category == "unknown"
    assert cls.proposed_employment_mode == "unknown"
    assert cls.proposed_sheet_purpose == "unknown"


def test_declaration_sheet_classification(tmp_xlsx) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "Врачи"
    ws1.append(["ФИО", "ИИН"])
    ws1.append(["Тестов Тест Тестович", "900101300123"])
    ws2 = wb.create_sheet("Декларация врачей")
    ws2.append(["ФИО", "ИИН"])
    ws2.append(["Тестов Тест Тестович", "900101300123"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    excluded = next(s for s in report["workbook"]["sheets"] if s["status"] == "excluded")
    cls = excluded["classification"]
    assert cls["proposed_sheet_purpose"] == "declaration"
    assert cls["proposed_employment_mode"] == "unknown"
    assert report["summary"]["rows_by_employment_mode"]["concurrent"] == 0


def test_employment_mode_aggregate_invariant(tmp_xlsx) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "врачи"
    ws1.append(["ФИО", "ИИН"])
    ws1.append(["Тестов Тест Тестович", "900101300123"])
    ws2 = wb.create_sheet("врачи совместители")
    ws2.append(["ФИО", "ИИН"])
    ws2.append(["Тестов Тест Тестович", "900101300123"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    summary = report["summary"]
    emp = summary["rows_by_employment_mode"]
    cat = summary["rows_by_personnel_category"]
    total = summary["probable_person_rows"]
    assert emp["primary"] + emp["concurrent"] + emp["unknown"] == total
    assert cat["doctor"] == total
    assert emp["primary"] == 1
    assert emp["concurrent"] == 1


def test_same_person_on_two_sheets_not_merged_as_person_attribute(tmp_xlsx) -> None:
    wb = Workbook()
    ws1 = wb.active
    ws1.title = "врачи"
    ws1.append(["ФИО", "ИИН"])
    ws1.append(["Тестов Тест Тестович", "900101300123"])
    ws2 = wb.create_sheet("врачи совместители")
    ws2.append(["ФИО", "ИИН"])
    ws2.append(["Тестов Тест Тестович", "900101300123"])
    path = _save(wb, tmp_xlsx())
    report = _profile(path)
    sheets = report["workbook"]["sheets"]
    primary = next(s for s in sheets if s["sheet_name"] == "врачи")
    concurrent = next(s for s in sheets if s["sheet_name"] == "врачи совместители")
    assert primary["classification"]["proposed_employment_mode"] == "primary"
    assert concurrent["classification"]["proposed_employment_mode"] == "concurrent"
    assert "employment_mode" not in (primary.get("statistics") or {})
    assert report["summary"]["probable_person_rows"] == 2
