"""Tests for ADR-040 Phase H — HR change events Excel export."""
from __future__ import annotations

from io import BytesIO
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook, load_workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_canonical_snapshot_service import build_canonical_snapshot_from_batch
from app.services.hr_change_events_export_service import (
    CHANGE_TYPE_CHANGED,
    CHANGE_TYPE_CONFLICT,
    CHANGE_TYPE_NEW,
    CHANGE_TYPE_REMOVED,
    EXPORT_COLUMNS,
    EXPORT_SHEETS,
    build_hr_change_events_export_rows,
    export_hr_change_events_xlsx,
)
from app.services.hr_import_monthly_diff_service import compute_batch_monthly_diff
from app.services.hr_import_service import import_control_list
from app.services.hr_snapshot_comparison_service import (
    EVENT_TYPE_NEW,
    EVENT_TYPE_REMOVED,
    hr_change_events_available,
)
from tests.test_hr_import_phase_040a_canonical_snapshot import (
    _cleanup_promotion_batch,
    _cyrillic_full_name,
    _ensure_roster_employee_metadata,
    _prepare_roster_row,
    _require_phase_040a,
    _test_iin,
)
from tests.test_hr_import_phase_040b_monthly_diff import _patch_snapshot_roster_correction
from tests.test_hr_import_phase_040f_hr_change_events import (
    _create_snapshot_from_batch,
    _get_snapshots_for_batches,
    _import_prepared_batch,
)
from tests.test_import_hr_control_list import _build_doctors_sheet


def _require_phase_040h() -> None:
    _require_phase_040a()
    with engine.begin() as conn:
        if not hr_change_events_available(conn):
            pytest.skip("ADR-040 Phase F migration not applied — run alembic upgrade head")


def _unique_iin(seed: str | None = None) -> str:
    token = seed or uuid4().hex
    return _test_iin(token)


def _load_workbook_rows(wb, sheet_name: str) -> list[list]:
    ws = wb[sheet_name]
    return [list(row) for row in ws.iter_rows(values_only=True)]


def _sheet_data_rows(wb, sheet_name: str) -> list[list]:
    rows = _load_workbook_rows(wb, sheet_name)
    if len(rows) <= 1:
        return []
    return rows[1:]


def _change_types_in_sheet(wb, sheet_name: str) -> set[str]:
    rows = _sheet_data_rows(wb, sheet_name)
    if not rows:
        return set()
    return {str(row[0]) for row in rows if row and row[0]}


def _seed_two_snapshot_scenario(seed, tmp_path: Path) -> tuple[int, int, int, int]:
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    org_unit_id = int(seed["unit_id"])

    stable_name = _cyrillic_full_name(f"Stable{suffix}")
    stable_iin = _unique_iin(f"1{suffix}")
    removed_name = _cyrillic_full_name(f"Removed{suffix}")
    removed_iin = _unique_iin(f"2{suffix}")
    new_name = _cyrillic_full_name(f"Added{suffix}")
    new_iin = _unique_iin(f"3{suffix}")

    batch_id_1 = None
    source_1 = tmp_path / f"phase040h_batch1_{uuid4().hex[:8]}.xlsx"
    wb1 = Workbook()
    wb1.remove(wb1.active)
    ws1 = wb1.create_sheet("врачи")
    _build_doctors_sheet(ws1)
    ws1.cell(row=8, column=3, value=stable_name)
    ws1.cell(row=8, column=5, value=stable_iin)
    ws1.cell(row=8, column=6, value=department)
    ws1.cell(row=8, column=7, value="Медицинский техник")
    ws1.cell(row=9, column=3, value=removed_name)
    ws1.cell(row=9, column=5, value=removed_iin)
    ws1.cell(row=9, column=6, value=department)
    wb1.save(source_1)
    wb1.close()

    with engine.begin() as conn:
        batch_id_1, _, _ = import_control_list(
            conn,
            file_path=source_1,
            imported_by=int(seed["initiator_user_id"]),
        )
        for row in conn.execute(
            text(
                """
                SELECT row_id, normalized_payload
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                """
            ),
            {"batch_id": batch_id_1},
        ).mappings().all():
            payload = dict(row["normalized_payload"] or {})
            row_id = int(row["row_id"])
            _prepare_roster_row(
                conn,
                batch_id=batch_id_1,
                row_id=row_id,
                full_name=str(payload.get("full_name") or ""),
                iin=str(payload.get("iin") or ""),
                department=department,
                org_unit_id=org_unit_id,
            )
            _ensure_roster_employee_metadata(conn, batch_id=batch_id_1, row_id=row_id)
            if str(payload.get("iin") or "") == stable_iin:
                conn.execute(
                    text(
                        """
                        UPDATE public.hr_import_rows
                        SET normalized_payload = jsonb_set(
                            normalized_payload,
                            '{position_raw}',
                            to_jsonb(CAST(:position AS text)),
                            true
                        )
                        WHERE row_id = :row_id
                        """
                    ),
                    {"row_id": row_id, "position": "Медицинский техник"},
                )
                _ensure_roster_employee_metadata(conn, batch_id=batch_id_1, row_id=row_id)

    source_2 = tmp_path / f"phase040h_batch2_{uuid4().hex[:8]}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=stable_name)
    ws.cell(row=8, column=5, value=stable_iin)
    ws.cell(row=8, column=6, value=department)
    ws.cell(row=8, column=7, value="Инженер по медицинскому оборудованию")
    ws.cell(row=9, column=3, value=new_name)
    ws.cell(row=9, column=5, value=new_iin)
    ws.cell(row=9, column=6, value=department)
    ws.cell(row=9, column=7, value="Медицинский техник")
    wb.save(source_2)
    wb.close()

    with engine.begin() as conn:
        batch_id_2, _, _ = import_control_list(
            conn,
            file_path=source_2,
            imported_by=int(seed["initiator_user_id"]),
        )
        for row in conn.execute(
            text(
                """
                SELECT row_id, normalized_payload
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                """
            ),
            {"batch_id": batch_id_2},
        ).mappings().all():
            row_id = int(row["row_id"])
            row_iin = str(payload.get("iin") or "")
            _prepare_roster_row(
                conn,
                batch_id=batch_id_2,
                row_id=row_id,
                full_name=str(payload.get("full_name") or ""),
                iin=row_iin,
                department=department,
                org_unit_id=org_unit_id,
            )
            _ensure_roster_employee_metadata(conn, batch_id=batch_id_2, row_id=row_id)
            position = None
            if row_iin == stable_iin:
                position = "Инженер по медицинскому оборудованию"
            elif row_iin == new_iin:
                position = "Медицинский техник"
            if position:
                conn.execute(
                    text(
                        """
                        UPDATE public.hr_import_rows
                        SET normalized_payload = jsonb_set(
                            normalized_payload,
                            '{position_raw}',
                            to_jsonb(CAST(:position AS text)),
                            true
                        )
                        WHERE row_id = :row_id
                        """
                    ),
                    {"row_id": row_id, "position": position},
                )
                _ensure_roster_employee_metadata(conn, batch_id=batch_id_2, row_id=row_id)

    _create_snapshot_from_batch(seed, batch_id_1)
    _create_snapshot_from_batch(seed, batch_id_2)
    prior_id, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
    return batch_id_1, batch_id_2, prior_id, new_id


def test_export_includes_new_changed_removed_sheets(seed, tmp_path: Path) -> None:
    _require_phase_040h()
    suffix = uuid4().hex[:8]
    department = f"Pytest Dept {suffix}"
    org_unit_id = int(seed["unit_id"])
    batch_id_1 = None
    batch_id_2 = None
    try:
        batch_id_1 = _import_prepared_batch(
            seed,
            tmp_path,
            full_name=_cyrillic_full_name(f"Base{suffix}"),
            iin=_unique_iin(f"1{suffix}"),
            department=department,
            org_unit_id=org_unit_id,
        )
        _create_snapshot_from_batch(seed, batch_id_1)
        batch_id_2 = _import_prepared_batch(
            seed,
            tmp_path,
            full_name=_cyrillic_full_name(f"NewOnly{suffix}"),
            iin=_unique_iin(f"2{suffix}"),
            department=department,
            org_unit_id=org_unit_id,
        )
        _create_snapshot_from_batch(seed, batch_id_2)
        _, new_id = _get_snapshots_for_batches(batch_id_1, batch_id_2)
        with engine.begin() as conn:
            rows = build_hr_change_events_export_rows(conn, new_snapshot_id=new_id)
        change_types = {row["change_type"] for row in rows}
        assert CHANGE_TYPE_NEW in change_types
        assert CHANGE_TYPE_REMOVED in change_types
    finally:
        if batch_id_1 is not None and batch_id_2 is not None:
            with engine.begin() as conn:
                _cleanup_promotion_batch(conn, batch_id_1)
                _cleanup_promotion_batch(conn, batch_id_2)


def test_export_workbook_has_expected_sheets(seed, tmp_path: Path) -> None:
    _require_phase_040h()
    batch_id_1, batch_id_2, prior_id, new_id = _seed_two_snapshot_scenario(seed, tmp_path)
    try:
        with engine.begin() as conn:
            content, filename = export_hr_change_events_xlsx(
                conn,
                new_snapshot_id=new_id,
            )
        assert filename.endswith(".xlsx")
        wb = load_workbook(BytesIO(content))
        assert wb.sheetnames == list(EXPORT_SHEETS)

        summary_rows = _load_workbook_rows(wb, "SUMMARY")
        assert summary_rows[0] == ["metric", "value"]
        assert any(row[0] == "NEW" for row in summary_rows)
        assert any(row[0] == "CHANGED" for row in summary_rows)
        assert any(row[0] == "REMOVED" for row in summary_rows)

        headers = [header for _, header in EXPORT_COLUMNS]
        for sheet_name in ("NEW", "CHANGED", "REMOVED", "CONFLICT"):
            assert list(_load_workbook_rows(wb, sheet_name)[0]) == headers
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_export_excludes_unchanged_and_includes_change_types(seed, tmp_path: Path) -> None:
    _require_phase_040h()
    batch_id_1, batch_id_2, prior_id, new_id = _seed_two_snapshot_scenario(seed, tmp_path)
    try:
        with engine.begin() as conn:
            rows = build_hr_change_events_export_rows(conn, new_snapshot_id=new_id)
            content, _ = export_hr_change_events_xlsx(conn, new_snapshot_id=new_id)

        wb = load_workbook(BytesIO(content))
        change_types = {row["change_type"] for row in rows}
        assert "UNCHANGED" not in change_types
        assert CHANGE_TYPE_CHANGED in change_types
        assert CHANGE_TYPE_REMOVED in change_types
        assert len(_sheet_data_rows(wb, "CHANGED")) >= 1
        assert len(_sheet_data_rows(wb, "REMOVED")) >= 1
        assert _change_types_in_sheet(wb, "CHANGED") == {CHANGE_TYPE_CHANGED}
        assert _change_types_in_sheet(wb, "REMOVED") == {CHANGE_TYPE_REMOVED}
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_export_filters_are_applied(seed, tmp_path: Path) -> None:
    _require_phase_040h()
    batch_id_1, batch_id_2, prior_id, new_id = _seed_two_snapshot_scenario(seed, tmp_path)
    try:
        with engine.begin() as conn:
            new_event = conn.execute(
                text(
                    """
                    SELECT department, full_name
                    FROM public.hr_change_events
                    WHERE new_snapshot_id = :new_snapshot_id
                      AND event_type = :event_type
                    LIMIT 1
                    """
                ),
                {"new_snapshot_id": new_id, "event_type": EVENT_TYPE_NEW},
            ).mappings().first()
            if new_event is None:
                removed_event = conn.execute(
                    text(
                        """
                        SELECT department, full_name
                        FROM public.hr_change_events
                        WHERE new_snapshot_id = :new_snapshot_id
                          AND event_type = :event_type
                        LIMIT 1
                        """
                    ),
                    {"new_snapshot_id": new_id, "event_type": EVENT_TYPE_REMOVED},
                ).mappings().one()
                filter_department = removed_event["department"]
                filtered_rows = build_hr_change_events_export_rows(
                    conn,
                    new_snapshot_id=new_id,
                    event_type=EVENT_TYPE_REMOVED,
                    department=filter_department,
                )
            else:
                filter_department = new_event["department"]
                filtered_rows = build_hr_change_events_export_rows(
                    conn,
                    new_snapshot_id=new_id,
                    event_type=EVENT_TYPE_NEW,
                    department=filter_department,
                )
            assert filtered_rows
            assert all(
                row["department"] == filter_department
                for row in filtered_rows
                if row.get("department")
            )

            search_name = filtered_rows[0]["fio"]
            search_rows = build_hr_change_events_export_rows(
                conn,
                new_snapshot_id=new_id,
                q=str(search_name).split()[0],
            )
            assert search_rows
            assert all(
                str(search_name).split()[0].lower() in str(row["fio"]).lower()
                or str(search_name).split()[0].lower() in str(row["iin"]).lower()
                for row in search_rows
            )
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_export_empty_result_returns_valid_workbook(seed, tmp_path: Path) -> None:
    _require_phase_040h()
    batch_id_1, batch_id_2, prior_id, new_id = _seed_two_snapshot_scenario(seed, tmp_path)
    try:
        with engine.begin() as conn:
            content, _ = export_hr_change_events_xlsx(
                conn,
                new_snapshot_id=new_id,
                department="Nonexistent Department 040h",
            )
        wb = load_workbook(BytesIO(content))
        assert wb.sheetnames == list(EXPORT_SHEETS)
        assert len(_sheet_data_rows(wb, "NEW")) == 0
        assert len(_sheet_data_rows(wb, "CHANGED")) == 0
        assert len(_sheet_data_rows(wb, "REMOVED")) == 0
        assert len(_sheet_data_rows(wb, "CONFLICT")) == 0
        summary_rows = _load_workbook_rows(wb, "SUMMARY")
        total_row = next(row for row in summary_rows if row[0] == "total_rows")
        assert total_row[1] == 0
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_export_includes_conflict_sheet_from_monthly_diff(seed, tmp_path: Path) -> None:
    _require_phase_040h()
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"Cfl{suffix}")
    wrong_iin = _unique_iin(f"1{suffix}")
    corrected_iin = _unique_iin(f"9{suffix}")
    department = f"Pytest Dept {suffix}"

    source = tmp_path / f"phase040h_{uuid4().hex[:8]}.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=wrong_iin)
    ws.cell(row=8, column=6, value=department)
    wb.save(source)
    wb.close()

    batch_id_1 = None
    batch_id_2 = None
    try:
        with engine.begin() as conn:
            batch_id_1, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id = conn.execute(
                text(
                    """
                    SELECT row_id
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id_1},
            ).scalar_one()
            _prepare_roster_row(
                conn,
                batch_id=batch_id_1,
                row_id=int(row_id),
                full_name=full_name,
                iin=wrong_iin,
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )

            batch_id_2, _, _ = import_control_list(
                conn,
                file_path=source,
                imported_by=int(seed["initiator_user_id"]),
            )
            row_id_2 = conn.execute(
                text(
                    """
                    SELECT row_id
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id_2},
            ).scalar_one()
            _prepare_roster_row(
                conn,
                batch_id=batch_id_2,
                row_id=int(row_id_2),
                full_name=full_name,
                iin=wrong_iin,
                department=department,
                org_unit_id=int(seed["unit_id"]),
            )

            build_canonical_snapshot_from_batch(
                conn,
                batch_id_1,
                promoted_by=int(seed["initiator_user_id"]),
            )
            _patch_snapshot_roster_correction(
                conn,
                batch_id=batch_id_1,
                corrected_iin=corrected_iin,
            )
            compute_batch_monthly_diff(conn, batch_id_2)

            rows = build_hr_change_events_export_rows(conn, source_batch_id=batch_id_2)
            content, _ = export_hr_change_events_xlsx(conn, source_batch_id=batch_id_2)

        conflict_rows = [row for row in rows if row["change_type"] == CHANGE_TYPE_CONFLICT]
        assert conflict_rows
        assert all(row["conflict_reason"] for row in conflict_rows)

        workbook = load_workbook(BytesIO(content))
        assert len(_sheet_data_rows(workbook, "CONFLICT")) >= 1
        assert _change_types_in_sheet(workbook, "CONFLICT") == {CHANGE_TYPE_CONFLICT}
    finally:
        if batch_id_1 is not None and batch_id_2 is not None:
            with engine.begin() as conn:
                _cleanup_promotion_batch(conn, batch_id_1)
                _cleanup_promotion_batch(conn, batch_id_2)
