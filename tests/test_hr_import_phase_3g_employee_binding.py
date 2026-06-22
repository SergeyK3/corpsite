"""Tests for ADR-039 Phase 3G — employee binding for normalized records."""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.main import app
from app.services.hr_import_employee_binding_service import (
    BINDING_METHOD_FULL_NAME,
    BINDING_METHOD_IIN,
    BINDING_STATUS_BOUND,
    BINDING_STATUS_CONFLICT,
    BINDING_STATUS_UNBOUND,
    auto_bind_import_row,
    repair_batch_employee_bindings,
    resolve_employee_binding,
)
from app.services.hr_import_normalized_record_service import normalized_records_available
from app.services.hr_import_promotion_service import BLOCKER_EMPLOYEE_REQUIRED, promote_normalized_records
from app.services.hr_import_service import import_control_list
from tests.conftest import auth_headers, insert_returning_id, table_exists
from tests.test_employee_documents_routes import _create_employee, _create_position, _phase_1a_available
from tests.test_import_hr_control_list import _build_doctors_sheet


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_3g_available() -> bool:
    with engine.begin() as conn:
        return normalized_records_available(conn) and table_exists(conn, "employee_identities")


def _require_phase_3g() -> None:
    if not _phase_3g_available():
        pytest.skip("ADR-039 Phase 3G prerequisites missing — run alembic upgrade head")


def _test_iin(seed: str) -> str:
    digits = "".join(ch for ch in seed if ch.isdigit())
    return f"{digits:0>12}"[-12:]


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _build_workbook(path: Path, *, full_name: str, iin: str, column_m: str = "КазНМУ, 1982") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=iin)
    ws.cell(row=8, column=13, value=column_m)
    wb.save(path)
    wb.close()


def _create_employee_with_iin(
    conn,
    *,
    full_name: str,
    iin: str,
    org_unit_id: int,
    created_by: int,
) -> int:
    pos_id = _create_position(conn, name=f"pytest_bind_{uuid4().hex[:8]}")
    emp_id = _create_employee(
        conn,
        full_name=full_name,
        org_unit_id=org_unit_id,
        position_id=pos_id,
        is_active=True,
    )
    insert_returning_id(
        conn,
        table="employee_identities",
        id_col="identity_id",
        values={
            "employee_id": emp_id,
            "identity_type": "IIN",
            "identity_value": iin,
            "is_primary": True,
            "created_by": created_by,
        },
    )
    return emp_id


def _import_batch(
    tmp_path: Path,
    seed,
    *,
    full_name: str,
    iin: str,
    column_m: str = "КазНМУ, 1982",
) -> int:
    source = tmp_path / f"phase3g_{uuid4().hex[:8]}.xlsx"
    _build_workbook(source, full_name=full_name, iin=iin, column_m=column_m)
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
    return int(batch_id)


def _first_row_id(conn, batch_id: int) -> int:
    row_id = conn.execute(
        text(
            """
            SELECT row_id
            FROM public.hr_import_rows
            WHERE batch_id = :batch_id
            ORDER BY row_id
            LIMIT 1
            """
        ),
        {"batch_id": batch_id},
    ).scalar_one()
    return int(row_id)


def _normalized_records_for_row(conn, row_id: int) -> list[dict]:
    rows = conn.execute(
        text(
            """
            SELECT normalized_record_id, employee_id
            FROM public.hr_import_normalized_records
            WHERE row_id = :row_id
            ORDER BY normalized_record_id
            """
        ),
        {"row_id": row_id},
    ).mappings().all()
    return [dict(row) for row in rows]


@pytest.fixture
def privileged_headers(seed, monkeypatch):
    monkeypatch.setenv("DIRECTORY_PRIVILEGED_USER_IDS", str(seed["initiator_user_id"]))
    return auth_headers(seed["initiator_user_id"])


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_resolve_employee_binding_by_iin(seed, tmp_path: Path):
    _require_phase_3g()
    if not _phase_1a_available():
        pytest.skip("employees tables missing")

    suffix = uuid4().hex[:8]
    employee_name = f"Bind IIN Employee {suffix}"
    import_name = f"Bind IIN Import Row {suffix}"
    iin = _test_iin(suffix)
    batch_id = None
    emp_id = None

    try:
        with engine.begin() as conn:
            emp_id = _create_employee_with_iin(
                conn,
                full_name=employee_name,
                iin=iin,
                org_unit_id=int(seed["unit_id"]),
                created_by=int(seed["initiator_user_id"]),
            )
            direct = resolve_employee_binding(
                conn,
                payload={"full_name": import_name, "iin": iin},
            )
            assert direct.status == BINDING_STATUS_BOUND
            assert direct.method == BINDING_METHOD_IIN
            assert direct.employee_id == emp_id

        batch_id = _import_batch(tmp_path, seed, full_name=import_name, iin=iin)

        with engine.begin() as conn:
            row_id = _first_row_id(conn, batch_id)
            row_employee_id = conn.execute(
                text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                {"row_id": row_id},
            ).scalar_one()
            # populate_normalized_records auto-binds during import when IIN matches directory.
            assert int(row_employee_id) == emp_id

            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET
                        employee_id = NULL,
                        match_status = 'NO_MATCH',
                        normalized_payload = jsonb_set(
                            jsonb_set(
                                normalized_payload,
                                '{iin}',
                                to_jsonb(CAST(:iin AS text)),
                                true
                            ),
                            '{full_name}',
                            to_jsonb(CAST(:full_name AS text)),
                            true
                        )
                    WHERE row_id = :row_id
                    """
                ),
                {"row_id": row_id, "iin": iin, "full_name": import_name},
            )
            binding = auto_bind_import_row(conn, row_id)
            records = _normalized_records_for_row(conn, row_id)
            row_employee_id = conn.execute(
                text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                {"row_id": row_id},
            ).scalar_one()

        assert binding.status == BINDING_STATUS_BOUND
        assert binding.method == BINDING_METHOD_IIN
        assert binding.employee_id == emp_id
        assert int(row_employee_id) == emp_id
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_resolve_employee_binding_fallback_by_full_name(seed, tmp_path: Path):
    _require_phase_3g()
    if not _phase_1a_available():
        pytest.skip("employees tables missing")

    suffix = uuid4().hex[:8]
    full_name = f"Bind Name Test {suffix}"
    import_iin = _test_iin(f"7{suffix}")
    batch_id = None
    emp_id = None

    try:
        with engine.begin() as conn:
            pos_id = _create_position(conn, name=f"pytest_name_{suffix}")
            emp_id = _create_employee(
                conn,
                full_name=full_name,
                org_unit_id=int(seed["unit_id"]),
                position_id=pos_id,
                is_active=True,
            )

        batch_id = _import_batch(tmp_path, seed, full_name=full_name, iin=import_iin)

        with engine.begin() as conn:
            row_id = _first_row_id(conn, batch_id)
            # Import populate may auto-bind by full_name when employee exists before import.
            conn.execute(
                text(
                    """
                    UPDATE public.hr_import_rows
                    SET employee_id = NULL, match_status = 'NO_MATCH'
                    WHERE row_id = :row_id
                    """
                ),
                {"row_id": row_id},
            )
            binding = auto_bind_import_row(conn, row_id)

        assert binding.status == BINDING_STATUS_BOUND
        assert binding.method == BINDING_METHOD_FULL_NAME
        assert binding.employee_id == emp_id
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_resolve_employee_binding_unbound_when_no_employee(seed, tmp_path: Path):
    _require_phase_3g()

    suffix = uuid4().hex[:8]
    full_name = f"Missing Employee {suffix}"
    iin = _test_iin(f"6{suffix}")
    batch_id = None

    try:
        batch_id = _import_batch(tmp_path, seed, full_name=full_name, iin=iin)

        with engine.begin() as conn:
            row_id = _first_row_id(conn, batch_id)
            binding = auto_bind_import_row(conn, row_id)
            records = _normalized_records_for_row(conn, row_id)

        assert binding.status == BINDING_STATUS_UNBOUND
        assert binding.employee_id is None
        assert binding.reason
        assert all(rec["employee_id"] is None for rec in records)
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_resolve_employee_binding_conflict_on_duplicate_full_name(seed, tmp_path: Path):
    _require_phase_3g()
    if not _phase_1a_available():
        pytest.skip("employees tables missing")

    suffix = uuid4().hex[:8]
    full_name = f"Conflict Name {suffix}"
    import_iin = _test_iin(f"5{suffix}")
    batch_id = None
    emp_ids: list[int] = []

    try:
        with engine.begin() as conn:
            for idx in range(2):
                pos_id = _create_position(conn, name=f"pytest_conflict_{suffix}_{idx}")
                emp_ids.append(
                    _create_employee(
                        conn,
                        full_name=full_name,
                        org_unit_id=int(seed["unit_id"]),
                        position_id=pos_id,
                        is_active=True,
                    )
                )

        batch_id = _import_batch(tmp_path, seed, full_name=full_name, iin=import_iin)

        with engine.begin() as conn:
            row_id = _first_row_id(conn, batch_id)
            binding = resolve_employee_binding(
                conn,
                payload={"full_name": full_name, "iin": import_iin},
            )
            auto_bind_import_row(conn, row_id)
            row_employee_id = conn.execute(
                text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                {"row_id": row_id},
            ).scalar_one()

        assert binding.status == BINDING_STATUS_CONFLICT
        assert binding.method == BINDING_METHOD_FULL_NAME
        assert len(binding.candidate_employee_ids) == 2
        assert row_employee_id is None
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            for emp_id in emp_ids:
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def _ensure_approved_normalized_record(conn, batch_id: int, *, reviewed_by: int) -> int:
    record_id = conn.execute(
        text(
            """
            SELECT normalized_record_id
            FROM public.hr_import_normalized_records
            WHERE batch_id = :batch_id
            ORDER BY normalized_record_id
            LIMIT 1
            """
        ),
        {"batch_id": batch_id},
    ).scalar_one_or_none()
    if record_id is not None:
        record_id = int(record_id)
    else:
        from app.services.hr_import_normalized_record_service import populate_normalized_records

        populate_normalized_records(conn, batch_id)
        row_id = conn.execute(
            text(
                """
                SELECT row_id
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).scalar_one()
        source_record_key = f"pytest-bind-{uuid4().hex}"
        record_id = int(
            conn.execute(
                text(
                    """
                    INSERT INTO public.hr_import_normalized_records (
                        batch_id, row_id, employee_id, fragment_index, source_field, source_text,
                        source_record_key, record_kind, document_type_id, document_type_code,
                        title, provider, issue_date, parse_method, review_status,
                        reviewed_at, reviewed_by
                    )
                    VALUES (
                        :batch_id, :row_id, NULL, 0, 'education_raw', 'КазНМУ, 1982',
                        :source_record_key, 'education',
                        (SELECT document_type_id FROM public.document_types WHERE code = 'EDUCATION_GRADUATION'),
                        'EDUCATION_GRADUATION', 'КазНМУ', 'КазНМУ', DATE '1982-01-01', 'manual',
                        'approved', NOW(), :reviewed_by
                    )
                    RETURNING normalized_record_id
                    """
                ),
                {
                    "batch_id": batch_id,
                    "row_id": int(row_id),
                    "source_record_key": source_record_key,
                    "reviewed_by": reviewed_by,
                },
            ).scalar_one()
        )
    conn.execute(
        text(
            """
            UPDATE public.hr_import_normalized_records
            SET review_status = 'approved',
                employee_id = NULL,
                reviewed_at = NOW(),
                reviewed_by = :reviewed_by
            WHERE normalized_record_id = :record_id
            """
        ),
        {"record_id": record_id, "reviewed_by": reviewed_by},
    )
    conn.execute(
        text(
            """
            UPDATE public.hr_import_rows
            SET employee_id = NULL
            WHERE batch_id = :batch_id
            """
        ),
        {"batch_id": batch_id},
    )
    return record_id


def test_dry_run_passes_after_employee_binding(seed, tmp_path: Path, privileged_headers):
    _require_phase_3g()
    if not _phase_1a_available():
        pytest.skip("employees tables missing")

    suffix = uuid4().hex[:8]
    full_name = f"DryRun Bind {suffix}"
    iin = _test_iin(f"4{suffix}")
    batch_id = None
    emp_id = None
    record_id = None

    try:
        batch_id = _import_batch(
            tmp_path,
            seed,
            full_name=full_name,
            iin=iin,
            column_m="КазНМУ, 1982; ПК 144 ч",
        )

        with engine.begin() as conn:
            record_id = _ensure_approved_normalized_record(
                conn,
                batch_id,
                reviewed_by=int(seed["initiator_user_id"]),
            )

        with engine.begin() as conn:
            blocked = promote_normalized_records(
                conn,
                promoted_by=int(seed["initiator_user_id"]),
                dry_run=True,
                record_ids=[record_id],
            )
        assert blocked["items"][0]["blockers"]
        assert blocked["items"][0]["blockers"][0]["code"] == BLOCKER_EMPLOYEE_REQUIRED

        with engine.begin() as conn:
            emp_id = _create_employee_with_iin(
                conn,
                full_name=full_name,
                iin=iin,
                org_unit_id=int(seed["unit_id"]),
                created_by=int(seed["initiator_user_id"]),
            )

        client = TestClient(app)
        bind_response = client.patch(
            f"/directory/personnel/import/normalized-records/{record_id}",
            headers=privileged_headers,
            json={"employee_id": emp_id},
        )
        assert bind_response.status_code == 200
        assert bind_response.json()["employee_id"] == emp_id

        with engine.begin() as conn:
            passed = promote_normalized_records(
                conn,
                promoted_by=int(seed["initiator_user_id"]),
                dry_run=True,
                record_ids=[record_id],
            )

        assert not any(
            blocker.get("code") == BLOCKER_EMPLOYEE_REQUIRED
            for blocker in passed["items"][0].get("blockers") or []
        )

        api_result = client.post(
            "/directory/personnel/import/normalized-records/promote",
            headers=privileged_headers,
            json={"record_ids": [record_id], "dry_run": True},
        )
        assert api_result.status_code == 200
        api_item = api_result.json()["items"][0]
        assert not any(
            blocker.get("code") == BLOCKER_EMPLOYEE_REQUIRED
            for blocker in api_item.get("blockers") or []
        )
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_manual_binding_api(seed, tmp_path: Path, privileged_headers):
    _require_phase_3g()
    if not _phase_1a_available():
        pytest.skip("employees tables missing")

    suffix = uuid4().hex[:8]
    full_name = f"Manual Bind {suffix}"
    iin = _test_iin(f"3{suffix}")
    batch_id = None
    emp_id = None

    try:
        with engine.begin() as conn:
            pos_id = _create_position(conn, name=f"pytest_manual_{suffix}")
            emp_id = _create_employee(
                conn,
                full_name=full_name,
                org_unit_id=int(seed["unit_id"]),
                position_id=pos_id,
                is_active=True,
            )

        batch_id = _import_batch(tmp_path, seed, full_name="Unknown Person", iin=iin)

        with engine.begin() as conn:
            record_id = int(
                conn.execute(
                    text(
                        """
                        SELECT normalized_record_id
                        FROM public.hr_import_normalized_records
                        WHERE batch_id = :batch_id
                        ORDER BY normalized_record_id
                        LIMIT 1
                        """
                    ),
                    {"batch_id": batch_id},
                ).scalar_one()
            )

        client = TestClient(app)
        response = client.patch(
            f"/directory/personnel/import/normalized-records/{record_id}",
            headers=privileged_headers,
            json={"employee_id": emp_id},
        )
        assert response.status_code == 200
        body = response.json()
        assert body["employee_id"] == emp_id
        assert body["employee_binding"]["status"] == BINDING_STATUS_BOUND
        assert body["employee_binding"]["method"] == "manual"
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_repair_bindings_api_uses_full_iin(seed, tmp_path: Path, privileged_headers):
    _require_phase_3g()
    if not _phase_1a_available():
        pytest.skip("employees tables missing")

    suffix = uuid4().hex[:8]
    employee_name = f"Repair Bind Employee {suffix}"
    import_name = f"Repair Bind Import {suffix}"
    iin = _test_iin(suffix)
    batch_id = None
    emp_id = None

    try:
        batch_id = _import_batch(tmp_path, seed, full_name=import_name, iin=iin)

        with engine.begin() as conn:
            emp_id = _create_employee_with_iin(
                conn,
                full_name=employee_name,
                iin=iin,
                org_unit_id=int(seed["unit_id"]),
                created_by=int(seed["initiator_user_id"]),
            )
            row_id = _first_row_id(conn, batch_id)
            current_employee_id = conn.execute(
                text("SELECT employee_id FROM public.hr_import_rows WHERE row_id = :row_id"),
                {"row_id": row_id},
            ).scalar_one()
            assert current_employee_id is None

        client = TestClient(app)
        repair_response = client.post(
            f"/directory/personnel/import/batches/{batch_id}/employee-bindings/repair",
            headers=privileged_headers,
        )
        assert repair_response.status_code == 200, repair_response.text
        repair_body = repair_response.json()
        assert repair_body["bound"] >= 1

        list_response = client.get(
            f"/directory/personnel/import/normalized-records?batch_id={batch_id}&limit=50",
            headers=privileged_headers,
        )
        assert list_response.status_code == 200, list_response.text
        for item in list_response.json()["items"]:
            assert item.get("iin") == iin
            assert "****" not in (item.get("iin") or "")
            if item.get("employee_binding", {}).get("status") == BINDING_STATUS_BOUND:
                assert item["employee_binding"]["method"] == BINDING_METHOD_IIN
                assert item["employee_id"] == emp_id
    finally:
        with engine.begin() as conn:
            if batch_id:
                _delete_batch(conn, batch_id)
            if emp_id:
                conn.execute(text("DELETE FROM public.employee_identities WHERE employee_id = :id"), {"id": emp_id})
                conn.execute(text("DELETE FROM public.employees WHERE employee_id = :id"), {"id": emp_id})
