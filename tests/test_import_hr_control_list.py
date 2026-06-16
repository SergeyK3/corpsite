"""Tests for HR control list staging parser (Phase 0B)."""
from __future__ import annotations

from datetime import datetime
from pathlib import Path

from openpyxl import Workbook

from scripts.import_hr_control_list import (
    build_audit,
    build_merged_section_lookup,
    clean_iin,
    export_outputs,
    get_layout_profile,
    mask_iin,
    parse_birth_date,
    parse_sheet_with_profile,
    parse_workbook,
    resolve_section_department,
    resolve_sheet_type,
)


def _build_doctors_sheet(ws) -> None:
    ws.append([""] * 20)
    ws.append([""] * 20)
    ws.append([""] * 20)
    ws.append([""] * 20)
    ws.append([""] * 20)
    ws.append([""] * 20)
    ws.append(
        [
            "№",
            "",
            "Фамилия, имя, отчество",
            "Год рождения",
            "ИИН",
            "пол",
            "",
            "",
            "",
            "Занимаемая должность",
            "",
            "",
            "Повышение квалификации",
            "Квалификационная категория",
        ]
        + [""] * 6
    )
    ws.append(
        [
            1,
            "АДМИНИСТРАТИВНЫЙ ПЕРСОНАЛ",
            "Иванов Иван Иванович",
            datetime(1990, 1, 1),
            "900101300123",
            "муж",
            "",
            "",
            "",
            "Директор",
            "",
            "",
            "2024 курс ПК",
            "высшая",
        ]
        + [""] * 6
    )
    ws.append(
        [
            2,
            "",
            "Петрова Анна Сергеевна",
            32874,
            "900101300123",
            "жен",
            "",
            "",
            "",
            "Врач-хирург",
            "",
            "",
            "",
            "",
        ]
        + [""] * 6
    )
    ws.append([3, "ХИРУРГИЯ", "", "", "", "", "", "", "", "", "", "", "", "", ""] + [""] * 6)
    ws.cell(row=11, column=2, value="ХИРУРГИЯ")
    ws.merge_cells("B11:B12")
    ws.append(
        [
            4,
            "",
            "Сидоров Петр",
            datetime(1985, 5, 5),
            "850505400456",
            "муж",
            "",
            "",
            "",
            "Хирург",
            "",
            "",
            "",
            "",
        ]
        + [""] * 6
    )


def _build_nurses_sheet(ws) -> None:
    ws.append([""] * 12)
    ws.append([""] * 12)
    ws.append([""] * 12)
    ws.append(
        [
            "",
            "№",
            "",
            "Фамилия, имя, отчество",
            "Год рождения",
            "ИИН",
            "пол",
            "",
            "",
            "",
            "Занимаемая должность, дата назнач.",
            "",
        ]
    )
    ws.append(
        [
            "",
            "",
            "ТЕРАПИЯ",
            "Сидорова Мария",
            datetime(1985, 5, 5),
            "850505400456",
            "жен",
            "",
            "",
            "",
            "Медсестра",
            "",
        ]
    )
    ws.append(["", 1, "", "Кузнецова Анна", datetime(1980, 3, 16), "800316400190", "жен", "", "", "", "М/с", ""])


def _build_junior_staff_sheet(ws) -> None:
    ws.append([""] * 10)
    ws.append([""] * 10)
    ws.append([""] * 10)
    ws.append([""] * 10)
    ws.append(
        [
            "",
            "",
            "Фамилия, имя, отчество",
            "Год рождения",
            "ИИН",
            "пол",
            "",
            "Занимаемая должность",
            "",
            "Телефон",
        ]
    )
    ws.append(
        [
            "",
            "ОБЩЕБОЛЬНИЧНЫЙ",
            "Амангельдинова Балжан",
            datetime(1974, 8, 26),
            "740826400893",
            "жен",
            "",
            "санитарка",
            "",
            "87025059904",
        ]
    )
    ws.append(["", "", "Шокумова Гульмира", datetime(1976, 1, 11), "760111450360", "жен", "", "санитарка", "", ""])


def _build_sample_workbook(path: Path) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    _build_doctors_sheet(wb.create_sheet("врачи"))
    _build_nurses_sheet(wb.create_sheet("медсестра"))
    _build_junior_staff_sheet(wb.create_sheet("санитарки"))
    wb.create_sheet("неизвестный лист").append(["ФИО", "ИИН"])
    wb.save(path)


def test_resolve_sheet_type():
    assert resolve_sheet_type("врачи") == "doctors"
    assert resolve_sheet_type("медсестра") == "nurses"
    assert resolve_sheet_type("санитарки") == "junior_staff"
    assert resolve_sheet_type("врачи совместители") == "part_time"
    assert resolve_sheet_type("декларационные листы") == "declaration"
    assert resolve_sheet_type("лист1") is None


def test_clean_iin_and_birth_date():
    assert clean_iin("900 101 300 123")[0:2] == ("900101300123", True)
    assert clean_iin("12345")[1] is False
    assert parse_birth_date(32874) == "1990-01-01"
    assert mask_iin("900101300123") == "9001****23"


def test_parses_doctors_layout_profile(tmp_path: Path):
    path = tmp_path / "doctors.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _build_doctors_sheet(wb.create_sheet("врачи"))
    wb.save(path)

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    rows = parse_sheet_with_profile(wb["врачи"], sheet_type="doctors", profile=get_layout_profile("doctors"))
    wb.close()

    first = rows[0]
    assert first.full_name == "Иванов Иван Иванович"
    assert first.department == "АДМИНИСТРАТИВНЫЙ ПЕРСОНАЛ"
    assert first.data["position_raw"] == "Директор"
    assert first.training_raw == "2024 курс ПК"
    assert first.certification_raw == "высшая"


def test_parses_nurses_layout_profile(tmp_path: Path):
    path = tmp_path / "nurses.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _build_nurses_sheet(wb.create_sheet("медсестра"))
    wb.save(path)

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    rows = parse_sheet_with_profile(wb["медсестра"], sheet_type="nurses", profile=get_layout_profile("nurses"))
    wb.close()

    assert rows[0].full_name == "Сидорова Мария"
    assert rows[0].department == "ТЕРАПИЯ"
    assert rows[1].department == "ТЕРАПИЯ"


def test_parses_junior_staff_layout_profile(tmp_path: Path):
    path = tmp_path / "junior.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _build_junior_staff_sheet(wb.create_sheet("санитарки"))
    wb.save(path)

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    rows = parse_sheet_with_profile(
        wb["санитарки"],
        sheet_type="junior_staff",
        profile=get_layout_profile("junior_staff"),
    )
    wb.close()

    assert rows[0].full_name == "Амангельдинова Балжан"
    assert rows[0].department == "ОБЩЕБОЛЬНИЧНЫЙ"
    assert rows[1].department == "ОБЩЕБОЛЬНИЧНЫЙ"


def test_resolves_department_from_merged_section(tmp_path: Path):
    path = tmp_path / "merged.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _build_doctors_sheet(wb.create_sheet("врачи"))
    wb.save(path)

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    ws = wb["врачи"]
    lookup = build_merged_section_lookup(ws, 2)
    assert lookup[11] == "ХИРУРГИЯ"
    assert lookup[12] == "ХИРУРГИЯ"
    assert resolve_section_department(ws, 12, 2, lookup, "PREV") == "ХИРУРГИЯ"

    rows = parse_sheet_with_profile(ws, sheet_type="doctors", profile=get_layout_profile("doctors"))
    surgeon = next(row for row in rows if row.full_name == "Сидоров Петр")
    assert surgeon.department == "ХИРУРГИЯ"
    wb.close()


def test_section_value_same_row_as_employee(tmp_path: Path):
    path = tmp_path / "same_row.xlsx"
    wb = Workbook()
    wb.remove(wb.active)
    _build_doctors_sheet(wb.create_sheet("врачи"))
    wb.save(path)

    wb = __import__("openpyxl").load_workbook(path, data_only=True)
    rows = parse_sheet_with_profile(wb["врачи"], sheet_type="doctors", profile=get_layout_profile("doctors"))
    wb.close()

    first = rows[0]
    assert first.full_name == "Иванов Иван Иванович"
    assert first.department == "АДМИНИСТРАТИВНЫЙ ПЕРСОНАЛ"
    assert first.full_name != first.department


def test_parse_workbook(tmp_path: Path):
    source = tmp_path / "control.xlsx"
    _build_sample_workbook(source)

    rows, warnings = parse_workbook(source)
    audit = build_audit(rows)

    assert any(w.startswith("skip_unknown_sheet:") for w in warnings)
    assert audit["total_rows"] >= 5
    assert audit["valid_iin"] >= 4
    assert audit["duplicate_iin"] >= 1
    assert audit["missing_full_name"] <= 1
    assert audit["missing_department"] == 0
    assert audit["with_training"] >= 1
    assert audit["with_certification"] >= 1


def test_export_outputs(tmp_path: Path):
    source = tmp_path / "control.xlsx"
    out_dir = tmp_path / "out"
    _build_sample_workbook(source)
    rows, _ = parse_workbook(source)
    audit = build_audit(rows)
    paths = export_outputs(rows, audit, out_dir)

    assert paths["preview"].exists()
    assert paths["errors"].exists()
    assert paths["duplicates"].exists()
