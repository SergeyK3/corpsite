"""Tests for ADR-039 Phase 3D — normalized records review API."""
from __future__ import annotations

from pathlib import Path
from uuid import uuid4

import pytest
from fastapi.testclient import TestClient
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_import_normalized_record_service import normalized_records_available
from app.services.hr_import_service import import_control_list
from tests.conftest import auth_headers, table_exists
from tests.test_import_hr_control_list import _build_doctors_sheet


def _db_available() -> bool:
    try:
        with engine.connect() as conn:
            conn.execute(text("SELECT 1"))
        return True
    except Exception:
        return False


def _phase_3d_available() -> bool:
    with engine.begin() as conn:
        return normalized_records_available(conn)


def _require_phase_3d() -> None:
    if not _phase_3d_available():
        pytest.skip("ADR-039 Phase 3B migration not applied — run alembic upgrade head")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _employee_documents_with_source_count(conn) -> int:
    if not table_exists(conn, "employee_documents"):
        return 0
    return int(
        conn.execute(
            text(
                """
                SELECT COUNT(*)
                FROM public.employee_documents
                WHERE source_normalized_record_id IS NOT NULL
                   OR source_record_key IS NOT NULL
                """
            )
        ).scalar_one()
    )


def _build_doctors_sheet_with_column_m(path: Path, column_m: str, column_n: str = "") -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=13, value=column_m)
    if column_n:
        ws.cell(row=8, column=14, value=column_n)
    wb.save(path)
    wb.close()


@pytest.fixture
def privileged_headers(seed, monkeypatch):
    monkeypatch.setenv("DIRECTORY_PRIVILEGED_USER_IDS", str(seed["initiator_user_id"]))
    return auth_headers(seed["initiator_user_id"])


@pytest.fixture
def staged_batch_with_records(seed, tmp_path: Path):
    _require_phase_3d()
    source = tmp_path / f"phase3d_review_{uuid4().hex[:8]}.xlsx"
    _build_doctors_sheet_with_column_m(
        source,
        "КазНМУ, 1982; ПК 144 ч",
        "сертификат специалиста, действ. до 01.01.2028",
    )
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
    yield batch_id
    with engine.begin() as conn:
        _delete_batch(conn, batch_id)


def _list_records(client: TestClient, headers: dict[str, str], **params: object) -> dict:
    query = "&".join(f"{key}={value}" for key, value in params.items() if value is not None)
    url = "/directory/personnel/import/normalized-records"
    if query:
        url = f"{url}?{query}"
    resp = client.get(url, headers=headers)
    assert resp.status_code == 200, resp.text
    return resp.json()


def _patch_review(
    client: TestClient,
    headers: dict[str, str],
    record_id: int,
    review_status: str,
    review_notes: str | None = None,
) -> dict:
    body: dict[str, object] = {"review_status": review_status}
    if review_notes is not None:
        body["review_notes"] = review_notes
    resp = client.patch(
        f"/directory/personnel/import/normalized-records/{record_id}",
        headers={**headers, "Content-Type": "application/json"},
        json=body,
    )
    return resp


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_summary_api_contract(client: TestClient, privileged_headers, staged_batch_with_records):
    batch_id = staged_batch_with_records
    resp = client.get(
        "/directory/personnel/import/normalized-records/summary",
        headers=privileged_headers,
    )
    assert resp.status_code == 200, resp.text
    body = resp.json()
    assert set(body.keys()) >= {
        "total",
        "pending",
        "approved",
        "rejected",
        "promoted",
        "superseded",
        "by_kind",
        "skipped",
    }
    assert set(body["by_kind"].keys()) == {"training", "certificate", "category", "education"}
    assert body["total"] >= 1
    assert body["pending"] >= 1

    scoped = client.get(
        f"/directory/personnel/import/normalized-records/summary?batch_id={batch_id}",
        headers=privileged_headers,
    )
    assert scoped.status_code == 200, scoped.text
    scoped_body = scoped.json()
    assert scoped_body["total"] >= 1
    assert scoped_body["pending"] >= 1


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_list_filters(client: TestClient, privileged_headers, staged_batch_with_records):
    batch_id = staged_batch_with_records
    all_items = _list_records(client, privileged_headers, batch_id=batch_id, limit=500)
    assert all_items["total"] >= 1

    training = _list_records(
        client,
        privileged_headers,
        batch_id=batch_id,
        record_kind="training",
        limit=500,
    )
    assert training["total"] >= 1
    assert all(item["record_kind"] == "training" for item in training["items"])

    pending = _list_records(
        client,
        privileged_headers,
        batch_id=batch_id,
        review_status="pending",
        limit=500,
    )
    assert pending["total"] >= 1
    assert all(item["review_status"] == "pending" for item in pending["items"])

    sample = all_items["items"][0]
    if sample.get("employee_id") is not None:
        by_employee = _list_records(
            client,
            privileged_headers,
            employee_id=sample["employee_id"],
            limit=500,
        )
        assert by_employee["total"] >= 1
        assert all(item["employee_id"] == sample["employee_id"] for item in by_employee["items"])

    first = training["items"][0]
    assert first["record_id"] == first["normalized_record_id"]
    assert "source_text" in first
    assert "source_record_key" in first
    assert "created_at" in first


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_list_pagination(client: TestClient, privileged_headers, staged_batch_with_records):
    batch_id = staged_batch_with_records
    first_page = _list_records(client, privileged_headers, batch_id=batch_id, limit=1, offset=0)
    second_page = _list_records(client, privileged_headers, batch_id=batch_id, limit=1, offset=1)

    assert first_page["limit"] == 1
    assert first_page["offset"] == 0
    assert len(first_page["items"]) == 1
    if first_page["total"] > 1:
        assert first_page["items"][0]["record_id"] != second_page["items"][0]["record_id"]


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_review_transitions(client: TestClient, privileged_headers, staged_batch_with_records, seed):
    batch_id = staged_batch_with_records
    record_id = _list_records(
        client,
        privileged_headers,
        batch_id=batch_id,
        review_status="pending",
        limit=1,
    )["items"][0]["record_id"]

    with engine.begin() as conn:
        before_docs = _employee_documents_with_source_count(conn)

    approved = _patch_review(client, privileged_headers, record_id, "approved", "ok")
    assert approved.status_code == 200, approved.text
    approved_body = approved.json()
    assert approved_body["review_status"] == "approved"
    assert approved_body["reviewed_by"] == int(seed["initiator_user_id"])
    assert approved_body["review_notes"] == "ok"
    assert approved_body["reviewed_at"] is not None

    back_pending = _patch_review(client, privileged_headers, record_id, "pending")
    assert back_pending.status_code == 200, back_pending.text
    pending_body = back_pending.json()
    assert pending_body["review_status"] == "pending"
    assert pending_body["reviewed_by"] is None
    assert pending_body["review_notes"] is None

    rejected = _patch_review(client, privileged_headers, record_id, "rejected", "no")
    assert rejected.status_code == 200, rejected.text
    rejected_body = rejected.json()
    assert rejected_body["review_status"] == "rejected"
    assert rejected_body["review_notes"] == "no"

    back_pending_again = _patch_review(client, privileged_headers, record_id, "pending")
    assert back_pending_again.status_code == 200, back_pending_again.text
    assert back_pending_again.json()["review_status"] == "pending"

    with engine.begin() as conn:
        after_docs = _employee_documents_with_source_count(conn)
    assert after_docs == before_docs


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_invalid_review_transitions(client: TestClient, privileged_headers, staged_batch_with_records):
    batch_id = staged_batch_with_records
    record_id = _list_records(
        client,
        privileged_headers,
        batch_id=batch_id,
        review_status="pending",
        limit=1,
    )["items"][0]["record_id"]

    approved = _patch_review(client, privileged_headers, record_id, "approved")
    assert approved.status_code == 200

    invalid = _patch_review(client, privileged_headers, record_id, "rejected")
    assert invalid.status_code == 400

    with engine.begin() as conn:
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET review_status = 'superseded', reviewed_at = NULL, reviewed_by = NULL, review_notes = NULL
                WHERE normalized_record_id = :record_id
                """
            ),
            {"record_id": record_id},
        )

    superseded_patch = _patch_review(client, privileged_headers, record_id, "pending")
    assert superseded_patch.status_code == 400


@pytest.mark.skipif(not _db_available(), reason="PostgreSQL not available")
def test_review_api_forbidden_without_privilege(client: TestClient, seed, staged_batch_with_records):
    headers = auth_headers(seed["executor_user_id"])
    resp = client.get("/directory/personnel/import/normalized-records/summary", headers=headers)
    assert resp.status_code == 403

    list_resp = client.get("/directory/personnel/import/normalized-records", headers=headers)
    assert list_resp.status_code == 403

    patch_resp = client.patch(
        "/directory/personnel/import/normalized-records/1",
        headers={**headers, "Content-Type": "application/json"},
        json={"review_status": "approved"},
    )
    assert patch_resp.status_code == 403
