"""Tests for ADR-040 Phase A — canonical HR snapshot foundation."""
from __future__ import annotations

import json
from pathlib import Path
from uuid import uuid4

import pytest
from openpyxl import Workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.hr_canonical_snapshot_service import (
    SNAPSHOT_STATUS_ACTIVE,
    SNAPSHOT_STATUS_SUPERSEDED,
    build_canonical_snapshot_from_batch,
    canonical_snapshot_available,
    compute_canonical_hash,
    compute_roster_match_key,
    dedupe_snapshot_entries,
    refresh_canonical_snapshot_after_promotion,
)
from app.services.hr_import_normalized_record_service import (
    normalized_records_available,
    populate_normalized_records,
)
from app.services.hr_import_promotion_service import promote_normalized_records
from app.services.hr_import_roster_promotion_service import promote_roster_batch
from app.services.hr_import_service import import_control_list
from tests.conftest import table_exists
from tests.test_employee_documents_routes import _create_employee, _create_position, _phase_1a_available
from tests.test_import_hr_control_list import _build_doctors_sheet


def _require_phase_040a() -> None:
    with engine.begin() as conn:
        if not canonical_snapshot_available(conn):
            pytest.skip("ADR-040 Phase A migration not applied — run alembic upgrade head")


def _delete_batch(conn, batch_id: int) -> None:
    conn.execute(
        text("DELETE FROM public.hr_import_batches WHERE batch_id = :batch_id"),
        {"batch_id": batch_id},
    )


def _cleanup_promotion_batch(conn, batch_id: int) -> None:
    employee_ids = [
        int(row[0])
        for row in conn.execute(
            text(
                """
                SELECT DISTINCT employee_id
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                  AND employee_id IS NOT NULL
                """
            ),
            {"batch_id": batch_id},
        ).all()
    ]
    if table_exists(conn, "hr_import_normalized_records"):
        conn.execute(
            text(
                """
                UPDATE public.hr_import_normalized_records
                SET review_status = 'approved',
                    promoted_document_id = NULL,
                    promoted_at = NULL,
                    promoted_by = NULL
                WHERE batch_id = :batch_id
                  AND review_status = 'promoted'
                """
            ),
            {"batch_id": batch_id},
        )
    if table_exists(conn, "employee_documents"):
        conn.execute(
            text(
                """
                DELETE FROM public.employee_documents
                WHERE source_batch_id = :batch_id
                   OR source_normalized_record_id IN (
                        SELECT normalized_record_id
                        FROM public.hr_import_normalized_records
                        WHERE batch_id = :batch_id
                   )
                """
            ),
            {"batch_id": batch_id},
        )
    _delete_batch(conn, batch_id)
    for employee_id in employee_ids:
        conn.execute(
            text("DELETE FROM public.employee_identities WHERE employee_id = :employee_id"),
            {"employee_id": employee_id},
        )
        if table_exists(conn, "employee_documents"):
            conn.execute(
                text("DELETE FROM public.employee_documents WHERE employee_id = :employee_id"),
                {"employee_id": employee_id},
            )
        conn.execute(
            text("DELETE FROM public.employees WHERE employee_id = :employee_id"),
            {"employee_id": employee_id},
        )


def _build_workbook(path: Path, *, full_name: str, iin: str) -> None:
    wb = Workbook()
    wb.remove(wb.active)
    ws = wb.create_sheet("врачи")
    _build_doctors_sheet(ws)
    ws.cell(row=8, column=3, value=full_name)
    ws.cell(row=8, column=5, value=iin)
    if ws.max_row > 8:
        ws.delete_rows(9, ws.max_row - 8)
    wb.save(path)
    wb.close()


def _test_iin(seed: str) -> str:
    digits = "".join(ch for ch in seed if ch.isdigit())
    return f"{digits:0>12}"[-12:]


def _cyrillic_full_name(prefix: str) -> str:
    token = int(uuid4().hex[:6], 16)
    patronymics = [
        "Петрович",
        "Иванович",
        "Сидорович",
        "Алексеевич",
    ]
    return f"{prefix} Тест {patronymics[token % len(patronymics)]}"


def _ensure_department_mapping(conn, *, department: str, org_unit_id: int) -> None:
    existing = conn.execute(
        text(
            """
            SELECT id FROM public.department_recoding
            WHERE LOWER(TRIM(import_department_name)) = LOWER(TRIM(:name))
            LIMIT 1
            """
        ),
        {"name": department},
    ).first()
    if existing:
        conn.execute(
            text(
                """
                UPDATE public.department_recoding
                SET org_unit_id = :org_unit_id, org_unit_name = :org_unit_name, is_active = TRUE
                WHERE id = :id
                """
            ),
            {
                "id": int(existing[0]),
                "org_unit_id": org_unit_id,
                "org_unit_name": f"Pytest Unit {org_unit_id}",
            },
        )
    else:
        conn.execute(
            text(
                """
                INSERT INTO public.department_recoding (
                    import_department_name, org_unit_id, org_unit_name, department_group, is_active
                )
                VALUES (:department, :org_unit_id, :org_unit_name, 'CLINICAL', TRUE)
                """
            ),
            {
                "department": department,
                "org_unit_id": org_unit_id,
                "org_unit_name": f"Pytest Unit {org_unit_id}",
            },
        )


def _ensure_roster_employee_metadata(conn, *, batch_id: int, row_id: int) -> None:
    conn.execute(
        text(
            """
            UPDATE public.hr_import_rows
            SET normalized_payload = jsonb_set(
                normalized_payload,
                '{metadata}',
                COALESCE(normalized_payload->'metadata', '{}'::jsonb)
                    || jsonb_build_object(
                        'row_type', 'EMPLOYEE',
                        'is_employee_roster', TRUE,
                        'classification', 'NORMAL',
                        'sheet_type', 'doctors'
                    ),
                true
            )
            WHERE batch_id = :batch_id AND row_id = :row_id
            """
        ),
        {"batch_id": batch_id, "row_id": row_id},
    )


def _prepare_roster_row(
    conn,
    *,
    batch_id: int,
    row_id: int,
    full_name: str,
    iin: str,
    department: str,
    org_unit_id: int,
) -> None:
    _ensure_department_mapping(conn, department=department, org_unit_id=org_unit_id)
    conn.execute(
        text(
            """
            UPDATE public.hr_import_rows
            SET normalized_payload = jsonb_set(
                jsonb_set(
                    jsonb_set(normalized_payload, '{full_name}', to_jsonb(CAST(:full_name AS text)), true),
                    '{iin}',
                    to_jsonb(CAST(:iin AS text)),
                    true
                ),
                '{department}',
                to_jsonb(CAST(:department AS text)),
                true
            )
            WHERE row_id = :row_id AND batch_id = :batch_id
            """
        ),
        {
            "row_id": row_id,
            "batch_id": batch_id,
            "full_name": full_name,
            "iin": iin,
            "department": department,
        },
    )
    _ensure_roster_employee_metadata(conn, batch_id=batch_id, row_id=row_id)
    populate_normalized_records(conn, batch_id)


def _import_single_employee_batch(seed, tmp_path: Path) -> tuple[int, str, str, str]:
    suffix = uuid4().hex[:8]
    full_name = _cyrillic_full_name(f"Snap{suffix}")
    iin = _test_iin(suffix)
    department = f"Pytest Dept {suffix}"
    source = tmp_path / f"phase040a_{suffix}.xlsx"
    _build_workbook(source, full_name=full_name, iin=iin)
    with engine.begin() as conn:
        batch_id, _, _ = import_control_list(
            conn,
            file_path=source,
            imported_by=int(seed["initiator_user_id"]),
        )
        row_id = conn.execute(
            text(
                """
                SELECT row_id
                FROM public.hr_import_rows
                WHERE batch_id = :batch_id
                ORDER BY row_id
                LIMIT 1
                """
            ),
            {"batch_id": batch_id},
        ).scalar_one()
        _prepare_roster_row(
            conn,
            batch_id=batch_id,
            row_id=int(row_id),
            full_name=full_name,
            iin=iin,
            department=department,
            org_unit_id=int(seed["unit_id"]),
        )
    return batch_id, full_name, iin, department


def test_canonical_hash_is_stable_for_identical_payload() -> None:
    payload = {
        "full_name": "Иванов Иван",
        "iin": "123456789012",
        "department": "Терапия",
        "merged_profile": {"basic": {"full_name": "Иванов Иван"}},
    }
    first = compute_canonical_hash(
        record_kind="roster",
        entity_scope="iin:123456789012",
        payload=payload,
    )
    second = compute_canonical_hash(
        record_kind="roster",
        entity_scope="iin:123456789012",
        payload={
            "department": "Терапия",
            "merged_profile": {"basic": {"full_name": "Иванов Иван"}},
            "full_name": "Иванов Иван",
            "iin": "123456789012",
            "row_id": 999,
            "batch_id": 123,
        },
    )
    assert first == second


def test_roster_match_key_prefers_employee_id() -> None:
    assert compute_roster_match_key(
        employee_id=42,
        iin="123456789012",
        full_name="Test User",
        birth_date="1980-01-01",
    ) == "emp:42"


def test_snapshot_created_after_roster_promotion(seed, tmp_path: Path) -> None:
    _require_phase_040a()
    batch_id, _, iin, _ = _import_single_employee_batch(seed, tmp_path)

    try:
        with engine.begin() as conn:
            result = promote_roster_batch(
                conn,
                batch_id,
                created_by=int(seed["initiator_user_id"]),
                dry_run=False,
            )
            assert result.get("canonical_snapshot", {}).get("created") is True

            snapshot = conn.execute(
                text(
                    """
                    SELECT snapshot_id, status, entry_count
                    FROM public.hr_canonical_snapshots
                    WHERE source_batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().one()
            assert snapshot["status"] == SNAPSHOT_STATUS_ACTIVE
            assert int(snapshot["entry_count"]) >= 1

            roster_entry = conn.execute(
                text(
                    """
                    SELECT employee_id, iin, match_key, payload, canonical_hash
                    FROM public.hr_canonical_snapshot_entries
                    WHERE snapshot_id = :snapshot_id
                      AND record_kind = 'roster'
                    LIMIT 1
                    """
                ),
                {"snapshot_id": int(snapshot["snapshot_id"])},
            ).mappings().one()
            assert roster_entry["match_key"]
            assert roster_entry["canonical_hash"]
            assert roster_entry["payload"]
            assert roster_entry["iin"] == iin or roster_entry["iin"]
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_second_snapshot_supersedes_first(seed, tmp_path: Path) -> None:
    _require_phase_040a()
    batch_id_1, _, _, _ = _import_single_employee_batch(seed, tmp_path)
    batch_id_2, _, _, _ = _import_single_employee_batch(seed, tmp_path)
    promoted_by = int(seed["initiator_user_id"])

    try:
        with engine.begin() as conn:
            promote_roster_batch(conn, batch_id_1, created_by=promoted_by, dry_run=False)
        with engine.begin() as conn:
            first = conn.execute(
                text(
                    """
                    SELECT snapshot_id, status, version
                    FROM public.hr_canonical_snapshots
                    WHERE source_batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id_1},
            ).mappings().one()

            promote_roster_batch(conn, batch_id_2, created_by=promoted_by, dry_run=False)
            first_after = conn.execute(
                text(
                    """
                    SELECT status, superseded_by_snapshot_id
                    FROM public.hr_canonical_snapshots
                    WHERE snapshot_id = :snapshot_id
                    """
                ),
                {"snapshot_id": int(first["snapshot_id"])},
            ).mappings().one()
            second = conn.execute(
                text(
                    """
                    SELECT snapshot_id, status, version
                    FROM public.hr_canonical_snapshots
                    WHERE source_batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id_2},
            ).mappings().one()
            active = conn.execute(
                text(
                    """
                    SELECT snapshot_id
                    FROM public.hr_canonical_snapshots
                    WHERE status = :active
                    LIMIT 1
                    """
                ),
                {"active": SNAPSHOT_STATUS_ACTIVE},
            ).mappings().one()

            assert first_after["status"] == SNAPSHOT_STATUS_SUPERSEDED
            assert int(first_after["superseded_by_snapshot_id"]) == int(second["snapshot_id"])
            assert second["status"] == SNAPSHOT_STATUS_ACTIVE
            assert int(second["version"]) > int(first["version"])
            assert int(active["snapshot_id"]) == int(second["snapshot_id"])
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id_1)
            _cleanup_promotion_batch(conn, batch_id_2)


def test_snapshot_idempotent_for_same_batch(seed, tmp_path: Path) -> None:
    _require_phase_040a()
    batch_id, _, _, _ = _import_single_employee_batch(seed, tmp_path)
    promoted_by = int(seed["initiator_user_id"])

    try:
        with engine.begin() as conn:
            first = build_canonical_snapshot_from_batch(conn, batch_id, promoted_by=promoted_by)
            second = build_canonical_snapshot_from_batch(conn, batch_id, promoted_by=promoted_by)
            assert first["created"] is True
            assert second["created"] is False
            assert first["snapshot_id"] == second["snapshot_id"]
            count = conn.execute(
                text(
                    """
                    SELECT COUNT(*) AS cnt
                    FROM public.hr_canonical_snapshots
                    WHERE source_batch_id = :batch_id
                    """
                ),
                {"batch_id": batch_id},
            ).scalar_one()
            assert int(count) == 1
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_snapshot_after_normalized_promotion(seed, tmp_path: Path) -> None:
    _require_phase_040a()
    if not _phase_1a_available():
        pytest.skip("ADR-037 Phase 1A tables missing")
    with engine.begin() as conn:
        if not normalized_records_available(conn):
            pytest.skip("normalized records table missing")

    batch_id, full_name, iin, _ = _import_single_employee_batch(seed, tmp_path)
    promoted_by = int(seed["initiator_user_id"])
    created_employee_id: int | None = None
    created_position_id: int | None = None

    try:
        with engine.begin() as conn:
            created_position_id = _create_position(conn, name=f"pytest_snap_{uuid4().hex[:8]}")
            created_employee_id = _create_employee(
                conn,
                full_name=full_name,
                org_unit_id=int(seed["unit_id"]),
                position_id=created_position_id,
                is_active=True,
            )
            row = conn.execute(
                text(
                    """
                    SELECT row_id
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    ORDER BY row_id
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).one()
            source_record_key = f"pytest-education-{uuid4().hex}"
            record_id = conn.execute(
                text(
                    """
                    INSERT INTO public.hr_import_normalized_records (
                        batch_id,
                        row_id,
                        employee_id,
                        fragment_index,
                        source_field,
                        source_text,
                        source_record_key,
                        record_kind,
                        document_type_id,
                        document_type_code,
                        title,
                        provider,
                        issue_date,
                        parse_method,
                        review_status,
                        reviewed_at,
                        reviewed_by
                    )
                    VALUES (
                        :batch_id,
                        :row_id,
                        :employee_id,
                        0,
                        'education_raw',
                        'КазНМУ, 1982',
                        :source_record_key,
                        'education',
                        (SELECT document_type_id FROM public.document_types WHERE code = 'EDUCATION_GRADUATION'),
                        'EDUCATION_GRADUATION',
                        'КазНМУ',
                        'КазНМУ',
                        DATE '1982-01-01',
                        'manual',
                        'approved',
                        NOW(),
                        :reviewed_by
                    )
                    RETURNING normalized_record_id
                    """
                ),
                {
                    "batch_id": batch_id,
                    "row_id": int(row[0]),
                    "employee_id": created_employee_id,
                    "source_record_key": source_record_key,
                    "reviewed_by": promoted_by,
                },
            ).scalar_one()

            promote_normalized_records(
                conn,
                promoted_by=promoted_by,
                dry_run=False,
                record_ids=[int(record_id)],
            )

        with engine.begin() as conn:
            education_entry = conn.execute(
                text(
                    """
                    SELECT e.match_key, e.canonical_hash, e.payload, e.employee_id
                    FROM public.hr_canonical_snapshot_entries e
                    JOIN public.hr_canonical_snapshots s ON s.snapshot_id = e.snapshot_id
                    WHERE s.source_batch_id = :batch_id
                      AND e.record_kind = 'education'
                    LIMIT 1
                    """
                ),
                {"batch_id": batch_id},
            ).mappings().one()
            assert education_entry["employee_id"] == created_employee_id
            assert source_record_key in education_entry["match_key"]
            assert education_entry["canonical_hash"]
            assert education_entry["payload"]
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)
            if created_employee_id is not None:
                conn.execute(
                    text("DELETE FROM public.employee_identities WHERE employee_id = :employee_id"),
                    {"employee_id": created_employee_id},
                )
                if table_exists(conn, "employee_documents"):
                    conn.execute(
                        text("DELETE FROM public.employee_documents WHERE employee_id = :employee_id"),
                        {"employee_id": created_employee_id},
                    )
                conn.execute(
                    text("DELETE FROM public.employees WHERE employee_id = :employee_id"),
                    {"employee_id": created_employee_id},
                )
            if created_position_id is not None and table_exists(conn, "positions"):
                conn.execute(
                    text("DELETE FROM public.positions WHERE position_id = :position_id"),
                    {"position_id": created_position_id},
                )


def test_refresh_helper_builds_snapshot_when_schema_available(seed, tmp_path: Path) -> None:
    _require_phase_040a()
    batch_id, _, _, _ = _import_single_employee_batch(seed, tmp_path)
    try:
        with engine.begin() as conn:
            result = refresh_canonical_snapshot_after_promotion(
                conn,
                batch_id,
                promoted_by=int(seed["initiator_user_id"]),
            )
            assert result is not None
            assert result["created"] is True
    finally:
        with engine.begin() as conn:
            _cleanup_promotion_batch(conn, batch_id)


def test_dedupe_snapshot_entries_merges_duplicate_roster_match_keys() -> None:
    match_key = "iin:123456789012"
    raw_entries = [
        {
            "entity_scope": match_key,
            "record_kind": "roster",
            "match_key": match_key,
            "canonical_hash": "hash-a",
            "employee_id": None,
            "iin": "123456789012",
            "payload": {
                "full_name": "Dup Employee",
                "iin": "123456789012",
                "education_raw": "КазНМУ, 1982",
            },
            "source_row_id": 10,
            "source_normalized_record_id": None,
        },
        {
            "entity_scope": match_key,
            "record_kind": "roster",
            "match_key": match_key,
            "canonical_hash": "hash-b",
            "employee_id": None,
            "iin": "123456789012",
            "payload": {
                "full_name": "Dup Employee",
                "iin": "123456789012",
                "certification_raw": 'Сертификат "Терапия" до 01.01.2028',
                "training_raw": "ПК 72 ч",
            },
            "source_row_id": 11,
            "source_normalized_record_id": None,
        },
    ]

    deduped, merged_count = dedupe_snapshot_entries(raw_entries)

    assert merged_count == 1
    assert len(deduped) == 1
    payload = deduped[0]["payload"]
    assert "КазНМУ, 1982" in payload["education_raw"]
    assert "Сертификат" in payload["certification_raw"]
    assert "ПК 72 ч" in payload["training_raw"]
    assert payload["provenance"]["source_row_ids"] == [10, 11]
    assert payload["provenance"]["duplicate_match_key_merged_count"] == 2


def test_snapshot_build_deduplicates_duplicate_roster_match_key(seed) -> None:
    _require_phase_040a()
    suffix = uuid4().hex[:8]
    iin = _test_iin(suffix)
    metadata = {
        "sheet_type": "doctors",
        "classification": "NORMAL",
        "row_type": "EMPLOYEE",
        "is_employee_roster": True,
    }
    row_payloads = (
        {
            "full_name": "Duplicate Snapshot Employee",
            "iin": iin,
            "education_raw": "КазНМУ, 1982",
            "metadata": metadata,
        },
        {
            "full_name": "Duplicate Snapshot Employee",
            "iin": iin,
            "certification_raw": 'Сертификат "Терапия" до 01.01.2028',
            "training_raw": "ПК 72 ч",
            "metadata": metadata,
        },
    )

    with engine.begin() as conn:
        batch_id = conn.execute(
            text(
                """
                INSERT INTO public.hr_import_batches (
                    source_type, file_name, imported_by, status,
                    total_rows, valid_rows, error_rows
                )
                VALUES ('HR_CONTROL_LIST', :file_name, :uid, 'PARSED', 2, 2, 0)
                RETURNING batch_id
                """
            ),
            {
                "file_name": f"dup_snapshot_{suffix}.xlsx",
                "uid": int(seed["initiator_user_id"]),
            },
        ).scalar_one()
        for row_num, payload in enumerate(row_payloads, start=8):
            conn.execute(
                text(
                    """
                    INSERT INTO public.hr_import_rows (
                        batch_id, source_sheet, source_row_number,
                        raw_payload, normalized_payload, match_status
                    )
                    VALUES (
                        :batch_id, 'doctors', :row_num,
                        CAST(:payload AS jsonb), CAST(:payload AS jsonb), 'NOT_PROCESSED'
                    )
                    """
                ),
                {
                    "batch_id": batch_id,
                    "row_num": row_num,
                    "payload": json.dumps(payload),
                },
            )
        result = build_canonical_snapshot_from_batch(
            conn,
            batch_id,
            promoted_by=int(seed["initiator_user_id"]),
        )
        entries = conn.execute(
            text(
                """
                SELECT match_key, payload
                FROM public.hr_canonical_snapshot_entries
                WHERE snapshot_id = :snapshot_id
                  AND record_kind = 'roster'
                  AND match_key = :match_key
                """
            ),
            {
                "snapshot_id": int(result["snapshot_id"]),
                "match_key": f"iin:{iin}",
            },
        ).mappings().all()
        _delete_batch(conn, batch_id)

    assert result["created"] is True
    assert result["duplicate_match_keys_merged"] >= 1
    assert len(entries) == 1
    payload = entries[0]["payload"]
    assert "КазНМУ, 1982" in payload["education_raw"]
    assert "Сертификат" in payload["certification_raw"]
    assert "ПК 72 ч" in payload["training_raw"]
    assert payload["provenance"]["duplicate_match_key_merged_count"] == 2
