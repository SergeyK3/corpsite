#!/usr/bin/env python3
"""ADR-040 Phase H — production verify: snapshot v2 + hr_change_events + Excel export.

Safe defaults: inspect/verify are read-only. Mutating commands require --execute.

Typical production flow (on VPS, from /opt/projects/corpsite/app):

  # 1) Current state
  .venv/bin/python3 scripts/adr040_production_phase_h_verify.py inspect

  # 2a) If a second import batch already exists (not yet promoted):
  .venv/bin/python3 scripts/adr040_production_phase_h_verify.py promote-batch \\
    --batch-id 40 --promoted-by 1 --execute

  # 2b) Or build controlled v2 from the original HR Excel + promote:
  .venv/bin/python3 scripts/adr040_production_phase_h_verify.py run-controlled-scenario \\
    --source-xlsx /path/to/control_list.xlsx --promoted-by 1 --execute

  # 3) Verify events + workbook
  .venv/bin/python3 scripts/adr040_production_phase_h_verify.py verify-events
  .venv/bin/python3 scripts/adr040_production_phase_h_verify.py verify-export \\
    --output /tmp/hr_registry_changes_verify.xlsx
"""
from __future__ import annotations

import argparse
import json
import shutil
import sys
from datetime import datetime, timezone
from io import BytesIO
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string
from sqlalchemy import text

ROOT = Path(__file__).resolve().parents[1]
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from app.db.engine import engine
from app.services.hr_canonical_snapshot_service import (
    SNAPSHOT_STATUS_ACTIVE,
    build_canonical_snapshot_from_batch,
    canonical_snapshot_available,
)
from app.services.hr_change_events_export_service import (
    CHANGE_TYPE_CHANGED,
    CHANGE_TYPE_CONFLICT,
    CHANGE_TYPE_NEW,
    CHANGE_TYPE_REMOVED,
    export_hr_change_events_xlsx,
)
from app.services.hr_import_monthly_diff_service import compute_batch_monthly_diff
from app.services.hr_import_service import import_control_list
from app.services.hr_snapshot_comparison_service import (
    hr_change_events_available,
    materialize_snapshot_change_events,
)

MARKER = "[ADR040-H]"
TEST_NEW_IIN = "990101300401"
TEST_CONFLICT_IIN = "990101300402"
CHANGED_POSITION_SUFFIX = " (ADR040-H CHANGED)"


def _print_json(payload: dict[str, Any]) -> None:
    print(json.dumps(payload, ensure_ascii=False, indent=2, default=str))


def inspect_state() -> dict[str, Any]:
    with engine.connect() as conn:
        snapshots = conn.execute(
            text(
                """
                SELECT snapshot_id, source_batch_id, version, status, entry_count, promoted_at
                FROM public.hr_canonical_snapshots
                ORDER BY snapshot_id
                """
            )
        ).mappings().all()
        events_total = 0
        events_by_type: dict[str, int] = {}
        if hr_change_events_available(conn):
            events_total = int(
                conn.execute(text("SELECT COUNT(*) FROM public.hr_change_events")).scalar_one()
            )
            rows = conn.execute(
                text(
                    """
                    SELECT event_type, COUNT(*) AS cnt
                    FROM public.hr_change_events
                    GROUP BY event_type
                    ORDER BY event_type
                    """
                )
            ).mappings().all()
            events_by_type = {str(r["event_type"]): int(r["cnt"]) for r in rows}

        batches = conn.execute(
            text(
                """
                SELECT b.batch_id, b.file_name, b.imported_at, b.status,
                       s.snapshot_id, s.version AS snapshot_version, s.status AS snapshot_status
                FROM public.hr_import_batches b
                LEFT JOIN public.hr_canonical_snapshots s
                  ON s.source_batch_id = b.batch_id
                ORDER BY b.batch_id DESC
                LIMIT 20
                """
            )
        ).mappings().all()

        diff_counts: dict[str, Any] = {}
        latest_unpromoted = None
        for batch in batches:
            if batch["snapshot_id"] is None:
                latest_unpromoted = int(batch["batch_id"])
                break

        if latest_unpromoted is not None:
            diff_rows = conn.execute(
                text(
                    """
                    SELECT COALESCE(diff_status, 'NULL') AS diff_status, COUNT(*) AS cnt
                    FROM public.hr_import_rows
                    WHERE batch_id = :batch_id
                    GROUP BY diff_status
                    ORDER BY diff_status
                    """
                ),
                {"batch_id": latest_unpromoted},
            ).mappings().all()
            diff_counts = {str(r["diff_status"]): int(r["cnt"]) for r in diff_rows}

    return {
        "snapshots": [dict(row) for row in snapshots],
        "hr_change_events_total": events_total,
        "hr_change_events_by_type": events_by_type,
        "recent_batches": [dict(row) for row in batches],
        "latest_unpromoted_batch_id": latest_unpromoted,
        "latest_unpromoted_diff_counts": diff_counts,
    }


def _active_snapshot(conn) -> Optional[dict[str, Any]]:
    row = conn.execute(
        text(
            """
            SELECT snapshot_id, source_batch_id, version, entry_count, promoted_at
            FROM public.hr_canonical_snapshots
            WHERE status = :status
            ORDER BY version DESC
            LIMIT 1
            """
        ),
        {"status": SNAPSHOT_STATUS_ACTIVE},
    ).mappings().first()
    return dict(row) if row else None


def _pick_roster_targets(conn, *, limit: int = 3) -> list[dict[str, Any]]:
    rows = conn.execute(
        text(
            """
            SELECT e.entry_id, e.match_key, e.iin, e.payload->>'full_name' AS full_name,
                   e.payload->>'position_raw' AS position_raw,
                   e.payload->>'department' AS department
            FROM public.hr_canonical_snapshot_entries e
            JOIN public.hr_canonical_snapshots s ON s.snapshot_id = e.snapshot_id
            WHERE s.status = :status
              AND e.record_kind = 'roster'
              AND COALESCE(e.iin, e.payload->>'iin', '') <> ''
            ORDER BY e.entry_id
            LIMIT :limit
            """
        ),
        {"status": SNAPSHOT_STATUS_ACTIVE, "limit": limit},
    ).mappings().all()
    return [dict(row) for row in rows]


def _find_iin_cells(ws) -> list[tuple[int, int]]:
    hits: list[tuple[int, int]] = []
    for row_idx in range(1, ws.max_row + 1):
        for col_idx in range(1, ws.max_column + 1):
            value = ws.cell(row=row_idx, column=col_idx).value
            if value is None:
                continue
            text_val = str(value).strip()
            if len(text_val) == 12 and text_val.isdigit():
                hits.append((row_idx, col_idx))
    return hits


def _row_has_iin(ws, target_iin: str) -> Optional[int]:
    target = str(target_iin).strip()
    for row_idx, col_idx in _find_iin_cells(ws):
        if str(ws.cell(row=row_idx, column=col_idx).value).strip() == target:
            return row_idx
    return None


def _position_col_near_iin(ws, iin_row: int, iin_col: int) -> int:
    for col_idx in range(iin_col + 1, min(iin_col + 8, ws.max_column + 1)):
        header = str(ws.cell(row=7, column=col_idx).value or "").lower()
        if "должност" in header or col_idx == column_index_from_string("J"):
            return col_idx
    return min(iin_col + 5, ws.max_column)


def prepare_controlled_v2_workbook(
    *,
    source_xlsx: Path,
    output_xlsx: Path,
    change_target: dict[str, Any],
    remove_target: dict[str, Any],
) -> dict[str, Any]:
    shutil.copy2(source_xlsx, output_xlsx)
    wb = load_workbook(output_xlsx)
    ws = wb[wb.sheetnames[0]]

    changed_row = _row_has_iin(ws, str(change_target["iin"]))
    removed_row = _row_has_iin(ws, str(remove_target["iin"]))
    if changed_row is None or removed_row is None:
        wb.close()
        raise RuntimeError(
            f"Could not locate change/remove IIN rows in workbook: "
            f"change={change_target['iin']} remove={remove_target['iin']}"
        )

    iin_col = None
    for row_idx, col_idx in _find_iin_cells(ws):
        if row_idx == changed_row:
            iin_col = col_idx
            break
    if iin_col is None:
        wb.close()
        raise RuntimeError("Could not resolve IIN column")

    pos_col = _position_col_near_iin(ws, changed_row, iin_col)
    old_position = str(ws.cell(row=changed_row, column=pos_col).value or change_target.get("position_raw") or "")
    new_position = (old_position + CHANGED_POSITION_SUFFIX).strip()
    ws.cell(row=changed_row, column=pos_col, value=new_position)

    ws.delete_rows(removed_row, 1)

    next_row = ws.max_row + 1
    ws.cell(row=next_row, column=3, value=f"{MARKER} NEW employee")
    ws.cell(row=next_row, column=iin_col, value=TEST_NEW_IIN)
    ws.cell(row=next_row, column=max(1, iin_col - 1), value=department_or_default(change_target))
    ws.cell(row=next_row, column=pos_col, value="ADR040-H test position")

    conflict_row_1 = next_row + 1
    conflict_row_2 = next_row + 2
    dept = department_or_default(change_target)
    for offset, name in ((conflict_row_1, f"{MARKER} CONFLICT A"), (conflict_row_2, f"{MARKER} CONFLICT B")):
        ws.cell(row=offset, column=3, value=name)
        ws.cell(row=offset, column=iin_col, value=TEST_CONFLICT_IIN)
        ws.cell(row=offset, column=max(1, iin_col - 1), value=dept)
        ws.cell(row=offset, column=pos_col, value="ADR040-H conflict position")

    wb.save(output_xlsx)
    wb.close()
    return {
        "output_xlsx": str(output_xlsx),
        "changed": {"iin": change_target["iin"], "row": changed_row, "new_position": new_position},
        "removed": {"iin": remove_target["iin"], "row": removed_row},
        "added_new_iin": TEST_NEW_IIN,
        "added_conflict_iin": TEST_CONFLICT_IIN,
    }


def department_or_default(target: dict[str, Any]) -> str:
    dept = str(target.get("department") or "").strip()
    return dept or "ADR040-H test department"


def compute_diff(*, batch_id: int, execute: bool) -> dict[str, Any]:
    if not execute:
        return {"dry_run": True, "would_compute_diff_for_batch_id": batch_id}
    with engine.begin() as conn:
        result = compute_batch_monthly_diff(conn, batch_id)
    return {"dry_run": False, "batch_id": batch_id, "diff_summary": result.get("summary"), "result": result}


def promote_batch(*, batch_id: int, promoted_by: int, execute: bool) -> dict[str, Any]:
    if not execute:
        return {"dry_run": True, "would_promote_batch_id": batch_id, "promoted_by": promoted_by}

    with engine.begin() as conn:
        if not canonical_snapshot_available(conn):
            raise RuntimeError("hr_canonical_snapshots is not available")
        existing = conn.execute(
            text(
                """
                SELECT snapshot_id, version, status
                FROM public.hr_canonical_snapshots
                WHERE source_batch_id = :batch_id
                """
            ),
            {"batch_id": batch_id},
        ).mappings().first()
        if existing:
            return {
                "dry_run": False,
                "skipped": True,
                "reason": "batch already promoted to canonical snapshot",
                "existing_snapshot": dict(existing),
                "hint": "Use a different import batch for v2, or run-controlled-scenario",
            }
        prior = _active_snapshot(conn)
        result = build_canonical_snapshot_from_batch(
            conn,
            batch_id,
            promoted_by=promoted_by,
        )
        after = _active_snapshot(conn)
        events = {}
        if hr_change_events_available(conn) and prior and after:
            events = conn.execute(
                text(
                    """
                    SELECT event_type, COUNT(*) AS cnt
                    FROM public.hr_change_events
                    WHERE prior_snapshot_id = :prior_id AND new_snapshot_id = :new_id
                    GROUP BY event_type
                    ORDER BY event_type
                    """
                ),
                {
                    "prior_id": int(prior["snapshot_id"]),
                    "new_id": int(after["snapshot_id"]),
                },
            ).mappings().all()
        return {
            "dry_run": False,
            "promotion_result": result,
            "active_snapshot": after,
            "materialized_events_by_type": {str(r["event_type"]): int(r["cnt"]) for r in events},
        }


def run_controlled_scenario(
    *,
    source_xlsx: Path,
    promoted_by: int,
    execute: bool,
    work_dir: Path,
) -> dict[str, Any]:
    work_dir.mkdir(parents=True, exist_ok=True)
    prepared = work_dir / f"adr040_v2_{datetime.now(timezone.utc).strftime('%Y%m%d_%H%M%S')}.xlsx"

    with engine.connect() as conn:
        targets = _pick_roster_targets(conn, limit=3)
        if len(targets) < 2:
            raise RuntimeError("Need at least 2 roster entries in active snapshot")

    prep_info = prepare_controlled_v2_workbook(
        source_xlsx=source_xlsx,
        output_xlsx=prepared,
        change_target=targets[0],
        remove_target=targets[1],
    )

    if not execute:
        return {
            "dry_run": True,
            "prepared_workbook": str(prepared),
            "prep_info": prep_info,
            "next_step": f"Re-run with --execute to import and promote",
        }

    with engine.begin() as conn:
        batch_id, summary, warnings = import_control_list(
            conn,
            file_path=prepared,
            imported_by=promoted_by,
        )
        diff = compute_batch_monthly_diff(conn, int(batch_id))
        diff_summary = diff.get("summary") or {}

    promotion = promote_batch(batch_id=int(batch_id), promoted_by=promoted_by, execute=True)
    return {
        "dry_run": False,
        "prepared_workbook": str(prepared),
        "prep_info": prep_info,
        "import_batch_id": int(batch_id),
        "import_summary": summary,
        "import_warnings": warnings,
        "monthly_diff_summary": diff_summary,
        "promotion": promotion,
    }


def verify_events() -> dict[str, Any]:
    with engine.connect() as conn:
        if not hr_change_events_available(conn):
            return {"ok": False, "reason": "hr_change_events table not available"}
        total = int(conn.execute(text("SELECT COUNT(*) FROM public.hr_change_events")).scalar_one())
        by_type = conn.execute(
            text(
                """
                SELECT event_type, COUNT(*) AS cnt
                FROM public.hr_change_events
                GROUP BY event_type
                ORDER BY event_type
                """
            )
        ).mappings().all()
        pairs = conn.execute(
            text(
                """
                SELECT prior_snapshot_id, new_snapshot_id, COUNT(*) AS cnt
                FROM public.hr_change_events
                GROUP BY prior_snapshot_id, new_snapshot_id
                ORDER BY cnt DESC
                """
            )
        ).mappings().all()
    mapped_changed = sum(
        int(row["cnt"])
        for row in by_type
        if str(row["event_type"]) not in {"NEW", "REMOVED"}
    )
    return {
        "ok": total > 0,
        "total": total,
        "by_event_type": {str(r["event_type"]): int(r["cnt"]) for r in by_type},
        "changed_like_events": mapped_changed,
        "snapshot_pairs": [dict(r) for r in pairs],
    }


def verify_export(*, output_path: Optional[Path]) -> dict[str, Any]:
    from openpyxl import load_workbook

    with engine.begin() as conn:
        content, filename = export_hr_change_events_xlsx(conn)

    path = output_path or (ROOT / "tmp" / filename)
    path.parent.mkdir(parents=True, exist_ok=True)
    path.write_bytes(content)

    wb = load_workbook(BytesIO(content), read_only=True)
    sheet_counts = {
        name: max(0, (wb[name].max_row or 1) - 1)
        for name in wb.sheetnames
    }
    wb.close()

    return {
        "ok": any(sheet_counts.get(s, 0) > 0 for s in ("NEW", "CHANGED", "REMOVED", "CONFLICT")),
        "filename": filename,
        "output_path": str(path),
        "sheet_counts": sheet_counts,
        "expected_change_types_present": {
            CHANGE_TYPE_NEW: sheet_counts.get("NEW", 0) > 0,
            CHANGE_TYPE_CHANGED: sheet_counts.get("CHANGED", 0) > 0,
            CHANGE_TYPE_REMOVED: sheet_counts.get("REMOVED", 0) > 0,
            CHANGE_TYPE_CONFLICT: sheet_counts.get("CONFLICT", 0) > 0,
        },
    }


def rematerialize_pair(*, prior_snapshot_id: int, new_snapshot_id: int, execute: bool) -> dict[str, Any]:
    if not execute:
        return {
            "dry_run": True,
            "would_materialize": {
                "prior_snapshot_id": prior_snapshot_id,
                "new_snapshot_id": new_snapshot_id,
            },
        }
    with engine.begin() as conn:
        result = materialize_snapshot_change_events(
            conn,
            prior_snapshot_id=prior_snapshot_id,
            new_snapshot_id=new_snapshot_id,
        )
    return {"dry_run": False, "result": result}


def main() -> int:
    parser = argparse.ArgumentParser(description="ADR-040 Phase H production verification")
    sub = parser.add_subparsers(dest="command", required=True)

    sub.add_parser("inspect", help="Read-only state summary")

    p_promote = sub.add_parser("promote-batch", help="Promote import batch to canonical snapshot vN+1")
    p_promote.add_argument("--batch-id", type=int, required=True)
    p_promote.add_argument("--promoted-by", type=int, required=True, help="users.user_id")
    p_promote.add_argument("--execute", action="store_true")

    p_diff = sub.add_parser("compute-diff", help="Run monthly diff for a batch")
    p_diff.add_argument("--batch-id", type=int, required=True)
    p_diff.add_argument("--execute", action="store_true")

    p_run = sub.add_parser("run-controlled-scenario", help="Prepare v2 xlsx, import, diff, promote")
    p_run.add_argument("--source-xlsx", type=Path, required=True)
    p_run.add_argument("--promoted-by", type=int, required=True)
    p_run.add_argument("--work-dir", type=Path, default=ROOT / "tmp" / "adr040_prod")
    p_run.add_argument("--execute", action="store_true")

    sub.add_parser("verify-events", help="Check hr_change_events counts")
    p_export = sub.add_parser("verify-export", help="Generate changes workbook and summarize sheets")
    p_export.add_argument("--output", type=Path, default=None)

    p_remat = sub.add_parser("rematerialize", help="Re-run materialize for an existing snapshot pair")
    p_remat.add_argument("--prior-snapshot-id", type=int, required=True)
    p_remat.add_argument("--new-snapshot-id", type=int, required=True)
    p_remat.add_argument("--execute", action="store_true")

    args = parser.parse_args()

    if args.command == "inspect":
        _print_json(inspect_state())
        return 0
    if args.command == "compute-diff":
        _print_json(compute_diff(batch_id=args.batch_id, execute=args.execute))
        return 0
    if args.command == "promote-batch":
        _print_json(promote_batch(batch_id=args.batch_id, promoted_by=args.promoted_by, execute=args.execute))
        return 0
    if args.command == "run-controlled-scenario":
        if not args.source_xlsx.is_file():
            print(f"source xlsx not found: {args.source_xlsx}", file=sys.stderr)
            return 2
        _print_json(
            run_controlled_scenario(
                source_xlsx=args.source_xlsx,
                promoted_by=args.promoted_by,
                execute=args.execute,
                work_dir=args.work_dir,
            )
        )
        return 0
    if args.command == "verify-events":
        _print_json(verify_events())
        return 0
    if args.command == "verify-export":
        _print_json(verify_export(output_path=args.output))
        return 0
    if args.command == "rematerialize":
        _print_json(
            rematerialize_pair(
                prior_snapshot_id=args.prior_snapshot_id,
                new_snapshot_id=args.new_snapshot_id,
                execute=args.execute,
            )
        )
        return 0
    return 1


if __name__ == "__main__":
    raise SystemExit(main())
