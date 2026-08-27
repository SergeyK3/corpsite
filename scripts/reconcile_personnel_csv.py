"""Read-only reconciliation of tmp/сотр слиян.csv against the local Corpsite DB.

The script never writes to PostgreSQL: all reads run in an explicit
READ ONLY transaction.  Its only output is tmp/personnel_reconciliation_report.csv.
"""
from __future__ import annotations

import csv
import re
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path
from typing import Any

from openpyxl import load_workbook
from sqlalchemy import text

from app.db.engine import engine


ROOT = Path(__file__).resolve().parents[1]
INPUT_CSV = ROOT / "tmp" / "сотр слиян.csv"
INPUT_XLSX = ROOT / "tmp" / "спрКорпсайт.xlsx"
OUTPUT_CSV = ROOT / "tmp" / "personnel_reconciliation_report.csv"
DEPARTMENT_SHEET = "конвер справ отделов"

STATUS_MATCH = "уже существует, данные совпадают"
STATUS_MISSING = "уже существует, имеются незаполненные поля"
STATUS_EXISTING = "существующий сотрудник"
STATUS_CONFLICT = "конфликт с существующими данными"
STATUS_AMBIGUOUS = "неоднозначное совпадение"
STATUS_POSSIBLE = "возможный существующий"
STATUS_NEW = "новый сотрудник"
STATUS_ERROR = "ошибка справочника или исходных данных"

REPORT_COLUMNS = (
    "csv_row",
    "full_name",
    "iin",
    "birth_date",
    "position",
    "department",
    "personnel_number",
    "phone",
    "date_from",
    "csv_duplicate_key",
    "employee_id",
    "person_id",
    "candidate_employee_ids",
    "candidate_person_ids",
    "fio_match_count",
    "match_method",
    "found_department_unit_id",
    "found_department_name",
    "found_department_code",
    "found_position_id",
    "found_position_name",
    "database_full_name",
    "database_iin",
    "database_birth_date",
    "database_org_unit_id",
    "database_department_name",
    "database_position_id",
    "database_position_name",
    "differences",
    "employment_differences",
    "missing_fields",
    "status",
    "notes",
)


def text_value(value: Any) -> str:
    return "" if value is None else str(value).strip()


def normalized(value: Any) -> str:
    value = unicodedata.normalize("NFKC", text_value(value)).casefold().replace("ё", "е")
    return re.sub(r"[^\w]+", "", value, flags=re.UNICODE)


def normalized_iin(value: Any) -> str:
    return re.sub(r"\D", "", text_value(value))


def parse_date(value: Any) -> date | None:
    raw = text_value(value)
    if not raw:
        return None
    for pattern in ("%d.%m.%Y", "%Y-%m-%d", "%d/%m/%Y"):
        try:
            return datetime.strptime(raw, pattern).date()
        except ValueError:
            pass
    return None


def date_text(value: Any) -> str:
    return value.isoformat() if isinstance(value, date) else text_value(value)


def load_department_map() -> tuple[dict[str, dict[str, Any]], set[str]]:
    workbook = load_workbook(INPUT_XLSX, read_only=True, data_only=True)
    try:
        if DEPARTMENT_SHEET not in workbook.sheetnames:
            raise RuntimeError(f"В XLSX отсутствует вкладка «{DEPARTMENT_SHEET}»")
        sheet = workbook[DEPARTMENT_SHEET]
        mapping_candidates: dict[str, list[dict[str, Any]]] = defaultdict(list)
        for values in sheet.iter_rows(min_row=2, values_only=True):
            if len(values) < 4:
                continue
            unit_id, name, code, source_department = values[:4]
            source_key = normalized(source_department)
            if not source_key:
                continue
            try:
                numeric_unit_id = int(unit_id)
            except (TypeError, ValueError):
                continue
            mapping_candidates[source_key].append(
                {
                    "unit_id": numeric_unit_id,
                    "name": text_value(name),
                    "code": text_value(code),
                    "source_department": text_value(source_department),
                }
            )
    finally:
        workbook.close()

    mapping: dict[str, dict[str, Any]] = {}
    ambiguous: set[str] = set()
    for key, candidates in mapping_candidates.items():
        distinct = {(row["unit_id"], row["name"], row["code"]) for row in candidates}
        if len(distinct) == 1:
            mapping[key] = candidates[0]
        else:
            ambiguous.add(key)
    return mapping, ambiguous


def choose_employee(candidates: list[dict[str, Any]]) -> dict[str, Any] | None:
    if not candidates:
        return None
    return sorted(
        candidates,
        key=lambda row: (
            bool(row.get("is_active")),
            str(row.get("updated_at") or ""),
            row["employee_id"],
        ),
        reverse=True,
    )[0]


def load_database() -> dict[str, Any]:
    with engine.connect() as connection:
        transaction = connection.begin()
        try:
            connection.execute(text("SET TRANSACTION READ ONLY"))
            if connection.execute(text("SHOW transaction_read_only")).scalar_one() != "on":
                raise RuntimeError("PostgreSQL transaction is not read-only")

            people = connection.execute(
                text(
                    """
                    SELECT person_id, iin, full_name, birth_date
                    FROM public.persons
                    ORDER BY person_id
                    """
                )
            ).mappings().all()
            employees = connection.execute(
                text(
                    """
                    SELECT employee_id, person_id, full_name, org_unit_id, position_id,
                           date_from, is_active, updated_at
                    FROM public.employees
                    ORDER BY employee_id
                    """
                )
            ).mappings().all()
            identities = connection.execute(
                text(
                    """
                    SELECT employee_id, identity_value
                    FROM public.employee_identities
                    WHERE identity_type = 'IIN' AND valid_to IS NULL
                    ORDER BY employee_id
                    """
                )
            ).mappings().all()
            units = connection.execute(
                text(
                    """
                    SELECT unit_id, name, code
                    FROM public.org_units
                    ORDER BY unit_id
                    """
                )
            ).mappings().all()
            positions = connection.execute(
                text("SELECT position_id, name FROM public.positions ORDER BY position_id")
            ).mappings().all()
        finally:
            transaction.rollback()

    person_by_id = {int(row["person_id"]): dict(row) for row in people}
    employees_by_person: dict[int, list[dict[str, Any]]] = defaultdict(list)
    employees_by_id = {int(row["employee_id"]): dict(row) for row in employees}
    for row in employees:
        item = dict(row)
        if item["person_id"] is not None:
            employees_by_person[int(item["person_id"])].append(item)

    people_by_iin: dict[str, set[int]] = defaultdict(set)
    people_by_name_birth: dict[tuple[str, date], set[int]] = defaultdict(set)
    people_by_name: dict[str, set[int]] = defaultdict(set)
    employees_by_name: dict[str, set[int]] = defaultdict(set)
    person_iins: dict[int, set[str]] = defaultdict(set)
    for person_id, person in person_by_id.items():
        iin = normalized_iin(person.get("iin"))
        if len(iin) == 12:
            people_by_iin[iin].add(person_id)
            person_iins[person_id].add(iin)
        key_name = normalized(person.get("full_name"))
        if key_name:
            people_by_name[key_name].add(person_id)
        if key_name and person.get("birth_date") is not None:
            people_by_name_birth[(key_name, person["birth_date"])].add(person_id)

    employee_to_person = {
        int(row["employee_id"]): int(row["person_id"])
        for row in employees
        if row["person_id"] is not None
    }
    for identity in identities:
        iin = normalized_iin(identity["identity_value"])
        person_id = employee_to_person.get(int(identity["employee_id"]))
        if len(iin) == 12 and person_id is not None:
            people_by_iin[iin].add(person_id)
            person_iins[person_id].add(iin)
    for employee_id, employee in employees_by_id.items():
        key_name = normalized(employee.get("full_name"))
        if key_name:
            employees_by_name[key_name].add(employee_id)

    units_by_id = {int(row["unit_id"]): dict(row) for row in units}
    positions_by_key: dict[str, list[dict[str, Any]]] = defaultdict(list)
    positions_by_id = {int(row["position_id"]): dict(row) for row in positions}
    for row in positions:
        key = normalized(row["name"])
        if key:
            positions_by_key[key].append(dict(row))

    return {
        "person_by_id": person_by_id,
        "employees_by_person": employees_by_person,
        "people_by_iin": people_by_iin,
        "people_by_name_birth": people_by_name_birth,
        "people_by_name": people_by_name,
        "employees_by_name": employees_by_name,
        "employees_by_id": employees_by_id,
        "person_iins": person_iins,
        "units_by_id": units_by_id,
        "positions_by_key": positions_by_key,
        "positions_by_id": positions_by_id,
    }


def csv_duplicate_keys(rows: list[dict[str, str]]) -> tuple[list[str], Counter[str]]:
    keys: list[str] = []
    for row in rows:
        iin = normalized_iin(row.get("ИИН"))
        birth_date = parse_date(row.get("Дата рождения"))
        name = normalized(row.get("ФИО"))
        if len(iin) == 12:
            keys.append(f"iin:{iin}")
        elif name and birth_date:
            keys.append(f"fio_birth:{name}|{birth_date.isoformat()}")
        else:
            keys.append("")
    counts = Counter(key for key in keys if key)
    return keys, counts


def reconcile_row(
    row: dict[str, str],
    row_number: int,
    duplicate_key: str,
    duplicate_counts: Counter[str],
    department_map: dict[str, dict[str, Any]],
    ambiguous_department_keys: set[str],
    database: dict[str, Any],
) -> dict[str, str]:
    report = {column: "" for column in REPORT_COLUMNS}
    report.update(
        {
            "csv_row": str(row_number),
            "full_name": text_value(row.get("ФИО")),
            "iin": text_value(row.get("ИИН")),
            "birth_date": text_value(row.get("Дата рождения")),
            "position": text_value(row.get("должность")),
            "department": text_value(row.get("отдел")),
            "personnel_number": text_value(row.get("ТабНомер")),
            "phone": text_value(row.get("телефон")),
            "date_from": text_value(row.get("date_from")),
        }
    )
    errors: list[str] = []
    notes: list[str] = []
    missing: list[str] = []
    identity_differences: list[str] = []
    employment_differences: list[str] = []

    name_key = normalized(report["full_name"])
    iin_key = normalized_iin(report["iin"])
    birth_date = parse_date(report["birth_date"])
    source_date_from = parse_date(report["date_from"])
    if not name_key:
        errors.append("не заполнено ФИО")
    if report["birth_date"] and birth_date is None:
        errors.append("некорректная дата рождения")
    if report["date_from"] and source_date_from is None:
        errors.append("некорректный date_from")
    if report["iin"] and len(iin_key) != 12:
        errors.append("ИИН должен содержать 12 цифр")

    department_key = normalized(report["department"])
    department = None
    if not department_key:
        errors.append("не заполнено отделение")
    elif department_key in ambiguous_department_keys:
        errors.append("неоднозначное сопоставление отделения в XLSX")
    elif department_key not in department_map:
        errors.append("отделение отсутствует в первых 4 столбцах XLSX")
    else:
        candidate = department_map[department_key]
        db_unit = database["units_by_id"].get(candidate["unit_id"])
        if db_unit is None:
            errors.append(f"unit_id={candidate['unit_id']} из XLSX отсутствует в public.org_units")
        else:
            department = db_unit
            report["found_department_unit_id"] = str(db_unit["unit_id"])
            report["found_department_name"] = text_value(db_unit["name"])
            report["found_department_code"] = text_value(db_unit["code"])
            if normalized(db_unit["code"]) != normalized(candidate["code"]):
                errors.append(f"код отделения XLSX не совпадает с public.org_units для unit_id={candidate['unit_id']}")

    position_key = normalized(report["position"])
    position = None
    if not position_key:
        errors.append("не заполнена должность")
    else:
        candidates = database["positions_by_key"].get(position_key, [])
        if not candidates:
            errors.append("должность отсутствует в public.positions")
        elif len(candidates) > 1:
            errors.append("должность неоднозначна в public.positions")
        else:
            position = candidates[0]
            report["found_position_id"] = str(position["position_id"])
            report["found_position_name"] = text_value(position["name"])

    if duplicate_key and duplicate_counts[duplicate_key] > 1:
        report["csv_duplicate_key"] = duplicate_key
        notes.append(f"дубль внутри CSV: {duplicate_counts[duplicate_key]} строк")

    person_ids: set[int] = set()
    fio_person_ids = set(database["people_by_name"].get(name_key, set())) if name_key else set()
    fio_employee_ids = set(database["employees_by_name"].get(name_key, set())) if name_key else set()
    fio_employee_person_ids = {
        int(database["employees_by_id"][employee_id]["person_id"])
        for employee_id in fio_employee_ids
        if database["employees_by_id"][employee_id]["person_id"] is not None
    }
    fio_person_ids.update(fio_employee_person_ids)
    report["fio_match_count"] = str(len(fio_employee_ids) or len(fio_person_ids))
    match_method = ""
    if len(iin_key) == 12:
        person_ids = set(database["people_by_iin"].get(iin_key, set()))
        if person_ids:
            match_method = "точный ИИН"
        elif fio_person_ids or fio_employee_ids:
            person_ids = set(fio_person_ids)
            match_method = "ФИО (возможное совпадение; ИИН не найден)"
        else:
            match_method = "точный ИИН: не найден"
    elif name_key:
        person_ids = set(fio_person_ids)
        if fio_person_ids or fio_employee_ids:
            match_method = "нормализованное ФИО (возможное совпадение)"
        elif birth_date is not None:
            match_method = "ФИО + дата рождения: не найден"
        else:
            match_method = "ФИО: не найден"
        notes.append("табельный номер не использован: в канонических таблицах БД нет такого поля")
    else:
        match_method = "недостаточно идентификаторов"
        if not report["iin"]:
            errors.append("без ИИН требуется ФИО и дата рождения; табельный номер в БД отсутствует")
    report["match_method"] = match_method

    person_candidates = [database["person_by_id"][person_id] for person_id in sorted(person_ids)]
    employee_candidates: list[dict[str, Any]] = []
    for person_id in sorted(person_ids):
        employee_candidates.extend(database["employees_by_person"].get(person_id, []))
    report["candidate_person_ids"] = ";".join(str(row["person_id"]) for row in person_candidates)
    report["candidate_employee_ids"] = ";".join(str(row["employee_id"]) for row in employee_candidates)
    if not employee_candidates and fio_employee_ids:
        employee_candidates = [database["employees_by_id"][employee_id] for employee_id in sorted(fio_employee_ids)]
        report["candidate_employee_ids"] = ";".join(str(row["employee_id"]) for row in employee_candidates)

    person = person_candidates[0] if len(person_ids) == 1 else None
    employee = choose_employee(employee_candidates)
    unique_fio_employee = len(fio_employee_ids) == 1
    existing_by_fio = match_method.startswith(("нормализованное ФИО", "ФИО (")) and (
        person is not None or unique_fio_employee
    )
    if person is not None:
        report["person_id"] = str(person["person_id"])
        report["database_full_name"] = text_value(person["full_name"])
        report["database_iin"] = text_value(person["iin"])
        report["database_birth_date"] = date_text(person["birth_date"])
        known_iins = database["person_iins"].get(int(person["person_id"]), set())
        if len(iin_key) == 12 and known_iins and iin_key not in known_iins:
            identity_differences.append("ИИН")
        if birth_date is not None and person["birth_date"] is not None and person["birth_date"] != birth_date:
            identity_differences.append("дата рождения")
    if employee is not None:
        report["employee_id"] = str(employee["employee_id"])
        report["database_full_name"] = report["database_full_name"] or text_value(employee["full_name"])
        report["database_org_unit_id"] = text_value(employee["org_unit_id"])
        report["database_position_id"] = text_value(employee["position_id"])
        if employee["org_unit_id"] is not None:
            db_unit = database["units_by_id"].get(int(employee["org_unit_id"]))
            report["database_department_name"] = text_value(db_unit["name"]) if db_unit else "<не найдено>"
        if employee["position_id"] is not None:
            db_position = database["positions_by_id"].get(int(employee["position_id"]))
            report["database_position_name"] = text_value(db_position["name"]) if db_position else "<не найдено>"
        if department is not None and employee["org_unit_id"] is not None and int(employee["org_unit_id"]) != int(department["unit_id"]):
            employment_differences.append("отделение")
        if position is not None and employee["position_id"] is not None and int(employee["position_id"]) != int(position["position_id"]):
            employment_differences.append("должность")
        if source_date_from is not None and employee["date_from"] is not None and employee["date_from"] != source_date_from:
            employment_differences.append("date_from")

    report["differences"] = "; ".join(identity_differences)
    report["employment_differences"] = "; ".join(employment_differences)
    report["missing_fields"] = "; ".join(missing)
    report["notes"] = "; ".join(notes)

    if duplicate_key and duplicate_counts[duplicate_key] > 1:
        report["status"] = STATUS_AMBIGUOUS
    elif len(person_ids) > 1 or len(fio_employee_ids) > 1:
        report["status"] = STATUS_AMBIGUOUS
        report["notes"] = "; ".join([report["notes"], "несколько сотрудников с одинаковым ФИО"]).strip("; ")
    elif person is not None and identity_differences:
        report["status"] = STATUS_CONFLICT
    elif existing_by_fio or len(person_ids) == 1:
        report["status"] = STATUS_EXISTING
    elif errors:
        report["status"] = STATUS_ERROR
        report["notes"] = "; ".join([*errors, *notes])
    else:
        report["status"] = STATUS_NEW
    return report


def main() -> None:
    if not INPUT_CSV.is_file() or not INPUT_XLSX.is_file():
        raise SystemExit("Не найдены входные файлы в tmp")
    with INPUT_CSV.open("r", encoding="utf-8-sig", newline="") as source:
        rows = list(csv.DictReader(source, delimiter=";"))
    required = {"ФИО", "ИИН", "Дата рождения", "должность", "отдел", "ТабНомер", "телефон", "date_from"}
    missing_headers = sorted(required.difference(rows[0].keys() if rows else set()))
    if missing_headers:
        raise SystemExit(f"В CSV отсутствуют столбцы: {', '.join(missing_headers)}")

    department_map, ambiguous_department_keys = load_department_map()
    database = load_database()
    duplicate_keys, duplicate_counts = csv_duplicate_keys(rows)
    report_rows = [
        reconcile_row(
            row,
            index + 2,
            duplicate_keys[index],
            duplicate_counts,
            department_map,
            ambiguous_department_keys,
            database,
        )
        for index, row in enumerate(rows)
    ]
    with OUTPUT_CSV.open("w", encoding="utf-8-sig", newline="") as output:
        writer = csv.DictWriter(output, fieldnames=REPORT_COLUMNS, extrasaction="ignore")
        writer.writeheader()
        writer.writerows(report_rows)

    counts = Counter(row["status"] for row in report_rows)
    fio_intersections = sum(1 for row in report_rows if int(row["fio_match_count"] or 0) > 0)
    print(f"Отчёт: {OUTPUT_CSV}")
    print("Строк CSV:", len(report_rows))
    print("Сотрудников в локальной БД:", len(database["employees_by_id"]))
    print("Точных пересечений по нормализованному ФИО:", fio_intersections)
    print("Причина прежних 0 совпадений: использовались только точный ИИН и ФИО + дата рождения; ФИО без даты рождения не сопоставлялось.")
    for status in (STATUS_EXISTING, STATUS_CONFLICT, STATUS_AMBIGUOUS, STATUS_NEW, STATUS_ERROR):
        print(f"{status}: {counts[status]}")


if __name__ == "__main__":
    main()
