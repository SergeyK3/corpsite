#!/usr/bin/env python3
"""HR control list Excel → staging preview + audit report (Phase 0, dry-run only)."""
from __future__ import annotations

import argparse
import csv
import re
import sys
from collections import Counter
from dataclasses import dataclass, field
from datetime import date, datetime, timedelta
from pathlib import Path
from typing import Any, Iterable, Optional

from openpyxl import load_workbook
from openpyxl.utils import column_index_from_string

PROJECT_ROOT = Path(__file__).resolve().parents[1]
DEFAULT_OUT_DIR = PROJECT_ROOT / "tmp"

OUTPUT_FIELDS = [
    "full_name",
    "iin",
    "birth_date",
    "sex",
    "nationality",
    "department",
    "position_raw",
    "education_raw",
    "diploma_specialty_raw",
    "qualification_raw",
    "experience_raw",
    "training_raw",
    "education_training_raw",
    "certification_raw",
    "degree_raw",
    "awards_raw",
    "note_raw",
    "phone_raw",
    "source_sheet",
    "source_row_number",
]

HEADER_ALIASES: dict[str, list[str]] = {
    "full_name": [
        "full_name",
        "фио",
        "ф.и.о.",
        "ф. и. о.",
        "фамилия имя отчество",
        "фамилия, имя, отчество",
        "сотрудник",
    ],
    "iin": ["iin", "иин", "иин/бин", "иин бин"],
    "birth_date": ["birth_date", "дата рождения", "год рождения", "д.р.", "др", "дата рожд"],
    "sex": ["sex", "пол"],
    "nationality": ["nationality", "национальность", "гражданство"],
    "department": [
        "department",
        "отделение",
        "подразделение",
        "структурное подразделение",
        "отдел",
    ],
    "position_raw": [
        "position_raw",
        "position",
        "должность",
        "занимаемая должность",
        "занимаемая должность, дата назнач.",
    ],
    "education_raw": [
        "education_raw",
        "education",
        "образование",
        "вуз, год окончания",
        "вуз год окончания",
    ],
    "diploma_specialty_raw": [
        "diploma_specialty_raw",
        "specialty",
        "специальность по диплому",
        "специальность",
    ],
    "qualification_raw": [
        "qualification_raw",
        "qualification",
        "квалификация",
        "категория должности",
        "квалификация по диплому",
    ],
    "experience_raw": [
        "experience_raw",
        "experience",
        "стаж",
        "стаж работы",
        "общий стаж",
        "трудовой стаж",
    ],
    "education_training_raw": [
        "education_training_raw",
        "education_training",
        "образование и обучение",
        "обучение и образование",
        "повышение квалификации",
        "повышения квалификации",
        "пк",
        "обучение",
    ],
    "training_raw": [
        "training_raw",
        "training",
    ],
    "certification_raw": [
        "certification_raw",
        "certification",
        "сертификат",
        "сертификация",
        "сертификаты",
        "квалификационная категория",
    ],
    "degree_raw": ["degree_raw", "degree", "ученая степень", "степень"],
    "awards_raw": ["awards_raw", "awards", "награды", "поощрения"],
    "note_raw": ["note_raw", "note", "примечание", "комментарий"],
    "phone_raw": ["phone_raw", "phone", "телефон", "контактный телефон"],
}

SHEET_TYPE_RULES: list[tuple[str, tuple[str, ...]]] = [
    ("declaration", ("декларац",)),
    ("part_time", ("совмест", "совм")),
    ("doctors", ("врач",)),
    ("nurses", ("медсестр", "медсестра", "м/с", "смр")),
    ("junior_staff", ("санитарк",)),
    ("other_staff", ("прочее", "прочие")),
]

ROW_TYPE_EMPLOYEE = "EMPLOYEE"
ROW_TYPE_CATEGORY_ROW = "CATEGORY_ROW"
ROW_TYPE_SUMMARY_ROW = "SUMMARY_ROW"
ROW_TYPE_DECLARATION_PERSON = "DECLARATION_PERSON"
ROW_TYPE_DECLARATION_ROW = "DECLARATION_ROW"

CATEGORY_ROW_LABELS_EXACT: frozenset[str] = frozenset(
    {
        "высшая",
        "первая",
        "вторая",
        "декрет",
        "пенсионеры",
        "пенсионер",
        "женщины",
        "мужчины",
        "всего",
        "итого",
        "доктор медицинских наук",
        "кандидат мед.наук",
        "кандидат медицинских наук",
        "кандидат мед наук",
        "без категории",
        "нет категории",
    }
)

CATEGORY_ROW_PREFIX_RE: tuple[re.Pattern[str], ...] = (
    re.compile(r"^женщины[-\s–—]", re.IGNORECASE),
    re.compile(r"^мужчины[-\s–—]", re.IGNORECASE),
    re.compile(r"^всего\b", re.IGNORECASE),
    re.compile(r"^итого\b", re.IGNORECASE),
    re.compile(r"^всего\s+\d", re.IGNORECASE),
    re.compile(r"^итого\s+\d", re.IGNORECASE),
)


@dataclass(frozen=True)
class SheetLayoutProfile:
    expected_header_row: int
    section_col: str
    columns: dict[str, str]


LAYOUT_PROFILES: dict[str, SheetLayoutProfile] = {
    "doctors": SheetLayoutProfile(
        expected_header_row=7,
        section_col="B",
        columns={
            "full_name": "C",
            "birth_date": "D",
            "iin": "E",
            "sex": "F",
            "nationality": "G",
            "education_raw": "H",
            "diploma_specialty_raw": "I",
            "position_raw": "J",
            "qualification_raw": "K",
            "experience_raw": "L",
            "education_training_raw": "M",
            "certification_raw": "N",
            "degree_raw": "O",
            "awards_raw": "P",
            "note_raw": "Q",
            "phone_raw": "R",
        },
    ),
    "nurses": SheetLayoutProfile(
        expected_header_row=4,
        section_col="C",
        columns={
            "full_name": "D",
            "birth_date": "E",
            "iin": "F",
            "sex": "G",
            "education_raw": "I",
            "diploma_specialty_raw": "J",
            "position_raw": "K",
            "training_raw": "N",
            "certification_raw": "O",
        },
    ),
    "junior_staff": SheetLayoutProfile(
        expected_header_row=5,
        section_col="B",
        columns={
            "full_name": "C",
            "birth_date": "D",
            "iin": "E",
            "sex": "F",
            "position_raw": "H",
        },
    ),
}


@dataclass
class ParsedRow:
    data: dict[str, str] = field(default_factory=dict)
    sheet_type: str = ""
    row_type: str = ROW_TYPE_EMPLOYEE
    declaration_group: str = ""
    is_employee_roster: bool = True
    iin_valid: bool = False
    iin_digits: str = ""
    errors: list[str] = field(default_factory=list)

    @property
    def full_name(self) -> str:
        return self.data.get("full_name", "")

    @property
    def department(self) -> str:
        return self.data.get("department", "")

    @property
    def training_raw(self) -> str:
        return self.data.get("training_raw", "") or self.data.get("education_training_raw", "")

    @property
    def education_training_raw(self) -> str:
        return self.data.get("education_training_raw", "") or self.data.get("training_raw", "")

    @property
    def certification_raw(self) -> str:
        return self.data.get("certification_raw", "")


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    text = str(value).replace("\u00a0", " ").strip()
    return " ".join(text.split())


def _norm_header(value: Any) -> str:
    text = _to_text(value).lstrip("\ufeff").lower()
    text = text.replace("ё", "е")
    return " ".join(text.split())


def _norm_header_match(value: Any) -> str:
    text = _norm_header(value)
    text = re.sub(r"[.,;:]", " ", text)
    return " ".join(text.split())


def _normalize_department(value: str) -> str:
    return _to_text(value)


def _col_idx(letter: str) -> int:
    return column_index_from_string(letter)


def get_layout_profile(sheet_type: str) -> Optional[SheetLayoutProfile]:
    return LAYOUT_PROFILES.get(sheet_type)


def _norm_sheet_name(name: str) -> str:
    text = _to_text(name).lower().replace("ё", "е")
    return " ".join(text.split())


def resolve_sheet_type(sheet_name: str) -> Optional[str]:
    normalized = _norm_sheet_name(sheet_name)
    for sheet_type, tokens in SHEET_TYPE_RULES:
        if any(token in normalized for token in tokens):
            return sheet_type
    return None


def resolve_declaration_group(sheet_name: str) -> str:
    """Map declaration sheet title → staff group (врач / медсестра / …)."""
    normalized = _norm_sheet_name(sheet_name)
    if "медсестр" in normalized or "смр" in normalized or "м/с" in normalized:
        return "nurses"
    if "санитарк" in normalized:
        return "junior_staff"
    if "проч" in normalized:
        return "other_staff"
    if "врач" in normalized:
        return "doctors"
    return "other_staff"


def _normalize_row_label(value: str) -> str:
    text = _norm_header(value)
    text = re.sub(r"\s*[-–—]\s*\d+\s*$", "", text)
    return text.strip()


def is_category_or_summary_label(name: str) -> bool:
    """True when full_name cell is a category/total label, not a person."""
    norm = _normalize_row_label(name)
    if not norm:
        return True
    if norm in CATEGORY_ROW_LABELS_EXACT:
        return True
    raw_norm = _norm_header(name)
    for pattern in CATEGORY_ROW_PREFIX_RE:
        if pattern.search(raw_norm):
            return True
    if re.fullmatch(r"(всего|итого)\s+\d+", raw_norm):
        return True
    return False


# Kazakh letters used in Kazakhstani FIO alongside Russian Cyrillic (ADR-039).
_KAZAKH_CYRILLIC_UPPER = "ӘҒҚҢӨҰҮҺІ"
_KAZAKH_CYRILLIC_LOWER = "әғқңөұүһі"
_NAME_PART_START_RE = re.compile(
    rf"^[А-ЯЁ{_KAZAKH_CYRILLIC_UPPER}A-Z]"
)
_NAME_PART_LOWERCASE_RE = re.compile(
    rf"[а-яё{_KAZAKH_CYRILLIC_LOWER}a-z]"
)


def _is_valid_iin_digits(iin_digits: str) -> bool:
    return len(iin_digits) == 12 and iin_digits.isdigit()


def _has_multi_word_name(name: str) -> bool:
    """True when full_name looks like a person name with at least two words."""
    text = _to_text(name)
    if not text:
        return False
    parts = text.split()
    if len(parts) < 2 or len(parts) > 5:
        return False
    return all(not re.fullmatch(r"\d+", part) for part in parts)


def looks_like_person_name(name: str) -> bool:
    text = _to_text(name)
    if not text or is_category_or_summary_label(text):
        return False
    parts = text.split()
    if len(parts) < 2 or len(parts) > 5:
        return False
    for part in parts:
        if re.fullmatch(r"\d+", part):
            return False
        if not _NAME_PART_START_RE.match(part):
            return False
        if not _NAME_PART_LOWERCASE_RE.search(part):
            return False
    return True


def infer_row_type(
    *,
    full_name: str,
    sheet_type: str,
    iin_digits: str,
) -> tuple[str, bool]:
    """Return (row_type, is_employee_roster)."""
    if sheet_type == "declaration":
        if looks_like_person_name(full_name):
            return ROW_TYPE_DECLARATION_PERSON, False
        return ROW_TYPE_DECLARATION_ROW, False

    if not _to_text(full_name) and not iin_digits:
        return ROW_TYPE_SUMMARY_ROW, False
    if is_category_or_summary_label(full_name):
        return ROW_TYPE_CATEGORY_ROW, False
    # ADR-039: valid IIN + multi-word FIO → EMPLOYEE even for non-Russian letters.
    if _is_valid_iin_digits(iin_digits) and _has_multi_word_name(full_name):
        return ROW_TYPE_EMPLOYEE, True
    if looks_like_person_name(full_name):
        return ROW_TYPE_EMPLOYEE, True
    if _to_text(full_name):
        return ROW_TYPE_CATEGORY_ROW, False
    return ROW_TYPE_SUMMARY_ROW, False


def _header_matches(normalized_header: str, alias: str) -> bool:
    header_key = _norm_header_match(normalized_header) if normalized_header else ""
    alias_key = _norm_header_match(alias)
    if not header_key or not alias_key:
        return False
    if header_key == alias_key:
        return True
    if len(alias_key) >= 4 and alias_key in header_key:
        return True
    return False


def _score_header_row(values: list[Any]) -> int:
    normalized = [_norm_header_match(v) for v in values if _norm_header_match(v)]
    if not normalized:
        return 0
    score = 0
    for aliases in HEADER_ALIASES.values():
        if any(
            _header_matches(header, alias)
            for header in normalized
            for alias in aliases
        ):
            score += 1
    return score


def find_header_row(ws, *, scan_limit: int = 40) -> tuple[int, list[Any]]:
    best_row = 0
    best_vals: list[Any] = []
    best_score = 0
    for row_idx in range(1, min(ws.max_row, scan_limit) + 1):
        vals = [ws.cell(row=row_idx, column=col).value for col in range(1, ws.max_column + 1)]
        score = _score_header_row(vals)
        if score > best_score:
            best_score = score
            best_row = row_idx
            best_vals = vals
    if best_score < 2:
        return 0, []
    return best_row, best_vals


def build_field_map(header_vals: list[Any]) -> dict[str, int]:
    normalized_to_col: dict[str, int] = {}
    for idx, header in enumerate(header_vals, start=1):
        normalized = _norm_header_match(header)
        if normalized:
            normalized_to_col[normalized] = idx

    field_map: dict[str, int] = {}
    for field_name, aliases in HEADER_ALIASES.items():
        for normalized, col_idx in normalized_to_col.items():
            if any(_header_matches(normalized, alias) for alias in aliases):
                field_map[field_name] = col_idx
                break
    return field_map


def build_merged_section_lookup(ws, section_col_idx: int) -> dict[int, str]:
    lookup: dict[int, str] = {}
    for rng in ws.merged_cells.ranges:
        if rng.min_col <= section_col_idx <= rng.max_col:
            value = _normalize_department(_to_text(ws.cell(row=rng.min_row, column=rng.min_col).value))
            if not value:
                continue
            for row_idx in range(rng.min_row, rng.max_row + 1):
                lookup[row_idx] = value
    return lookup


def resolve_section_department(
    ws,
    row_idx: int,
    section_col_idx: int,
    merged_lookup: dict[int, str],
    current_department: str,
) -> str:
    direct = _normalize_department(_to_text(ws.cell(row=row_idx, column=section_col_idx).value))
    if direct:
        return direct

    merged = merged_lookup.get(row_idx, "")
    if merged:
        return merged

    return current_department


def _cell_by_letter(ws, row_idx: int, letter: str) -> Any:
    return ws.cell(row=row_idx, column=_col_idx(letter)).value


def _cell_for_field(
    ws,
    row_idx: int,
    field_name: str,
    profile: SheetLayoutProfile,
    field_map: dict[str, int],
) -> Any:
    """Prefer header-detected column; fall back to layout profile letter."""
    col_idx = field_map.get(field_name)
    if col_idx:
        return ws.cell(row=row_idx, column=col_idx).value
    letter = profile.columns.get(field_name)
    if letter:
        return _cell_by_letter(ws, row_idx, letter)
    return None


def _is_employee_row_profile(
    ws,
    row_idx: int,
    profile: SheetLayoutProfile,
    field_map: dict[str, int],
) -> bool:
    name_raw = _cell_for_field(ws, row_idx, "full_name", profile, field_map)
    if _to_text(name_raw):
        return True

    iin_raw = _cell_for_field(ws, row_idx, "iin", profile, field_map)
    digits, _, _ = clean_iin(iin_raw)
    if digits:
        return True

    return False


def _pick_cell(ws, row_idx: int, field_map: dict[str, int], field_name: str) -> Any:
    col = field_map.get(field_name)
    if not col:
        return None
    return ws.cell(row=row_idx, column=col).value


def _excel_serial_to_date(value: float) -> Optional[date]:
    if value < 1 or value > 100000:
        return None
    base = datetime(1899, 12, 30)
    try:
        return (base + timedelta(days=float(value))).date()
    except (OverflowError, ValueError):
        return None


def parse_birth_date(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    if isinstance(value, (int, float)):
        parsed = _excel_serial_to_date(float(value))
        return parsed.isoformat() if parsed else ""
    text = _to_text(value)
    if not text:
        return ""
    for fmt in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y", "%d-%m-%Y"):
        try:
            return datetime.strptime(text, fmt).date().isoformat()
        except ValueError:
            continue
    if re.fullmatch(r"\d+(?:[.,]\d+)?", text):
        parsed = _excel_serial_to_date(float(text.replace(",", ".")))
        return parsed.isoformat() if parsed else ""
    return ""


def _coerce_iin_text(value: Any) -> str:
    """Normalize Excel/scalar IIN values to a digit-friendly string."""
    if value is None or value == "":
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, int):
        return str(value)
    if isinstance(value, float):
        if value.is_integer() or abs(value - round(value)) < 1e-6:
            return str(int(round(value)))
        return format(value, ".0f")
    text = _to_text(value)
    if re.fullmatch(r"\d+\.0+", text):
        return text.split(".", 1)[0]
    lowered = text.lower()
    if "e" in lowered:
        try:
            as_float = float(text.replace(",", "."))
            if as_float.is_integer() or abs(as_float - round(as_float)) < 1e-6:
                return str(int(round(as_float)))
        except ValueError:
            pass
    return text


def clean_iin(value: Any) -> tuple[str, bool, list[str]]:
    raw_text = _coerce_iin_text(value)
    digits = re.sub(r"\D", "", raw_text)
    if not digits:
        return "", False, ["missing_iin"]
    if len(digits) > 12 and digits.endswith("0") and re.search(r"\.0+\s*$", raw_text.replace(" ", "")):
        digits = digits[:12]
    if len(digits) == 11:
        digits = f"0{digits}"
    if len(digits) != 12:
        return digits, False, [f"invalid_iin_length:{len(digits)}"]
    return digits, True, []


def _row_has_content(ws, row_idx: int, field_map: dict[str, int]) -> bool:
    for field_name in ("full_name", "iin", "department", "position_raw"):
        if _to_text(_pick_cell(ws, row_idx, field_map, field_name)):
            return True
    return False


def _build_parsed_row(
    *,
    data: dict[str, str],
    sheet_type: str,
    iin_raw: Any,
) -> ParsedRow:
    iin_digits, iin_valid, iin_errors = clean_iin(iin_raw)
    data["iin"] = iin_digits

    row_type, is_employee_roster = infer_row_type(
        full_name=data.get("full_name", ""),
        sheet_type=sheet_type,
        iin_digits=iin_digits,
    )
    declaration_group = ""
    if sheet_type == "declaration":
        declaration_group = resolve_declaration_group(data.get("source_sheet", ""))

    errors: list[str] = []
    if is_employee_roster:
        errors.extend(iin_errors)
        if not data["full_name"]:
            errors.append("missing_full_name")
        if not data["department"]:
            errors.append("missing_department")

    return ParsedRow(
        data=data,
        sheet_type=sheet_type,
        row_type=row_type,
        declaration_group=declaration_group,
        is_employee_roster=is_employee_roster,
        iin_valid=iin_valid,
        iin_digits=iin_digits,
        errors=errors,
    )


def parse_sheet_with_profile(ws, *, sheet_type: str, profile: SheetLayoutProfile) -> list[ParsedRow]:
    header_row_idx, header_vals = find_header_row(ws)
    if not header_row_idx:
        return []

    field_map = build_field_map(header_vals)
    section_col_idx = _col_idx(profile.section_col)
    merged_lookup = build_merged_section_lookup(ws, section_col_idx)
    current_department = ""
    parsed_rows: list[ParsedRow] = []

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        if not _is_employee_row_profile(ws, row_idx, profile, field_map):
            section_only = resolve_section_department(
                ws,
                row_idx,
                section_col_idx,
                merged_lookup,
                current_department,
            )
            if section_only and section_only != current_department:
                direct = _normalize_department(_to_text(ws.cell(row=row_idx, column=section_col_idx).value))
                merged = merged_lookup.get(row_idx, "")
                if direct or merged:
                    current_department = section_only
            continue

        current_department = resolve_section_department(
            ws,
            row_idx,
            section_col_idx,
            merged_lookup,
            current_department,
        )

        data = {name: "" for name in OUTPUT_FIELDS}
        data["source_sheet"] = ws.title
        data["source_row_number"] = str(row_idx)
        data["department"] = current_department

        iin_raw = None
        for field_name in profile.columns:
            raw = _cell_for_field(ws, row_idx, field_name, profile, field_map)
            if field_name == "iin":
                iin_raw = raw
            if field_name == "birth_date":
                data[field_name] = parse_birth_date(raw)
            else:
                data[field_name] = _to_text(raw)

        if data.get("education_training_raw"):
            data["training_raw"] = data["education_training_raw"]
        elif data.get("training_raw") and not data.get("education_training_raw"):
            data["education_training_raw"] = data["training_raw"]

        parsed_rows.append(
            _build_parsed_row(data=data, sheet_type=sheet_type, iin_raw=iin_raw)
        )

    return parsed_rows


def parse_sheet_generic(ws, *, sheet_type: str) -> list[ParsedRow]:
    header_row_idx, header_vals = find_header_row(ws)
    if not header_row_idx:
        return []

    field_map = build_field_map(header_vals)
    if "full_name" not in field_map and "iin" not in field_map:
        return []

    parsed_rows: list[ParsedRow] = []
    last_department = ""

    for row_idx in range(header_row_idx + 1, ws.max_row + 1):
        if not _row_has_content(ws, row_idx, field_map):
            continue

        data = {name: "" for name in OUTPUT_FIELDS}
        data["source_sheet"] = ws.title
        data["source_row_number"] = str(row_idx)

        for field_name in OUTPUT_FIELDS:
            if field_name in ("source_sheet", "source_row_number", "department"):
                continue
            raw = _pick_cell(ws, row_idx, field_map, field_name)
            if field_name == "birth_date":
                data[field_name] = parse_birth_date(raw)
            else:
                data[field_name] = _to_text(raw)

        if data["department"]:
            last_department = data["department"]
        elif last_department:
            data["department"] = last_department

        if data.get("education_training_raw"):
            data["training_raw"] = data["education_training_raw"]
        elif data.get("training_raw") and not data.get("education_training_raw"):
            data["education_training_raw"] = data["training_raw"]

        parsed_rows.append(
            _build_parsed_row(
                data=data,
                sheet_type=sheet_type,
                iin_raw=_pick_cell(ws, row_idx, field_map, "iin"),
            )
        )

    return parsed_rows


def parse_sheet(ws, *, sheet_type: str) -> list[ParsedRow]:
    profile = get_layout_profile(sheet_type)
    if profile:
        return parse_sheet_with_profile(ws, sheet_type=sheet_type, profile=profile)
    return parse_sheet_generic(ws, sheet_type=sheet_type)


def parse_workbook(path: Path) -> tuple[list[ParsedRow], list[str]]:
    wb = load_workbook(path, data_only=True, read_only=False)
    all_rows: list[ParsedRow] = []
    warnings: list[str] = []

    for sheet_name in wb.sheetnames:
        sheet_type = resolve_sheet_type(sheet_name)
        if not sheet_type:
            warnings.append(f"skip_unknown_sheet:{sheet_name}")
            continue
        ws = wb[sheet_name]
        rows = parse_sheet(ws, sheet_type=sheet_type)
        if not rows:
            warnings.append(f"no_data_rows:{sheet_name}")
        all_rows.extend(rows)

    wb.close()
    return all_rows, warnings


def build_audit(rows: list[ParsedRow]) -> dict[str, Any]:
    iin_counts = Counter(row.iin_digits for row in rows if row.iin_digits)
    duplicate_iins = {iin for iin, count in iin_counts.items() if count > 1}

    return {
        "total_rows": len(rows),
        "valid_iin": sum(1 for row in rows if row.iin_valid),
        "invalid_iin": sum(1 for row in rows if row.iin_digits and not row.iin_valid),
        "missing_iin": sum(1 for row in rows if not row.iin_digits),
        "duplicate_iin": len(duplicate_iins),
        "duplicate_iin_rows": sum(1 for row in rows if row.iin_digits in duplicate_iins),
        "missing_full_name": sum(1 for row in rows if not row.full_name),
        "missing_department": sum(1 for row in rows if not row.department),
        "with_training": sum(1 for row in rows if row.training_raw),
        "with_certification": sum(1 for row in rows if row.certification_raw),
        "duplicate_iins": sorted(duplicate_iins),
    }


def mask_iin(iin: str) -> str:
    if len(iin) == 12:
        return f"{iin[:4]}****{iin[-2:]}"
    if len(iin) > 4:
        return f"{iin[:2]}***"
    return "***"


def write_csv(path: Path, rows: Iterable[dict[str, str]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", encoding="utf-8-sig", newline="") as fh:
        writer = csv.DictWriter(fh, fieldnames=fieldnames, extrasaction="ignore")
        writer.writeheader()
        for row in rows:
            writer.writerow(row)


def export_outputs(rows: list[ParsedRow], audit: dict[str, Any], out_dir: Path) -> dict[str, Path]:
    duplicate_iins = set(audit["duplicate_iins"])

    preview_rows = [row.data for row in rows]
    error_rows = [
        {**row.data, "error_reason": ";".join(row.errors)}
        for row in rows
        if row.errors
    ]
    duplicate_rows = [
        {**row.data, "duplicate_group": row.iin_digits}
        for row in rows
        if row.iin_digits in duplicate_iins
    ]

    paths = {
        "preview": out_dir / "hr_import_preview.csv",
        "errors": out_dir / "hr_import_errors.csv",
        "duplicates": out_dir / "hr_import_duplicates.csv",
    }
    write_csv(paths["preview"], preview_rows, OUTPUT_FIELDS)
    write_csv(paths["errors"], error_rows, OUTPUT_FIELDS + ["error_reason"])
    write_csv(paths["duplicates"], duplicate_rows, OUTPUT_FIELDS + ["duplicate_group"])
    return paths


def print_summary(audit: dict[str, Any], warnings: list[str], paths: dict[str, Path], *, dry_run: bool) -> None:
    print("HR Control List Import — audit summary")
    print("=" * 40)
    print(f"mode: {'dry-run' if dry_run else 'apply (not implemented)'}")
    print(f"total_rows: {audit['total_rows']}")
    print(f"valid_iin: {audit['valid_iin']}")
    print(f"invalid_iin: {audit['invalid_iin']}")
    print(f"missing_iin: {audit['missing_iin']}")
    print(f"duplicate_iin_groups: {audit['duplicate_iin']}")
    print(f"duplicate_iin_rows: {audit['duplicate_iin_rows']}")
    print(f"missing_full_name: {audit['missing_full_name']}")
    print(f"missing_department: {audit['missing_department']}")
    print(f"with_training: {audit['with_training']}")
    print(f"with_certification: {audit['with_certification']}")
    if warnings:
        print("warnings:")
        for warning in warnings:
            print(f"  - {warning}")
    print("outputs:")
    for label, path in paths.items():
        print(f"  - {label}: {path}")


def print_preview(rows: list[ParsedRow], limit: int = 10) -> None:
    print()
    print(f"Preview (first {limit} rows, masked IIN):")
    print("-" * 40)
    for row in rows[:limit]:
        masked = {**row.data, "iin": mask_iin(row.data.get("iin", ""))}
        print(
            f"{masked['source_sheet']}:{masked['source_row_number']} | "
            f"{masked['full_name']} | iin={masked['iin']} | "
            f"dept={masked['department']} | pos={masked['position_raw']}"
        )


def print_stage_summary(
    batch_id: int,
    summary: dict[str, Any],
    warnings: list[str],
) -> None:
    print("HR Control List Import — staging summary")
    print("=" * 40)
    print(f"mode: stage")
    print(f"batch_id: {batch_id}")
    print(f"total_rows: {summary['total_rows']}")
    print(f"valid_iin: {summary['valid_iin']}")
    print(f"invalid_iin: {summary['invalid_iin']}")
    print(f"duplicate_iin_groups: {summary['duplicate_iin_groups']}")
    print(f"duplicate_iin_rows: {summary['duplicate_iin_rows']}")
    print(f"missing_full_name: {summary['missing_full_name']}")
    print(f"missing_department: {summary['missing_department']}")
    print(f"with_training: {summary['with_training']}")
    print(f"with_certification: {summary['with_certification']}")
    if warnings:
        print("warnings:")
        for warning in warnings:
            print(f"  - {warning}")


def parse_args(argv: Optional[list[str]] = None) -> argparse.Namespace:
    parser = argparse.ArgumentParser(description="Parse HR control list Excel into staging preview CSVs.")
    parser.add_argument("--file", required=True, help="Path to control list .xlsx")
    parser.add_argument(
        "--out-dir",
        default=str(DEFAULT_OUT_DIR),
        help="Directory for preview CSV outputs (default: tmp/)",
    )
    parser.add_argument(
        "--dry-run",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Parse and export preview only (default: true)",
    )
    parser.add_argument(
        "--stage",
        action="store_true",
        help="Parse and persist batch + rows to hr_import_* staging tables",
    )
    parser.add_argument(
        "--imported-by",
        type=int,
        default=None,
        help="users.user_id for imported_by (default: lowest user_id in DB)",
    )
    parser.add_argument(
        "--apply",
        action="store_true",
        help="Reserved for a future apply phase (currently disabled)",
    )
    return parser.parse_args(argv)


def main(argv: Optional[list[str]] = None) -> int:
    args = parse_args(argv)
    if args.apply:
        print("Apply mode is disabled until Design Review.", file=sys.stderr)
        return 2

    source_path = Path(args.file)
    if not source_path.is_file():
        print(f"File not found: {source_path}", file=sys.stderr)
        return 1

    if args.stage:
        if str(PROJECT_ROOT) not in sys.path:
            sys.path.insert(0, str(PROJECT_ROOT))
        from app.db.engine import engine
        from app.services.hr_import_service import import_control_list, resolve_default_imported_by

        with engine.begin() as conn:
            imported_by = args.imported_by
            if imported_by is None:
                imported_by = resolve_default_imported_by(conn)
            batch_id, summary, warnings = import_control_list(
                conn,
                file_path=source_path,
                imported_by=imported_by,
            )
        print_stage_summary(batch_id, summary, warnings)
        return 0

    rows, warnings = parse_workbook(source_path)
    audit = build_audit(rows)
    paths = export_outputs(rows, audit, Path(args.out_dir))
    print_summary(audit, warnings, paths, dry_run=args.dry_run)
    print_preview(rows)
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
