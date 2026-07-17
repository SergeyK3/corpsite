"""Read-only Control List workbook profiling."""
from __future__ import annotations

import re
from collections import Counter
from datetime import datetime, timezone
from pathlib import Path
from typing import Any, Optional

from openpyxl import load_workbook
from openpyxl.utils import get_column_letter

from scripts.ops.control_list_import import SCHEMA_VERSION
from scripts.ops.control_list_import.diagnostics import (
    IssueRegistry,
    analyze_department_source,
    analyze_iin_column_stats,
    build_diagnostics,
)
from scripts.ops.control_list_import.sheet_classification import classify_sheet_name
from scripts.ops.control_list_import.header_aliases import (
    header_alias_tokens,
    match_semantic_field,
    normalize_header,
)
from scripts.ops.control_list_import.value_types import (
    analyze_composite_cell,
    classify_value_type,
    detect_value_issues,
    extract_iin_digits,
    extract_phone_digits,
    is_iin_column,
    looks_like_fio,
    mask_sample_value,
    to_text,
)

INFLATED_ROW_THRESHOLD = 5
INFLATED_COL_THRESHOLD = 3

_FOOTER_RE = re.compile(r"^(итого|всего|подпис|signature)", re.IGNORECASE)


def normalize_sheet_name_for_exclusion(name: str) -> str:
    text = to_text(name).lower().replace("ё", "е")
    return re.sub(r"\s+", " ", text).strip()


def sheet_exclusion_match(
    sheet_name: str,
    exclusion_terms: list[str],
) -> tuple[bool, Optional[str], Optional[str]]:
    normalized_name = normalize_sheet_name_for_exclusion(sheet_name)
    for term in exclusion_terms:
        norm_term = normalize_sheet_name_for_exclusion(term)
        if norm_term and norm_term in normalized_name:
            return True, "sheet_name_declaration", term
    return False, None, None


def _cell_is_empty(value: Any) -> bool:
    if value is None:
        return True
    if isinstance(value, str) and not value.strip():
        return True
    return False


def _iter_sheet_rows(ws, max_row: int, max_col: int) -> list[tuple[int, list[Any]]]:
    rows: list[tuple[int, list[Any]]] = []
    for row_idx, row in enumerate(
        ws.iter_rows(min_row=1, max_row=max_row, min_col=1, max_col=max_col, values_only=False),
        start=1,
    ):
        values = [cell.value for cell in row]
        rows.append((row_idx, values))
    return rows


def _actual_bounds(rows: list[tuple[int, list[Any]]]) -> tuple[int, int, int, int]:
    last_row = 0
    last_col = 0
    for row_idx, values in rows:
        for col_idx, value in enumerate(values, start=1):
            if not _cell_is_empty(value):
                last_row = max(last_row, row_idx)
                last_col = max(last_col, col_idx)
    return 1, 1, last_row, last_col


def _row_non_empty_count(values: list[Any]) -> int:
    return sum(1 for v in values if not _cell_is_empty(v))


def _score_header_row(values: list[Any]) -> tuple[float, list[str]]:
    matched_aliases: list[str] = []
    score = 0.0
    normalized_cells = [normalize_header(v) for v in values if not _cell_is_empty(v)]
    if not normalized_cells:
        return 0.0, []

    score += min(len(normalized_cells), 12) * 0.4

    alias_hits = 0
    for cell in normalized_cells:
        field, alias, confidence = match_semantic_field(cell)
        if field and confidence >= 0.55:
            alias_hits += 1
            if alias:
                matched_aliases.append(alias)
            score += confidence * 2.0

    known = header_alias_tokens()
    for cell in normalized_cells:
        if cell in known:
            alias_hits += 1
            score += 1.5

    joined = " ".join(normalized_cells)
    for token in ("иин", "фио", "фамилия", "должност", "подраздел", "образован", "телефон"):
        if token in joined:
            score += 0.8

    score += alias_hits * 0.5
    confidence = min(1.0, score / 12.0)
    return round(confidence, 3), sorted(set(matched_aliases))


def detect_probable_header_row(
    rows: list[tuple[int, list[Any]]],
    *,
    scan_limit: int,
) -> tuple[Optional[int], float, list[str], Optional[int]]:
    candidates: list[tuple[int, float, list[str]]] = []
    for row_idx, values in rows[:scan_limit]:
        non_empty = _row_non_empty_count(values)
        confidence, aliases = _score_header_row(values)
        if non_empty < 2:
            if not aliases or confidence < 0.35:
                continue
        elif confidence < 0.25:
            continue
        candidates.append((row_idx, confidence, aliases))

    if not candidates:
        return None, 0.0, [], None

    candidates.sort(key=lambda item: (item[1], _row_non_empty_count(rows[item[0] - 1][1])), reverse=True)
    header_row, confidence, aliases = candidates[0]
    first_data_row = header_row + 1
    return header_row, confidence, aliases, first_data_row


def classify_row(
    values: list[Any],
    *,
    row_index: int,
    header_row: Optional[int],
    column_semantics: dict[int, str],
) -> str:
    if _row_non_empty_count(values) == 0:
        return "empty"

    if header_row is not None and row_index == header_row:
        return "header"

    texts = [to_text(v) for v in values if not _cell_is_empty(v)]
    joined = " ".join(texts).lower()

    if header_row is not None and row_index < header_row:
        return "title"

    if any(_FOOTER_RE.search(t) for t in texts):
        return "footer"

    has_fio = False
    has_iin = False
    has_birth = False
    has_position = False

    for col_idx, value in enumerate(values, start=1):
        if _cell_is_empty(value):
            continue
        sf = column_semantics.get(col_idx, "")
        if sf.endswith(".full_name") or looks_like_fio(value):
            has_fio = True
        if sf.endswith(".iin"):
            digits, _ = extract_iin_digits(value, apply_issues=False)
            if len(digits) == 12:
                has_iin = True
        if sf.endswith(".birth_date"):
            has_birth = True
        if sf.endswith(".position_title"):
            has_position = True

    if not has_fio:
        for value in values:
            if looks_like_fio(value):
                has_fio = True
                break

    iin_col = next((idx for idx, sf in column_semantics.items() if sf == "person.iin"), None)
    if not has_iin and iin_col is not None:
        value = values[iin_col - 1] if iin_col - 1 < len(values) else None
        if not _cell_is_empty(value):
            digits, issues = extract_iin_digits(value, apply_issues=False)
            if len(digits) == 12 and "iin_not_12_digits" not in issues:
                has_iin = True

    if has_fio or has_iin or (has_birth and has_position):
        return "data"

    dept_cols = [idx for idx, sf in column_semantics.items() if sf.endswith(".department_name")]
    if dept_cols:
        for col_idx in dept_cols:
            text = to_text(values[col_idx - 1] if col_idx - 1 < len(values) else "")
            if text and not looks_like_fio(text) and len(text) > 3:
                if not has_fio and not has_iin:
                    return "section_header"

    if len(texts) <= 2 and len(joined) > 3 and not has_iin:
        return "section_header"

    return "unknown"


def _column_issue_counts(registry: IssueRegistry, sheet_name: str, col_idx: int) -> dict[str, int]:
    counts: Counter[str] = Counter()
    for (s, _row, c), issues in registry.cell_issues.items():
        if s == sheet_name and c == col_idx:
            for issue in issues:
                counts[issue] += 1
    key = (sheet_name, col_idx)
    for issue in registry.column_issues.get(key, set()):
        counts[issue] += 1
    return dict(counts)


def _sheet_issue_counts(registry: IssueRegistry, sheet_name: str) -> dict[str, int]:
    counts: Counter[str] = Counter()
    for (s, _row, _col), issues in registry.cell_issues.items():
        if s == sheet_name:
            for issue in issues:
                counts[issue] += 1
    for (s, _col), issues in registry.column_issues.items():
        if s == sheet_name:
            for issue in issues:
                counts[issue] += 1
    counts.update(registry.sheet_issues.get(sheet_name, {}))
    return dict(counts)


def profile_workbook(
    input_path: Path,
    *,
    exclusion_terms: list[str],
    max_samples_per_column: int = 5,
    header_scan_limit: int = 30,
) -> dict[str, Any]:
    input_path = input_path.resolve()
    wb = load_workbook(input_path, read_only=False, data_only=True)
    registry = IssueRegistry()

    try:
        analyzed_at = datetime.now(timezone.utc).replace(microsecond=0).isoformat()
        sheets_out: list[dict[str, Any]] = []
        iin_column_stats: list[dict[str, Any]] = []
        department_diagnostics: list[dict[str, Any]] = []

        summary_probable_person_rows = 0
        summary_rows_with_iin = 0
        summary_rows_without_iin = 0
        summary_semantic_fields: Counter[str] = Counter()
        rows_by_employment_mode: Counter[str] = Counter()
        rows_by_personnel_category: Counter[str] = Counter()

        for sheet_index, sheet_name in enumerate(wb.sheetnames):
            ws = wb[sheet_name]
            excel_max_row = ws.max_row or 0
            excel_max_col = ws.max_column or 0
            hidden_state = getattr(ws, "sheet_state", "visible") or "visible"

            excluded, exclusion_reason, matched_term = sheet_exclusion_match(sheet_name, exclusion_terms)
            merged_ranges = [str(rng) for rng in ws.merged_cells.ranges]
            sheet_classification = classify_sheet_name(
                sheet_name,
                is_declaration_excluded=excluded,
            )

            if excluded:
                sheets_out.append(
                    {
                        "sheet_name": sheet_name,
                        "sheet_index": sheet_index,
                        "hidden_state": hidden_state,
                        "excel_max_row": excel_max_row,
                        "excel_max_column": excel_max_col,
                        "status": "excluded",
                        "exclusion_reason": exclusion_reason,
                        "matched_exclusion_term": matched_term,
                        "applied_exclusion_rule": "exclude_sheet_name_contains",
                        "merged_ranges": merged_ranges,
                        "classification": sheet_classification.as_dict(),
                    }
                )
                continue

            scan_row_limit = max(excel_max_row, header_scan_limit)
            scan_col_limit = max(excel_max_col, 1)
            rows = _iter_sheet_rows(ws, scan_row_limit, scan_col_limit)
            _, _, actual_last_row, actual_last_col = _actual_bounds(rows)

            if excel_max_row - actual_last_row >= INFLATED_ROW_THRESHOLD:
                registry.emit_sheet(sheet_name, "inflated_excel_used_range")
            if excel_max_col - actual_last_col >= INFLATED_COL_THRESHOLD:
                registry.emit_sheet(sheet_name, "inflated_excel_used_range")

            trimmed_rows = [(idx, vals[:actual_last_col]) for idx, vals in rows if idx <= actual_last_row]

            header_row, header_confidence, matched_aliases, first_data_row = detect_probable_header_row(
                trimmed_rows,
                scan_limit=header_scan_limit,
            )
            if header_confidence < 0.45:
                registry.emit_sheet(sheet_name, "header_not_confident")

            column_semantics: dict[int, str] = {}
            column_confidence: dict[int, float] = {}
            columns_out: list[dict[str, Any]] = []

            if header_row is not None:
                header_values = trimmed_rows[header_row - 1][1]
                for col_idx in range(1, actual_last_col + 1):
                    raw_header = header_values[col_idx - 1] if col_idx - 1 < len(header_values) else None
                    normalized_header = normalize_header(raw_header)
                    semantic_field, matched_alias, semantic_confidence = match_semantic_field(raw_header)
                    if semantic_field:
                        column_semantics[col_idx] = semantic_field
                        column_confidence[col_idx] = semantic_confidence
                        summary_semantic_fields[semantic_field] += 1

                    type_counter: Counter[str] = Counter()
                    samples: list[str] = []
                    non_empty = 0
                    empty = 0

                    for row_idx, values in trimmed_rows:
                        if row_idx <= header_row:
                            continue
                        value = values[col_idx - 1] if col_idx - 1 < len(values) else None
                        if _cell_is_empty(value):
                            empty += 1
                            continue
                        non_empty += 1

                        number_format = ""
                        try:
                            cell = ws.cell(row=row_idx, column=col_idx)
                            number_format = cell.number_format or ""
                        except Exception:
                            number_format = ""

                        is_formula = isinstance(value, str) and value.startswith("=")
                        sf = semantic_field
                        conf = semantic_confidence
                        value_type = classify_value_type(
                            value,
                            number_format=number_format,
                            semantic_field=sf,
                            semantic_confidence=conf,
                            is_formula=is_formula,
                        )
                        type_counter[value_type] += 1

                        cell_issues = detect_value_issues(
                            value,
                            value_type=value_type,
                            number_format=number_format,
                            semantic_field=sf,
                            semantic_confidence=conf,
                            is_formula=is_formula,
                        )
                        record_count = None
                        if value_type == "composite_text":
                            record_count = analyze_composite_cell(to_text(value)).record_count

                        registry.emit_cell(
                            sheet_name,
                            row_idx,
                            col_idx,
                            cell_issues,
                            semantic_field=sf,
                            record_count=record_count,
                        )

                        if len(samples) < max_samples_per_column:
                            masked = mask_sample_value(
                                value,
                                value_type=value_type,
                                semantic_field=semantic_field,
                            )
                            if masked and masked not in samples:
                                samples.append(masked)

                    if non_empty > 1 and len(type_counter) >= 3:
                        registry.emit_column(sheet_name, col_idx, "mixed_value_types_in_column")

                    col_meta = {
                        "column_index": col_idx,
                        "column_letter": get_column_letter(col_idx),
                        "raw_header": to_text(raw_header),
                        "normalized_header": normalized_header,
                        "proposed_semantic_field": semantic_field,
                        "semantic_confidence": semantic_confidence,
                        "matched_alias": matched_alias,
                        "non_empty_count": non_empty,
                        "empty_count": empty,
                        "value_type_distribution": dict(type_counter),
                        "issue_counts": _column_issue_counts(registry, sheet_name, col_idx),
                        "samples": samples,
                    }
                    columns_out.append(col_meta)

                    if is_iin_column(semantic_field, semantic_confidence):
                        iin_column_stats.append(analyze_iin_column_stats(sheet_name, col_meta, registry))

            row_classifications: dict[str, int] = Counter()
            probable_person_rows = 0
            rows_with_iin = 0
            rows_without_iin = 0
            rows_with_phone = 0
            rows_with_composite_education = 0
            rows_with_composite_training = 0
            probable_inherited_section_rows = 0

            dept_col = next(
                (idx for idx, sf in column_semantics.items() if sf == "employment.department_name"),
                None,
            )
            last_dept_value = ""

            for row_idx, values in trimmed_rows:
                if header_row is not None and row_idx <= header_row:
                    continue
                row_class = classify_row(
                    values,
                    row_index=row_idx,
                    header_row=header_row,
                    column_semantics=column_semantics,
                )
                row_classifications[row_class] += 1

                if row_class != "data":
                    if row_class == "section_header" and dept_col is not None:
                        dept_val = to_text(values[dept_col - 1] if dept_col - 1 < len(values) else "")
                        if dept_val:
                            last_dept_value = dept_val
                    continue

                probable_person_rows += 1
                has_iin = False
                has_phone = False

                for col_idx, sf in column_semantics.items():
                    value = values[col_idx - 1] if col_idx - 1 < len(values) else None
                    if _cell_is_empty(value):
                        continue
                    if sf.endswith(".iin"):
                        digits, issues = extract_iin_digits(value, apply_issues=False)
                        if len(digits) == 12 and "iin_not_12_digits" not in issues:
                            has_iin = True
                    if sf.endswith(".phone"):
                        phone_digits, _ = extract_phone_digits(value)
                        if phone_digits:
                            has_phone = True
                    if sf == "education.records":
                        comp = analyze_composite_cell(to_text(value))
                        if comp.is_composite:
                            rows_with_composite_education += 1
                    if sf == "training.records":
                        comp = analyze_composite_cell(to_text(value))
                        if comp.is_composite:
                            rows_with_composite_training += 1

                if not has_iin:
                    iin_col = next((idx for idx, sf in column_semantics.items() if sf == "person.iin"), None)
                    if iin_col is not None:
                        value = values[iin_col - 1] if iin_col - 1 < len(values) else None
                        digits, issues = extract_iin_digits(value, apply_issues=False)
                        if len(digits) == 12 and "iin_not_12_digits" not in issues:
                            has_iin = True

                if has_iin:
                    rows_with_iin += 1
                else:
                    rows_without_iin += 1
                if has_phone:
                    rows_with_phone += 1

                if dept_col is not None:
                    dept_val = to_text(values[dept_col - 1] if dept_col - 1 < len(values) else "")
                    if not dept_val and last_dept_value:
                        probable_inherited_section_rows += 1
                        registry.emit_cell(
                            sheet_name,
                            row_idx,
                            dept_col,
                            ["probable_inherited_section_value"],
                        )
                    elif dept_val:
                        last_dept_value = dept_val

            summary_probable_person_rows += probable_person_rows
            summary_rows_with_iin += rows_with_iin
            summary_rows_without_iin += rows_without_iin

            employment_mode = sheet_classification.proposed_employment_mode
            personnel_category = sheet_classification.proposed_personnel_category
            rows_by_employment_mode[employment_mode] += probable_person_rows
            rows_by_personnel_category[personnel_category] += probable_person_rows

            dept_diag = analyze_department_source(
                sheet_name=sheet_name,
                column_semantics=column_semantics,
                columns_out=columns_out,
                row_classifications=dict(row_classifications),
                merged_ranges=merged_ranges,
                probable_inherited_section_rows=probable_inherited_section_rows,
            )
            department_diagnostics.append(dept_diag)

            sheets_out.append(
                {
                    "sheet_name": sheet_name,
                    "sheet_index": sheet_index,
                    "hidden_state": hidden_state,
                    "excel_max_row": excel_max_row,
                    "excel_max_column": excel_max_col,
                    "actual_last_row": actual_last_row,
                    "actual_last_column": actual_last_col,
                    "merged_ranges": merged_ranges,
                    "status": "analyzed",
                    "exclusion_reason": None,
                    "probable_header_row": header_row,
                    "header_confidence": header_confidence,
                    "matched_header_aliases": matched_aliases,
                    "probable_first_data_row": first_data_row,
                    "columns": columns_out,
                    "row_classifications": dict(row_classifications),
                    "department_source": dept_diag,
                    "classification": sheet_classification.as_dict(),
                    "statistics": {
                        "probable_person_rows": probable_person_rows,
                        "rows_with_iin": rows_with_iin,
                        "rows_without_iin": rows_without_iin,
                        "rows_with_phone": rows_with_phone,
                        "rows_with_composite_education": rows_with_composite_education,
                        "rows_with_composite_training": rows_with_composite_training,
                        "probable_inherited_section_rows": probable_inherited_section_rows,
                        "detected_semantic_fields": sorted(set(column_semantics.values())),
                        "issue_counts": _sheet_issue_counts(registry, sheet_name),
                    },
                }
            )

        analyzed_count = sum(1 for s in sheets_out if s.get("status") == "analyzed")
        excluded_count = sum(1 for s in sheets_out if s.get("status") == "excluded")
        unique_issue_counts = registry.unique_cell_issue_counts()
        diagnostics = build_diagnostics(
            registry,
            sheets_out,
            iin_column_stats,
            department_diagnostics,
        )

        return {
            "schema_version": SCHEMA_VERSION,
            "source": {
                "filename": input_path.name,
                "absolute_path": str(input_path),
                "size_bytes": input_path.stat().st_size,
            },
            "analyzed_at": analyzed_at,
            "configuration": {
                "excluded_sheet_name_contains": exclusion_terms,
                "max_samples_per_column": max_samples_per_column,
                "header_scan_limit": header_scan_limit,
            },
            "workbook": {
                "sheet_count": len(sheets_out),
                "analyzed_sheet_count": analyzed_count,
                "excluded_sheet_count": excluded_count,
                "sheets": sheets_out,
            },
            "summary": {
                "probable_person_rows": summary_probable_person_rows,
                "rows_with_iin": summary_rows_with_iin,
                "rows_without_iin": summary_rows_without_iin,
                "composite_cell_count": diagnostics["composite_cell_count_unique"],
                "issues_by_code": dict(unique_issue_counts),
                "semantic_fields_detected": dict(summary_semantic_fields),
                "rows_by_employment_mode": {
                    "primary": rows_by_employment_mode.get("primary", 0),
                    "concurrent": rows_by_employment_mode.get("concurrent", 0),
                    "unknown": rows_by_employment_mode.get("unknown", 0),
                },
                "rows_by_personnel_category": {
                    "doctor": rows_by_personnel_category.get("doctor", 0),
                    "nursing_staff": rows_by_personnel_category.get("nursing_staff", 0),
                    "junior_medical_staff": rows_by_personnel_category.get("junior_medical_staff", 0),
                    "other_staff": rows_by_personnel_category.get("other_staff", 0),
                    "unknown": rows_by_personnel_category.get("unknown", 0),
                },
            },
            "diagnostics": diagnostics,
        }
    finally:
        wb.close()
