"""Diagnostic aggregation for workbook profiling."""
from __future__ import annotations

from collections import Counter, defaultdict
from dataclasses import dataclass, field
from typing import Any, Optional

from openpyxl.utils import get_column_letter



@dataclass
class IssueRegistry:
    """Track unique (sheet, row, col, issue) emissions."""

    cell_issues: dict[tuple[str, int, int], set[str]] = field(default_factory=dict)
    sheet_issues: dict[str, Counter[str]] = field(default_factory=lambda: defaultdict(Counter))
    column_issues: dict[tuple[str, int], set[str]] = field(default_factory=dict)
    duplicate_emissions: Counter[str] = field(default_factory=Counter)
    composite_by_semantic: Counter[str] = field(default_factory=Counter)
    composite_record_distribution: Counter[int] = field(default_factory=Counter)

    def emit_cell(
        self,
        sheet_name: str,
        row: int,
        col: int,
        issues: list[str],
        *,
        semantic_field: Optional[str] = None,
        record_count: Optional[int] = None,
    ) -> None:
        key = (sheet_name, row, col)
        existing = self.cell_issues.setdefault(key, set())
        composite_counted = False
        for issue in issues:
            if issue in existing:
                self.duplicate_emissions[f"{sheet_name}!{get_column_letter(col)}{row}:{issue}"] += 1
                continue
            existing.add(issue)
            if issue in {"multiple_numbered_records_in_cell", "composite_text_without_standard_line_breaks"}:
                if not composite_counted:
                    sf = semantic_field or "unmapped"
                    self.composite_by_semantic[sf] += 1
                    if record_count is not None:
                        self.composite_record_distribution[record_count] += 1
                    composite_counted = True

    def emit_sheet(self, sheet_name: str, issue: str, count: int = 1) -> None:
        self.sheet_issues[sheet_name][issue] += count

    def emit_column(self, sheet_name: str, col: int, issue: str) -> None:
        key = (sheet_name, col)
        col_set = self.column_issues.setdefault(key, set())
        if issue not in col_set:
            col_set.add(issue)

    def unique_cell_issue_counts(self) -> Counter[str]:
        counts: Counter[str] = Counter()
        for issues in self.cell_issues.values():
            for issue in issues:
                counts[issue] += 1
        for sheet_counts in self.sheet_issues.values():
            counts.update(sheet_counts)
        for (_sheet, _col), issues in self.column_issues.items():
            for issue in issues:
                counts[issue] += 1
        return counts

    def composite_cell_count(self) -> int:
        composite_issues = {"multiple_numbered_records_in_cell", "composite_text_without_standard_line_breaks"}
        cells: set[tuple[str, int, int]] = set()
        for key, issues in self.cell_issues.items():
            if issues & composite_issues:
                cells.add(key)
        return len(cells)


def build_diagnostics(
    registry: IssueRegistry,
    sheets_out: list[dict[str, Any]],
    iin_column_stats: list[dict[str, Any]],
    department_diagnostics: list[dict[str, Any]],
) -> dict[str, Any]:
    unique_counts = registry.unique_cell_issue_counts()
    issues_by_sheet: dict[str, dict[str, int]] = defaultdict(dict)
    issues_by_column: dict[str, dict[str, int]] = defaultdict(dict)

    for (sheet, row, col), issues in registry.cell_issues.items():
        col_label = f"{sheet}!{get_column_letter(col)}"
        for issue in issues:
            issues_by_sheet[sheet][issue] = issues_by_sheet[sheet].get(issue, 0) + 1
            issues_by_column[col_label][issue] = issues_by_column[col_label].get(issue, 0) + 1

    for sheet_name, sheet_counts in registry.sheet_issues.items():
        for issue, count in sheet_counts.items():
            issues_by_sheet[sheet_name][issue] = issues_by_sheet[sheet_name].get(issue, 0) + count

    for (sheet, _col), issues in registry.column_issues.items():
        for issue in issues:
            issues_by_sheet[sheet][issue] = issues_by_sheet[sheet].get(issue, 0) + 1

    unique_cells_by_issue = dict(unique_counts)

    return {
        "issues_by_sheet": dict(issues_by_sheet),
        "issues_by_column": dict(issues_by_column),
        "composite_by_semantic_field": dict(registry.composite_by_semantic),
        "composite_record_count_distribution": dict(registry.composite_record_distribution),
        "unique_cells_by_issue": unique_cells_by_issue,
        "duplicate_issue_emissions": dict(registry.duplicate_emissions),
        "iin_column_diagnostics": iin_column_stats,
        "department_diagnostics": department_diagnostics,
        "composite_cell_count_unique": registry.composite_cell_count(),
    }


def analyze_department_source(
    *,
    sheet_name: str,
    column_semantics: dict[int, str],
    columns_out: list[dict[str, Any]],
    row_classifications: dict[str, int],
    merged_ranges: list[str],
    probable_inherited_section_rows: int,
) -> dict[str, Any]:
    dept_cols = [
        col
        for col, sf in column_semantics.items()
        if sf == "employment.department_name"
    ]
    section_header_count = row_classifications.get("section_header", 0)
    has_department_alias = bool(dept_cols)

    dept_col_info = None
    if dept_cols:
        col_idx = dept_cols[0]
        col_meta = next((c for c in columns_out if c["column_index"] == col_idx), None)
        dept_col_info = {
            "column_index": col_idx,
            "column_letter": col_meta["column_letter"] if col_meta else get_column_letter(col_idx),
            "raw_header": col_meta["raw_header"] if col_meta else "",
        }

    if has_department_alias:
        mode = "column"
        reason = "header alias matched employment.department_name"
    elif section_header_count > 0 or merged_ranges:
        mode = "section_rows"
        reason = (
            "department conveyed via section header rows and/or merged cells, "
            "not a dedicated department column"
        )
    else:
        mode = "unresolved"
        reason = "no department header alias and no section-row pattern detected"

    return {
        "sheet_name": sheet_name,
        "department_source_mode": mode,
        "proposed_department_column": dept_col_info,
        "section_header_count": section_header_count,
        "merged_range_count": len(merged_ranges),
        "probable_inherited_section_rows": probable_inherited_section_rows,
        "has_department_header_alias": has_department_alias,
        "reason": reason,
    }


def analyze_iin_column_stats(
    sheet_name: str,
    col_meta: dict[str, Any],
    registry: IssueRegistry,
) -> dict[str, Any]:
    col_idx = col_meta["column_index"]
    col_letter = col_meta["column_letter"]
    non_empty = col_meta["non_empty_count"]
    type_dist = col_meta.get("value_type_distribution", {})

    issue_cells = Counter()
    for (s, row, c), issues in registry.cell_issues.items():
        if s == sheet_name and c == col_idx:
            for issue in issues:
                if issue.startswith("iin_"):
                    issue_cells[issue] += 1

    numeric_values = type_dist.get("iin_candidate", 0) + type_dist.get("integer", 0)
    text_values = non_empty - numeric_values

    return {
        "sheet_name": sheet_name,
        "column_letter": col_letter,
        "non_empty_cells": non_empty,
        "numeric_values": numeric_values,
        "text_values": max(text_values, 0),
        "valid_normalized_12_digits": type_dist.get("iin_candidate", 0),
        "unique_issue_cells": dict(issue_cells),
        "issue_counts_in_column_meta": {
            k: v for k, v in col_meta.get("issue_counts", {}).items() if k.startswith("iin_")
        },
    }
