"""In-memory XLSX renderer for the canonical control-list projection."""
from __future__ import annotations

import json
import re
from hashlib import sha256
from dataclasses import dataclass
from datetime import date
from io import BytesIO
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.control_list_projection.schemas import (
    ControlListAcademicDegreeItem,
    ControlListAwardItem,
    ControlListEducationItem,
    ControlListProjectionResponse,
    ControlListProjectionRow,
    ControlListTrainingItem,
)

XLSX_MEDIA_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
CONTROL_LIST_EXPORT_MAX_ROWS = 10_000
CONTROL_LIST_EXPORT_MAX_BYTES = 25 * 1024 * 1024
CONTROL_LIST_EXPORT_MAX_TEXT_CHARACTERS = 8_000_000
EXCEL_CELL_MAX_CHARACTERS = 32_767
EXCEL_DATE_FORMAT = "DD.MM.YYYY"
EXCEL_DATETIME_FORMAT = "DD.MM.YYYY HH:MM:SS"

CONTROL_LIST_HEADERS = (
    "№",
    "Группа подразделений",
    "Подразделение",
    "Фамилия, имя, отчество",
    "Дата рождения",
    "ИИН",
    "Занимаемая должность",
    "Категория должности",
    "Ставка",
    "Дата начала в должности",
    "Образование",
    "Год окончания",
    "Специальность по диплому",
    "Повышение квалификации",
    "Учёная степень",
    "Награды",
    "Телефоны",
    "ID сотрудника",
)

_COLUMN_WIDTHS = (
    7,
    25,
    30,
    36,
    15,
    17,
    31,
    23,
    11,
    20,
    38,
    17,
    38,
    48,
    39,
    45,
    25,
    17,
)
_ILLEGAL_XML_CHARACTERS = re.compile(r"[\x00-\x08\x0B\x0C\x0E-\x1F]")
_FORMULA_PREFIXES = frozenset("=+-@")


class ControlListWorkbookError(RuntimeError):
    """The projection could not be rendered as a safe workbook."""


class ControlListExportLimitError(ControlListWorkbookError):
    """The complete export exceeds an explicit resource or Excel limit."""


@dataclass(frozen=True)
class ControlListWorkbookArtifact:
    content: bytes
    filename: str
    sha256: str
    media_type: str = XLSX_MEDIA_TYPE


def _plain_text(value: Any) -> str:
    text = "" if value is None else str(value)
    if _ILLEGAL_XML_CHARACTERS.search(text):
        raise ControlListWorkbookError("Workbook text contains unsupported characters.")
    return text


def _excel_safe_text(value: Any) -> str | None:
    """Keep display text while forcing formula-like values to remain strings."""

    if value is None:
        return None
    text = _plain_text(value)
    if not text:
        return None
    stripped = text.lstrip()
    if stripped and stripped[0] in _FORMULA_PREFIXES:
        return "'" + text
    return text


def _indexed_lines(values: Iterable[str | None]) -> str | None:
    lines = [f"[{index}] {value or ''}" for index, value in enumerate(values, start=1)]
    return "\n".join(lines) if lines else None


def _date_text(value: date | None) -> str | None:
    return value.strftime("%d.%m.%Y") if value is not None else None


def _education_columns(
    items: list[ControlListEducationItem],
) -> tuple[str | None, str | None, str | None]:
    return (
        _indexed_lines(
            _plain_text(item.institution_name) if item.institution_name else None
            for item in items
        ),
        _indexed_lines(
            str(item.graduation_year)
            if item.graduation_year is not None
            else None
            for item in items
        ),
        _indexed_lines(
            _plain_text(item.specialty) if item.specialty else None for item in items
        ),
    )


def _training_line(item: ControlListTrainingItem, index: int) -> str:
    parts: list[str] = []
    if item.title:
        parts.append(_plain_text(item.title))
    if item.organization_name:
        parts.append(_plain_text(item.organization_name))
    if item.hours is not None:
        parts.append(f"{item.hours} ч.")
    period_start = _date_text(item.started_at)
    period_end = _date_text(item.completed_at)
    if period_start or period_end:
        parts.append(f"{period_start or '…'} — {period_end or '…'}")
    if item.certificate_number:
        parts.append(f"сертификат № {_plain_text(item.certificate_number)}")
    return f"[{index}] " + " — ".join(parts)


def _degree_line(item: ControlListAcademicDegreeItem, index: int) -> str:
    primary = item.label or item.degree_other or item.degree or item.degree_type
    parts = [_plain_text(primary)] if primary else []
    if item.field_of_science:
        parts.append(_plain_text(item.field_of_science))
    if item.completed_at:
        parts.append(_plain_text(item.completed_at))
    if item.document_number:
        parts.append(f"документ № {_plain_text(item.document_number)}")
    return f"[{index}] " + " — ".join(parts)


def _award_line(item: ControlListAwardItem, index: int) -> str:
    parts: list[str] = []
    if item.name:
        parts.append(_plain_text(item.name))
    if item.category:
        parts.append(_plain_text(item.category))
    if item.issued_by:
        parts.append(_plain_text(item.issued_by))
    if item.awarded_at:
        parts.append(_plain_text(item.awarded_at))
    if item.document_number:
        parts.append(f"документ № {_plain_text(item.document_number)}")
    return f"[{index}] " + " — ".join(parts)


def _multiline(lines: Iterable[str]) -> str | None:
    materialized = list(lines)
    return "\n".join(materialized) if materialized else None


def _row_values(row: ControlListProjectionRow) -> tuple[Any, ...]:
    institutions, years, specialties = _education_columns(row.education)
    training = _multiline(
        _training_line(item, index) for index, item in enumerate(row.training, start=1)
    )
    degrees = _multiline(
        _degree_line(item, index)
        for index, item in enumerate(row.academic_degrees, start=1)
    )
    awards = _multiline(
        _award_line(item, index) for index, item in enumerate(row.awards, start=1)
    )
    phones = _multiline(_plain_text(item.value) for item in row.phones)
    return (
        row.number,
        _excel_safe_text(row.org_group),
        _excel_safe_text(row.org_unit),
        _excel_safe_text(row.full_name),
        row.birth_date,
        _excel_safe_text(row.iin),
        _excel_safe_text(row.position),
        _excel_safe_text(row.position_category),
        float(row.employment_rate) if row.employment_rate is not None else None,
        row.assignment_start_date,
        _excel_safe_text(institutions),
        _excel_safe_text(years),
        _excel_safe_text(specialties),
        _excel_safe_text(training),
        _excel_safe_text(degrees),
        _excel_safe_text(awards),
        _excel_safe_text(phones),
        str(row.employee_id),
    )


def _validate_materialized_rows(rows: list[tuple[Any, ...]]) -> None:
    total_text = 0
    for row in rows:
        for value in row:
            if not isinstance(value, str):
                continue
            length = len(value)
            if length > EXCEL_CELL_MAX_CHARACTERS:
                raise ControlListExportLimitError("An Excel cell exceeds its supported size.")
            total_text += length
            if total_text > CONTROL_LIST_EXPORT_MAX_TEXT_CHARACTERS:
                raise ControlListExportLimitError("The export exceeds its text-size limit.")


def _scope_text(projection: ControlListProjectionResponse) -> str:
    scope = projection.metadata.scope
    if scope.organization_wide:
        return "Вся организация"
    return json.dumps(scope.org_unit_ids or [], ensure_ascii=False, separators=(",", ":"))


def _filters_text(projection: ControlListProjectionResponse) -> str:
    values = projection.metadata.filters.model_dump(mode="json")
    return json.dumps(values, ensure_ascii=False, sort_keys=True, separators=(",", ":"))


def _metadata_rows(projection: ControlListProjectionResponse) -> list[tuple[str, Any]]:
    generated_at = projection.metadata.generated_at
    excel_generated_at = generated_at.replace(tzinfo=None)
    return [
        ("Версия схемы", projection.metadata.schema_version),
        ("Дата среза", projection.metadata.as_of_date),
        ("Время формирования", excel_generated_at),
        ("Часовой пояс", projection.metadata.timezone),
        ("Организационный scope", _scope_text(projection)),
        ("Применённые фильтры", _filters_text(projection)),
        ("Инициатор экспорта", str(projection.metadata.initiator_user_id)),
        ("Количество сотрудников", projection.total),
        ("Request ID", None),
        (
            "Классификация",
            "Файл содержит персональные данные. Требуется защищённая обработка.",
        ),
    ]


def _style_main_sheet(workbook: Workbook, row_count: int) -> None:
    sheet = workbook["Контрольный список"]
    thin_gray = Side(style="thin", color="808080")
    border = Border(left=thin_gray, right=thin_gray, top=thin_gray, bottom=thin_gray)
    header_fill = PatternFill("solid", fgColor="1F4E78")
    header_font = Font(bold=True, color="FFFFFF")

    for cell in sheet[1]:
        cell.font = header_font
        cell.fill = header_fill
        cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    sheet.row_dimensions[1].height = 36

    for row_index in range(2, row_count + 2):
        max_lines = 1
        for cell in sheet[row_index]:
            cell.border = border
            cell.alignment = Alignment(vertical="top", wrap_text=True)
            if isinstance(cell.value, str):
                max_lines = max(max_lines, cell.value.count("\n") + 1)
        sheet.row_dimensions[row_index].height = min(90, 18 * max_lines)

    for index, width in enumerate(_COLUMN_WIDTHS, start=1):
        sheet.column_dimensions[get_column_letter(index)].width = width
    sheet.column_dimensions["R"].hidden = True
    sheet.freeze_panes = "A2"
    sheet.auto_filter.ref = f"A1:R{max(1, row_count + 1)}"
    sheet.sheet_view.showGridLines = True

    for row_index in range(2, row_count + 2):
        sheet.cell(row=row_index, column=5).number_format = EXCEL_DATE_FORMAT
        sheet.cell(row=row_index, column=6).number_format = "@"
        sheet.cell(row=row_index, column=9).number_format = "0.00##"
        sheet.cell(row=row_index, column=10).number_format = EXCEL_DATE_FORMAT
        sheet.cell(row=row_index, column=17).number_format = "@"
        sheet.cell(row=row_index, column=18).number_format = "@"
        for column in (1, 5, 6, 9, 10):
            sheet.cell(row=row_index, column=column).alignment = Alignment(
                horizontal="center", vertical="top", wrap_text=True
            )


def _style_metadata_sheet(workbook: Workbook) -> None:
    sheet = workbook["Метаданные"]
    sheet.column_dimensions["A"].width = 30
    sheet.column_dimensions["B"].width = 80
    sheet.freeze_panes = "A2"
    for row_index in range(1, sheet.max_row + 1):
        sheet.cell(row=row_index, column=1).font = Font(bold=True)
        for column in (1, 2):
            sheet.cell(row=row_index, column=column).alignment = Alignment(
                vertical="top", wrap_text=True
            )
    sheet["B2"].number_format = EXCEL_DATE_FORMAT
    sheet["B3"].number_format = EXCEL_DATETIME_FORMAT


def build_control_list_workbook(
    projection: ControlListProjectionResponse,
    *,
    request_id: str,
) -> ControlListWorkbookArtifact:
    """Render the complete projection to memory, or fail without partial output."""

    if projection.total != len(projection.items):
        raise ControlListWorkbookError("Projection row count is inconsistent.")
    if projection.total > CONTROL_LIST_EXPORT_MAX_ROWS:
        raise ControlListExportLimitError("The export exceeds its row limit.")

    rows = [_row_values(row) for row in projection.items]
    _validate_materialized_rows(rows)

    try:
        workbook = Workbook()
        main_sheet = workbook.active
        main_sheet.title = "Контрольный список"
        main_sheet.append(CONTROL_LIST_HEADERS)
        for row in rows:
            main_sheet.append(row)

        metadata_sheet = workbook.create_sheet("Метаданные")
        for label, value in _metadata_rows(projection):
            metadata_sheet.append(
                (
                    label,
                    _excel_safe_text(request_id) if label == "Request ID" else value,
                )
            )

        _style_main_sheet(workbook, projection.total)
        _style_metadata_sheet(workbook)
        created_at = projection.metadata.generated_at.replace(tzinfo=None)
        workbook.properties.creator = "Corpsite"
        workbook.properties.title = "Контрольный список персонала"
        workbook.properties.subject = "Персональные данные"
        workbook.properties.created = created_at
        workbook.properties.modified = created_at

        stream = BytesIO()
        workbook.save(stream)
        workbook.close()
        content = stream.getvalue()
    except ControlListWorkbookError:
        raise
    except Exception as exc:
        raise ControlListWorkbookError("The Excel workbook could not be created.") from exc

    if len(content) > CONTROL_LIST_EXPORT_MAX_BYTES:
        raise ControlListExportLimitError("The export exceeds its file-size limit.")

    filename = f"Контрольный_список_{projection.metadata.as_of_date.isoformat()}.xlsx"
    return ControlListWorkbookArtifact(
        content=content,
        filename=filename,
        sha256=sha256(content).hexdigest(),
    )
