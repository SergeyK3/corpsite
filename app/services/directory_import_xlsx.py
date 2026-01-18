# app/services/directory_import_xlsx.py
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Dict, Optional

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import text

from app.db.engine import engine
from app.services.directory_import_csv import (
    _get_or_create_department_id,
    _get_or_create_position_id,
)


def _to_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.replace("\u00a0", " ").strip()


def _norm_header_any(h: Any) -> str:
    s = _to_text(h).lstrip("\ufeff").lower()
    s = " ".join(s.split())
    return s


def _parse_date_any(s: str) -> Optional[date]:
    s = (s or "").strip()
    if not s:
        return None
    for fmt in ("%Y-%m-%d", "%d.%m.%Y"):
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            pass
    return None


def _parse_date_cell(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _to_text(v)
    if not s:
        return None
    return _parse_date_any(s)


def _parse_rate_cell(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    s = _to_text(v).replace(",", ".")
    try:
        return float(s)
    except Exception:
        return None


def _parse_bool_cell(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        try:
            return bool(int(v))
        except Exception:
            return None
    s = _to_text(v).strip().lower()
    if s == "":
        return None
    if s in ("1", "true", "yes", "y", "on", "да"):
        return True
    if s in ("0", "false", "no", "n", "off", "нет"):
        return False
    return None


def _pick_header(field_map: Dict[str, int], names: list[str]) -> Optional[int]:
    for n in names:
        k = _norm_header_any(n)
        if k in field_map:
            return field_map[k]
    return None


def import_employees_xlsx_bytes(*, raw: bytes) -> Dict[str, Any]:
    if not raw:
        raise HTTPException(status_code=400, detail="Empty XLSX body.")

    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    ws = wb.active

    header_row_idx: Optional[int] = None
    header_vals: list[Any] = []
    for r in range(1, min(ws.max_row, 50) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(_to_text(v) for v in vals):
            header_row_idx = r
            header_vals = vals
            break

    if not header_row_idx:
        raise HTTPException(status_code=400, detail="XLSX has no header row.")

    field_map: Dict[str, int] = {}
    for idx, h in enumerate(header_vals, start=1):
        hn = _norm_header_any(h)
        if hn:
            field_map[hn] = idx

    c_emp = _pick_header(field_map, ["employee_id", "табельный номер", "таб. №", "таб номер", "тн", "id"])
    c_name = _pick_header(field_map, ["full_name", "фио", "ф.и.о.", "сотрудник", "фамилия имя отчество"])
    c_dept = _pick_header(field_map, ["department_name", "отдел", "подразделение", "отделение"])
    c_pos = _pick_header(field_map, ["position_name", "должность", "позиция"])

    if not (c_emp and c_name and c_dept and c_pos):
        raise HTTPException(
            status_code=400,
            detail="XLSX header must include: employee_id/full_name/department_name/position_name (or RU equivalents).",
        )

    c_from = _pick_header(field_map, ["date_from", "дата с", "дата_с", "начало", "date from"])
    c_to = _pick_header(field_map, ["date_to", "дата по", "дата_по", "окончание", "date to"])
    c_rate = _pick_header(field_map, ["employment_rate", "ставка", "rate", "fte"])
    c_active = _pick_header(field_map, ["is_active", "работает", "активен", "active"])

    emp_upserted = 0
    rows_seen = 0

    with engine.begin() as conn:
        dep_cache: Dict[str, int] = {}
        pos_cache: Dict[str, int] = {}

        for r in range(header_row_idx + 1, ws.max_row + 1):
            rows_seen += 1

            employee_id = _to_text(ws.cell(row=r, column=c_emp).value)
            full_name = _to_text(ws.cell(row=r, column=c_name).value)
            dept_name = _to_text(ws.cell(row=r, column=c_dept).value)
            pos_name = _to_text(ws.cell(row=r, column=c_pos).value)

            if not employee_id or not full_name or not dept_name or not pos_name:
                continue

            if dept_name not in dep_cache:
                dep_cache[dept_name] = _get_or_create_department_id(conn, dept_name)

            if pos_name not in pos_cache:
                pos_cache[pos_name] = _get_or_create_position_id(conn, pos_name)

            department_id = dep_cache[dept_name]
            position_id = pos_cache[pos_name]

            date_from_v = _parse_date_cell(ws.cell(row=r, column=c_from).value) if c_from else None
            date_to_v = _parse_date_cell(ws.cell(row=r, column=c_to).value) if c_to else None
            rate_v = _parse_rate_cell(ws.cell(row=r, column=c_rate).value) if c_rate else None
            active_v = _parse_bool_cell(ws.cell(row=r, column=c_active).value) if c_active else None

            if rate_v is None:
                rate_v = 1.00
            if active_v is None:
                active_v = True
            if date_to_v is not None:
                active_v = False

            conn.execute(
                text(
                    """
                    INSERT INTO public.employees
                      (employee_id, full_name, department_id, position_id, date_from, date_to, employment_rate, is_active)
                    VALUES
                      (:employee_id, :full_name, :department_id, :position_id, :date_from, :date_to, :employment_rate, :is_active)
                    ON CONFLICT (employee_id) DO UPDATE SET
                      full_name = EXCLUDED.full_name,
                      department_id = EXCLUDED.department_id,
                      position_id = EXCLUDED.position_id,
                      date_from = EXCLUDED.date_from,
                      date_to = EXCLUDED.date_to,
                      employment_rate = EXCLUDED.employment_rate,
                      is_active = EXCLUDED.is_active
                    """
                ),
                {
                    "employee_id": employee_id,
                    "full_name": full_name,
                    "department_id": department_id,
                    "position_id": position_id,
                    "date_from": date_from_v,
                    "date_to": date_to_v,
                    "employment_rate": rate_v,
                    "is_active": active_v,
                },
            )
            emp_upserted += 1

    return {
        "rows_seen": rows_seen,
        "departments_touched": len(set(dep_cache.keys())),
        "positions_touched": len(set(pos_cache.keys())),
        "employees_upserted": emp_upserted,
        "sheet": ws.title,
        "header_row": header_row_idx,
    }
