"""HR import row review detail and declaration export (Phase 2F)."""
from __future__ import annotations

import io
import re
from datetime import date
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text
from sqlalchemy.engine import Connection

from app.services.department_recoding_service import lookup_recoding
from app.services.hr_import_profile_service import build_import_profile
from app.services.hr_import_analytics_service import (
    BatchNotFoundError,
    _classify_certification,
    _ensure_batch_exists,
    _format_category_date,
    calc_category_validity_note,
    _infer_staff_type,
    _medical_category_entries,
    is_declaration_row,
    is_real_employee_row,
    load_row_payload,
)
from app.services.hr_import_document_parser import (
    parse_certification_raw,
    parse_education_raw,
    parse_education_training_raw,
    split_raw_fragments,
)

AWARD_SPLIT_RE = re.compile(r"[\n;]+")
DEGREE_PATTERNS = (
    ("candidate", r"кандидат\s+мед"),
    ("doctor", r"доктор\s+мед"),
)


def _parse_birth_date(value: str) -> Optional[date]:
    text_val = (value or "").strip()
    if not text_val:
        return None
    try:
        return date.fromisoformat(text_val[:10])
    except ValueError:
        return None


def _format_date(value: Optional[date]) -> str:
    if not value:
        return ""
    if value.month == 1 and value.day == 1 and value.year >= 1900:
        return str(value.year)
    return value.isoformat()


def _parse_awards(raw: str) -> list[dict[str, str]]:
    items: list[dict[str, str]] = []
    for piece in split_raw_fragments(raw):
        text_val = piece.strip()
        if not text_val:
            continue
        year_match = re.search(r"\b(19|20)\d{2}\b", text_val)
        award_date = year_match.group(0) if year_match else ""
        title = re.sub(r"\b(19|20)\d{2}\b", "", text_val).strip(" ,;.-")
        items.append({"title": title or text_val, "date": award_date})
    return items


def _parse_degrees(raw: str) -> dict[str, bool]:
    lowered = (raw or "").lower()
    return {
        "candidate_medical_sciences": bool(re.search(DEGREE_PATTERNS[0][1], lowered)),
        "doctor_medical_sciences": bool(re.search(DEGREE_PATTERNS[1][1], lowered)),
        "raw_text": raw or "",
    }


def _parse_notes(raw: str) -> list[str]:
    if not raw.strip():
        return []
    return [p.strip() for p in re.split(r"[\n;]+", raw) if p.strip()]


def get_row_review_detail(conn: Connection, batch_id: int, row_id: int) -> dict[str, Any]:
    from app.services.hr_import_ai_extraction_service import load_ai_extraction_draft

    row = load_row_payload(conn, batch_id, row_id)
    payload = row["payload"]
    metadata = row["metadata"]
    iin = str(payload.get("iin", "") or "").strip()
    department = str(payload.get("department", "") or "").strip()
    recoding = lookup_recoding(conn, department)

    education_items = []
    for frag in parse_education_raw(
        str(payload.get("education_raw", "") or ""),
        str(payload.get("diploma_specialty_raw", "") or ""),
    ):
        education_items.append(
            {
                "institution": frag.organization or frag.title or "",
                "year": _format_date(frag.parsed_issued_at),
                "specialty": frag.specialty or "",
                "raw_text": frag.raw_text,
            }
        )

    training_items = []
    qual_category_items = []
    certificate_items = []
    for frag in parse_education_training_raw(
        str(payload.get("education_training_raw", "") or payload.get("training_raw", "") or "")
    ):
        if frag.document_kind == "training":
            training_items.append(
                {
                    "title": frag.title or frag.raw_text,
                    "year": _format_date(frag.parsed_issued_at),
                    "hours": float(frag.parsed_hours) if frag.parsed_hours is not None else None,
                    "raw_text": frag.raw_text,
                }
            )
    for frag in parse_certification_raw(str(payload.get("certification_raw", "") or "")):
        item = {
            "category": frag.category or frag.title or "",
            "date": _format_date(frag.parsed_issued_at or frag.parsed_valid_until),
            "specialty": frag.specialty or "",
            "raw_text": frag.raw_text,
            "document_type": frag.proposed_document_type or "",
        }
        if frag.proposed_document_type == "QUALIFICATION_CATEGORY":
            qual_category_items.append(item)
        else:
            certificate_items.append(
                {
                    "kind": frag.category or frag.proposed_document_type or "Сертификат",
                    "topic": frag.title or frag.raw_text,
                    "date": item["date"],
                    "valid_until": _format_date(frag.parsed_valid_until),
                    "hours": float(frag.parsed_hours) if frag.parsed_hours is not None else None,
                    "link": "",
                    "raw_text": frag.raw_text,
                }
            )

    staff_type = _infer_staff_type(
        {
            "sheet_type": metadata.get("sheet_type", ""),
            "position_raw": str(payload.get("position_raw", "") or ""),
            "is_part_time": metadata.get("is_part_time", False),
        }
    )

    profile = build_import_profile(payload)
    ai_draft = load_ai_extraction_draft(conn, batch_id, row["row_id"])

    return {
        "batch_id": batch_id,
        "row_id": row["row_id"],
        "source_sheet": row["source_sheet"],
        "source_row_number": row["source_row_number"],
        "employee_id": row["employee_id"],
        "full_name": str(payload.get("full_name", "") or ""),
        "iin": iin,
        "birth_date": str(payload.get("birth_date", "") or ""),
        "sex": profile["basic"]["sex"],
        "employment_rate": profile["basic"]["employment_rate"],
        "department": department,
        "department_source": department,
        "department_recoding": {
            "org_unit_id": int(recoding["org_unit_id"]) if recoding and recoding.get("org_unit_id") else None,
            "org_unit_name": recoding["org_unit_name"] if recoding else "",
            "department_group": recoding["department_group"] if recoding else "",
        }
        if recoding
        else None,
        "position_raw": str(payload.get("position_raw", "") or ""),
        "staff_type": staff_type,
        "is_part_time": bool(metadata.get("is_part_time", False)),
        "sheet_type": metadata.get("sheet_type", ""),
        "classification": metadata.get("classification", ""),
        "declaration_group": metadata.get("declaration_group", ""),
        "profile": profile,
        "education": education_items,
        "experience_raw": str(payload.get("experience_raw", "") or ""),
        "training": training_items,
        "qualification_categories": qual_category_items,
        "certificates": certificate_items,
        "degrees": _parse_degrees(str(payload.get("degree_raw", "") or "")),
        "awards": _parse_awards(str(payload.get("awards_raw", "") or "")),
        "notes": _parse_notes(str(payload.get("note_raw", "") or "")),
        "ai_extraction": ai_draft,
        **_load_row_monthly_diff_fields(conn, batch_id, row_id),
    }


def _load_row_monthly_diff_fields(conn: Connection, batch_id: int, row_id: int) -> dict[str, Any]:
    try:
        from app.services.hr_import_monthly_diff_service import load_row_diff_fields

        return load_row_diff_fields(conn, batch_id).get(int(row_id), {})
    except Exception:
        return {}


MEDICAL_CATEGORY_LABELS = {
    "highest": "высшая",
    "first": "первая",
    "second": "вторая",
}


def get_row_medical_category_history(conn: Connection, batch_id: int, row_id: int) -> dict[str, Any]:
    row = load_row_payload(conn, batch_id, row_id)
    payload = row["payload"]
    recoding = lookup_recoding(conn, str(payload.get("department", "") or "").strip())
    entries = _medical_category_entries(str(payload.get("certification_raw", "") or ""))
    sorted_entries = sorted(
        entries,
        key=lambda entry: (entry["date"] or date.min, entry["category"]),
        reverse=True,
    )
    return {
        "batch_id": batch_id,
        "row_id": row_id,
        "full_name": str(payload.get("full_name", "") or "").strip(),
        "position_raw": str(payload.get("position_raw", "") or "").strip(),
        "department": str(payload.get("department", "") or "").strip(),
        "org_unit_name": recoding["org_unit_name"] if recoding else "",
        "items": [
            {
                "date": _format_category_date(entry["date"]),
                "category": entry["category"],
                "category_label": MEDICAL_CATEGORY_LABELS.get(entry["category"], ""),
                "specialty": entry["specialty"],
                "validity_note": calc_category_validity_note(entry["date"]),
            }
            for entry in sorted_entries
        ],
    }


def export_declarations_excel(
    conn: Connection,
    batch_id: int,
    *,
    department_group: Optional[str] = None,
    org_group_id: Optional[int] = None,
    org_unit_id: Optional[int] = None,
    org_unit_name: Optional[str] = None,
    staff_type: Optional[str] = None,
    q_name: Optional[str] = None,
) -> bytes:
    from app.services.hr_import_analytics_service import _canonical_department_label, _load_staging_rows

    rows = [r for r in _load_staging_rows(conn, batch_id) if is_declaration_row(r)]

    effective_org_group_id = org_group_id
    if effective_org_group_id is None and department_group:
        try:
            parsed = int(str(department_group).strip())
            if parsed >= 1:
                effective_org_group_id = parsed
        except ValueError:
            pass

    filtered: list[dict[str, Any]] = []
    for row in rows:
        if effective_org_group_id is not None and row.get("org_group_id") != effective_org_group_id:
            continue
        if org_unit_id is not None:
            if row.get("org_unit_id") != org_unit_id and not (
                org_unit_name
                and (row.get("org_unit_name") or "").strip().lower()
                == org_unit_name.strip().lower()
            ):
                continue
        elif org_unit_name:
            if (row.get("org_unit_name") or "").strip().lower() != org_unit_name.strip().lower():
                continue
        if staff_type and _infer_staff_type(row) != staff_type:
            continue
        if q_name and q_name.strip().lower() not in row["full_name"].lower():
            continue
        filtered.append(row)

    grouped: dict[str, list[dict[str, Any]]] = {}
    for row in filtered:
        dept_label = _canonical_department_label(row)
        grouped.setdefault(dept_label, []).append(row)

    wb = Workbook()
    ws = wb.active
    ws.title = "Декларации"
    row_idx = 1
    header_font = Font(bold=True)
    for dept_name in sorted(grouped.keys()):
        ws.cell(row=row_idx, column=1, value=dept_name.upper()).font = header_font
        row_idx += 1
        ws.cell(row=row_idx, column=1, value="ФИО").font = header_font
        ws.cell(row=row_idx, column=2, value="Тип декларации").font = header_font
        row_idx += 1
        for item in grouped[dept_name]:
            decl_type = item.get("declaration_group") or item.get("sheet_type") or "—"
            ws.cell(row=row_idx, column=1, value=item.get("full_name") or "—")
            ws.cell(row=row_idx, column=2, value=decl_type)
            row_idx += 1
        row_idx += 1

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()
