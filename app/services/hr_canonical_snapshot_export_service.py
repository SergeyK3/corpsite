"""ADR-040 Phase D — export canonical HR snapshot roster to Excel."""
from __future__ import annotations

import io
from datetime import date, datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text
from sqlalchemy.engine import Connection

from app.db.models.hr_import import SOURCE_TYPE_HR_CONTROL_LIST
from app.services.hr_canonical_snapshot_service import (
    RECORD_KIND_ROSTER,
    SNAPSHOT_STATUS_ACTIVE,
    canonical_snapshot_available,
    get_active_snapshot,
)

BASE_COLUMNS: list[tuple[str, str]] = [
    ("iin", "ИИН"),
    ("full_name", "ФИО"),
    ("birth_date", "Дата рождения"),
    ("position_raw", "Должность"),
    ("department", "Отделение"),
    ("certification_raw", "Категория персонала"),
    ("education_raw", "Образование"),
    ("certificate_text", "Сертификат"),
    ("training_hours", "Часы обучения"),
    ("status", "Статус"),
    ("employee_id", "Employee ID"),
    ("match_key", "Match key"),
]

METADATA_COLUMNS: list[tuple[str, str]] = [
    ("snapshot_id", "snapshot_id"),
    ("snapshot_version", "snapshot_version"),
    ("canonical_hash", "canonical_hash"),
    ("source_batch_id", "source_batch_id"),
    ("updated_at", "updated_at"),
]


class CanonicalSnapshotExportError(Exception):
    def __init__(self, message: str, *, status_code: int = 400) -> None:
        self.message = message
        self.status_code = status_code
        super().__init__(message)


def normalize_export_source_type(source_type: Optional[str]) -> str:
    normalized = (source_type or "roster").strip().upper()
    if normalized in {"", "ROSTER", "HR_CONTROL_LIST"}:
        return SOURCE_TYPE_HR_CONTROL_LIST
    return normalized


def _display_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    if isinstance(value, bool):
        return "true" if value else "false"
    return value


def _load_snapshot(conn: Connection, snapshot_id: int) -> dict[str, Any]:
    row = conn.execute(
        text(
            """
            SELECT snapshot_id, source_batch_id, version, source_type, status,
                   entry_count, promoted_by, promoted_at
            FROM public.hr_canonical_snapshots
            WHERE snapshot_id = :snapshot_id
            """
        ),
        {"snapshot_id": snapshot_id},
    ).mappings().first()
    if not row:
        raise CanonicalSnapshotExportError(
            f"snapshot_id={snapshot_id} not found",
            status_code=404,
        )
    return dict(row)


def resolve_export_snapshot(
    conn: Connection,
    *,
    source_type: Optional[str] = None,
    snapshot_id: Optional[int] = None,
) -> dict[str, Any]:
    if not canonical_snapshot_available(conn):
        raise CanonicalSnapshotExportError(
            "hr_canonical_snapshots is not available",
            status_code=503,
        )

    if snapshot_id is not None:
        snapshot = _load_snapshot(conn, int(snapshot_id))
        expected_source_type = normalize_export_source_type(source_type)
        if str(snapshot["source_type"]) != expected_source_type:
            raise CanonicalSnapshotExportError(
                f"snapshot_id={snapshot_id} has source_type={snapshot['source_type']}, "
                f"expected {expected_source_type}",
                status_code=404,
            )
        return snapshot

    resolved_source_type = normalize_export_source_type(source_type)
    active = get_active_snapshot(conn, source_type=resolved_source_type)
    if active is None:
        raise CanonicalSnapshotExportError(
            f"no active canonical snapshot for source_type={resolved_source_type}",
            status_code=404,
        )
    return active


def _load_roster_entries(conn: Connection, snapshot_id: int) -> list[dict[str, Any]]:
    rows = conn.execute(
        text(
            """
            SELECT
                entry_id,
                match_key,
                canonical_hash,
                employee_id,
                payload
            FROM public.hr_canonical_snapshot_entries
            WHERE snapshot_id = :snapshot_id
              AND record_kind = :record_kind
            ORDER BY match_key, entry_id
            """
        ),
        {"snapshot_id": snapshot_id, "record_kind": RECORD_KIND_ROSTER},
    ).mappings().all()
    return [dict(row) for row in rows]


def _load_normalized_enrichment(conn: Connection, snapshot_id: int) -> dict[str, dict[str, Any]]:
    rows = conn.execute(
        text(
            """
            SELECT match_key, record_kind, payload
            FROM public.hr_canonical_snapshot_entries
            WHERE snapshot_id = :snapshot_id
              AND record_kind IN ('training', 'certificate')
            ORDER BY entry_id
            """
        ),
        {"snapshot_id": snapshot_id},
    ).mappings().all()

    enrichment: dict[str, dict[str, Any]] = {}
    for row in rows:
        match_key = str(row["match_key"])
        roster_key = match_key.split("|", 1)[0]
        bucket = enrichment.setdefault(
            roster_key,
            {"certificate_titles": [], "training_hours": 0.0},
        )
        payload = dict(row["payload"] or {})
        record_kind = str(row["record_kind"])
        if record_kind == "certificate":
            title = str(payload.get("title") or "").strip()
            if title:
                bucket["certificate_titles"].append(title)
        elif record_kind == "training":
            hours = payload.get("hours")
            if hours is not None:
                try:
                    bucket["training_hours"] += float(hours)
                except (TypeError, ValueError):
                    pass
    return enrichment


def _build_export_row(
    *,
    entry: dict[str, Any],
    snapshot: dict[str, Any],
    enrichment: dict[str, dict[str, Any]],
    include_metadata: bool,
) -> dict[str, Any]:
    payload = dict(entry.get("payload") or {})
    roster_key = str(entry.get("match_key") or "")
    extra = enrichment.get(roster_key, {})
    certificate_titles = extra.get("certificate_titles") or []
    training_hours = extra.get("training_hours")

    hours_value: Any = ""
    if training_hours:
        hours_value = training_hours
    elif payload.get("training_raw"):
        hours_value = payload.get("training_raw")

    employee_id = entry.get("employee_id")
    status = "BOUND" if employee_id is not None else "UNBOUND"

    row: dict[str, Any] = {
        "iin": payload.get("iin") or "",
        "full_name": payload.get("full_name") or "",
        "birth_date": payload.get("birth_date") or "",
        "position_raw": payload.get("position_raw") or "",
        "department": payload.get("department") or "",
        "certification_raw": payload.get("certification_raw") or "",
        "education_raw": payload.get("education_raw") or "",
        "certificate_text": "; ".join(certificate_titles) if certificate_titles else "",
        "training_hours": hours_value,
        "status": status,
        "employee_id": employee_id if employee_id is not None else "",
        "match_key": roster_key,
    }

    if include_metadata:
        promoted_at = snapshot.get("promoted_at")
        row.update(
            {
                "snapshot_id": int(snapshot["snapshot_id"]),
                "snapshot_version": int(snapshot["version"]),
                "canonical_hash": entry.get("canonical_hash") or "",
                "source_batch_id": int(snapshot["source_batch_id"]),
                "updated_at": promoted_at.isoformat() if isinstance(promoted_at, datetime) else promoted_at,
            }
        )
    return row


def build_canonical_snapshot_export_rows(
    conn: Connection,
    *,
    source_type: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    include_metadata: bool = False,
) -> tuple[dict[str, Any], list[dict[str, Any]]]:
    snapshot = resolve_export_snapshot(
        conn,
        source_type=source_type,
        snapshot_id=snapshot_id,
    )
    resolved_snapshot_id = int(snapshot["snapshot_id"])
    entries = _load_roster_entries(conn, resolved_snapshot_id)
    enrichment = _load_normalized_enrichment(conn, resolved_snapshot_id)
    rows = [
        _build_export_row(
            entry=entry,
            snapshot=snapshot,
            enrichment=enrichment,
            include_metadata=include_metadata,
        )
        for entry in entries
    ]
    return snapshot, rows


def export_canonical_snapshot_xlsx(
    conn: Connection,
    *,
    source_type: Optional[str] = None,
    snapshot_id: Optional[int] = None,
    include_metadata: bool = False,
) -> tuple[bytes, str]:
    snapshot, rows = build_canonical_snapshot_export_rows(
        conn,
        source_type=source_type,
        snapshot_id=snapshot_id,
        include_metadata=include_metadata,
    )

    columns = list(BASE_COLUMNS)
    if include_metadata:
        columns.extend(METADATA_COLUMNS)

    wb = Workbook()
    ws = wb.active
    ws.title = "Канонический реестр"

    header_font = Font(bold=True)
    for col_idx, (_, header) in enumerate(columns, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(columns, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_display_cell(row.get(key)))

    buf = io.BytesIO()
    wb.save(buf)

    filename = (
        f"canonical_snapshot_{int(snapshot['snapshot_id'])}"
        f"_v{int(snapshot['version'])}.xlsx"
    )
    return buf.getvalue(), filename


def get_canonical_snapshot_export_metadata(
    conn: Connection,
    *,
    source_type: Optional[str] = None,
    snapshot_id: Optional[int] = None,
) -> dict[str, Any]:
    snapshot = resolve_export_snapshot(
        conn,
        source_type=source_type,
        snapshot_id=snapshot_id,
    )
    return {
        "snapshot_id": int(snapshot["snapshot_id"]),
        "source_batch_id": int(snapshot["source_batch_id"]),
        "version": int(snapshot["version"]),
        "source_type": snapshot["source_type"],
        "status": snapshot["status"],
        "entry_count": int(snapshot["entry_count"]),
        "promoted_at": snapshot["promoted_at"],
        "is_active": snapshot["status"] == SNAPSHOT_STATUS_ACTIVE,
    }
