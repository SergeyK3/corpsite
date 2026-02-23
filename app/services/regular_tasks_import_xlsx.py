# FILE: app/services/regular_tasks_import_xlsx.py
from __future__ import annotations

import io
import json
from typing import Any, Dict, Optional, List, Tuple

from fastapi import HTTPException
from openpyxl import load_workbook
from sqlalchemy import text
from sqlalchemy.engine import Connection

from app.db.engine import engine


def _to_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.replace("\u00a0", " ").strip()


def _norm_header_any(h: Any) -> str:
    s = _to_text(h).lstrip("\ufeff").lower()
    s = " ".join(s.split())
    return s


def _as_int_or_none(v: Any) -> Optional[int]:
    if v is None:
        return None
    if isinstance(v, bool):
        return None
    if isinstance(v, (int, float)):
        try:
            return int(v)
        except Exception:
            return None
    s = _to_text(v)
    if not s:
        return None
    try:
        return int(float(s.replace(",", ".")))
    except Exception:
        return None


def _as_bool_or_none(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)) and v in (0, 1):
        return bool(int(v))
    s = _to_text(v).lower()
    if s == "":
        return None
    if s in ("1", "true", "yes", "y", "on", "да"):
        return True
    if s in ("0", "false", "no", "n", "off", "нет"):
        return False
    return None


def _as_json_obj(v: Any) -> Dict[str, Any]:
    if v is None:
        return {}
    if isinstance(v, dict):
        return v
    s = _to_text(v)
    if not s:
        return {}
    try:
        obj = json.loads(s)
        return obj if isinstance(obj, dict) else {}
    except Exception:
        return {}


def _pick_header(field_map: Dict[str, int], names: list[str]) -> Optional[int]:
    for n in names:
        k = _norm_header_any(n)
        if k in field_map:
            return field_map[k]
    return None


# -----------------------------
# roles lookup (adaptive)
# -----------------------------
def _get_columns(conn: Connection, table: str, schema: str = "public") -> List[str]:
    q = text(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = :schema AND table_name = :table
        ORDER BY ordinal_position
        """
    )
    rows = conn.execute(q, {"schema": schema, "table": table}).fetchall()
    return [r[0] for r in rows]


def _pick_first(existing: List[str], candidates: List[str]) -> Optional[str]:
    s = set(existing)
    for c in candidates:
        if c in s:
            return c
    return None


def _roles_table_meta(conn: Connection) -> Tuple[str, Optional[str]]:
    """
    Returns:
      (id_col, code_col_or_none)

    code_col_or_none will be one of common candidates if exists:
      code / slug / name / title / role / key / role_name / role_code ...
    """
    cols = _get_columns(conn, "roles", "public")
    if not cols:
        raise HTTPException(status_code=500, detail="roles table not found")

    id_col = _pick_first(cols, ["role_id", "id"])
    if not id_col:
        raise HTTPException(status_code=500, detail="roles table has no recognizable id column (role_id/id)")

    # Try to find a "code-like" column automatically
    code_col = _pick_first(
        cols,
        [
            "role_code",
            "code",
            "slug",
            "key",
            "name",
            "title",
            "role",
            "role_name",
            "name_ru",
            "display_name",
            "display_name_ru",
        ],
    )
    # code_col may be None; then we will only accept numeric ids in Excel.
    if code_col == id_col:
        code_col = None

    return id_col, code_col


def _resolve_role_id(conn: Connection, *, role_token: str) -> Optional[int]:
    """
    role_token comes from Excel (ExecutorRoleCode).
    Resolution priority:
      1) if token looks like int -> treat as role_id (only if exists)
      2) if roles has code-like column -> lookup by exact match
    """
    role_token = (role_token or "").strip()
    if not role_token:
        return None

    id_col, code_col = _roles_table_meta(conn)

    # 1) numeric -> role_id
    as_int = _as_int_or_none(role_token)
    if as_int is not None:
        row = conn.execute(
            text(f"SELECT {id_col} FROM public.roles WHERE {id_col} = :rid"),
            {"rid": int(as_int)},
        ).first()
        if row:
            return int(row[0])

    # 2) by code-like column
    if code_col:
        row = conn.execute(
            text(f"SELECT {id_col} FROM public.roles WHERE {code_col} = :v"),
            {"v": role_token},
        ).first()
        if row:
            return int(row[0])

    return None


def import_regular_task_templates_xlsx_bytes(*, raw: bytes) -> Dict[str, Any]:
    if not raw:
        raise HTTPException(status_code=400, detail="Empty XLSX body.")

    wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
    ws = wb["RegularTaskTemplates"] if "RegularTaskTemplates" in wb.sheetnames else wb.active

    header_row_idx: Optional[int] = None
    header_vals: list[Any] = []

    for r in range(1, min(ws.max_row, 50) + 1):
        vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
        if any(_to_text(v) for v in vals):
            header_row_idx = r
            header_vals = vals
            break

    if not header_row_idx:
        raise HTTPException(status_code=400, detail="XLSX has no header row.")

    field_map: Dict[str, int] = {}
    for idx, h in enumerate(header_vals, start=1):
        hn = _norm_header_any(h)
        if hn:
            field_map[hn] = idx

    # ---- map your real headers to (code, title) ----
    # code  <- ReportCode
    # title <- Deliverable
    c_code = _pick_header(field_map, ["reportcode", "report_code", "report code", "code", "код"])
    c_title = _pick_header(field_map, ["deliverable", "title", "название", "наименование"])

    if not (c_code and c_title):
        raise HTTPException(
            status_code=400,
            detail="RegularTaskTemplates must include headers: ReportCode + Deliverable (or code + title).",
        )

    c_active = _pick_header(field_map, ["is_active", "активен", "активность", "active"])

    # IMPORTANT: Excel has ExecutorRoleCode (token), not executor_role_id
    c_exec_role_token = _pick_header(
        field_map,
        ["executorrolecode", "executor_role_code", "executor role code", "executorrole", "executor_role", "роль исполнителя", "роль исполнителя код"],
    )

    c_schedule_type = _pick_header(field_map, ["periodicity", "schedule_type", "тип расписания", "schedule"])
    c_schedule_params = _pick_header(field_map, ["schedule_params", "параметры расписания", "params", "schedule json"])
    c_create_offset = _pick_header(field_map, ["createoffsetdays", "create_offset_days", "create offset days", "createoffset"])
    c_due_offset = _pick_header(field_map, ["dueoffsetdays", "due_offset_days", "due offset days", "dueoffset"])

    rows_seen = 0
    upserted = 0
    skipped = 0
    missing_roles = 0

    sql_upsert = text(
        """
        INSERT INTO public.regular_tasks (
          code,
          title,
          is_active,
          executor_role_id,
          schedule_type,
          schedule_params,
          create_offset_days,
          due_offset_days,
          updated_at
        ) VALUES (
          :code,
          :title,
          :is_active,
          :executor_role_id,
          :schedule_type,
          CAST(:schedule_params AS jsonb),
          :create_offset_days,
          :due_offset_days,
          now()
        )
        ON CONFLICT (code) DO UPDATE SET
          title = EXCLUDED.title,
          is_active = EXCLUDED.is_active,
          executor_role_id = EXCLUDED.executor_role_id,
          schedule_type = EXCLUDED.schedule_type,
          schedule_params = EXCLUDED.schedule_params,
          create_offset_days = EXCLUDED.create_offset_days,
          due_offset_days = EXCLUDED.due_offset_days,
          updated_at = now()
        """
    )

    with engine.begin() as conn:
        # Ensure roles meta is readable early (will raise 500 if roles missing)
        _roles_table_meta(conn)

        for r in range(header_row_idx + 1, ws.max_row + 1):
            rows_seen += 1

            code = _to_text(ws.cell(row=r, column=c_code).value)
            title = _to_text(ws.cell(row=r, column=c_title).value)

            if not code or not title:
                skipped += 1
                continue

            is_active = _as_bool_or_none(ws.cell(row=r, column=c_active).value) if c_active else None
            if is_active is None:
                is_active = True

            # Resolve executor_role_id safely
            executor_role_id: Optional[int] = None
            if c_exec_role_token:
                role_token = _to_text(ws.cell(row=r, column=c_exec_role_token).value)
                if role_token:
                    executor_role_id = _resolve_role_id(conn, role_token=role_token)
                    if executor_role_id is None:
                        # strict: if role specified but not found => skip row, do not violate FK
                        missing_roles += 1
                        skipped += 1
                        continue

            schedule_type = _to_text(ws.cell(row=r, column=c_schedule_type).value) if c_schedule_type else ""
            schedule_type = schedule_type or None

            schedule_params_obj = _as_json_obj(ws.cell(row=r, column=c_schedule_params).value) if c_schedule_params else {}
            create_offset_days = _as_int_or_none(ws.cell(row=r, column=c_create_offset).value) if c_create_offset else None
            due_offset_days = _as_int_or_none(ws.cell(row=r, column=c_due_offset).value) if c_due_offset else None

            conn.execute(
                sql_upsert,
                {
                    "code": code,
                    "title": title,
                    "is_active": bool(is_active),
                    "executor_role_id": executor_role_id,
                    "schedule_type": schedule_type,
                    "schedule_params": json.dumps(schedule_params_obj, ensure_ascii=False),
                    "create_offset_days": int(create_offset_days or 0),
                    "due_offset_days": int(due_offset_days or 0),
                },
            )
            upserted += 1

    return {
        "sheet": ws.title,
        "header_row": header_row_idx,
        "rows_seen": rows_seen,
        "upserted": upserted,
        "skipped": skipped,
        "missing_roles": missing_roles,
    }