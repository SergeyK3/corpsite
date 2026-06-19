"""ADR-040 Phase H — export materialized HR change events workbook to Excel."""
from __future__ import annotations

import io
import json
from datetime import UTC, date, datetime
from typing import Any, Optional

from openpyxl import Workbook
from openpyxl.styles import Font
from sqlalchemy import text
from sqlalchemy.engine import Connection

from app.services.hr_import_monthly_diff_service import DIFF_STATUS_CONFLICT
from app.services.hr_snapshot_comparison_service import (
    EVENT_TYPE_NEW,
    EVENT_TYPE_REMOVED,
    EVENT_TYPES,
    _build_hr_change_events_filters,
    hr_change_events_available,
)

CHANGE_TYPE_NEW = "NEW"
CHANGE_TYPE_CHANGED = "CHANGED"
CHANGE_TYPE_REMOVED = "REMOVED"
CHANGE_TYPE_CONFLICT = "CONFLICT"

CHANGE_EXPORT_TYPES = frozenset(
    {
        CHANGE_TYPE_NEW,
        CHANGE_TYPE_CHANGED,
        CHANGE_TYPE_REMOVED,
        CHANGE_TYPE_CONFLICT,
    }
)

CHANGED_EVENT_TYPES = frozenset(EVENT_TYPES - {EVENT_TYPE_NEW, EVENT_TYPE_REMOVED})

EXPORT_SHEETS = (
    "SUMMARY",
    CHANGE_TYPE_NEW,
    CHANGE_TYPE_CHANGED,
    CHANGE_TYPE_REMOVED,
    CHANGE_TYPE_CONFLICT,
)

EXPORT_COLUMNS: list[tuple[str, str]] = [
    ("change_type", "change_type"),
    ("iin", "iin"),
    ("fio", "fio"),
    ("employee_id", "employee_id"),
    ("match_key", "match_key"),
    ("department", "department"),
    ("position", "position"),
    ("field", "field"),
    ("old_value", "old_value"),
    ("new_value", "new_value"),
    ("source_batch_id", "source_batch_id"),
    ("snapshot_from", "snapshot_from"),
    ("snapshot_to", "snapshot_to"),
    ("conflict_reason", "conflict_reason"),
    ("updated_at", "updated_at"),
]


class HrChangeEventsExportError(Exception):
    def __init__(self, message: str, *, status_code: int = 400) -> None:
        self.message = message
        self.status_code = status_code
        super().__init__(message)


def _display_cell(value: Any) -> Any:
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.isoformat()
    return value


def _event_type_to_change_type(event_type: str) -> str:
    normalized = event_type.strip().upper()
    if normalized == EVENT_TYPE_NEW:
        return CHANGE_TYPE_NEW
    if normalized == EVENT_TYPE_REMOVED:
        return CHANGE_TYPE_REMOVED
    if normalized in CHANGED_EVENT_TYPES:
        return CHANGE_TYPE_CHANGED
    return CHANGE_TYPE_CHANGED


def _snapshot_label(snapshot_id: Any, version: Any) -> str:
    if snapshot_id is None:
        return ""
    if version is None:
        return str(snapshot_id)
    return f"{snapshot_id}/v{version}"


def _extract_position(
    *,
    field_name: Optional[str],
    new_value: Optional[str],
    old_value: Optional[str],
    payload: Any,
) -> str:
    if field_name == "position_raw":
        return str(new_value or old_value or "")
    if isinstance(payload, dict):
        position = payload.get("position_raw")
        if position:
            return str(position)
    return ""


def _matches_search(
    *,
    q: Optional[str],
    full_name: Optional[str],
    iin: Optional[str],
    employee_id: Optional[int],
) -> bool:
    if not q or not q.strip():
        return True
    needle = q.strip().lower()
    name = str(full_name or "").lower()
    iin_text = str(iin or "").lower()
    employee_text = str(employee_id) if employee_id is not None else ""
    return needle in name or needle in iin_text or needle in employee_text


def _conflict_reason_from_field_diffs(field_diffs: Any, payload: Any) -> str:
    if isinstance(field_diffs, str):
        try:
            field_diffs = json.loads(field_diffs)
        except json.JSONDecodeError:
            field_diffs = None
    if isinstance(payload, str):
        try:
            payload = json.loads(payload)
        except json.JSONDecodeError:
            payload = None

    corrected_fields: set[str] = set()
    if isinstance(payload, dict):
        canonical_payload = payload.get("_canonical_correction_fields") or []
        if isinstance(canonical_payload, list):
            corrected_fields = {str(item) for item in canonical_payload}

    if isinstance(field_diffs, dict) and corrected_fields.intersection(field_diffs):
        overlap = ", ".join(sorted(corrected_fields.intersection(field_diffs)))
        return f"canonical_correction_conflict:{overlap}"

    if isinstance(field_diffs, dict) and field_diffs:
        return "import_field_conflict"

    return "import_conflict"


def _fetch_change_event_rows(
    conn: Connection,
    *,
    employee_id: Optional[int] = None,
    department: Optional[str] = None,
    org_unit_id: Optional[int] = None,
    event_type: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    prior_snapshot_id: Optional[int] = None,
    new_snapshot_id: Optional[int] = None,
    source_batch_id: Optional[int] = None,
    q: Optional[str] = None,
) -> list[dict[str, Any]]:
    where_sql, params = _build_hr_change_events_filters(
        employee_id=employee_id,
        department=department,
        org_unit_id=org_unit_id,
        event_type=event_type,
        date_from=date_from,
        date_to=date_to,
        prior_snapshot_id=prior_snapshot_id,
        new_snapshot_id=new_snapshot_id,
        source_batch_id=source_batch_id,
    )

    rows = conn.execute(
        text(
            f"""
            SELECT
                e.change_event_id,
                e.prior_snapshot_id,
                e.new_snapshot_id,
                e.event_type,
                e.event_at,
                e.employee_id,
                e.match_key,
                e.field_name,
                e.old_value,
                e.new_value,
                e.department,
                e.full_name,
                e.iin,
                e.new_entry_id,
                e.prior_entry_id,
                ps.version AS prior_snapshot_version,
                ns.version AS new_snapshot_version,
                ns.source_batch_id,
                COALESCE(ne.payload, pe.payload) AS roster_payload
            FROM public.hr_change_events e
            JOIN public.hr_canonical_snapshots ps
              ON ps.snapshot_id = e.prior_snapshot_id
            JOIN public.hr_canonical_snapshots ns
              ON ns.snapshot_id = e.new_snapshot_id
            LEFT JOIN public.hr_canonical_snapshot_entries ne
              ON ne.entry_id = e.new_entry_id
            LEFT JOIN public.hr_canonical_snapshot_entries pe
              ON pe.entry_id = e.prior_entry_id
            WHERE {where_sql}
            ORDER BY e.event_at DESC, e.change_event_id DESC
            """
        ),
        params,
    ).mappings().all()

    export_rows: list[dict[str, Any]] = []
    for row in rows:
        if not _matches_search(
            q=q,
            full_name=row.get("full_name"),
            iin=row.get("iin"),
            employee_id=int(row["employee_id"]) if row.get("employee_id") is not None else None,
        ):
            continue

        change_type = _event_type_to_change_type(str(row["event_type"]))
        if event_type is not None and event_type.strip().upper() == CHANGE_TYPE_CONFLICT:
            continue
        if event_type is not None and change_type == CHANGE_TYPE_CONFLICT:
            continue

        event_at = row.get("event_at")
        export_rows.append(
            {
                "change_type": change_type,
                "iin": row.get("iin") or "",
                "fio": row.get("full_name") or "",
                "employee_id": row.get("employee_id") if row.get("employee_id") is not None else "",
                "match_key": row.get("match_key") or "",
                "department": row.get("department") or "",
                "position": _extract_position(
                    field_name=row.get("field_name"),
                    new_value=row.get("new_value"),
                    old_value=row.get("old_value"),
                    payload=row.get("roster_payload"),
                ),
                "field": row.get("field_name") or "",
                "old_value": row.get("old_value") or "",
                "new_value": row.get("new_value") or "",
                "source_batch_id": row.get("source_batch_id") or "",
                "snapshot_from": _snapshot_label(row.get("prior_snapshot_id"), row.get("prior_snapshot_version")),
                "snapshot_to": _snapshot_label(row.get("new_snapshot_id"), row.get("new_snapshot_version")),
                "conflict_reason": "",
                "updated_at": event_at.isoformat() if isinstance(event_at, datetime) else event_at,
            }
        )
    return export_rows


def _resolve_conflict_batch_ids(
    conn: Connection,
    *,
    source_batch_id: Optional[int],
    new_snapshot_id: Optional[int],
    change_event_rows: list[dict[str, Any]],
) -> list[int]:
    if source_batch_id is not None:
        return [int(source_batch_id)]

    if new_snapshot_id is not None:
        batch_id = conn.execute(
            text(
                """
                SELECT source_batch_id
                FROM public.hr_canonical_snapshots
                WHERE snapshot_id = :snapshot_id
                """
            ),
            {"snapshot_id": int(new_snapshot_id)},
        ).scalar_one_or_none()
        if batch_id is not None:
            return [int(batch_id)]

    batch_ids = {
        int(row["source_batch_id"])
        for row in change_event_rows
        if row.get("source_batch_id") not in (None, "")
    }
    return sorted(batch_ids)


def _fetch_conflict_rows(
    conn: Connection,
    *,
    batch_ids: list[int],
    employee_id: Optional[int] = None,
    department: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    q: Optional[str] = None,
) -> list[dict[str, Any]]:
    if not batch_ids:
        return []

    where_parts = [
        "r.diff_status = :diff_status",
        "r.batch_id = ANY(:batch_ids)",
    ]
    params: dict[str, Any] = {
        "diff_status": DIFF_STATUS_CONFLICT,
        "batch_ids": batch_ids,
    }

    if employee_id is not None:
        where_parts.append("r.employee_id = :employee_id")
        params["employee_id"] = int(employee_id)
    if department is not None:
        where_parts.append(
            "LOWER(TRIM(COALESCE(r.normalized_payload->>'department', ''))) = LOWER(TRIM(:department))"
        )
        params["department"] = department.strip()
    if date_from is not None:
        where_parts.append("COALESCE(r.diff_computed_at, b.imported_at)::date >= :date_from")
        params["date_from"] = date_from
    if date_to is not None:
        where_parts.append("COALESCE(r.diff_computed_at, b.imported_at)::date <= :date_to")
        params["date_to"] = date_to

    where_sql = " AND ".join(where_parts)
    rows = conn.execute(
        text(
            f"""
            SELECT
                r.row_id,
                r.batch_id,
                r.employee_id,
                r.field_diffs,
                r.normalized_payload,
                r.diff_computed_at,
                r.canonical_snapshot_id,
                cs.version AS canonical_snapshot_version,
                ns.snapshot_id AS new_snapshot_id,
                ns.version AS new_snapshot_version
            FROM public.hr_import_rows r
            JOIN public.hr_import_batches b ON b.batch_id = r.batch_id
            LEFT JOIN public.hr_canonical_snapshots cs
              ON cs.snapshot_id = r.canonical_snapshot_id
            LEFT JOIN public.hr_canonical_snapshots ns
              ON ns.source_batch_id = r.batch_id
            WHERE {where_sql}
            ORDER BY r.diff_computed_at DESC NULLS LAST, r.row_id DESC
            """
        ),
        params,
    ).mappings().all()

    export_rows: list[dict[str, Any]] = []
    for row in rows:
        payload = row.get("normalized_payload") or {}
        if isinstance(payload, str):
            try:
                payload = json.loads(payload)
            except json.JSONDecodeError:
                payload = {}

        full_name = payload.get("full_name") if isinstance(payload, dict) else None
        iin = payload.get("iin") if isinstance(payload, dict) else None
        row_employee_id = int(row["employee_id"]) if row.get("employee_id") is not None else None
        if not _matches_search(q=q, full_name=full_name, iin=iin, employee_id=row_employee_id):
            continue

        field_diffs = row.get("field_diffs")
        first_field = ""
        old_value = ""
        new_value = ""
        if isinstance(field_diffs, dict) and field_diffs:
            first_field = next(iter(field_diffs))
            diff = field_diffs[first_field]
            if isinstance(diff, dict):
                old_value = str(diff.get("canonical") or "")
                new_value = str(diff.get("incoming") or "")

        updated_at = row.get("diff_computed_at")
        export_rows.append(
            {
                "change_type": CHANGE_TYPE_CONFLICT,
                "iin": iin or "",
                "fio": full_name or "",
                "employee_id": row_employee_id if row_employee_id is not None else "",
                "match_key": f"iin:{iin}" if iin else f"row:{row['row_id']}",
                "department": (payload.get("department") if isinstance(payload, dict) else None) or "",
                "position": _extract_position(
                    field_name="position_raw",
                    new_value=None,
                    old_value=None,
                    payload=payload,
                ),
                "field": first_field,
                "old_value": old_value,
                "new_value": new_value,
                "source_batch_id": int(row["batch_id"]),
                "snapshot_from": _snapshot_label(
                    row.get("canonical_snapshot_id"),
                    row.get("canonical_snapshot_version"),
                ),
                "snapshot_to": _snapshot_label(row.get("new_snapshot_id"), row.get("new_snapshot_version")),
                "conflict_reason": _conflict_reason_from_field_diffs(field_diffs, payload),
                "updated_at": updated_at.isoformat() if isinstance(updated_at, datetime) else updated_at,
            }
        )
    return export_rows


def _should_include_conflicts(event_type: Optional[str]) -> bool:
    if event_type is None or not event_type.strip():
        return True
    normalized = event_type.strip().upper()
    if normalized == CHANGE_TYPE_CONFLICT:
        return True
    if normalized in EVENT_TYPES:
        return False
    return True


def build_hr_change_events_export_rows(
    conn: Connection,
    *,
    employee_id: Optional[int] = None,
    department: Optional[str] = None,
    org_unit_id: Optional[int] = None,
    event_type: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    prior_snapshot_id: Optional[int] = None,
    new_snapshot_id: Optional[int] = None,
    source_batch_id: Optional[int] = None,
    q: Optional[str] = None,
) -> list[dict[str, Any]]:
    if not hr_change_events_available(conn):
        raise HrChangeEventsExportError(
            "hr_change_events is not available",
            status_code=503,
        )

    normalized_event_type: Optional[str] = None
    if event_type is not None and event_type.strip():
        normalized_event_type = event_type.strip().upper()
        if normalized_event_type == CHANGE_TYPE_CONFLICT:
            normalized_event_type = CHANGE_TYPE_CONFLICT
        elif normalized_event_type not in EVENT_TYPES:
            raise HrChangeEventsExportError("Invalid event_type filter.", status_code=422)

    event_rows: list[dict[str, Any]] = []
    if normalized_event_type != CHANGE_TYPE_CONFLICT:
        event_rows = _fetch_change_event_rows(
            conn,
            employee_id=employee_id,
            department=department,
            org_unit_id=org_unit_id,
            event_type=normalized_event_type,
            date_from=date_from,
            date_to=date_to,
            prior_snapshot_id=prior_snapshot_id,
            new_snapshot_id=new_snapshot_id,
            source_batch_id=source_batch_id,
            q=q,
        )

    conflict_rows: list[dict[str, Any]] = []
    if _should_include_conflicts(normalized_event_type):
        batch_ids = _resolve_conflict_batch_ids(
            conn,
            source_batch_id=source_batch_id,
            new_snapshot_id=new_snapshot_id,
            change_event_rows=event_rows,
        )
        conflict_rows = _fetch_conflict_rows(
            conn,
            batch_ids=batch_ids,
            employee_id=employee_id,
            department=department,
            date_from=date_from,
            date_to=date_to,
            q=q,
        )

    if normalized_event_type == CHANGE_TYPE_CONFLICT:
        return conflict_rows

    return event_rows + conflict_rows


def _group_rows_by_change_type(rows: list[dict[str, Any]]) -> dict[str, list[dict[str, Any]]]:
    grouped = {change_type: [] for change_type in CHANGE_EXPORT_TYPES}
    for row in rows:
        change_type = str(row.get("change_type") or "").upper()
        if change_type in grouped:
            grouped[change_type].append(row)
    return grouped


def _write_sheet(
    ws,
    *,
    rows: list[dict[str, Any]],
    header_font: Font,
) -> None:
    for col_idx, (_, header) in enumerate(EXPORT_COLUMNS, start=1):
        cell = ws.cell(row=1, column=col_idx, value=header)
        cell.font = header_font

    for row_idx, row in enumerate(rows, start=2):
        for col_idx, (key, _) in enumerate(EXPORT_COLUMNS, start=1):
            ws.cell(row=row_idx, column=col_idx, value=_display_cell(row.get(key)))


def _write_summary_sheet(
    ws,
    *,
    grouped: dict[str, list[dict[str, Any]]],
    filters: dict[str, Any],
    header_font: Font,
) -> None:
    ws.cell(row=1, column=1, value="metric").font = header_font
    ws.cell(row=1, column=2, value="value").font = header_font

    metrics = [
        ("exported_at", datetime.now(UTC).isoformat(timespec="seconds").replace("+00:00", "Z")),
        ("total_rows", sum(len(items) for items in grouped.values())),
        ("NEW", len(grouped[CHANGE_TYPE_NEW])),
        ("CHANGED", len(grouped[CHANGE_TYPE_CHANGED])),
        ("REMOVED", len(grouped[CHANGE_TYPE_REMOVED])),
        ("CONFLICT", len(grouped[CHANGE_TYPE_CONFLICT])),
    ]
    for idx, (metric, value) in enumerate(metrics, start=2):
        ws.cell(row=idx, column=1, value=metric)
        ws.cell(row=idx, column=2, value=value)

    start = len(metrics) + 3
    ws.cell(row=start, column=1, value="filters").font = header_font
    ws.cell(row=start, column=2, value="value").font = header_font
    filter_items = [(key, value) for key, value in filters.items() if value not in (None, "")]
    for offset, (key, value) in enumerate(filter_items, start=1):
        ws.cell(row=start + offset, column=1, value=key)
        ws.cell(row=start + offset, column=2, value=str(value))


def export_hr_change_events_xlsx(
    conn: Connection,
    *,
    employee_id: Optional[int] = None,
    department: Optional[str] = None,
    org_unit_id: Optional[int] = None,
    event_type: Optional[str] = None,
    date_from: Optional[date] = None,
    date_to: Optional[date] = None,
    prior_snapshot_id: Optional[int] = None,
    new_snapshot_id: Optional[int] = None,
    source_batch_id: Optional[int] = None,
    q: Optional[str] = None,
) -> tuple[bytes, str]:
    rows = build_hr_change_events_export_rows(
        conn,
        employee_id=employee_id,
        department=department,
        org_unit_id=org_unit_id,
        event_type=event_type,
        date_from=date_from,
        date_to=date_to,
        prior_snapshot_id=prior_snapshot_id,
        new_snapshot_id=new_snapshot_id,
        source_batch_id=source_batch_id,
        q=q,
    )

    grouped = _group_rows_by_change_type(rows)
    filters = {
        "employee_id": employee_id,
        "department": department,
        "org_unit_id": org_unit_id,
        "event_type": event_type,
        "date_from": date_from,
        "date_to": date_to,
        "prior_snapshot_id": prior_snapshot_id,
        "new_snapshot_id": new_snapshot_id,
        "source_batch_id": source_batch_id,
        "q": q,
    }

    wb = Workbook()
    wb.remove(wb.active)
    header_font = Font(bold=True)

    summary_ws = wb.create_sheet("SUMMARY")
    _write_summary_sheet(summary_ws, grouped=grouped, filters=filters, header_font=header_font)

    for sheet_name in (CHANGE_TYPE_NEW, CHANGE_TYPE_CHANGED, CHANGE_TYPE_REMOVED, CHANGE_TYPE_CONFLICT):
        ws = wb.create_sheet(sheet_name)
        _write_sheet(ws, rows=grouped[sheet_name], header_font=header_font)

    buf = io.BytesIO()
    wb.save(buf)

    stamp = datetime.now(UTC).strftime("%Y%m%d")
    filename = f"hr_registry_changes_{stamp}.xlsx"
    return buf.getvalue(), filename
