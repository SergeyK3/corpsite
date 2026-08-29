"""Read-only personnel reports built from the operational personnel contour."""
from __future__ import annotations

from datetime import datetime, timezone
from decimal import Decimal, InvalidOperation
from io import BytesIO
import re
from typing import Any, Iterable

from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.pagebreak import Break
from sqlalchemy import text
from sqlalchemy.engine import Connection, Engine


NOT_SPECIFIED = "Не указано"
ALL_GROUPS = "Все группы"
ALL_DEPARTMENTS = "Все отделения"
CLINICAL_GROUP_ID = 1
PARACLINICAL_GROUP_ID = 2
ADMINISTRATIVE_GROUP_ID = 3
MEDICAL_GROUP_IDS = frozenset({CLINICAL_GROUP_ID, PARACLINICAL_GROUP_ID})
LEADER_POSITION_CATEGORY = "leaders"


class PersonnelReportAccessError(ValueError):
    """The requested organization scope is outside the user's visibility."""


class PersonnelReportFilterError(ValueError):
    """The requested organization filters contradict each other."""


def normalize_position_name(value: str | None) -> str:
    """Normalize a displayed position name for centralized roster ranking."""
    normalized = str(value or "").casefold().replace("ё", "е")
    normalized = re.sub(r"[-‐‑‒–—−]+", " ", normalized)
    normalized = re.sub(r"[^\w\s]+", " ", normalized, flags=re.UNICODE)
    return " ".join(normalized.split())


def _is_nurse(position_name: str) -> bool:
    return bool(
        re.search(r"\bмедсестр\w*\b", position_name)
        or re.search(r"\bмедбрат\w*\b", position_name)
        or re.search(r"\bмедицинск\w*\s+(?:сестр\w*|брат\w*)\b", position_name)
    )


def personnel_position_rank(
    *,
    group_id: int,
    position_name: str | None,
    position_category: str | None,
) -> int:
    """Return the report-specific position rank; lower values are shown first."""
    normalized = normalize_position_name(position_name)
    category = str(position_category or "").strip().casefold()

    if group_id == ADMINISTRATIVE_GROUP_ID:
        return 0 if category == LEADER_POSITION_CATEGORY else 1
    if group_id not in MEDICAL_GROUP_IDS:
        return 0

    if category == LEADER_POSITION_CATEGORY or re.search(r"\bзаведующ\w*\b", normalized):
        return 0
    if re.search(r"\bврач\w*\b", normalized):
        return 1
    if re.search(r"\bстарш\w*\b", normalized) and _is_nurse(normalized):
        return 2
    if re.search(r"\bсестр\w*\s+хозяйк\w*\b", normalized):
        return 4
    if _is_nurse(normalized):
        return 3
    if re.search(r"\bсанитарк\w*\b|\bсанитар\w*\b", normalized):
        return 5
    return 6


def _load_report_org_options(
    conn: Connection,
    *,
    scope_unit_ids: list[int] | None,
) -> dict[str, list[dict[str, Any]]]:
    params: dict[str, Any] = {}
    scope_sql = ""
    if scope_unit_ids is not None:
        if not scope_unit_ids:
            return {"groups": [], "departments": []}
        scope_sql = "AND ou.unit_id = ANY(:scope_unit_ids)"
        params["scope_unit_ids"] = list(scope_unit_ids)

    rows = conn.execute(
        text(
            f"""
            SELECT ou.unit_id, ou.name AS unit_name, ou.group_id,
                   COALESCE(dg.group_name, 'Группа #' || ou.group_id::text) AS group_name
            FROM public.org_units ou
            LEFT JOIN public.deps_group dg ON dg.group_id = ou.group_id
            WHERE COALESCE(ou.is_active, TRUE) = TRUE
              AND ou.group_id IS NOT NULL
              {scope_sql}
            ORDER BY ou.group_id, LOWER(ou.name), ou.unit_id
            """
        ),
        params,
    ).mappings().all()

    departments = [
        {
            "unit_id": int(row["unit_id"]),
            "unit_name": str(row["unit_name"]),
            "group_id": int(row["group_id"]),
        }
        for row in rows
    ]
    group_names: dict[int, str] = {}
    for row in rows:
        group_names.setdefault(int(row["group_id"]), str(row["group_name"]))
    groups = [
        {"group_id": group_id, "group_name": group_names[group_id]}
        for group_id in sorted(group_names)
    ]
    return {"groups": groups, "departments": departments}


def list_report_org_options(
    engine: Engine,
    *,
    scope_unit_ids: list[int] | None,
) -> dict[str, list[dict[str, Any]]]:
    with engine.connect() as conn:
        return _load_report_org_options(conn, scope_unit_ids=scope_unit_ids)


def _select_report_departments(
    options: dict[str, list[dict[str, Any]]],
    *,
    group_id: int | None,
    org_unit_id: int | None,
) -> tuple[list[dict[str, Any]], dict[str, Any] | None, dict[str, Any] | None]:
    groups_by_id = {int(row["group_id"]): row for row in options["groups"]}
    departments_by_id = {int(row["unit_id"]): row for row in options["departments"]}

    selected_group = groups_by_id.get(group_id) if group_id is not None else None
    if group_id is not None and selected_group is None:
        raise PersonnelReportAccessError("Группа отделений недоступна или не найдена.")

    selected_department = departments_by_id.get(org_unit_id) if org_unit_id is not None else None
    if org_unit_id is not None and selected_department is None:
        raise PersonnelReportAccessError("Отделение недоступно или не найдено.")
    if (
        selected_group is not None
        and selected_department is not None
        and int(selected_department["group_id"]) != int(selected_group["group_id"])
    ):
        raise PersonnelReportFilterError("Отделение не относится к выбранной группе.")

    departments = options["departments"]
    if selected_group is not None:
        departments = [
            row for row in departments if int(row["group_id"]) == int(selected_group["group_id"])
        ]
    if selected_department is not None:
        departments = [selected_department]
    return departments, selected_group, selected_department


def build_personnel_roster(
    engine: Engine,
    *,
    scope_unit_ids: list[int] | None,
    group_id: int | None = None,
    org_unit_id: int | None = None,
    generated_at: datetime | None = None,
) -> dict[str, Any]:
    """Build the single structured source used by both preview and Excel."""
    formed_at = generated_at or datetime.now(timezone.utc)
    with engine.connect() as conn:
        options = _load_report_org_options(conn, scope_unit_ids=scope_unit_ids)
        departments, selected_group, selected_department = _select_report_departments(
            options,
            group_id=group_id,
            org_unit_id=org_unit_id,
        )
        department_ids = [int(row["unit_id"]) for row in departments]
        if department_ids:
            rows = conn.execute(
                text(
                    """
                    WITH current_roster AS (
                        SELECT e.employee_id,
                               ou.group_id,
                               COALESCE(dg.group_name, 'Группа #' || ou.group_id::text) AS group_name,
                               ou.unit_id,
                               ou.name AS unit_name,
                               COALESCE(NULLIF(BTRIM(p.full_name), ''), NULLIF(BTRIM(e.full_name), '')) AS full_name,
                               NULLIF(BTRIM(pos.name), '') AS position_name,
                               NULLIF(BTRIM(pos.category), '') AS position_category,
                               pa.rate,
                               ROW_NUMBER() OVER (
                                   PARTITION BY e.employee_id
                                   ORDER BY ou.group_id, LOWER(ou.name), ou.unit_id, pa.assignment_id
                               ) AS employee_row
                        FROM public.employees e
                        JOIN public.persons p ON p.person_id = e.person_id
                        JOIN public.person_assignments pa
                          ON pa.person_id = e.person_id
                         AND pa.org_unit_id = ANY(:org_unit_ids)
                         AND pa.active_flag IS TRUE
                         AND pa.is_primary IS TRUE
                         AND pa.lifecycle_status = 'active'
                         AND pa.start_date <= CURRENT_DATE
                         AND (pa.end_date IS NULL OR pa.end_date >= CURRENT_DATE)
                        JOIN public.org_units ou ON ou.unit_id = pa.org_unit_id
                        LEFT JOIN public.deps_group dg ON dg.group_id = ou.group_id
                        LEFT JOIN public.positions pos ON pos.position_id = pa.position_id
                        WHERE COALESCE(e.is_active, TRUE) IS TRUE
                          AND e.operational_status = 'active'
                          AND COALESCE(e.date_from, CURRENT_DATE) <= CURRENT_DATE
                          AND (e.date_to IS NULL OR e.date_to >= CURRENT_DATE)
                          AND COALESCE(p.person_status, 'active') = 'active'
                    )
                    SELECT employee_id, group_id, group_name, unit_id, unit_name,
                           full_name, position_name, position_category, rate
                    FROM current_roster
                    WHERE employee_row = 1
                    ORDER BY group_id, LOWER(unit_name), unit_id,
                             LOWER(full_name), full_name, employee_id
                    """
                ),
                {"org_unit_ids": department_ids},
            ).mappings().all()
        else:
            rows = []

    group_names = {int(row["group_id"]): str(row["group_name"]) for row in options["groups"]}
    department_names = {
        int(row["unit_id"]): str(row["unit_name"]) for row in options["departments"]
    }
    grouped: dict[int, dict[int, list[dict[str, Any]]]] = {}
    for row in rows:
        group_key = int(row["group_id"])
        department_key = int(row["unit_id"])
        grouped.setdefault(group_key, {}).setdefault(department_key, []).append(dict(row))

    report_groups: list[dict[str, Any]] = []
    summary: list[dict[str, Any]] = []
    flat_items: list[dict[str, Any]] = []
    summary_number = 1
    total_rate = Decimal("0")
    missing_rate_count = 0
    for group_key in sorted(grouped):
        report_departments: list[dict[str, Any]] = []
        department_rows = grouped[group_key]
        for department_key in sorted(
            department_rows,
            key=lambda value: (department_names[value].casefold(), department_names[value], value),
        ):
            items: list[dict[str, Any]] = []
            sorted_department_rows = sorted(
                department_rows[department_key],
                key=lambda row: (
                    personnel_position_rank(
                        group_id=group_key,
                        position_name=row.get("position_name"),
                        position_category=row.get("position_category"),
                    ),
                    str(row.get("full_name") or NOT_SPECIFIED).casefold(),
                    str(row.get("full_name") or NOT_SPECIFIED),
                    int(row["employee_id"]),
                ),
            )
            department_rate_total = Decimal("0")
            for number, row in enumerate(sorted_department_rows, start=1):
                rate_decimal = _decimal_rate(row.get("rate"))
                item = {
                    "employee_id": int(row["employee_id"]),
                    "number": number,
                    "full_name": str(row.get("full_name") or NOT_SPECIFIED),
                    "position": str(row.get("position_name") or NOT_SPECIFIED),
                    "rate": _format_rate(rate_decimal) if rate_decimal is not None else NOT_SPECIFIED,
                    "rate_value": float(rate_decimal) if rate_decimal is not None else None,
                }
                if rate_decimal is not None:
                    department_rate_total += rate_decimal
                    total_rate += rate_decimal
                else:
                    missing_rate_count += 1
                items.append(item)
                flat_items.append(item)
            department_name = department_names[department_key]
            report_departments.append(
                {"id": department_key, "name": department_name, "items": items}
            )
            summary.append(
                {
                    "number": summary_number,
                    "group": {"id": group_key, "name": group_names[group_key]},
                    "department": {"id": department_key, "name": department_name},
                    "employee_count": len(items),
                    "rate_total": float(department_rate_total),
                }
            )
            summary_number += 1
        report_groups.append(
            {
                "id": group_key,
                "name": group_names[group_key],
                "departments": report_departments,
            }
        )

    return {
        "report_code": "personnel_roster",
        "report_name": "Личный состав",
        "generated_at": formed_at.isoformat(),
        "filters": {
            "group": (
                {"id": int(selected_group["group_id"]), "name": str(selected_group["group_name"])}
                if selected_group
                else None
            ),
            "department": (
                {"id": int(selected_department["unit_id"]), "name": str(selected_department["unit_name"])}
                if selected_department
                else None
            ),
        },
        "summary": summary,
        "total": len({item["employee_id"] for item in flat_items}),
        "total_rate": float(total_rate),
        "missing_rate_count": missing_rate_count,
        "groups": report_groups,
        "items": flat_items,
    }


def _decimal_rate(value: Any) -> Decimal | None:
    if value is None:
        return None
    try:
        return Decimal(str(value))
    except (InvalidOperation, ValueError):
        return None


def _format_rate(value: Decimal) -> str:
    return format(value.normalize(), "f").replace(".", ",")


def _write_table_header(ws: Any, row: int, headers: Iterable[str]) -> None:
    for column, value in enumerate(headers, start=1):
        cell = ws.cell(row, column, value)
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="2563EB")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)


def build_personnel_roster_xlsx(report: dict[str, Any]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Личный состав"
    ws.sheet_view.showGridLines = False

    ws.merge_cells("A1:E1")
    ws["A1"] = report["report_name"]
    ws["A1"].font = Font(bold=True, size=16)
    ws["A1"].alignment = Alignment(horizontal="center")
    ws["A3"] = "Группа отделений"
    ws["B3"] = report["filters"]["group"]["name"] if report["filters"]["group"] else ALL_GROUPS
    ws["A4"] = "Отделение"
    ws["B4"] = report["filters"]["department"]["name"] if report["filters"]["department"] else ALL_DEPARTMENTS
    ws["A5"] = "Дата и время формирования"
    ws["B5"] = datetime.fromisoformat(report["generated_at"]).astimezone().strftime("%d.%m.%Y %H:%M")

    row = 7
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "Сводный состав по отделениям").font = Font(bold=True, size=13)
    row += 1
    summary_header_row = row
    _write_table_header(
        ws,
        row,
        ("№", "Группа отделений", "Отделение", "Количество человек", "Количество ставок"),
    )
    for summary_item in report["summary"]:
        row += 1
        values = (
            summary_item["number"],
            summary_item["group"]["name"],
            summary_item["department"]["name"],
            summary_item["employee_count"],
            summary_item["rate_total"],
        )
        for column, value in enumerate(values, start=1):
            cell = ws.cell(row, column, value)
            if column == 5:
                cell.number_format = "0.##"
    row += 1
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=3)
    ws.cell(row, 1, "ВСЕГО").font = Font(bold=True)
    ws.cell(row, 4, report["total"]).font = Font(bold=True)
    ws.cell(row, 5, report["total_rate"]).font = Font(bold=True)
    ws.cell(row, 5).number_format = "0.##"

    if report["missing_rate_count"]:
        row += 1
        ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
        ws.cell(
            row,
            1,
            f"Ставка не указана у {report['missing_rate_count']} сотрудников",
        ).font = Font(italic=True, color="92400E")

    row += 2
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
    ws.cell(row, 1, "Личный состав").font = Font(bold=True, size=13)

    first_department = True
    for group in report["groups"]:
        for department_index, department in enumerate(group["departments"]):
            row += 1
            if not first_department:
                ws.row_breaks.append(Break(id=row - 1))
            first_department = False
            if department_index == 0:
                ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
                group_cell = ws.cell(row, 1, group["name"])
                group_cell.font = Font(bold=True, size=12, color="FFFFFF")
                group_cell.fill = PatternFill("solid", fgColor="334155")
                row += 1
            ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=5)
            department_cell = ws.cell(row, 1, department["name"])
            department_cell.font = Font(bold=True)
            department_cell.fill = PatternFill("solid", fgColor="DBEAFE")
            row += 1
            _write_table_header(ws, row, ("№", "ФИО", "Должность", "Ставка"))
            for item in department["items"]:
                row += 1
                values = (
                    item["number"],
                    item["full_name"],
                    item["position"],
                    item["rate_value"] if item["rate_value"] is not None else NOT_SPECIFIED,
                )
                for column, value in enumerate(values, start=1):
                    cell = ws.cell(row, column, value)
                    if column == 4 and item["rate_value"] is not None:
                        cell.number_format = "0.##"
                    cell.alignment = Alignment(
                        horizontal="center" if column in (1, 4) else "left",
                        vertical="top",
                        wrap_text=True,
                    )

    widths = (7, 34, 34, 20, 18)
    for column, width in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(column)].width = width
    ws.freeze_panes = f"A{summary_header_row + 1}"
    ws.page_setup.orientation = "portrait"
    ws.page_setup.fitToWidth = 1
    ws.page_setup.fitToHeight = 0
    ws.sheet_properties.pageSetUpPr.fitToPage = True
    ws.print_area = f"A1:E{max(row, summary_header_row)}"

    stream = BytesIO()
    wb.save(stream)
    return stream.getvalue()


def roster_filename(report: dict[str, Any]) -> str:
    filters = report["filters"]
    if filters["department"]:
        scope_name = filters["department"]["name"]
    elif filters["group"]:
        scope_name = filters["group"]["name"]
    else:
        scope_name = "Все_доступные"
    safe_scope = re.sub(r"[^\w\-.]+", "_", str(scope_name), flags=re.UNICODE).strip("_")
    date_part = datetime.fromisoformat(report["generated_at"]).strftime("%Y-%m-%d")
    return f"Личный_состав_{safe_scope or 'отчёт'}_{date_part}.xlsx"
