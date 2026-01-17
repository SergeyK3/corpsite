# app/directory.py
from __future__ import annotations

import os
import csv
import io
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Tuple, Set

from openpyxl import load_workbook
from fastapi import APIRouter, Header, HTTPException, Query, Path, Request
from sqlalchemy import text

from app.db.engine import engine

router = APIRouter(prefix="/directory", tags=["directory"])


# ---------------------------
# Config
# ---------------------------
def _rbac_mode() -> str:
    # off | dept
    v = (os.getenv("DIRECTORY_RBAC_MODE") or "dept").strip().lower()
    return v if v in ("off", "dept") else "dept"


def _parse_int_set_env(name: str) -> Set[int]:
    raw = (os.getenv(name) or "").strip()
    if not raw:
        return set()
    out: Set[int] = set()
    for part in raw.split(","):
        s = part.strip()
        if not s:
            continue
        try:
            out.add(int(s))
        except ValueError:
            continue
    return out


def _privileged_user_ids() -> Set[int]:
    # Preferred: DIRECTORY_PRIVILEGED_USER_IDS
    # Backward-compat: DIRECTORY_PRIVILEGED_IDS
    s = set()
    s |= _parse_int_set_env("DIRECTORY_PRIVILEGED_USER_IDS")
    s |= _parse_int_set_env("DIRECTORY_PRIVILEGED_IDS")
    return s


def _privileged_role_ids() -> Set[int]:
    # Preferred: DIRECTORY_PRIVILEGED_ROLE_IDS
    return _parse_int_set_env("DIRECTORY_PRIVILEGED_ROLE_IDS")


def _as_http500(e: Exception) -> HTTPException:
    return HTTPException(
        status_code=500,
        detail=f"directory error: {type(e).__name__}: {str(e)}",
    )


# ---------------------------
# Requester context (RBAC)
# ---------------------------
def _require_user_id(x_user_id: Optional[str]) -> int:
    if not x_user_id:
        raise HTTPException(status_code=401, detail="Missing X-User-Id header.")
    s = str(x_user_id).strip()
    if not s:
        raise HTTPException(status_code=400, detail="Invalid X-User-Id header.")
    try:
        return int(s)
    except ValueError:
        raise HTTPException(status_code=400, detail="Invalid X-User-Id header.")


def _load_user_ctx(user_id: int) -> Dict[str, Any]:
    q = text(
        """
        SELECT user_id, role_id, unit_id, is_active
        FROM public.users
        WHERE user_id = :uid
        LIMIT 1
        """
    )
    with engine.begin() as conn:
        row = conn.execute(q, {"uid": user_id}).mappings().first()

    if not row:
        raise HTTPException(status_code=401, detail="Unknown user.")
    if not bool(row.get("is_active")):
        raise HTTPException(status_code=403, detail="User is inactive.")
    return dict(row)


def _is_privileged(user_ctx: Dict[str, Any]) -> bool:
    uid = int(user_ctx["user_id"])
    rid = int(user_ctx["role_id"]) if user_ctx.get("role_id") is not None else -1
    if uid in _privileged_user_ids():
        return True
    if rid in _privileged_role_ids():
        return True
    return False


def _require_dept_scope(user_ctx: Dict[str, Any]) -> int:
    """
    For RBAC_MODE=dept: non-privileged users must have unit_id.
    unit_id is treated as department scope.
    """
    unit_id = user_ctx.get("unit_id")
    if unit_id is None:
        raise HTTPException(
            status_code=403,
            detail="directory: cannot determine department scope for user (unit_id is null).",
        )
    try:
        return int(unit_id)
    except Exception:
        raise HTTPException(
            status_code=403,
            detail="directory: invalid unit_id for department scope.",
        )


# ---------------------------
# Introspection helpers
# ---------------------------
def _list_relations(schema: str = "public") -> List[Tuple[str, str]]:
    q = text(
        """
        SELECT table_name AS name, 'table' AS kind
        FROM information_schema.tables
        WHERE table_schema = :schema AND table_type = 'BASE TABLE'
        UNION ALL
        SELECT table_name AS name, 'view' AS kind
        FROM information_schema.views
        WHERE table_schema = :schema
        ORDER BY name
        """
    )
    with engine.begin() as conn:
        rows = conn.execute(q, {"schema": schema}).fetchall()
        return [(r[0], r[1]) for r in rows]


def _get_columns(rel: str, schema: str = "public") -> List[str]:
    q = text(
        """
        SELECT column_name
        FROM information_schema.columns
        WHERE table_schema = :schema AND table_name = :rel
        ORDER BY ordinal_position
        """
    )
    with engine.begin() as conn:
        rows = conn.execute(q, {"schema": schema, "rel": rel}).fetchall()
        return [r[0] for r in rows]


def _pick_first(existing: List[str], candidates: List[str]) -> Optional[str]:
    s = set(existing)
    for c in candidates:
        if c in s:
            return c
    return None


# ---------------------------
# Relation scoring (auto-detect)
# ---------------------------
def _score_employees(name: str, cols: List[str]) -> int:
    lc = {c.lower() for c in cols}
    n = name.lower()
    score = 0

    if "employee" in n or "employees" in n:
        score += 25
    if "person" in n or "people" in n:
        score += 10
    if "staff" in n or "workers" in n:
        score += 10
    if "dir" in n or "directory" in n:
        score += 8

    if lc.intersection({"employee_id", "tab_no", "tab_number", "personnel_no", "tn", "tabel", "id"}):
        score += 25

    if lc.intersection({"fio", "full_name", "name_ru", "employee_name"}):
        score += 20
    if lc.intersection({"last_name", "surname"}):
        score += 10

    if lc.intersection({"department_id", "dept_id", "org_unit_id", "department"}):
        score += 10
    if lc.intersection({"position_id", "pos_id", "position"}):
        score += 6

    return score


def _score_departments(name: str, cols: List[str]) -> int:
    lc = {c.lower() for c in cols}
    n = name.lower()
    score = 0

    if "department" in n or "departments" in n:
        score += 25
    if "dept" in n:
        score += 18
    if "org_unit" in n or "orgunit" in n or "unit" in n:
        score += 10

    if lc.intersection({"id", "department_id", "dept_id", "org_unit_id"}):
        score += 20
    if lc.intersection({"name", "name_ru", "dept_name", "department_name", "org_unit_name"}):
        score += 20

    return score


def _score_positions(name: str, cols: List[str]) -> int:
    lc = {c.lower() for c in cols}
    n = name.lower()
    score = 0

    if "position" in n or "positions" in n:
        score += 25
    if "post" in n or "job" in n:
        score += 10

    if lc.intersection({"id", "position_id", "pos_id"}):
        score += 20
    if lc.intersection({"name", "name_ru", "pos_name", "position_name"}):
        score += 20

    return score


def _best_relation(score_fn, min_score: int) -> Tuple[Optional[str], List[str], int]:
    rels = _list_relations("public")
    best_name: Optional[str] = None
    best_cols: List[str] = []
    best_score = -1

    for name, _kind in rels:
        cols = _get_columns(name, "public")
        if not cols:
            continue
        score = score_fn(name, cols)
        if score > best_score:
            best_name, best_cols, best_score = name, cols, score

    if best_name is None or best_score < min_score:
        return None, [], best_score

    return best_name, best_cols, best_score


def _employees_relation() -> Tuple[str, List[str]]:
    rel, cols, score = _best_relation(_score_employees, min_score=20)
    if not rel:
        raise HTTPException(status_code=500, detail=f"Cannot auto-detect employees relation (score={score}).")
    return rel, cols


def _departments_relation() -> Tuple[Optional[str], List[str]]:
    rel, cols, _ = _best_relation(_score_departments, min_score=30)
    return rel, cols


def _positions_relation() -> Tuple[Optional[str], List[str]]:
    rel, cols, _ = _best_relation(_score_positions, min_score=30)
    return rel, cols


# ---------------------------
# Column mapping
# ---------------------------
def _employees_id_col(cols: List[str]) -> str:
    c = _pick_first(cols, ["employee_id", "id", "tab_no", "tab_number", "personnel_no", "tn", "tabel"])
    if not c:
        raise HTTPException(status_code=500, detail="Employees relation has no recognizable id column.")
    return c


def _dept_fk_col(cols: List[str]) -> Optional[str]:
    return _pick_first(cols, ["department_id", "dept_id", "org_unit_id"])


def _pos_fk_col(cols: List[str]) -> Optional[str]:
    return _pick_first(cols, ["position_id", "pos_id"])


def _dict_id_col(cols: List[str], preferred: List[str]) -> Optional[str]:
    return _pick_first(cols, preferred + ["id"])


def _dict_name_col(cols: List[str], preferred: List[str]) -> Optional[str]:
    return _pick_first(cols, preferred + ["name", "name_ru"])


def _normalize_employee_id_text(v: Any) -> str:
    # IMPORTANT: preserve leading zeros
    if v is None:
        return ""
    return str(v).strip()


# ---------------------------
# Dictionary endpoints
# ---------------------------
@router.get("/departments")
def list_departments(
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    x_user_id: Optional[str] = Header(default=None, alias="X-User-Id"),
) -> Dict[str, Any]:
    try:
        uid = _require_user_id(x_user_id)
        user_ctx = _load_user_ctx(uid)
        privileged = _is_privileged(user_ctx)

        rel, cols = _departments_relation()
        if not rel:
            return {"items": [], "total": 0}

        id_col = _dict_id_col(cols, ["department_id", "dept_id", "org_unit_id"])
        name_col = _dict_name_col(cols, ["department_name", "dept_name", "org_unit_name"])

        if not id_col or not name_col:
            return {"items": [], "total": 0}

        where = "TRUE"
        params: Dict[str, Any] = {"limit": limit, "offset": offset}

        if _rbac_mode() == "dept" and not privileged:
            dept_id = _require_dept_scope(user_ctx)
            where = f"CAST({id_col} AS TEXT) = :dept_id_text"
            params["dept_id_text"] = str(dept_id)

        q_total = text(f"SELECT COUNT(*) AS cnt FROM public.{rel} WHERE {where}")
        q_list = text(
            f"""
            SELECT {id_col} AS id, {name_col} AS name
            FROM public.{rel}
            WHERE {where}
            ORDER BY CAST({id_col} AS TEXT) ASC
            LIMIT :limit OFFSET :offset
            """
        )

        with engine.begin() as conn:
            total = int(conn.execute(q_total, params).mappings().first()["cnt"])
            rows = conn.execute(q_list, params).mappings().all()

        items = [{"id": r["id"], "name": r["name"]} for r in rows]
        return {"items": items, "total": total}

    except HTTPException:
        raise
    except Exception as e:
        raise _as_http500(e)


@router.get("/positions")
def list_positions(
    limit: int = Query(default=200, ge=1, le=1000),
    offset: int = Query(default=0, ge=0),
    x_user_id: Optional[str] = Header(default=None, alias="X-User-Id"),
) -> Dict[str, Any]:
    try:
        # positions usually not sensitive; keep open under dept-rbac too
        _ = _require_user_id(x_user_id)

        rel, cols = _positions_relation()
        if not rel:
            return {"items": [], "total": 0}

        id_col = _dict_id_col(cols, ["position_id", "pos_id"])
        name_col = _dict_name_col(cols, ["position_name", "pos_name"])

        if not id_col or not name_col:
            return {"items": [], "total": 0}

        q_total = text(f"SELECT COUNT(*) AS cnt FROM public.{rel}")
        q_list = text(
            f"""
            SELECT {id_col} AS id, {name_col} AS name
            FROM public.{rel}
            ORDER BY CAST({id_col} AS TEXT) ASC
            LIMIT :limit OFFSET :offset
            """
        )

        with engine.begin() as conn:
            total = int(conn.execute(q_total).mappings().first()["cnt"])
            rows = conn.execute(q_list, {"limit": limit, "offset": offset}).mappings().all()

        items = [{"id": r["id"], "name": r["name"]} for r in rows]
        return {"items": items, "total": total}

    except HTTPException:
        raise
    except Exception as e:
        raise _as_http500(e)


# ---------------------------
# Employees: list + card (with JOIN-enrichment)
# ---------------------------
def _employee_select_sql(emp_rel: str, emp_cols: List[str]) -> Tuple[str, Dict[str, str]]:
    emp_id_col = _employees_id_col(emp_cols)

    fio_col = _pick_first(emp_cols, ["fio", "full_name", "name", "name_ru", "employee_name"])
    last_col = _pick_first(emp_cols, ["last_name", "surname"])
    first_col = _pick_first(emp_cols, ["first_name", "name_first", "given_name"])
    mid_col = _pick_first(emp_cols, ["middle_name", "patronymic"])

    status_col = _pick_first(emp_cols, ["status", "is_active", "active"])
    date_from_col = _pick_first(emp_cols, ["date_from", "start_date", "employment_date", "hired_at"])
    date_to_col = _pick_first(emp_cols, ["date_to", "end_date", "terminated_at", "fired_at"])
    rate_col = _pick_first(emp_cols, ["rate", "stavka", "fte", "workload", "employment_rate"])

    dept_fk = _dept_fk_col(emp_cols)
    pos_fk = _pos_fk_col(emp_cols)

    dep_rel, dep_cols = _departments_relation()
    pos_rel, pos_cols = _positions_relation()

    dep_id_col = _dict_id_col(dep_cols, ["department_id", "dept_id", "org_unit_id"]) if dep_rel else None
    dep_name_col = _dict_name_col(dep_cols, ["department_name", "dept_name", "org_unit_name"]) if dep_rel else None

    pos_id_col = _dict_id_col(pos_cols, ["position_id", "pos_id"]) if pos_rel else None
    pos_name_col = _dict_name_col(pos_cols, ["position_name", "pos_name"]) if pos_rel else None

    select_parts = [
        f"e.{emp_id_col} AS e_id",
        (f"e.{fio_col} AS e_fio" if fio_col else "NULL AS e_fio"),
        (f"e.{last_col} AS e_last" if last_col else "NULL AS e_last"),
        (f"e.{first_col} AS e_first" if first_col else "NULL AS e_first"),
        (f"e.{mid_col} AS e_mid" if mid_col else "NULL AS e_mid"),
        (f"e.{rate_col} AS e_rate" if rate_col else "NULL AS e_rate"),
        (f"e.{status_col} AS e_status" if status_col else "NULL AS e_status"),
        (f"e.{date_from_col} AS e_date_from" if date_from_col else "NULL AS e_date_from"),
        (f"e.{date_to_col} AS e_date_to" if date_to_col else "NULL AS e_date_to"),
        (f"e.{dept_fk} AS e_dept_id" if dept_fk else "NULL AS e_dept_id"),
        (f"e.{pos_fk} AS e_pos_id" if pos_fk else "NULL AS e_pos_id"),
    ]

    join_sql = ""

    if dep_rel and dept_fk and dep_id_col and dep_name_col:
        select_parts.append(f"d.{dep_name_col} AS dept_name")
        join_sql += (
            f" LEFT JOIN public.{dep_rel} d"
            f" ON CAST(d.{dep_id_col} AS TEXT) = CAST(e.{dept_fk} AS TEXT)"
        )
    else:
        select_parts.append("NULL AS dept_name")

    if pos_rel and pos_fk and pos_id_col and pos_name_col:
        select_parts.append(f"p.{pos_name_col} AS pos_name")
        join_sql += (
            f" LEFT JOIN public.{pos_rel} p"
            f" ON CAST(p.{pos_id_col} AS TEXT) = CAST(e.{pos_fk} AS TEXT)"
        )
    else:
        select_parts.append("NULL AS pos_name")

    sql = "SELECT " + ", ".join(select_parts) + f" FROM public.{emp_rel} e" + join_sql
    return sql, {"emp_id_col": emp_id_col}


def _normalize_employee_joined(row: Dict[str, Any], emp_rel: str) -> Dict[str, Any]:
    fio = (str(row.get("e_fio")).strip() if row.get("e_fio") else "").strip()
    if not fio:
        parts = []
        for k in ("e_last", "e_first", "e_mid"):
            if row.get(k):
                parts.append(str(row.get(k)).strip())
        fio = " ".join([p for p in parts if p]) or None

    status_norm = "unknown"
    v = row.get("e_status")
    if isinstance(v, bool):
        status_norm = "active" if v else "inactive"
    elif v is None:
        status_norm = "unknown"
    else:
        s = str(v).strip().lower()
        if s in ("1", "true", "yes", "y", "on", "active", "работает", "в штате"):
            status_norm = "active"
        elif s in ("0", "false", "no", "n", "off", "inactive", "уволен", "не работает"):
            status_norm = "inactive"
        else:
            status_norm = s

    return {
        "id": str(row.get("e_id")) if row.get("e_id") is not None else None,
        "fio": fio,
        "department": {"id": row.get("e_dept_id"), "name": row.get("dept_name")},
        "position": {"id": row.get("e_pos_id"), "name": row.get("pos_name")},
        "rate": row.get("e_rate"),
        "status": status_norm,
        "date_from": row.get("e_date_from"),
        "date_to": row.get("e_date_to"),
        "source": {"relation": emp_rel},
    }


@router.get("/employees")
def list_employees(
    x_user_id: Optional[str] = Header(default=None, alias="X-User-Id"),
    status: str = Query(default="active", pattern="^(active|inactive|all)$"),
    q: Optional[str] = Query(default=None),
    department_id: Optional[int] = Query(default=None, ge=1),
    position_id: Optional[int] = Query(default=None, ge=1),
    limit: int = Query(default=50, ge=1, le=200),
    offset: int = Query(default=0, ge=0),
    sort: Optional[str] = Query(default=None),
    order: Optional[str] = Query(default=None),
) -> Dict[str, Any]:
    try:
        uid = _require_user_id(x_user_id)
        user_ctx = _load_user_ctx(uid)
        privileged = _is_privileged(user_ctx)

        emp_rel, emp_cols = _employees_relation()
        emp_id_col = _employees_id_col(emp_cols)

        base_sql, _ = _employee_select_sql(emp_rel, emp_cols)

        status_col = _pick_first(emp_cols, ["status", "is_active", "active"])
        fio_col = _pick_first(emp_cols, ["fio", "full_name", "name", "name_ru", "employee_name"])
        last_col = _pick_first(emp_cols, ["last_name", "surname"])
        first_col = _pick_first(emp_cols, ["first_name", "name_first", "given_name"])
        mid_col = _pick_first(emp_cols, ["middle_name", "patronymic"])

        dept_fk = _dept_fk_col(emp_cols)
        pos_fk = _pos_fk_col(emp_cols)

        where: List[str] = []
        params: Dict[str, Any] = {"limit": limit, "offset": offset}

        # Status
        if status != "all" and status_col:
            if status_col in ("is_active", "active"):
                where.append(f"e.{status_col} = :is_active")
                params["is_active"] = True if status == "active" else False
            else:
                if status == "active":
                    where.append(
                        f"LOWER(CAST(e.{status_col} AS TEXT)) IN ('active','работает','в штате','1','true','yes')"
                    )
                else:
                    where.append(
                        f"LOWER(CAST(e.{status_col} AS TEXT)) IN ('inactive','уволен','не работает','0','false','no')"
                    )

        # Search
        if q:
            qq = q.strip().lower()
            params["q"] = f"%{qq}%"
            parts = []
            if fio_col:
                parts.append(f"LOWER(CAST(e.{fio_col} AS TEXT)) LIKE :q")
            if last_col:
                parts.append(f"LOWER(CAST(e.{last_col} AS TEXT)) LIKE :q")
            if first_col:
                parts.append(f"LOWER(CAST(e.{first_col} AS TEXT)) LIKE :q")
            if mid_col:
                parts.append(f"LOWER(CAST(e.{mid_col} AS TEXT)) LIKE :q")
            if parts:
                where.append("(" + " OR ".join(parts) + ")")

        # RBAC: dept scope
        if _rbac_mode() == "dept" and not privileged:
            dept_id = _require_dept_scope(user_ctx)
            if dept_fk:
                where.append(f"CAST(e.{dept_fk} AS TEXT) = :rbac_dept_id_text")
                params["rbac_dept_id_text"] = str(dept_id)
            else:
                # Employees table without department FK -> cannot enforce dept RBAC correctly
                raise HTTPException(status_code=500, detail="directory: employees has no department_id column.")

        # Explicit filter by dept/pos
        # IMPORTANT: under RBAC=dept for non-privileged, department_id filter must not allow escaping scope.
        if department_id is not None and dept_fk:
            if _rbac_mode() == "dept" and not privileged:
                # Only allow same dept as scope
                dept_id = _require_dept_scope(user_ctx)
                if int(department_id) != int(dept_id):
                    # Return empty list (safer than 403 for list)
                    return {"items": [], "total": 0}
            where.append(f"CAST(e.{dept_fk} AS TEXT) = :department_id_text")
            params["department_id_text"] = str(int(department_id))

        if position_id is not None and pos_fk:
            where.append(f"CAST(e.{pos_fk} AS TEXT) = :position_id_text")
            params["position_id_text"] = str(int(position_id))

        where_sql = " AND ".join(where) if where else "TRUE"

        ord_dir = "ASC" if (order or "").lower() != "desc" else "DESC"
        if (sort or "").lower() in ("full_name", "fio", "name") and fio_col:
            order_sql = f"LOWER(CAST(e.{fio_col} AS TEXT)) {ord_dir}, CAST(e.{emp_id_col} AS TEXT) ASC"
        elif (sort or "").lower() in ("full_name", "fio", "name") and last_col:
            if first_col:
                order_sql = f"LOWER(CAST(e.{last_col} AS TEXT)) {ord_dir}, LOWER(CAST(e.{first_col} AS TEXT)) {ord_dir}"
            else:
                order_sql = f"LOWER(CAST(e.{last_col} AS TEXT)) {ord_dir}"
        else:
            order_sql = f"CAST(e.{emp_id_col} AS TEXT) ASC"

        q_list = text(
            f"""
            {base_sql}
            WHERE {where_sql}
            ORDER BY {order_sql}
            LIMIT :limit OFFSET :offset
            """
        )

        q_total = text(
            f"""
            SELECT COUNT(*) AS cnt
            FROM public.{emp_rel} e
            WHERE {where_sql}
            """
        )

        with engine.begin() as conn:
            total = int(conn.execute(q_total, params).mappings().first()["cnt"])
            rows = conn.execute(q_list, params).mappings().all()

        items = [_normalize_employee_joined(dict(r), emp_rel) for r in rows]
        return {"items": items, "total": total}

    except HTTPException:
        raise
    except Exception as e:
        raise _as_http500(e)


@router.get("/employees/{employee_id}")
def get_employee(
    employee_id: str = Path(..., min_length=1),
    x_user_id: Optional[str] = Header(default=None, alias="X-User-Id"),
) -> Dict[str, Any]:
    """
    IMPORTANT: employee_id is TEXT. Preserve leading zeros.
    """
    try:
        uid = _require_user_id(x_user_id)
        user_ctx = _load_user_ctx(uid)
        privileged = _is_privileged(user_ctx)

        target_id_text = _normalize_employee_id_text(employee_id)
        if not target_id_text:
            raise HTTPException(status_code=404, detail="Employee not found.")

        emp_rel, emp_cols = _employees_relation()
        emp_id_col = _employees_id_col(emp_cols)
        dept_fk = _dept_fk_col(emp_cols)

        base_sql, _ = _employee_select_sql(emp_rel, emp_cols)

        where = [f"CAST(e.{emp_id_col} AS TEXT) = :id_text"]
        params: Dict[str, Any] = {"id_text": target_id_text}

        if _rbac_mode() == "dept" and not privileged:
            dept_id = _require_dept_scope(user_ctx)
            if not dept_fk:
                raise HTTPException(status_code=500, detail="directory: employees has no department_id column.")
            where.append(f"CAST(e.{dept_fk} AS TEXT) = :rbac_dept_id_text")
            params["rbac_dept_id_text"] = str(dept_id)

        q_one = text(
            f"""
            {base_sql}
            WHERE {' AND '.join(where)}
            LIMIT 1
            """
        )

        with engine.begin() as conn:
            row = conn.execute(q_one, params).mappings().first()

        if not row:
            # 404 is intentional (does not leak existence outside scope)
            raise HTTPException(status_code=404, detail="Employee not found.")

        return _normalize_employee_joined(dict(row), emp_rel)

    except HTTPException:
        raise
    except Exception as e:
        raise _as_http500(e)


# ============================================================
# IMPORT helpers (use real PK/name columns for departments/positions)
# ============================================================
def _dept_table_meta() -> Tuple[str, str]:
    cols = _get_columns("departments", "public")
    if not cols:
        raise HTTPException(status_code=500, detail="departments table not found.")
    id_col = _pick_first(cols, ["department_id", "id"])
    name_col = _pick_first(cols, ["name", "name_ru", "department_name", "dept_name"])
    if not id_col or not name_col:
        raise HTTPException(status_code=500, detail="departments table has no recognizable id/name columns.")
    return id_col, name_col


def _pos_table_meta() -> Tuple[str, str]:
    cols = _get_columns("positions", "public")
    if not cols:
        raise HTTPException(status_code=500, detail="positions table not found.")
    id_col = _pick_first(cols, ["position_id", "id"])
    name_col = _pick_first(cols, ["name", "name_ru", "position_name", "pos_name"])
    if not id_col or not name_col:
        raise HTTPException(status_code=500, detail="positions table has no recognizable id/name columns.")
    return id_col, name_col


def _get_or_create_department_id(conn, name: str) -> int:
    name = name.strip()
    if not name:
        raise ValueError("department_name is empty")

    id_col, name_col = _dept_table_meta()

    row = conn.execute(
        text(f"SELECT {id_col} AS id FROM public.departments WHERE {name_col} = :name"),
        {"name": name},
    ).mappings().first()
    if row:
        return int(row["id"])

    row2 = conn.execute(
        text(
            f"""
            INSERT INTO public.departments ({name_col})
            VALUES (:name)
            ON CONFLICT ({name_col}) DO NOTHING
            RETURNING {id_col} AS id
            """
        ),
        {"name": name},
    ).mappings().first()

    if row2 and row2.get("id") is not None:
        return int(row2["id"])

    row3 = conn.execute(
        text(f"SELECT {id_col} AS id FROM public.departments WHERE {name_col} = :name"),
        {"name": name},
    ).mappings().first()
    if not row3:
        raise ValueError(f"cannot resolve department id for name={name}")
    return int(row3["id"])


def _get_or_create_position_id(conn, name: str) -> int:
    name = name.strip()
    if not name:
        raise ValueError("position_name is empty")

    id_col, name_col = _pos_table_meta()

    row = conn.execute(
        text(f"SELECT {id_col} AS id FROM public.positions WHERE {name_col} = :name"),
        {"name": name},
    ).mappings().first()
    if row:
        return int(row["id"])

    row2 = conn.execute(
        text(
            f"""
            INSERT INTO public.positions ({name_col})
            VALUES (:name)
            ON CONFLICT ({name_col}) DO NOTHING
            RETURNING {id_col} AS id
            """
        ),
        {"name": name},
    ).mappings().first()

    if row2 and row2.get("id") is not None:
        return int(row2["id"])

    row3 = conn.execute(
        text(f"SELECT {id_col} AS id FROM public.positions WHERE {name_col} = :name"),
        {"name": name},
    ).mappings().first()
    if not row3:
        raise ValueError(f"cannot resolve position id for name={name}")
    return int(row3["id"])


# ============================================================
# IMPORT: employees CSV (privileged only)
# ============================================================
def _decode_csv_bytes(b: bytes) -> str:
    if not b:
        return ""

    if b.startswith(b"\xef\xbb\xbf"):
        return b.decode("utf-8-sig")
    if b.startswith(b"\xff\xfe") or b.startswith(b"\xfe\xff"):
        return b.decode("utf-16")

    if b[:200].count(b"\x00") > 10:
        for enc in ("utf-16le", "utf-16be", "utf-16"):
            try:
                return b.decode(enc)
            except UnicodeDecodeError:
                pass

    for enc in ("utf-8", "utf-8-sig", "cp1251", "cp866", "koi8-r"):
        try:
            return b.decode(enc)
        except UnicodeDecodeError:
            continue

    return b.decode("utf-8", errors="replace")


def _sniff_delimiter(sample: str) -> str:
    semi = sample.count(";")
    comma = sample.count(",")
    return ";" if semi > comma else ","


def _norm_header(h: str) -> str:
    if h is None:
        return ""
    s = str(h).strip().lstrip("\ufeff")
    s = s.replace("\u00a0", " ")
    return s.strip().lower()


def _parse_bool(v: Optional[str]) -> Optional[bool]:
    if v is None:
        return None
    s = str(v).strip().lower()
    if s == "":
        return None
    if s in ("1", "true", "yes", "y", "on", "да"):
        return True
    if s in ("0", "false", "no", "n", "off", "нет"):
        return False
    return None


def _parse_rate(v: Optional[str]) -> Optional[float]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None
    s = s.replace(",", ".")
    try:
        return float(s)
    except ValueError:
        return None


def _parse_date_any(v: Optional[str]) -> Optional[date]:
    if v is None:
        return None
    s = str(v).strip()
    if not s:
        return None

    try:
        return datetime.strptime(s, "%Y-%m-%d").date()
    except ValueError:
        pass
    try:
        return datetime.strptime(s, "%d.%m.%Y").date()
    except ValueError:
        pass

    return None


@router.post("/import/employees_csv")
async def import_employees_csv(
    request: Request,
    x_user_id: Optional[str] = Header(default=None, alias="X-User-Id"),
) -> Dict[str, Any]:
    uid = _require_user_id(x_user_id)
    user_ctx = _load_user_ctx(uid)
    if not _is_privileged(user_ctx):
        raise HTTPException(status_code=403, detail="Forbidden.")

    try:
        raw = await request.body()
        text_csv = _decode_csv_bytes(raw)

        if not text_csv.strip():
            raise HTTPException(status_code=400, detail="Empty CSV body.")

        delim = _sniff_delimiter(text_csv[:4096])
        f = io.StringIO(text_csv)
        reader = csv.DictReader(f, delimiter=delim)

        if not reader.fieldnames:
            raise HTTPException(status_code=400, detail="CSV has no header row.")

        field_map: Dict[str, str] = {}
        for h in reader.fieldnames:
            field_map[_norm_header(h)] = h

        def col(name: str) -> Optional[str]:
            return field_map.get(name)

        c_emp = col("employee_id")
        c_name = col("full_name")
        c_dept = col("department_name")
        c_pos = col("position_name")

        if not (c_emp and c_name and c_dept and c_pos):
            raise HTTPException(
                status_code=400,
                detail="CSV header must include: employee_id, full_name, department_name, position_name.",
            )

        c_from = col("date_from")
        c_to = col("date_to")
        c_rate = col("employment_rate")
        c_active = col("is_active")

        rows_seen = 0
        emp_upserted = 0

        with engine.begin() as conn:
            dep_cache: Dict[str, int] = {}
            pos_cache: Dict[str, int] = {}

            for r in reader:
                rows_seen += 1

                employee_id = str((r.get(c_emp) or "")).strip()
                full_name = str((r.get(c_name) or "")).strip()
                dept_name = str((r.get(c_dept) or "")).strip()
                pos_name = str((r.get(c_pos) or "")).strip()

                if not employee_id or not full_name or not dept_name or not pos_name:
                    continue

                if dept_name not in dep_cache:
                    dep_cache[dept_name] = _get_or_create_department_id(conn, dept_name)

                if pos_name not in pos_cache:
                    pos_cache[pos_name] = _get_or_create_position_id(conn, pos_name)

                department_id = dep_cache[dept_name]
                position_id = pos_cache[pos_name]

                date_from_v = _parse_date_any(r.get(c_from)) if c_from else None
                date_to_v = _parse_date_any(r.get(c_to)) if c_to else None
                rate_v = _parse_rate(r.get(c_rate)) if c_rate else None
                active_v = _parse_bool(r.get(c_active)) if c_active else None

                if rate_v is None:
                    rate_v = 1.00
                if active_v is None:
                    active_v = True
                if date_to_v is not None:
                    active_v = False

                conn.execute(
                    text(
                        """
                        INSERT INTO public.employees
                          (employee_id, full_name, department_id, position_id, date_from, date_to, employment_rate, is_active)
                        VALUES
                          (:employee_id, :full_name, :department_id, :position_id, :date_from, :date_to, :employment_rate, :is_active)
                        ON CONFLICT (employee_id) DO UPDATE SET
                          full_name = EXCLUDED.full_name,
                          department_id = EXCLUDED.department_id,
                          position_id = EXCLUDED.position_id,
                          date_from = EXCLUDED.date_from,
                          date_to = EXCLUDED.date_to,
                          employment_rate = EXCLUDED.employment_rate,
                          is_active = EXCLUDED.is_active
                        """
                    ),
                    {
                        "employee_id": employee_id,
                        "full_name": full_name,
                        "department_id": department_id,
                        "position_id": position_id,
                        "date_from": date_from_v,
                        "date_to": date_to_v,
                        "employment_rate": rate_v,
                        "is_active": active_v,
                    },
                )
                emp_upserted += 1

        return {
            "rows_seen": rows_seen,
            "departments_touched": len(dep_cache),
            "positions_touched": len(pos_cache),
            "employees_upserted": emp_upserted,
            "delimiter": delim,
            "encoding": "auto(utf-8/cp1251)",
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"import error: {type(e).__name__}: {str(e)}")


# ============================================================
# IMPORT: employees XLSX (recommended)
# ============================================================
def _to_text(v: Any) -> str:
    if v is None:
        return ""
    s = str(v)
    return s.replace("\u00a0", " ").strip()


def _norm_header_any(h: Any) -> str:
    s = _to_text(h).lstrip("\ufeff").lower()
    s = " ".join(s.split())
    return s


def _parse_date_cell(v: Any) -> Optional[date]:
    if v is None:
        return None
    if isinstance(v, datetime):
        return v.date()
    if isinstance(v, date):
        return v
    s = _to_text(v)
    if not s:
        return None
    return _parse_date_any(s)


def _parse_rate_cell(v: Any) -> Optional[float]:
    if v is None:
        return None
    if isinstance(v, (int, float)):
        try:
            return float(v)
        except Exception:
            return None
    return _parse_rate(_to_text(v))


def _parse_bool_cell(v: Any) -> Optional[bool]:
    if v is None:
        return None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        return bool(int(v))
    return _parse_bool(_to_text(v))


def _pick_header(field_map: Dict[str, int], names: List[str]) -> Optional[int]:
    for n in names:
        k = _norm_header_any(n)
        if k in field_map:
            return field_map[k]
    return None


@router.post("/import/employees_xlsx")
async def import_employees_xlsx(
    request: Request,
    x_user_id: Optional[str] = Header(default=None, alias="X-User-Id"),
) -> Dict[str, Any]:
    uid = _require_user_id(x_user_id)
    user_ctx = _load_user_ctx(uid)
    if not _is_privileged(user_ctx):
        raise HTTPException(status_code=403, detail="Forbidden.")

    try:
        raw = await request.body()
        if not raw:
            raise HTTPException(status_code=400, detail="Empty XLSX body.")

        wb = load_workbook(filename=io.BytesIO(raw), data_only=True)
        ws = wb.active

        header_row_idx = None
        header_vals: List[Any] = []
        for r in range(1, min(ws.max_row, 50) + 1):
            vals = [ws.cell(row=r, column=c).value for c in range(1, ws.max_column + 1)]
            if any(_to_text(v) for v in vals):
                header_row_idx = r
                header_vals = vals
                break

        if not header_row_idx:
            raise HTTPException(status_code=400, detail="XLSX has no header row.")

        field_map: Dict[str, int] = {}
        for idx, h in enumerate(header_vals, start=1):
            hn = _norm_header_any(h)
            if hn:
                field_map[hn] = idx

        c_emp = _pick_header(field_map, ["employee_id", "табельный номер", "таб. №", "таб номер", "тн", "id"])
        c_name = _pick_header(field_map, ["full_name", "фио", "ф.и.о.", "сотрудник", "фамилия имя отчество"])
        c_dept = _pick_header(field_map, ["department_name", "отдел", "подразделение", "отделение"])
        c_pos = _pick_header(field_map, ["position_name", "должность", "позиция"])

        if not (c_emp and c_name and c_dept and c_pos):
            raise HTTPException(
                status_code=400,
                detail="XLSX header must include: employee_id/full_name/department_name/position_name (or RU equivalents).",
            )

        c_from = _pick_header(field_map, ["date_from", "дата с", "дата_с", "начало", "date from"])
        c_to = _pick_header(field_map, ["date_to", "дата по", "дата_по", "окончание", "date to"])
        c_rate = _pick_header(field_map, ["employment_rate", "ставка", "rate", "fte"])
        c_active = _pick_header(field_map, ["is_active", "работает", "активен", "active"])

        emp_upserted = 0
        rows_seen = 0

        with engine.begin() as conn:
            dep_cache: Dict[str, int] = {}
            pos_cache: Dict[str, int] = {}

            for r in range(header_row_idx + 1, ws.max_row + 1):
                rows_seen += 1

                employee_id = _to_text(ws.cell(row=r, column=c_emp).value)
                full_name = _to_text(ws.cell(row=r, column=c_name).value)
                dept_name = _to_text(ws.cell(row=r, column=c_dept).value)
                pos_name = _to_text(ws.cell(row=r, column=c_pos).value)

                if not employee_id or not full_name or not dept_name or not pos_name:
                    continue

                if dept_name not in dep_cache:
                    dep_cache[dept_name] = _get_or_create_department_id(conn, dept_name)

                if pos_name not in pos_cache:
                    pos_cache[pos_name] = _get_or_create_position_id(conn, pos_name)

                department_id = dep_cache[dept_name]
                position_id = pos_cache[pos_name]

                date_from_v = _parse_date_cell(ws.cell(row=r, column=c_from).value) if c_from else None
                date_to_v = _parse_date_cell(ws.cell(row=r, column=c_to).value) if c_to else None
                rate_v = _parse_rate_cell(ws.cell(row=r, column=c_rate).value) if c_rate else None
                active_v = _parse_bool_cell(ws.cell(row=r, column=c_active).value) if c_active else None

                if rate_v is None:
                    rate_v = 1.00
                if active_v is None:
                    active_v = True
                if date_to_v is not None:
                    active_v = False

                conn.execute(
                    text(
                        """
                        INSERT INTO public.employees
                          (employee_id, full_name, department_id, position_id, date_from, date_to, employment_rate, is_active)
                        VALUES
                          (:employee_id, :full_name, :department_id, :position_id, :date_from, :date_to, :employment_rate, :is_active)
                        ON CONFLICT (employee_id) DO UPDATE SET
                          full_name = EXCLUDED.full_name,
                          department_id = EXCLUDED.department_id,
                          position_id = EXCLUDED.position_id,
                          date_from = EXCLUDED.date_from,
                          date_to = EXCLUDED.date_to,
                          employment_rate = EXCLUDED.employment_rate,
                          is_active = EXCLUDED.is_active
                        """
                    ),
                    {
                        "employee_id": employee_id,
                        "full_name": full_name,
                        "department_id": department_id,
                        "position_id": position_id,
                        "date_from": date_from_v,
                        "date_to": date_to_v,
                        "employment_rate": rate_v,
                        "is_active": active_v,
                    },
                )
                emp_upserted += 1

        return {
            "rows_seen": rows_seen,
            "departments_touched": len(set(dep_cache.keys())),
            "positions_touched": len(set(pos_cache.keys())),
            "employees_upserted": emp_upserted,
            "sheet": ws.title,
            "header_row": header_row_idx,
        }

    except HTTPException:
        raise
    except Exception as e:
        raise HTTPException(status_code=500, detail=f"import xlsx error: {type(e).__name__}: {str(e)}")
