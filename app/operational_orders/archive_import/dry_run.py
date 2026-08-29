"""Read-only validation of an Operational Orders XLSX archive manifest."""
from __future__ import annotations

import hashlib
import os
import unicodedata
from collections import Counter, defaultdict
from datetime import date, datetime
from pathlib import Path, PurePosixPath, PureWindowsPath
from typing import Any, Iterable

from openpyxl import load_workbook

from app.operational_orders.archive_import.models import (
    DryRunReport,
    DryRunSummary,
    SourceOrderRow,
    ValidationIssue,
    ValidationRowResult,
)


DEFAULT_SHEET_NAME = "Производственные приказы"
ROOT_ARCHIVE_SECTION = "Корень архива"
ALLOWED_EXTENSIONS = frozenset({".doc", ".docx", ".pdf"})
HASH_CHUNK_SIZE = 1024 * 1024

REQUIRED_COLUMNS = (
    "№ п/п",
    "Имя файла (Word/PDF)",
    "Тип документа",
    "Статус",
    "Тип события / предмет приказа",
    "Номер приказа",
    "Дата приказа",
    "Примечание",
    "Исходная папка",
    "Раздел архива",
    "Относительный путь к файлу",
)


class DryRunTechnicalError(RuntimeError):
    """The manifest could not be inspected safely."""

    def __init__(self, code: str, message: str):
        super().__init__(message)
        self.code = code


def normalize_header(value: Any) -> str:
    text = _to_text(value).lstrip("\ufeff")
    return " ".join(unicodedata.normalize("NFC", text).casefold().split())


def run_dry_run(
    *,
    xlsx_path: str | Path,
    archive_root: str | Path,
    sheet_name: str = DEFAULT_SHEET_NAME,
    expected_rows: int | None = None,
) -> DryRunReport:
    xlsx = Path(xlsx_path)
    root = Path(archive_root)
    if not xlsx.is_file():
        raise DryRunTechnicalError("XLSX_NOT_FOUND", f"XLSX file not found: {xlsx}")
    if not root.is_dir():
        raise DryRunTechnicalError("ARCHIVE_ROOT_NOT_FOUND", f"Archive root not found: {root}")

    try:
        workbook = load_workbook(xlsx, read_only=True, data_only=True)
    except Exception as exc:
        raise DryRunTechnicalError("XLSX_OPEN_FAILED", f"Could not open XLSX: {exc}") from exc
    try:
        if sheet_name not in workbook.sheetnames:
            raise DryRunTechnicalError("SHEET_NOT_FOUND", f"Sheet not found: {sheet_name}")
        worksheet = workbook[sheet_name]
        header_row, column_map, source_headers = _find_headers(worksheet)
        source_rows = _read_source_rows(worksheet, header_row, column_map)
    finally:
        workbook.close()

    global_errors: list[ValidationIssue] = []
    if expected_rows is not None and len(source_rows) != expected_rows:
        global_errors.append(
            ValidationIssue(
                code="EXPECTED_ROWS_MISMATCH",
                severity="ERROR",
                message=f"Expected {expected_rows} rows, found {len(source_rows)}.",
            )
        )

    duplicate_source_numbers = _duplicate_rows(source_rows, lambda row: _identity_key(row.source_number))
    duplicate_relative_paths = _duplicate_rows(source_rows, lambda row: _path_identity(row.relative_path))
    root_resolved = root.resolve(strict=True)
    results = tuple(
        _validate_row(
            row,
            root_resolved=root_resolved,
            duplicate_source_rows=duplicate_source_numbers.get(_identity_key(row.source_number), ()),
            duplicate_path_rows=duplicate_relative_paths.get(_path_identity(row.relative_path), ()),
        )
        for row in source_rows
    )

    sha_rows: dict[str, list[int]] = defaultdict(list)
    for result in results:
        if result.sha256:
            sha_rows[result.sha256].append(result.source.excel_row)
    duplicate_sha = {
        digest: tuple(rows)
        for digest, rows in sorted(sha_rows.items())
        if len(rows) > 1
    }
    warnings = tuple(
        ValidationIssue(
            code="DUPLICATE_SHA256",
            severity="WARNING",
            message=f"Identical SHA-256 for rows: {', '.join(map(str, rows))}.",
            related_rows=rows,
        )
        for rows in duplicate_sha.values()
    )
    row_errors = tuple(
        issue
        for result in results
        for issue in result.issues
        if issue.severity == "ERROR"
    )
    errors = tuple(global_errors) + row_errors
    summary = _build_summary(results, duplicate_sha)
    return DryRunReport(
        xlsx_path=str(xlsx.resolve()),
        archive_root=str(root_resolved),
        sheet_name=sheet_name,
        header_row=header_row,
        source_headers=source_headers,
        rows=results,
        summary=summary,
        warnings=warnings,
        errors=errors,
    )


def _find_headers(worksheet) -> tuple[int, dict[str, int], dict[str, str]]:
    required_by_key = {normalize_header(name): name for name in REQUIRED_COLUMNS}
    scan_limit = min(int(worksheet.max_row or 0), 50)
    for row_number in range(1, scan_limit + 1):
        normalized_to_columns: dict[str, list[int]] = defaultdict(list)
        originals: dict[str, str] = {}
        for column in range(1, int(worksheet.max_column or 0) + 1):
            value = worksheet.cell(row=row_number, column=column).value
            normalized = normalize_header(value)
            if normalized:
                normalized_to_columns[normalized].append(column)
                originals.setdefault(normalized, _to_text(value))
        if all(key in normalized_to_columns for key in required_by_key):
            duplicate_required = {
                required_by_key[key]: tuple(normalized_to_columns[key])
                for key in required_by_key
                if len(normalized_to_columns[key]) > 1
            }
            if duplicate_required:
                details = ", ".join(
                    f"{name}: columns {columns}"
                    for name, columns in duplicate_required.items()
                )
                raise DryRunTechnicalError(
                    "REQUIRED_COLUMN_DUPLICATE",
                    f"Required columns are duplicated in header row {row_number}: {details}",
                )
            column_map = {
                canonical: normalized_to_columns[key][0]
                for key, canonical in required_by_key.items()
            }
            source_headers = {
                canonical: originals[key]
                for key, canonical in required_by_key.items()
            }
            return row_number, column_map, source_headers

    found: set[str] = set()
    for row_number in range(1, scan_limit + 1):
        for column in range(1, int(worksheet.max_column or 0) + 1):
            found.add(normalize_header(worksheet.cell(row=row_number, column=column).value))
    missing = [name for name in REQUIRED_COLUMNS if normalize_header(name) not in found]
    raise DryRunTechnicalError(
        "REQUIRED_COLUMNS_MISSING",
        f"Required columns missing: {', '.join(missing)}",
    )


def _read_source_rows(worksheet, header_row: int, column_map: dict[str, int]) -> list[SourceOrderRow]:
    rows: list[SourceOrderRow] = []
    for excel_row in range(header_row + 1, int(worksheet.max_row or header_row) + 1):
        values = {
            name: worksheet.cell(row=excel_row, column=column).value
            for name, column in column_map.items()
        }
        if not any(_to_text(value) for value in values.values()):
            continue
        rows.append(
            SourceOrderRow(
                excel_row=excel_row,
                source_number=_to_text(values["№ п/п"]),
                file_name=_to_text(values["Имя файла (Word/PDF)"]),
                document_type=_to_text(values["Тип документа"]),
                source_status=_to_text(values["Статус"]),
                event_type=_to_text(values["Тип события / предмет приказа"]),
                order_number=_to_text(values["Номер приказа"]),
                order_date_raw=_date_to_text(values["Дата приказа"]),
                note=_to_text(values["Примечание"]),
                source_folder=_to_text(values["Исходная папка"]),
                archive_section=_to_text(values["Раздел архива"]),
                relative_path=_to_text(values["Относительный путь к файлу"]),
            )
        )
    return rows


def _validate_row(
    row: SourceOrderRow,
    *,
    root_resolved: Path,
    duplicate_source_rows: tuple[int, ...],
    duplicate_path_rows: tuple[int, ...],
) -> ValidationRowResult:
    issues: list[ValidationIssue] = []

    def error(code: str, message: str, field: str, related: tuple[int, ...] = ()) -> None:
        issues.append(
            ValidationIssue(
                code=code,
                severity="ERROR",
                message=message,
                excel_row=row.excel_row,
                field=field,
                related_rows=related,
            )
        )

    if not row.source_number:
        error("SOURCE_NUMBER_EMPTY", "Source row number is empty.", "№ п/п")
    elif len(duplicate_source_rows) > 1:
        error(
            "SOURCE_NUMBER_DUPLICATE",
            f"Source row number is duplicated in Excel rows {duplicate_source_rows}.",
            "№ п/п",
            duplicate_source_rows,
        )

    parsed_date, date_error = _parse_date(row.order_date_raw)
    if date_error:
        error("ORDER_DATE_INVALID", date_error, "Дата приказа")

    relative = row.relative_path
    extension = Path(relative.replace("\\", "/")).suffix.casefold() if relative else ""
    resolved_path: Path | None = None
    file_exists = False
    size: int | None = None
    digest: str | None = None

    if not relative:
        error("RELATIVE_PATH_EMPTY", "Relative path is empty.", "Относительный путь к файлу")
    else:
        windows_path = PureWindowsPath(relative)
        posix_path = PurePosixPath(relative)
        normalized_parts = tuple(part for part in relative.replace("\\", "/").split("/") if part not in {"", "."})
        absolute = windows_path.is_absolute() or bool(windows_path.drive) or posix_path.is_absolute()
        has_parent_escape = ".." in normalized_parts
        if absolute:
            error("PATH_ABSOLUTE", "Path must be relative.", "Относительный путь к файлу")
        if has_parent_escape:
            error("PATH_PARENT_ESCAPE", "Path contains '..'.", "Относительный путь к файлу")
        if len(duplicate_path_rows) > 1:
            error(
                "RELATIVE_PATH_DUPLICATE",
                f"Relative path is duplicated in Excel rows {duplicate_path_rows}.",
                "Относительный путь к файлу",
                duplicate_path_rows,
            )
        if extension not in ALLOWED_EXTENSIONS:
            error(
                "EXTENSION_NOT_ALLOWED",
                f"Extension is not allowed: {extension or '<empty>'}.",
                "Относительный путь к файлу",
            )

        leaf_name = normalized_parts[-1] if normalized_parts else ""
        if _identity_key(leaf_name) != _identity_key(row.file_name):
            error(
                "FILE_NAME_MISMATCH",
                f"Path filename {leaf_name!r} does not match manifest filename {row.file_name!r}.",
                "Имя файла (Word/PDF)",
            )

        expected_section = normalized_parts[0] if len(normalized_parts) > 1 else ROOT_ARCHIVE_SECTION
        if row.archive_section != expected_section:
            error(
                "ARCHIVE_SECTION_MISMATCH",
                f"Expected archive section {expected_section!r}, found {row.archive_section!r}.",
                "Раздел архива",
            )

        if not absolute and not has_parent_escape and normalized_parts:
            candidate = root_resolved.joinpath(*normalized_parts)
            try:
                resolved_path = candidate.resolve(strict=False)
                if not _is_within_root(resolved_path, root_resolved):
                    error(
                        "PATH_OUTSIDE_ARCHIVE_ROOT",
                        "Resolved path is outside archive root.",
                        "Относительный путь к файлу",
                    )
                    resolved_path = None
                elif not resolved_path.is_file():
                    error("FILE_NOT_FOUND", f"File not found: {relative}", "Относительный путь к файлу")
                else:
                    file_exists = True
                    if _identity_key(resolved_path.name) != _identity_key(row.file_name):
                        error(
                            "PHYSICAL_FILE_NAME_MISMATCH",
                            f"Physical filename {resolved_path.name!r} does not match manifest filename {row.file_name!r}.",
                            "Имя файла (Word/PDF)",
                        )
                    try:
                        size = resolved_path.stat().st_size
                    except OSError as exc:
                        error("FILE_SIZE_READ_FAILED", f"Could not read file size: {exc}", "Относительный путь к файлу")
                    try:
                        digest = _sha256_file(resolved_path)
                    except OSError as exc:
                        error("FILE_HASH_READ_FAILED", f"Could not calculate SHA-256: {exc}", "Относительный путь к файлу")
            except OSError as exc:
                error("PATH_RESOLVE_FAILED", f"Could not resolve path: {exc}", "Относительный путь к файлу")

    return ValidationRowResult(
        source=row,
        resolved_path=str(resolved_path) if resolved_path is not None else None,
        parsed_order_date=parsed_date,
        extension=extension,
        file_exists=file_exists,
        file_size_bytes=size,
        sha256=digest,
        issues=tuple(issues),
    )


def _build_summary(
    results: tuple[ValidationRowResult, ...],
    duplicate_sha: dict[str, tuple[int, ...]],
) -> DryRunSummary:
    source_rows = [result.source for result in results]
    extension_counts = Counter(result.extension or "<empty>" for result in results)
    status_counts = Counter(row.source_status or "<empty>" for row in source_rows)
    section_counts = Counter(row.archive_section or "<empty>" for row in source_rows)
    number_rows = _value_rows(source_rows, lambda row: row.order_number)
    date_rows = _value_rows(
        source_rows,
        lambda row: (_parse_date(row.order_date_raw)[0] or row.order_date_raw),
    )
    duplicate_numbers = {str(value): rows for value, rows in number_rows.items() if value and len(rows) > 1}
    duplicate_dates = {
        value.isoformat() if isinstance(value, date) else str(value): rows
        for value, rows in date_rows.items()
        if value and len(rows) > 1
    }
    valid_rows = sum(result.is_valid for result in results)
    return DryRunSummary(
        total_rows=len(results),
        valid_rows=valid_rows,
        error_rows=len(results) - valid_rows,
        existing_files=sum(result.file_exists for result in results),
        extension_counts=dict(sorted(extension_counts.items())),
        source_status_counts=dict(sorted(status_counts.items(), key=lambda item: item[0].casefold())),
        unique_archive_sections=len(section_counts),
        root_archive_files=section_counts.get(ROOT_ARCHIVE_SECTION, 0),
        archive_section_counts=dict(sorted(section_counts.items(), key=lambda item: item[0].casefold())),
        filled_order_numbers=sum(bool(row.order_number) for row in source_rows),
        empty_order_numbers=sum(not row.order_number for row in source_rows),
        filled_order_dates=sum(bool(row.order_date_raw) for row in source_rows),
        empty_order_dates=sum(not row.order_date_raw for row in source_rows),
        duplicate_order_numbers=dict(sorted(duplicate_numbers.items(), key=lambda item: item[0].casefold())),
        duplicate_order_dates=dict(sorted(duplicate_dates.items())),
        duplicate_sha256=duplicate_sha,
    )


def _to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    return str(value).replace("\u00a0", " ").strip()


def _date_to_text(value: Any) -> str:
    if value is None:
        return ""
    if isinstance(value, datetime):
        return value.date().isoformat()
    if isinstance(value, date):
        return value.isoformat()
    return _to_text(value)


def _parse_date(value: str) -> tuple[date | None, str | None]:
    text = str(value or "").strip()
    if not text:
        return None, None
    for pattern in ("%Y-%m-%d", "%d.%m.%Y", "%d/%m/%Y"):
        try:
            return datetime.strptime(text, pattern).date(), None
        except ValueError:
            continue
    return None, f"Unsupported order date: {text!r}."


def _identity_key(value: str) -> str:
    return unicodedata.normalize("NFC", str(value or "")).casefold()


def _path_identity(value: str) -> str:
    normalized = str(value or "").replace("/", "\\")
    return _identity_key(normalized)


def _duplicate_rows(
    rows: Iterable[SourceOrderRow],
    key_fn,
) -> dict[str, tuple[int, ...]]:
    grouped: dict[str, list[int]] = defaultdict(list)
    for row in rows:
        key = key_fn(row)
        if key:
            grouped[key].append(row.excel_row)
    return {key: tuple(items) for key, items in grouped.items() if len(items) > 1}


def _value_rows(rows: Iterable[SourceOrderRow], value_fn) -> dict[Any, tuple[int, ...]]:
    grouped: dict[Any, list[int]] = defaultdict(list)
    for row in rows:
        value = value_fn(row)
        if value:
            grouped[value].append(row.excel_row)
    return {key: tuple(items) for key, items in grouped.items()}


def _is_within_root(candidate: Path, root: Path) -> bool:
    try:
        common = os.path.commonpath((str(candidate), str(root)))
        return os.path.normcase(common) == os.path.normcase(str(root))
    except ValueError:
        return False


def _sha256_file(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as source:
        while chunk := source.read(HASH_CHUNK_SIZE):
            digest.update(chunk)
    return digest.hexdigest()
